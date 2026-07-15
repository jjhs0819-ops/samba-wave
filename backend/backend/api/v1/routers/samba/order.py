"""SambaWave Order API router."""

import asyncio
import re
from datetime import datetime
from typing import Any, Optional

from fastapi import APIRouter, Depends, File, Header, HTTPException, Query, UploadFile
from pydantic import BaseModel
from sqlmodel.ext.asyncio.session import AsyncSession
from sqlmodel import select
from backend.db.orm import (
    get_read_session,
    get_read_session_dependency,
    get_write_session_dependency,
)
from backend.domain.samba.cache import cache
from backend.domain.samba.tenant.middleware import get_optional_tenant_id
from backend.domain.samba.order.model import SambaOrder
from backend.domain.samba.order.playauto_alias import (
    normalize_playauto_alias_code,
    parse_playauto_alias_entry,
)
from backend.domain.samba.order.repository import SambaOrderRepository
from backend.domain.samba.order.service import SambaOrderService
from backend.dtos.samba.order import (
    FetchProductImageRequest,
    OrderCreate,
    OrderStatusUpdate,
    OrderUpdate,
)
from backend.utils.logger import logger

router = APIRouter(prefix="/orders", tags=["samba-orders"])
public_router = APIRouter(prefix="/orders", tags=["samba-orders-public"])

# ── 매칭 캐시(_mpn_cache) 모듈 전역 — 증분 갱신 ──
# 과거: 호출마다 등록상품 전체(~10만건, 1GB) 풀스캔 빌드 → 빌드(150초)>TTL 이라
# 캐시가 안 채워지고 무한 재스캔 → read 풀 고갈 사고.
# 현재: updated_at(ix_scp_updated_at_desc) 변경분만 증분 머지 + 주기적 전체 재빌드.
_MPN_CACHE_TTL_SEC = 180.0  # 증분 적용 최소 간격(초)
_MPN_FULL_REBUILD_SEC = 21600.0  # 전체 재빌드 주기(초, 6h) — 삭제·등록해제 staleness 정리. 증분이 신선도 담당하므로 드물게
# (by_global, by_account) 튜플 — by_account는 정확 매칭(account_id, product_no) 인덱스
_mpn_cache_data: tuple[dict[str, dict], dict[str, dict]] | None = None
_mpn_cache_built_at: float = 0.0  # 마지막 빌드/증분 monotonic
_mpn_cache_full_built_at: float = 0.0  # 마지막 전체빌드 monotonic
_mpn_cache_delta_since = None  # 증분 쿼리 기준 wall-clock(datetime, UTC)
_mpn_cache_lock = asyncio.Lock()


def _index_mpn_row(_row, by_global: dict, by_account: dict, sourcing_urls: dict) -> int:
    """수집상품 1행을 by_global / by_account 인덱스에 반영. ambiguous 신규 발생 수 반환.

    전체 빌드와 증분 머지 둘 다 이 함수를 재사용 — 인덱싱 규칙을 1곳에 모은다.
    """
    _cpid, _site, _spid, _thumb_raw, _mpnos, _src_url, _cat, _cost = _row
    if not (_mpnos and isinstance(_mpnos, dict)):
        return 0
    # cp 단가 원가(주문 cost 보강용, issue #365) — float 컬럼이라 TOAST 부담 없음
    _cp_cost = float(_cost) if _cost else 0.0
    # 썸네일은 (images->>0)로 첫 URL만 추출 — TOAST 전체 fetch 회피하면서 표시용 확보.
    _thumb = _thumb_raw or ""
    _olink = _src_url or (
        sourcing_urls.get(_site, "").format(_spid)
        if _site in sourcing_urls and _spid
        else ""
    )
    # account_id별 등록된 site_ids 모음 — `{account_id}_sites` 키 패턴.
    _sites_by_account: dict[str, list[str]] = {}
    for _k, _v in _mpnos.items():
        if _k.endswith("_sites") and isinstance(_v, list):
            _account_id = _k[: -len("_sites")]
            _sites_by_account[_account_id] = [str(s) for s in _v if s]

    _ambiguous_new = 0
    for _k, _v in _mpnos.items():
        if (
            not _v
            or _k.endswith("_qa")
            or _k.endswith("_sites")
            or _k.endswith("_master")
        ):
            continue
        # _origin 키도 인덱싱한다 — 스마트스토어 주문 product_id 에는
        # channelProductNo 대신 originProductNo 가 들어오는 케이스가 있어
        # 매칭 실패 → source_site/source_url 공란 저장 사고가 반복되어 추가.
        if _k.endswith("_origin"):
            _account_key = _k[: -len("_origin")]
        else:
            _account_key = str(_k)
        if isinstance(_v, dict):
            _values = [
                _v.get("smartstoreChannelProductNo"),
                _v.get("originProductNo"),
                _v.get("channelProductNo"),
            ]
        else:
            _values = [_v]
        for _sub_v in _values:
            if not _sub_v:
                continue
            _key = str(_sub_v)
            # __claiming__<epoch> 등록중 임시 마커 — 실제 상품번호 아님 (이슈 #579)
            if _key.startswith("__claiming__"):
                continue
            # 글로벌 인덱스 — 충돌 감지 (다른 cp가 같은 키 차지 시 ambiguous)
            _existing_global = by_global.get(_key)
            if not _existing_global:
                by_global[_key] = {
                    "collected_product_id": _cpid,
                    "source_site": _site,
                    "product_image": _thumb,
                    "original_link": _olink,
                    "category": _cat or "",
                    "cost": _cp_cost,
                    "site_ids_by_account": dict(_sites_by_account),
                }
            elif _existing_global.get("collected_product_id") != _cpid:
                if not _existing_global.get("ambiguous"):
                    _ambiguous_new += 1
                _existing_global["ambiguous"] = True
            else:
                # 같은 cp 재반영(증분 포함) — site_ids만 보강
                for acc, sites in _sites_by_account.items():
                    _existing_global["site_ids_by_account"].setdefault(acc, []).extend(
                        s
                        for s in sites
                        if s not in _existing_global["site_ids_by_account"].get(acc, [])
                    )
            # 정확 매칭 인덱스 — (account_id, product_no). 증분 시 동일 cp는 갱신,
            # 다른 cp가 이미 점유 중이면 가장 오래된 것 우선(덮어쓰기 안 함).
            _acc_key = f"{_account_key}:{_key}"
            _existing_acc = by_account.get(_acc_key)
            if (
                _existing_acc is None
                or _existing_acc.get("collected_product_id") == _cpid
            ):
                # 신규 or 같은 cp 재반영 — 기존 ambiguous 플래그는 보존
                _prev_ambig = bool(_existing_acc and _existing_acc.get("ambiguous"))
                by_account[_acc_key] = {
                    "collected_product_id": _cpid,
                    "source_site": _site,
                    "product_image": _thumb,
                    "original_link": _olink,
                    "category": _cat or "",
                    "cost": _cp_cost,
                    "site_ids_by_account": dict(_sites_by_account),
                }
                if _prev_ambig:
                    by_account[_acc_key]["ambiguous"] = True
            else:
                # #534 — 다른 cp가 같은 (account_id, product_no) 점유 = 진짜 identity 충돌.
                # 한 마켓 리스팅이 두 수집상품을 가리킴 → 판매링크≠소싱대상 오연결 사고.
                # 오래된 엔트리 유지하되 ambiguous 표시 → 주문 매칭에서 거부(오연결 방지).
                _existing_acc["ambiguous"] = True
    return _ambiguous_new


# images(JSON 배열, TOAST)는 SELECT에서 제외 — 포함 시 전체 스캔이 61초→337초로 폭증해
# 빌드가 per-account 타임아웃(180~300초)에 매번 killed → 캐시 영영 미생성 사고.
# product_image는 표시용일 뿐(마켓 자동채움 + /fetch-product-image 지연조회 존재)이라 빈값으로 둔다.
_MPN_SELECT_COLS = (
    "SELECT id, source_site, site_product_id, (images->>0) AS thumb, "
    "market_product_nos, source_url, category, cost FROM samba_collected_product "
    "WHERE market_product_nos IS NOT NULL"
)


async def _get_mpn_cache(
    session, sourcing_urls: dict
) -> tuple[dict[str, dict], dict[str, dict]]:
    """market_product_no → collected_product 인덱스 (증분 갱신).

    리턴: (by_global, by_account)
      - by_global[product_no]            = entry  (기존 호환 키, 충돌 시 entry["ambiguous"]=True)
      - by_account[f"{account_id}:{no}"] = entry  (정확 매칭용)

    [성능] 과거엔 호출마다 등록상품 전체(~10만건, 1GB 테이블)를 풀스캔해 빌드 →
    빌드(150초)가 TTL보다 길어 캐시가 안 채워지고 무한 재스캔 → read 풀 고갈 사고.
    이제 증분 방식:
      - 콜드스타트 / 전체 재빌드 주기(_MPN_FULL_REBUILD_SEC) 경과 시: 전체 빌드
      - 그 외: updated_at >= 직전빌드 변경분만(ix_scp_updated_at_desc) 가져와 기존 캐시에 머지
    삭제·등록해제로 사라진 키는 증분에서 안 지워지나, 전체 재빌드 주기마다 정리됨.
    매칭 키는 정확 키만 쓰므로 staleness가 오매칭을 만들지 않음.

    SELECT 전용이라 별도 read session을 연다. 인자 ``session``은 호환용(미사용).
    """
    import time as _t
    from datetime import datetime, timezone

    from sqlalchemy import text as _sa_text

    global _mpn_cache_data, _mpn_cache_built_at, _mpn_cache_full_built_at
    global _mpn_cache_delta_since
    async with _mpn_cache_lock:
        now = _t.monotonic()
        if (
            _mpn_cache_data is not None
            and (now - _mpn_cache_built_at) < _MPN_CACHE_TTL_SEC
        ):
            return _mpn_cache_data

        now_wall = datetime.now(timezone.utc)
        _full_rebuild = (
            _mpn_cache_data is None
            or (now - _mpn_cache_full_built_at) >= _MPN_FULL_REBUILD_SEC
        )

        if _full_rebuild:
            by_global: dict[str, dict] = {}
            by_account: dict[str, dict] = {}
            _ambiguous = 0
            async with get_read_session() as _read_sess:
                _cp_result = await _read_sess.execute(_sa_text(_MPN_SELECT_COLS))
                _cp_rows = _cp_result.fetchall()
            for _row in _cp_rows:
                _ambiguous += _index_mpn_row(_row, by_global, by_account, sourcing_urls)
            _mpn_cache_data = (by_global, by_account)
            _mpn_cache_full_built_at = now
            _mpn_cache_delta_since = now_wall
            _mpn_cache_built_at = now
            logger.info(
                f"[주문동기화] _mpn_cache 전체빌드 — global={len(by_global):,} "
                f"by_account={len(by_account):,} ambiguous={_ambiguous:,} "
                f"행={len(_cp_rows):,}"
            )
        else:
            # 증분 머지 — 변경분만. 시계 오차/경계 누락 방지 위해 10초 여유
            by_global, by_account = _mpn_cache_data
            _since = _mpn_cache_delta_since or now_wall
            from datetime import timedelta

            _since_q = _since - timedelta(seconds=10)
            async with get_read_session() as _read_sess:
                _cp_result = await _read_sess.execute(
                    _sa_text(_MPN_SELECT_COLS + " AND updated_at >= :since"),
                    {"since": _since_q},
                )
                _cp_rows = _cp_result.fetchall()
            _ambiguous = 0
            for _row in _cp_rows:
                _ambiguous += _index_mpn_row(_row, by_global, by_account, sourcing_urls)
            _mpn_cache_delta_since = now_wall
            _mpn_cache_built_at = now
            logger.info(
                f"[주문동기화] _mpn_cache 증분머지 — 변경 {len(_cp_rows):,}건 "
                f"global={len(by_global):,} ambiguous신규={_ambiguous:,}"
            )
        return _mpn_cache_data


ACTIVE_ORDER_STATUSES = (
    "new_order",
    "invoice_printed",
    "pending",
    "preparing",
    "wait_ship",
    "arrived",
)
EXCLUDED_ORDER_STATUSES = (
    "cancel_requested",
    "cancelling",
    "cancelled",
    "return_requested",
    "returning",
    "returned",
    "return_completed",
    "exchange_requested",
    "exchanging",
    "exchanged",
    "exchange_pending",
    "exchange_done",
    "ship_failed",
    "undeliverable",
    "shipping",
    "delivered",
    "confirmed",
)
PENDING_ORDER_STATUSES = (
    "pending",
    "preparing",
    "wait_ship",
    "arrived",
    "ship_failed",
    "undeliverable",
)

# 프론트 STATUS_MAP 라벨 → 내부 enum 키 역매핑.
# 플레이오토 미등록 주문에서 status를 shipping_status(한글 라벨)와 의미적으로 맞출 때 사용.
SHIPPING_LABEL_TO_STATUS_KEY = {
    "주문접수": "pending",
    "상품준비중": "pending",  # preparing 제거 — 미등록 주문엔 준비중 의미 없음
    "배송대기중": "wait_ship",
    "송장전송완료": "wait_ship",
    "상품도착": "arrived",
    "사무실도착": "arrived",
    "송장전송실패": "ship_failed",
    "국내배송중": "shipping",
    "출고완료": "shipping",
    "배송완료": "delivered",
    "구매확정": "delivered",
    "취소중": "cancelling",
    "취소요청": "cancel_requested",
    "취소완료": "cancelled",
    "반품중": "returning",
    "반품요청": "return_requested",
    "반품완료": "returned",
    "교환중": "exchanging",
    "교환완료": "exchanged",
    "회수확정": "return_completed",
    "발송불가": "undeliverable",
}

# 취소요청 알람 — 마켓에서 취소 신호(shipping_status='취소요청'/'취소완료')가 들어왔지만
# 우리 내부 status가 아직 '송장 나가기 전' 단계라 발주·송장 등록 사고 위험이 있는 케이스.
# UI 라벨 기준: 주문접수/상품준비중/배송대기중/송장전송실패 (= pending/preparing/wait_ship/ship_failed)
# [2026-06-18] arrived(사무실도착)/shipping(국내배송중)/delivered(배송완료) 제거 — 이미 송장이
#   나간 뒤라 '발주·송장 막기'가 불가능해 알람 취지에 안 맞음. 특히 배송완료 건의 마켓 '취소요청'은
#   배송 후 반품요청(롯데온 등)이라 발주사고 방지와 무관 → 오탐 제거(배송완료 건 알람에서 빠짐).
CANCEL_ALERT_SHIPPING_STATUSES = ("취소요청", "취소완료")
CANCEL_ALERT_TARGET_STATUSES = (
    "pending",
    "preparing",
    "wait_ship",
    "ship_failed",
)


def _build_cancel_alert_clause():
    """알람 카운트와 알람 필터에서 공통으로 쓰는 WHERE 조각.

    조건: 마켓 shipping_status 가 '취소요청'/'취소완료' + 우리 내부 status는 아직 처리/배송 단계
      → 발주·송장 등록 사고 위험. 운영자가 보고 막아야 할 미처리 케이스.

    내부 status='cancel_requested'는 운영자가 이미 인지하고 드롭박스를 전환한 상태라
    더 이상 발주/송장이 나가지 않으므로 알람 대상에서 제외.
    """
    from sqlalchemy import and_

    return and_(
        SambaOrder.shipping_status.in_(CANCEL_ALERT_SHIPPING_STATUSES),
        SambaOrder.status.in_(CANCEL_ALERT_TARGET_STATUSES),
    )


def _build_action_tag_filter(action_tag: str):
    from sqlalchemy import func, or_

    normalized = action_tag.strip()
    if not normalized:
        return None

    padded = f",{normalized},"
    action_expr = func.concat(",", func.coalesce(SambaOrder.action_tag, ""), ",")
    return or_(
        SambaOrder.action_tag == normalized,
        action_expr.like(f"{padded}%"),
        action_expr.like(f"%{padded}"),
        action_expr.like(f"%{padded}%"),
    )


class PaginatedOrdersResponse(BaseModel):
    items: list[SambaOrder]
    total_count: int
    total_sale: float
    pending_count: int
    # 상품메모(#535) — {collected_product_id: memo}. 주문의 collected_product_id로
    # 현재 상품 memo를 live-join(스냅샷 아님). 빈 메모는 제외.
    product_memos: dict[str, str] = {}


def _read_service(session: AsyncSession) -> SambaOrderService:
    return SambaOrderService(SambaOrderRepository(session))


def _write_service(session: AsyncSession) -> SambaOrderService:
    return SambaOrderService(SambaOrderRepository(session))


async def _resolve_market_filter_channel_ids(
    session: AsyncSession,
    market_filter: Optional[str],
    tenant_id: Optional[str],
) -> list[str]:
    if not market_filter or not market_filter.startswith("type:"):
        return []

    from sqlalchemy import or_, select

    from backend.domain.samba.account.model import SambaMarketAccount

    market_type = market_filter[5:]
    stmt = select(SambaMarketAccount.id).where(
        SambaMarketAccount.market_type == market_type
    )
    if tenant_id is not None:
        stmt = stmt.where(
            or_(
                SambaMarketAccount.tenant_id == tenant_id,
                SambaMarketAccount.tenant_id == None,  # noqa: E711
            )
        )
    result = await session.execute(stmt)
    return [row[0] for row in result.all() if row[0]]


async def _build_order_filters(
    session: AsyncSession,
    tenant_id: Optional[str],
    *,
    market_filter: str = "",
    site_filter: str = "",
    account_filter: str = "",
    market_status: str = "",
    status_filter: str = "",
    input_filter: str = "",
    invoice_filter: str = "",
    registration_filter: str = "",
    search_text: str = "",
    search_category: str = "customer",
) -> list[Any]:
    from sqlalchemy import and_, func, or_, select

    filters: list[Any] = []

    if tenant_id is not None:
        filters.append(
            or_(
                SambaOrder.tenant_id == tenant_id,
                SambaOrder.tenant_id == None,  # noqa: E711
            )
        )

    if market_filter:
        if market_filter.startswith("acc:"):
            filters.append(SambaOrder.channel_id == market_filter[4:])
        elif market_filter.startswith("type:"):
            channel_ids = await _resolve_market_filter_channel_ids(
                session, market_filter, tenant_id
            )
            if channel_ids:
                filters.append(SambaOrder.channel_id.in_(channel_ids))
            else:
                filters.append(SambaOrder.channel_id == "__no_matching_channel__")

    if site_filter:
        normalized_site_filter = site_filter.replace(" ", "")
        normalized_source_site = func.replace(
            func.coalesce(SambaOrder.source_site, ""), " ", ""
        )
        # GSSHOP 통합 필터 — DB에는 GSShop/GS이숍/GS이숍(고경) 등 변형 혼재 → 모두 매칭
        gs_aliases = {"GSSHOP", "GSShop", "GS이숍", "GS이샵", "GS샵"}
        if normalized_site_filter.upper() == "SNKRDUNK":
            # 소싱처=스니덩크: 크림 판매 주문의 실제 소싱처 — 소싱계정이 SNKRDUNK인 주문
            from backend.domain.samba.sourcing_account.model import SambaSourcingAccount

            snkr_acc_subq = select(SambaSourcingAccount.id).where(
                func.upper(SambaSourcingAccount.site_name) == "SNKRDUNK"
            )
            filters.append(
                or_(
                    normalized_source_site == "SNKRDUNK",
                    SambaOrder.sourcing_account_id.in_(snkr_acc_subq),
                )
            )
        elif normalized_site_filter in gs_aliases:
            from backend.core.sql_safe import escape_like

            gs_filters = []
            for alias in gs_aliases:
                safe_alias = escape_like(alias)
                gs_filters.append(normalized_source_site == alias)
                gs_filters.append(
                    normalized_source_site.like(f"{safe_alias}(%", escape="\\")
                )
            filters.append(or_(*gs_filters))
        elif "(" in normalized_site_filter:
            filters.append(normalized_source_site == normalized_site_filter)
        else:
            # site_filter 는 외부 입력 — `%`/`_` 메타 escape 후 ESCAPE '\\' 명시.
            # `(%` 는 의도된 wildcard 이므로 보존, escape 는 site_filter 부분만 적용.
            from backend.core.sql_safe import escape_like

            safe_site = escape_like(normalized_site_filter)
            filters.append(
                or_(
                    normalized_source_site == normalized_site_filter,
                    normalized_source_site.like(f"{safe_site}(%", escape="\\"),
                )
            )
    if account_filter:
        # '기타'(account_filter='etc') = 주문계정을 명시적으로 '기타'로 지정한 주문만.
        # NULL(소싱계정 미지정)은 '기타'가 아니므로 포함하지 않는다 — exact match.
        filters.append(SambaOrder.sourcing_account_id == account_filter)
    if market_status:
        filters.append(SambaOrder.shipping_status == market_status)

    if status_filter:
        if status_filter == "active":
            filters.append(SambaOrder.status.in_(ACTIVE_ORDER_STATUSES))
        elif status_filter == "cancel_return_excluded":
            # status 컬럼만 기준 — shipping_status 는 일절 관여 금지
            filters.append(~SambaOrder.status.in_(EXCLUDED_ORDER_STATUSES))
        elif status_filter == "cancel_alert":
            # 알람 카운트와 동일한 조건 — 발주·송장 사고 위험 케이스
            filters.append(_build_cancel_alert_clause())
        else:
            filters.append(SambaOrder.status == status_filter)

    if input_filter == "has_order":
        filters.append(
            and_(
                SambaOrder.sourcing_order_number != None,  # noqa: E711
                SambaOrder.sourcing_order_number != "",
            )
        )
    elif input_filter == "no_order":
        filters.append(
            or_(
                SambaOrder.sourcing_order_number == None,  # noqa: E711
                SambaOrder.sourcing_order_number == "",
            )
        )
    elif input_filter == "has_invoice":
        filters.append(
            and_(
                SambaOrder.tracking_number != None,  # noqa: E711
                SambaOrder.tracking_number != "",
            )
        )
    elif input_filter == "no_invoice":
        filters.append(
            or_(
                SambaOrder.tracking_number == None,  # noqa: E711
                SambaOrder.tracking_number == "",
            )
        )
    elif input_filter in {
        "no_price",
        "no_stock",
        "direct",
        "kkadaegi",
        "gift",
        "staff_a",
        "staff_b",
    }:
        action_filter = _build_action_tag_filter(input_filter)
        if action_filter is not None:
            filters.append(action_filter)

    # 송장필터 — 입력필터와 독립적으로 동작 (이중 선택 가능)
    # 크림(KREAM) 주문은 tracking_number(허브넷 HBL)가 주문 생성 시부터 채워지므로
    # 송장 유무 판정을 해외송장번호(overseas_tracking_number, 스니덩크→사무국 발송) 기준으로 함
    _is_kream = func.upper(func.coalesce(SambaOrder.source_site, "")) == "KREAM"
    if invoice_filter == "has_invoice":
        filters.append(
            or_(
                and_(
                    _is_kream,
                    SambaOrder.overseas_tracking_number != None,  # noqa: E711
                    SambaOrder.overseas_tracking_number != "",
                ),
                and_(
                    ~_is_kream,
                    SambaOrder.tracking_number != None,  # noqa: E711
                    SambaOrder.tracking_number != "",
                ),
            )
        )
    elif invoice_filter == "no_invoice":
        filters.append(
            or_(
                and_(
                    _is_kream,
                    or_(
                        SambaOrder.overseas_tracking_number == None,  # noqa: E711
                        SambaOrder.overseas_tracking_number == "",
                    ),
                ),
                and_(
                    ~_is_kream,
                    or_(
                        SambaOrder.tracking_number == None,  # noqa: E711
                        SambaOrder.tracking_number == "",
                    ),
                ),
            )
        )

    # 등록필터
    # - product_image: SSG/스마트스토어/플레이오토가 매칭 없이도 자동으로 채워주므로 판정 기준에서 제외
    # - source_url: SSG는 itemId 기반으로 주문 수집 시 자동 채워주므로 SSG 계정 주문에서는 판정 기준에서 제외
    # - 타 마켓은 "미등록 입력" UX(사용자가 직접 source_url 채움)로 source_url 채우면 등록으로 간주 (기존 동작 유지)
    if registration_filter in ("registered", "unregistered"):
        from backend.domain.samba.account.model import SambaMarketAccount

        _ssg_stmt = select(SambaMarketAccount.id).where(
            SambaMarketAccount.market_type == "ssg"
        )
        if tenant_id is not None:
            _ssg_stmt = _ssg_stmt.where(
                or_(
                    SambaMarketAccount.tenant_id == tenant_id,
                    SambaMarketAccount.tenant_id == None,  # noqa: E711
                )
            )
        _ssg_rows = (await session.execute(_ssg_stmt)).all()
        ssg_channel_ids = [r[0] for r in _ssg_rows if r[0]]

        has_source_url = and_(
            SambaOrder.source_url != None,  # noqa: E711
            SambaOrder.source_url != "",
        )
        no_source_url = or_(
            SambaOrder.source_url == None,  # noqa: E711
            SambaOrder.source_url == "",
        )
        if ssg_channel_ids:
            is_ssg = SambaOrder.channel_id.in_(ssg_channel_ids)
            not_ssg = SambaOrder.channel_id.notin_(ssg_channel_ids)
        else:
            from sqlalchemy import false, true

            is_ssg = false()
            not_ssg = true()

        if registration_filter == "registered":
            filters.append(
                or_(
                    SambaOrder.collected_product_id != None,  # noqa: E711
                    and_(not_ssg, has_source_url),
                )
            )
        else:
            filters.append(
                and_(
                    SambaOrder.collected_product_id == None,  # noqa: E711
                    or_(is_ssg, no_source_url),
                )
            )

    normalized_search = search_text.strip()
    if normalized_search:
        # search_text 는 외부 입력 — `%`/`_` 메타 escape 후 ESCAPE '\\' 명시.
        from backend.core.sql_safe import escape_like

        safe_q = escape_like(normalized_search.lower())
        lower_q = f"%{safe_q}%"
        if search_category == "product":
            filters.append(SambaOrder.product_name.ilike(lower_q, escape="\\"))
        elif search_category == "product_id":
            filters.append(SambaOrder.product_id.ilike(lower_q, escape="\\"))
        elif search_category == "order_number":
            # 상품주문번호(order_number) + 묶음주문번호(shipment_id) + 외부주문번호(ext_order_number) 모두 매칭
            filters.append(
                or_(
                    SambaOrder.order_number.ilike(lower_q, escape="\\"),
                    SambaOrder.shipment_id.ilike(lower_q, escape="\\"),
                    SambaOrder.ext_order_number.ilike(lower_q, escape="\\"),
                )
            )
        elif search_category == "sourcing_order_number":
            filters.append(SambaOrder.sourcing_order_number.ilike(lower_q, escape="\\"))
        elif search_category == "tracking_number":
            # 국내송장(tracking_number) + 해외송장(overseas_tracking_number, 크림) 모두 매칭
            filters.append(
                or_(
                    SambaOrder.tracking_number.ilike(lower_q, escape="\\"),
                    SambaOrder.overseas_tracking_number.ilike(lower_q, escape="\\"),
                )
            )
        else:
            # 고객명(수령인) + 주문자명 모두 매칭 — 선물하기 등 수령인≠주문자 케이스 대응
            filters.append(
                or_(
                    SambaOrder.customer_name.ilike(lower_q, escape="\\"),
                    SambaOrder.orderer_name.ilike(lower_q, escape="\\"),
                )
            )

    return filters


def _build_order_sort(sort_by: str):
    from sqlalchemy import func

    date_col = func.coalesce(SambaOrder.paid_at, SambaOrder.created_at)
    # 결제시간 동일 시 수집시간(created_at) 최신순 보조정렬
    sort_map = {
        "date_asc": [date_col.asc(), SambaOrder.created_at.desc()],
        "profit_desc": [SambaOrder.profit.desc(), SambaOrder.created_at.desc()],
        "profit_asc": [SambaOrder.profit.asc(), SambaOrder.created_at.desc()],
        "price_desc": [SambaOrder.sale_price.desc(), SambaOrder.created_at.desc()],
        "price_asc": [SambaOrder.sale_price.asc(), SambaOrder.created_at.desc()],
    }
    return sort_map.get(sort_by, [date_col.desc(), SambaOrder.created_at.desc()])


async def _run_paginated_order_query(
    session: AsyncSession,
    base_filters: list[Any],
    *,
    skip: int,
    limit: int,
    sort_by: str,
    extra_filters: Optional[list[Any]] = None,
) -> PaginatedOrdersResponse:
    from sqlalchemy import case, func, select

    sale_expr = func.coalesce(SambaOrder.total_payment_amount, SambaOrder.sale_price, 0)
    query_filters = [*base_filters, *(extra_filters or [])]

    total_stmt = select(
        func.count().label("total_count"),
        func.coalesce(func.sum(sale_expr), 0).label("total_sale"),
        func.coalesce(
            func.sum(case((SambaOrder.status.in_(PENDING_ORDER_STATUSES), 1), else_=0)),
            0,
        ).label("pending_count"),
    )
    if query_filters:
        total_stmt = total_stmt.where(*query_filters)
    total_row = (await session.execute(total_stmt)).one()

    items_stmt = select(SambaOrder)
    if query_filters:
        items_stmt = items_stmt.where(*query_filters)
    items_stmt = (
        items_stmt.order_by(*_build_order_sort(sort_by)).offset(skip).limit(limit)
    )
    items = list((await session.execute(items_stmt)).scalars().all())

    # KREAM 주문 한글 상품명 보강 — collected_product.name(한글)으로 오버라이드
    _kream_cp_ids = [
        o.collected_product_id
        for o in items
        if o.source_site == "KREAM" and o.collected_product_id
    ]
    if _kream_cp_ids:
        from backend.domain.samba.collector.model import SambaCollectedProduct as _CP

        _cp_rows = (
            await session.execute(
                select(_CP.id, _CP.name, _CP.images, _CP.source_url).where(
                    _CP.id.in_(_kream_cp_ids)
                )
            )
        ).all()
        import json as _json

        _cp_data_map = {r[0]: (r[1], r[2], r[3]) for r in _cp_rows}
        for o in items:
            if o.source_site == "KREAM" and o.collected_product_id:
                _name, _imgs, _cp_src_url = _cp_data_map.get(
                    o.collected_product_id, (None, None, None)
                )
                if _name:
                    o.product_name = _name
                if not o.product_image and _imgs:
                    _img_list = _json.loads(_imgs) if isinstance(_imgs, str) else _imgs
                    if _img_list:
                        o.product_image = _img_list[0]
                if _cp_src_url:
                    o.source_url = _cp_src_url

    # 상품메모(#535) live-join — 주문의 collected_product_id로 현재 상품 memo 조회.
    # cp_id는 전역 유니크라 tenant 필터 불요. 빈 메모는 맵에서 제외.
    product_memos: dict[str, str] = {}
    _memo_cp_ids = [
        o.collected_product_id
        for o in items
        if o.collected_product_id and o.collected_product_id != "DELETED"
    ]
    if _memo_cp_ids:
        from backend.domain.samba.collector.model import SambaCollectedProduct as _CPM

        _memo_rows = (
            await session.execute(
                select(_CPM.id, _CPM.memo).where(
                    _CPM.id.in_(_memo_cp_ids), _CPM.memo.isnot(None)
                )
            )
        ).all()
        for _cid, _memo in _memo_rows:
            if _memo and str(_memo).strip():
                product_memos[_cid] = _memo

    return PaginatedOrdersResponse(
        items=items,
        total_count=int(total_row.total_count or 0),
        total_sale=float(total_row.total_sale or 0),
        pending_count=int(total_row.pending_count or 0),
        product_memos=product_memos,
    )


@router.get("", response_model=list[SambaOrder])
async def list_orders(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    status: Optional[str] = None,
    session: AsyncSession = Depends(get_read_session_dependency),
    tenant_id: Optional[str] = Depends(get_optional_tenant_id),
):
    from sqlmodel import select

    # tenant_id가 있으면 해당 테넌트 주문만 조회
    if tenant_id is not None:
        stmt = (
            select(SambaOrder)
            .order_by(SambaOrder.created_at.desc())
            .offset(skip)
            .limit(limit)
        )
        from sqlalchemy import or_

        stmt = stmt.where(
            or_(
                SambaOrder.tenant_id == tenant_id,
                SambaOrder.tenant_id == None,  # noqa: E711
            )
        )
        if status:
            stmt = stmt.where(SambaOrder.status == status)
        result = await session.execute(stmt)
        return result.scalars().all()
    svc = _read_service(session)
    return await svc.list_orders(skip=skip, limit=limit, status=status)


@router.get("/dashboard-stats")
async def dashboard_stats(
    session: AsyncSession = Depends(get_read_session_dependency),
    tenant_id: Optional[str] = Depends(get_optional_tenant_id),
):
    """대시보드 집계 — DB에서 SUM/COUNT 후 결과만 반환 (빠름)."""
    # 캐시 조회 (TTL 60초, tenant별 키)
    _cache_key = f"order:dashboard-stats-v3:{tenant_id or '_global'}"
    _cached = await cache.get(_cache_key)
    if _cached:
        return _cached

    from sqlalchemy import select, func, case, and_, extract, text, or_
    from datetime import datetime, timedelta, timezone as tz

    # 이행매출 대상 상태 (주문상태 드롭박스 기준)
    FULFILLMENT_STATUSES = (
        "pending",
        "wait_ship",
        "processing",
        "arrived",
        "ship_failed",
        "shipping",
        "shipped",
        "delivered",
        "exchanged",
        "exchanging",
        "exchange_requested",
    )

    # KST 기준 (UTC+9)
    KST = tz(timedelta(hours=9))
    now = datetime.now(KST).replace(tzinfo=None)
    this_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    if now.month == 1:
        last_month_start = this_month_start.replace(year=now.year - 1, month=12)
    else:
        last_month_start = this_month_start.replace(month=now.month - 1)
    week_ago = (now - timedelta(days=6)).replace(
        hour=0, minute=0, second=0, microsecond=0
    )

    # 날짜 기준: 고객결제일(paid_at)만 사용, KST 변환
    order_date = SambaOrder.paid_at + text("INTERVAL '9 hours'")

    # 금월 집계
    this_month_q = select(
        func.count().label("count"),
        func.coalesce(func.sum(SambaOrder.sale_price), 0).label("sales"),
        func.coalesce(
            func.sum(
                case(
                    (
                        SambaOrder.status.in_(FULFILLMENT_STATUSES),
                        SambaOrder.sale_price,
                    ),
                    else_=0,
                )
            ),
            0,
        ).label("fulfillment_sales"),
        func.sum(
            case(
                (SambaOrder.status.in_(FULFILLMENT_STATUSES), 1),
                else_=0,
            )
        ).label("fulfillment_count"),
    ).where(SambaOrder.paid_at != None, order_date >= this_month_start)  # noqa: E711
    if tenant_id is not None:
        this_month_q = this_month_q.where(
            or_(
                SambaOrder.tenant_id == tenant_id,
                SambaOrder.tenant_id == None,  # noqa: E711
            )
        )
    tm = (await session.execute(this_month_q)).one()

    # 전월 집계
    last_month_q = select(
        func.count().label("count"),
        func.coalesce(func.sum(SambaOrder.sale_price), 0).label("sales"),
        func.coalesce(
            func.sum(
                case(
                    (
                        SambaOrder.status.in_(FULFILLMENT_STATUSES),
                        SambaOrder.sale_price,
                    ),
                    else_=0,
                )
            ),
            0,
        ).label("fulfillment_sales"),
        func.sum(
            case(
                (SambaOrder.status.in_(FULFILLMENT_STATUSES), 1),
                else_=0,
            )
        ).label("fulfillment_count"),
    ).where(
        SambaOrder.paid_at != None,
        and_(order_date >= last_month_start, order_date < this_month_start),
    )  # noqa: E711
    if tenant_id is not None:
        last_month_q = last_month_q.where(
            or_(
                SambaOrder.tenant_id == tenant_id,
                SambaOrder.tenant_id == None,  # noqa: E711
            )
        )
    lm = (await session.execute(last_month_q)).one()

    # 최근 7일 일별 집계
    # 미발송(송장 미입력) 조건 — 송장수집 모달 카운트와 동일 기준:
    #   tracking_number 비어있음 + 취소/반품/교환 상태 제외 + 배송완료 키워드 제외
    from backend.domain.samba.order.model import (
        EXCLUDED_ORDER_STATUSES,
        SHIPPED_SHIPPING_STATUS_KEYWORDS,
    )

    _ship_col = func.coalesce(SambaOrder.shipping_status, "")
    # 발송 조건 — 운송장 입력됨 또는 배송완료 키워드 + 취소/반품/교환 상태 제외
    shipped_cond = and_(
        or_(
            and_(
                SambaOrder.tracking_number != None,  # noqa: E711
                SambaOrder.tracking_number != "",
            ),
            *[_ship_col.like(f"%{kw}%") for kw in SHIPPED_SHIPPING_STATUS_KEYWORDS],
        ),
        or_(
            SambaOrder.status == None,  # noqa: E711
            SambaOrder.status.notin_(EXCLUDED_ORDER_STATUSES),
        ),
    )
    daily_q = (
        select(
            func.date(order_date).label("day"),
            func.count().label("count"),
            func.coalesce(func.sum(SambaOrder.sale_price), 0).label("sales"),
            func.coalesce(
                func.sum(
                    case(
                        (
                            SambaOrder.status.in_(FULFILLMENT_STATUSES),
                            SambaOrder.sale_price,
                        ),
                        else_=0,
                    )
                ),
                0,
            ).label("fulfillment_sales"),
            func.sum(
                case(
                    (SambaOrder.status.in_(FULFILLMENT_STATUSES), 1),
                    else_=0,
                )
            ).label("fulfillment_count"),
            func.sum(
                case(
                    (shipped_cond, 1),
                    else_=0,
                )
            ).label("shipped_count"),
        )
        .where(SambaOrder.paid_at != None, order_date >= week_ago)  # noqa: E711
        .group_by(func.date(order_date))
    )
    if tenant_id is not None:
        daily_q = daily_q.where(
            or_(
                SambaOrder.tenant_id == tenant_id,
                SambaOrder.tenant_id == None,  # noqa: E711
            )
        )
    daily_rows = (await session.execute(daily_q)).all()

    # 미발송(unshippedCount) = "송장 수집대상 주문 수"(송장 진행현황 모달이 일괄수집할 대상).
    #   주의: 모달 '대기'(PENDING 잡)는 잡이 처리되면 곧바로 빠지는 순간값 → 스냅샷 부적합.
    #   대신 enqueue_pending_orders WHERE(SambaOrder 직접) + 모달 배송키워드 제외 = 결정적 집합.
    #   오늘 행: 라이브 계산(트레일링 7일). 과거 행: samba_daily_unshipped_snapshot(매일 0시 cron).
    #   스냅샷 없는 과거일은 None("-") — 거짓 0 채움 금지.
    from backend.domain.samba.order.model import SambaDailyUnshippedSnapshot

    _action_tag_expr = func.concat(",", func.coalesce(SambaOrder.action_tag, ""), ",")
    # order_date = paid_at(폴백 created_at) + 9h(KST). week_ago = KST 오늘-6일 00:00.
    #   윈도우 = [week_ago, week_ago+7일) = enqueue_pending_orders 의 [since, until) 와 동일.
    _unshipped_target_q = (
        select(func.count())
        .select_from(SambaOrder)
        .where(
            or_(
                SambaOrder.tracking_number == None,  # noqa: E711
                SambaOrder.tracking_number == "",
            ),
            SambaOrder.sourcing_order_number != None,  # noqa: E711
            SambaOrder.sourcing_order_number != "",
            or_(
                and_(
                    SambaOrder.source_site != None,  # noqa: E711
                    SambaOrder.source_site != "",
                ),
                and_(
                    SambaOrder.source_url != None,  # noqa: E711
                    SambaOrder.source_url != "",
                ),
                SambaOrder.collected_product_id != None,  # noqa: E711
            ),
            order_date >= week_ago,
            order_date < (week_ago + timedelta(days=7)),
            or_(
                SambaOrder.status == None,  # noqa: E711
                SambaOrder.status.notin_(EXCLUDED_ORDER_STATUSES),
            ),
            *[_ship_col.notlike(f"%{kw}%") for kw in SHIPPED_SHIPPING_STATUS_KEYWORDS],
            ~_action_tag_expr.like("%,kkadaegi,%"),
        )
    )
    if tenant_id is not None:
        _unshipped_target_q = _unshipped_target_q.where(
            or_(
                SambaOrder.tenant_id == tenant_id,
                SambaOrder.tenant_id == None,  # noqa: E711
            )
        )
    unshipped_live_total = int(
        (await session.execute(_unshipped_target_q)).scalar() or 0
    )

    # 과거 6일 미발송 스냅샷 (오늘 행은 위 라이브값 사용)
    _unshipped_today_str = (week_ago + timedelta(days=6)).strftime("%Y-%m-%d")
    _unshipped_past_dates = [
        (week_ago + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(6)
    ]
    unshipped_snap_rows = (
        await session.execute(
            select(
                SambaDailyUnshippedSnapshot.snapshot_date,
                SambaDailyUnshippedSnapshot.unshipped_count,
            ).where(
                SambaDailyUnshippedSnapshot.snapshot_date.in_(_unshipped_past_dates)
            )
        )
    ).all()
    unshipped_snap_map = {
        r.snapshot_date: int(r.unshipped_count) for r in unshipped_snap_rows
    }

    weekly = []
    for i in range(7):
        d = week_ago + timedelta(days=i)
        day_str = d.strftime("%Y-%m-%d")
        row = next((r for r in daily_rows if str(r.day) == day_str), None)
        if day_str == _unshipped_today_str:
            _unshipped = unshipped_live_total
        else:
            # 스냅샷 없으면 None → 프론트 "-" 표시 (거짓 0 채움 금지)
            _unshipped = unshipped_snap_map.get(day_str)
        weekly.append(
            {
                "date": day_str,
                "sales": float(row.sales) if row else 0,
                "count": int(row.count) if row else 0,
                "fulfillmentSales": float(row.fulfillment_sales) if row else 0,
                "fulfillmentCount": int(row.fulfillment_count) if row else 0,
                "shippedCount": int(row.shipped_count) if row else 0,
                "unshippedCount": _unshipped,
            }
        )

    # 월별 집계 (연간 12개월)
    year_start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    monthly_q = (
        select(
            extract("month", order_date).label("month"),
            func.coalesce(func.sum(SambaOrder.sale_price), 0).label("sales"),
            func.coalesce(
                func.sum(
                    case(
                        (
                            SambaOrder.status.in_(FULFILLMENT_STATUSES),
                            SambaOrder.sale_price,
                        ),
                        else_=0,
                    )
                ),
                0,
            ).label("fulfillment_sales"),
        )
        .where(
            SambaOrder.paid_at != None,  # noqa: E711
            and_(
                order_date >= year_start,
                extract("year", order_date) == now.year,
            ),
        )
        .group_by(extract("month", order_date))
    )
    if tenant_id is not None:
        monthly_q = monthly_q.where(
            or_(
                SambaOrder.tenant_id == tenant_id,
                SambaOrder.tenant_id == None,  # noqa: E711
            )
        )
    monthly_rows = (await session.execute(monthly_q)).all()
    monthly = []
    for m in range(1, 13):
        row = next((r for r in monthly_rows if int(r.month) == m), None)
        monthly.append(
            {
                "month": m,
                "sales": float(row.sales) if row else 0,
                "fulfillmentSales": float(row.fulfillment_sales) if row else 0,
            }
        )

    # 최근 7일 신규등록/마켓삭제 상품 단위 일별 카운트 (KST 기준)
    # 신규등록: registered_accounts 0→≥1 전환 시점 (first_market_registered_at)
    # 마켓삭제: 품절 인식 이벤트(sold_out) 기준 — 1상품/1일 중복 제거
    from backend.api.v1.routers.samba.collector_common import (
        build_market_registered_conditions,
    )
    from backend.domain.samba.collector.model import SambaCollectedProduct

    reg_date = SambaCollectedProduct.first_market_registered_at + text(
        "INTERVAL '9 hours'"
    )
    new_reg_q = (
        select(
            func.date(reg_date).label("day"),
            func.count().label("cnt"),
        )
        .where(
            SambaCollectedProduct.first_market_registered_at != None,  # noqa: E711
            reg_date >= week_ago,
        )
        .group_by(func.date(reg_date))
    )
    if tenant_id is not None:
        new_reg_q = new_reg_q.where(
            or_(
                SambaCollectedProduct.tenant_id == tenant_id,
                SambaCollectedProduct.tenant_id == None,  # noqa: E711
            )
        )
    new_reg_rows = (await session.execute(new_reg_q)).all()
    new_reg_map = {str(r.day): int(r.cnt) for r in new_reg_rows}

    # (del_q 제거 — sold_out 이벤트만 잡아 마켓삭제가 과소계산됨)
    # 마켓삭제는 등록상품수 스냅샷 역산으로 구하므로 별도 쿼리 불필요

    # 일별 누적 등록상품수: "지금 마켓에 1개 이상 등록된 상품수" 정의로 통일
    #   - 오늘(today_str): 실시간 build_market_registered_conditions 계산값
    #   - 과거 6일: samba_daily_registered_snapshot 테이블의 그날 0시 스냅샷
    #   - 스냅샷이 없는 과거일은 None(프론트에서 "-" 표시) — 거짓 평탄 채움 금지
    from backend.domain.samba.collector.model import SambaDailyRegisteredSnapshot

    today_str = (week_ago + timedelta(days=6)).strftime("%Y-%m-%d")
    reg_count_map: dict[str, Optional[int]] = {}

    # 마켓 1개 이상 등록된 상품수 (현재 시점) — KPI + 오늘 행에 사용
    market_registered_q = select(func.count(SambaCollectedProduct.id)).where(
        *build_market_registered_conditions(SambaCollectedProduct)
    )
    if tenant_id is not None:
        market_registered_q = market_registered_q.where(
            or_(
                SambaCollectedProduct.tenant_id == tenant_id,
                SambaCollectedProduct.tenant_id == None,  # noqa: E711
            )
        )
    market_registered_count = (await session.execute(market_registered_q)).scalar() or 0
    reg_count_map[today_str] = int(market_registered_count)

    # 과거 6일 스냅샷 일괄 조회
    past_dates = [(week_ago + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(6)]
    snap_q = select(
        SambaDailyRegisteredSnapshot.snapshot_date,
        SambaDailyRegisteredSnapshot.registered_count,
    ).where(SambaDailyRegisteredSnapshot.snapshot_date.in_(past_dates))
    snap_rows = (await session.execute(snap_q)).all()
    snap_map = {r.snapshot_date: int(r.registered_count) for r in snap_rows}

    # 스냅샷이 있으면 사용, 없으면 None(프론트 "-" 표시)
    # — 매일 0시 TASK 6 누적되면 자연스럽게 진짜 스냅샷으로 대체됨
    for d_str in past_dates:
        reg_count_map[d_str] = snap_map.get(d_str)

    # 일별 누적 수집상품수 = "그 날(말) 시점 삼바에 저장되어있는 전체 상품수"
    # 구현: 현재 total 에서 그 다음날 이후 created 된 행수를 빼서 역산 (1풀스캔 + 1범위스캔)
    total_collected_q = select(func.count(SambaCollectedProduct.id))
    if tenant_id is not None:
        total_collected_q = total_collected_q.where(
            or_(
                SambaCollectedProduct.tenant_id == tenant_id,
                SambaCollectedProduct.tenant_id == None,  # noqa: E711
            )
        )
    total_collected = int((await session.execute(total_collected_q)).scalar() or 0)

    created_kst = SambaCollectedProduct.created_at + text("INTERVAL '9 hours'")
    daily_new_q = (
        select(
            func.date(created_kst).label("day"),
            func.count().label("cnt"),
        )
        .where(
            SambaCollectedProduct.created_at != None,  # noqa: E711
            created_kst >= week_ago,
        )
        .group_by(func.date(created_kst))
    )
    if tenant_id is not None:
        daily_new_q = daily_new_q.where(
            or_(
                SambaCollectedProduct.tenant_id == tenant_id,
                SambaCollectedProduct.tenant_id == None,  # noqa: E711
            )
        )
    daily_new_rows = (await session.execute(daily_new_q)).all()
    daily_new_map = {str(r.day): int(r.cnt) for r in daily_new_rows}

    # 7일 누적 카운트: 오늘=total, 어제=total-(오늘신규), 그저께=어제-(어제신규) ...
    collected_count_map: dict[str, int] = {today_str: total_collected}
    running_total = total_collected
    for i in range(5, -1, -1):
        d_str = past_dates[i]
        next_d_str = past_dates[i + 1] if i + 1 < 6 else today_str
        running_total -= daily_new_map.get(next_d_str, 0)
        collected_count_map[d_str] = max(running_total, 0)

    # 7일 이전(week_ago - 1d) 스냅샷 추가 조회 — 첫 행 신규등록 역산용
    prev_day_str = (week_ago - timedelta(days=1)).strftime("%Y-%m-%d")
    prev_snap_row = (
        await session.execute(
            select(SambaDailyRegisteredSnapshot.registered_count).where(
                SambaDailyRegisteredSnapshot.snapshot_date == prev_day_str
            )
        )
    ).scalar()
    reg_count_map[prev_day_str] = (
        int(prev_snap_row) if prev_snap_row is not None else None
    )

    # 신규등록 = first_market_registered_at 기준 (0→≥1 최초 전환 날짜). 상품당 최초 1회만
    #   찍히므로 재등록(품절삭제 후 재등록)은 제외 — "진짜 신규 상품"만 카운트.
    # 순증감(±) = 등록상품수[d] - 등록상품수[d-1] (스냅샷 델타).
    #   구 "마켓삭제"는 (전일 + 신규등록 - 금일) 역산 + max(...,0) 클램프였으나,
    #   재등록을 신규로 못 세어 방정식이 음수로 깨지면 거짓 0 이 나와 폐기(#dashboard-fix).
    #   실제 삭제 이벤트가 DB에 안 남아 삭제 건수 소급 불가 → 정직하게 순증감만 노출.
    #   스냅샷 없는 날은 순증감 None(프론트 "—").
    all_dates = [(week_ago + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(7)]
    for idx, w in enumerate(weekly):
        d_str = w["date"]
        prev_str = prev_day_str if idx == 0 else all_dates[idx - 1]
        reg_today = reg_count_map.get(d_str)
        reg_prev = reg_count_map.get(prev_str)
        new_reg = int(new_reg_map.get(d_str, 0))
        w["newRegistered"] = new_reg
        if reg_today is not None and reg_prev is not None:
            w["netChange"] = reg_today - reg_prev
        else:
            w["netChange"] = None
        w["registeredCount"] = reg_today
        w["collectedCount"] = int(collected_count_map.get(d_str, 0))

    tm_fulfillment_rate = (
        round(int(tm.fulfillment_count or 0) / int(tm.count) * 100) if tm.count else 0
    )
    lm_fulfillment_rate = (
        round(int(lm.fulfillment_count or 0) / int(lm.count) * 100) if lm.count else 0
    )
    sales_change = (
        round(((float(tm.sales) - float(lm.sales)) / float(lm.sales)) * 100, 1)
        if lm.sales
        else 0
    )

    result = {
        "thisMonth": {
            "count": int(tm.count),
            "sales": float(tm.sales),
            "fulfillmentSales": float(tm.fulfillment_sales or 0),
            "fulfillmentCount": int(tm.fulfillment_count or 0),
            "fulfillment": tm_fulfillment_rate,
        },
        "lastMonth": {
            "count": int(lm.count),
            "sales": float(lm.sales),
            "fulfillmentSales": float(lm.fulfillment_sales or 0),
            "fulfillmentCount": int(lm.fulfillment_count or 0),
            "fulfillment": lm_fulfillment_rate,
        },
        "salesChange": sales_change,
        "weekly": weekly,
        "monthly": monthly,
        "marketRegisteredCount": int(market_registered_count),
    }
    # 캐시 TTL 5분 — 첫 로드는 무거우나 후속 로드는 즉시. 매출 집계는 1분 단위
    # 변화 의미 없고, 매 새로고침마다 풀스캔 도는 게 더 큰 비용.
    await cache.set(_cache_key, result, ttl=300)
    return result


@router.get("/search", response_model=list[SambaOrder])
async def search_orders(
    q: str = Query(..., min_length=1),
    session: AsyncSession = Depends(get_read_session_dependency),
):
    svc = _read_service(session)
    return await svc.search_orders(q)


@router.get("/by-date-range-paged", response_model=PaginatedOrdersResponse)
async def list_orders_by_date_range_paged(
    start: str = Query(..., description="start date YYYY-MM-DD"),
    end: str = Query(..., description="end date YYYY-MM-DD"),
    skip: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=500),
    market_filter: str = Query(""),
    site_filter: str = Query(""),
    account_filter: str = Query(""),
    market_status: str = Query(""),
    status_filter: str = Query(""),
    input_filter: str = Query(""),
    invoice_filter: str = Query(""),
    registration_filter: str = Query(""),
    search_text: str = Query(""),
    search_category: str = Query("customer"),
    sort_by: str = Query("date_desc"),
    session: AsyncSession = Depends(get_read_session_dependency),
    tenant_id: Optional[str] = Depends(get_optional_tenant_id),
):
    from backend.utils import kst_date_range_to_utc

    start_dt, end_dt = kst_date_range_to_utc(start, end)
    filters = await _build_order_filters(
        session,
        tenant_id,
        market_filter=market_filter,
        site_filter=site_filter,
        account_filter=account_filter,
        market_status=market_status,
        status_filter=status_filter,
        input_filter=input_filter,
        invoice_filter=invoice_filter,
        registration_filter=registration_filter,
        search_text=search_text,
        search_category=search_category,
    )
    if status_filter == "cancel_alert":
        date_extra: list[Any] = []
    else:
        date_extra = [
            SambaOrder.paid_at != None,  # noqa: E711
            SambaOrder.paid_at >= start_dt,
            SambaOrder.paid_at <= end_dt,
        ]
    return await _run_paginated_order_query(
        session,
        filters,
        skip=skip,
        limit=limit,
        sort_by=sort_by,
        extra_filters=date_extra,
    )


@router.get("/by-collected-product-paged", response_model=PaginatedOrdersResponse)
async def list_orders_by_collected_product_paged(
    collected_product_id: str = Query(..., description="collected product ID"),
    skip: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=500),
    market_filter: str = Query(""),
    site_filter: str = Query(""),
    account_filter: str = Query(""),
    market_status: str = Query(""),
    status_filter: str = Query(""),
    input_filter: str = Query(""),
    invoice_filter: str = Query(""),
    registration_filter: str = Query(""),
    search_text: str = Query(""),
    search_category: str = Query("customer"),
    sort_by: str = Query("date_desc"),
    session: AsyncSession = Depends(get_read_session_dependency),
    tenant_id: Optional[str] = Depends(get_optional_tenant_id),
):
    filters = await _build_order_filters(
        session,
        tenant_id,
        market_filter=market_filter,
        site_filter=site_filter,
        account_filter=account_filter,
        market_status=market_status,
        status_filter=status_filter,
        input_filter=input_filter,
        invoice_filter=invoice_filter,
        registration_filter=registration_filter,
        search_text=search_text,
        search_category=search_category,
    )
    return await _run_paginated_order_query(
        session,
        filters,
        skip=skip,
        limit=limit,
        sort_by=sort_by,
        extra_filters=[SambaOrder.collected_product_id == collected_product_id],
    )


class ExcelExportRequest(BaseModel):
    """엑셀 다운로드 요청 — 선택ID 우선, 없으면 필터 전체.

    format 분기:
      - 'ub1' (default): 소싱처 발주 양식 — 마켓주문일자/마켓명/.../옵션1 (10컬럼)
      - 'lotte': 롯데택배 양식 (수령자명/전화번호/우편번호/주소/상품명/수량/배송메세지)
                 플레이오토 다운로드 양식과 동일 헤더·순서.
    """

    order_ids: Optional[list[str]] = None
    start: Optional[str] = None
    end: Optional[str] = None
    market_filter: str = ""
    site_filter: str = ""
    account_filter: str = ""
    market_status: str = ""
    status_filter: str = ""
    input_filter: str = ""
    invoice_filter: str = ""
    registration_filter: str = ""
    search_text: str = ""
    search_category: str = "customer"
    sort_by: str = "date_desc"
    format: str = "ub1"  # 'ub1' | 'lotte'


@router.post("/excel-export")
async def export_orders_excel(
    payload: ExcelExportRequest,
    session: AsyncSession = Depends(get_read_session_dependency),
    tenant_id: Optional[str] = Depends(get_optional_tenant_id),
):
    """주문 엑셀 다운로드 — 사진 UB1 포맷 (10컬럼).

    - 체크박스 선택 ID 우선 (`order_ids`), 없으면 필터 전체.
    - 필터 모드는 50,000건 상한, 초과 시 400.
    """
    from datetime import datetime, timedelta, timezone
    from io import BytesIO
    from urllib.parse import quote

    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from sqlalchemy import select

    from backend.utils import kst_date_range_to_utc

    MAX_FILTER_ROWS = 50_000

    if payload.order_ids:
        ids = [oid for oid in payload.order_ids if oid]
        if not ids:
            raise HTTPException(status_code=400, detail="선택된 주문이 없습니다.")
        stmt = select(SambaOrder).where(SambaOrder.id.in_(ids))
        if tenant_id is not None:
            from sqlalchemy import or_

            stmt = stmt.where(
                or_(
                    SambaOrder.tenant_id == tenant_id,
                    SambaOrder.tenant_id == None,  # noqa: E711
                )
            )
        stmt = stmt.order_by(*_build_order_sort(payload.sort_by))
        rows = list((await session.execute(stmt)).scalars().all())
    else:
        if not payload.start or not payload.end:
            raise HTTPException(
                status_code=400, detail="start/end 또는 order_ids 가 필요합니다."
            )
        start_dt, end_dt = kst_date_range_to_utc(payload.start, payload.end)
        filters = await _build_order_filters(
            session,
            tenant_id,
            market_filter=payload.market_filter,
            site_filter=payload.site_filter,
            account_filter=payload.account_filter,
            market_status=payload.market_status,
            status_filter=payload.status_filter,
            input_filter=payload.input_filter,
            invoice_filter=payload.invoice_filter,
            registration_filter=payload.registration_filter,
            search_text=payload.search_text,
            search_category=payload.search_category,
        )
        extra = [
            SambaOrder.paid_at != None,  # noqa: E711
            SambaOrder.paid_at >= start_dt,
            SambaOrder.paid_at <= end_dt,
        ]
        stmt = (
            select(SambaOrder)
            .where(*filters, *extra)
            .order_by(*_build_order_sort(payload.sort_by))
            .limit(MAX_FILTER_ROWS + 1)
        )
        rows = list((await session.execute(stmt)).scalars().all())
        if len(rows) > MAX_FILTER_ROWS:
            raise HTTPException(
                status_code=400,
                detail=f"결과 {len(rows):,}건이 상한 {MAX_FILTER_ROWS:,}건을 초과했습니다. 필터를 좁혀주세요.",
            )

    # 채널명 괄호 파싱: "11번가(sogyung)" -> ("11번가", "sogyung")
    paren_re = re.compile(r"^(.*?)\s*\(([^()]+)\)\s*$")

    def split_channel(name: Optional[str]) -> tuple[str, str]:
        if not name:
            return ("", "")
        m = paren_re.match(name)
        if not m:
            return (name, "")
        return (m.group(1).strip(), m.group(2).strip())

    KST = timezone(timedelta(hours=9))

    wb = Workbook()
    ws = wb.active

    bold = Font(bold=True)

    fmt = (payload.format or "ub1").strip().lower()

    if fmt == "lotte":
        # 롯데택배 송장 발송용 양식 — 플레이오토 다운 양식과 동일 헤더·순서 (2026-06-08 사용자 캡처).
        # 합포장 주소/연락처/배송메시지 그대로 1행 1주문. 헤더는 굵게만 (배경색 X — 양식 원본 따라감).
        today_kst = datetime.now(timezone.utc).astimezone(KST).strftime("%Y-%m-%d")
        ws.title = today_kst
        headers = [
            "수령자명",
            "수령자전화번호",
            "배송지우편번호",
            "배송지주소",
            "상품명",
            "주문수량",
            "배송메세지",
        ]
        ws.append(headers)
        for c in ws[1]:
            c.font = bold

        def _join_addr(addr: Optional[str], detail: Optional[str]) -> str:
            a = (addr or "").strip()
            d = (detail or "").strip()
            if a and d:
                return f"{a} {d}"
            return a or d

        for o in rows:
            ws.append(
                [
                    o.customer_name or "",
                    o.customer_phone or "",
                    o.customer_postal_code or "",
                    _join_addr(o.customer_address, o.customer_address_detail),
                    o.product_name or "",
                    int(o.quantity or 0),
                    o.customer_note or "",
                ]
            )

        # 컬럼 너비 — 캡처에서 실측 유사 비율
        widths = [10, 18, 12, 50, 50, 8, 30]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        if payload.order_ids:
            fname = f"롯데택배_선택{len(rows)}건.xlsx"
        else:
            fname = f"롯데택배_{payload.start}_{payload.end}.xlsx"
        quoted = quote(fname)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=\"lotte_courier.xlsx\"; filename*=UTF-8''{quoted}"
            },
        )

    # ── 기본: UB1 소싱처 발주 양식 (10컬럼) ──
    ws.title = "orders"
    headers = [
        "마켓주문일자",
        "마켓명",
        "마켓아이디",
        "수령인명",
        "마켓상품명",
        "마켓주문번호",
        "구매가격",
        "국제운송료",
        "사이트주문번호",
        "옵션1",
    ]
    ws.append(headers)
    yellow = PatternFill("solid", fgColor="FFFF00")
    center = Alignment(horizontal="center")
    for c in ws[1]:
        c.font = bold
        c.fill = yellow
        c.alignment = center

    # 마켓별 마켓주문번호 컬럼 보정:
    #   - 롯데ON: order_number 형식 "{odNo}_{odSeq}" → 끝의 "_숫자" suffix 제거하여 odNo만 노출
    #   - 쿠팡: order_number 에 shipmentBoxId(=배송번호)가 들어 있으나, 발주서의 "주문번호"는 orderId.
    #          orderId 는 shipment_id 컬럼에 저장되어 있으므로 shipment_id 를 우선 사용.
    lotteon_suffix_re = re.compile(r"_\d+$")

    def excel_market_order_no(o: SambaOrder, market_name: str) -> str:
        raw = (o.order_number or "").strip()
        if market_name == "쿠팡":
            return (o.shipment_id or raw or "").strip()
        if market_name == "롯데ON":
            return lotteon_suffix_re.sub("", raw)
        return raw

    for o in rows:
        market_name, market_account = split_channel(o.channel_name)
        paid_kst = ""
        if o.paid_at:
            paid = o.paid_at
            if paid.tzinfo is None:
                paid = paid.replace(tzinfo=timezone.utc)
            paid_kst = paid.astimezone(KST).strftime("%Y-%m-%d")
        ws.append(
            [
                paid_kst,
                market_name,
                market_account,
                o.customer_name or "",
                o.product_name or "",
                excel_market_order_no(o, market_name),
                int(o.cost or 0),
                int(o.shipping_fee or 0),
                o.sourcing_order_number or "",
                o.product_option or "",
            ]
        )

    widths = [13, 10, 12, 10, 55, 22, 10, 10, 22, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    if payload.order_ids:
        fname = f"주문_선택{len(rows)}건.xlsx"
    else:
        fname = f"주문_{payload.start}_{payload.end}.xlsx"
    quoted = quote(fname)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=\"orders.xlsx\"; filename*=UTF-8''{quoted}"
        },
    )


@router.get("/analytics-aggregate")
async def analytics_aggregate(
    start: str = Query(..., description="시작일 YYYY-MM-DD"),
    end: str = Query(..., description="종료일 YYYY-MM-DD"),
    session: AsyncSession = Depends(get_read_session_dependency),
    tenant_id: Optional[str] = Depends(get_optional_tenant_id),
):
    """매출통계 페이지 전용 사전집계 엔드포인트.

    paid_at(KST) 일자 × channel_name × source_site × status 단위로
    sum(sale_price), count(*)를 미리 집계해서 반환한다.
    매출통계 페이지가 raw 주문 4천+건(6MB)을 통째 받아 클라이언트에서
    필터링하던 구조를 대체 — 페이로드 99% 축소, 무음 실패 회귀 방지.
    """
    from sqlalchemy import select as sa_select, func as sa_func, or_
    from backend.domain.samba.account.model import SambaMarketAccount
    from backend.utils import kst_date_range_to_utc

    start_dt, end_dt = kst_date_range_to_utc(start, end)

    # paid_at이 있으면 paid_at 기준, 없으면 created_at 기준으로 집계
    # (쿠팡/롯데홈쇼핑 등 paid_at 미설정 주문 누락 방지)
    effective_at = sa_func.coalesce(SambaOrder.paid_at, SambaOrder.created_at)
    kst_date = sa_func.date(sa_func.timezone("Asia/Seoul", effective_at))

    # 마켓 그룹키 — samba_market_account.market_name(G마켓/옥션/11번가/...) 우선,
    # 매칭 안 되면 channel_name 사용. channel_name(=계정 닉네임 "가디(...)")으로
    # 그룹화하면 마켓별 통계에 계정 닉네임이 노출되는 문제 발생(2026-05-26).
    market_key = sa_func.coalesce(
        SambaMarketAccount.market_name, SambaOrder.channel_name
    )

    stmt = (
        sa_select(
            kst_date.label("date"),
            market_key.label("channel_name"),
            SambaOrder.source_site,
            SambaOrder.status,
            sa_func.coalesce(sa_func.sum(SambaOrder.sale_price), 0).label("sales"),
            sa_func.count().label("orders"),
            sa_func.coalesce(sa_func.sum(SambaOrder.profit), 0).label("profit"),
            sa_func.coalesce(
                sa_func.sum(SambaOrder.cost * SambaOrder.quantity), 0
            ).label("cost"),
        )
        .select_from(SambaOrder)
        .outerjoin(SambaMarketAccount, SambaMarketAccount.id == SambaOrder.channel_id)
        .where(
            effective_at >= start_dt,
            effective_at <= end_dt,
        )
        .group_by(
            kst_date,
            market_key,
            SambaOrder.source_site,
            SambaOrder.status,
        )
    )
    if tenant_id is not None:
        stmt = stmt.where(
            or_(
                SambaOrder.tenant_id == tenant_id,
                SambaOrder.tenant_id == None,  # noqa: E711
            )
        )
    result = await session.execute(stmt)
    rows = [
        {
            "date": str(r.date),
            "channel_name": r.channel_name or "",
            "source_site": r.source_site or "",
            "status": r.status or "",
            "sales": float(r.sales or 0),
            "orders": int(r.orders or 0),
            "profit": float(r.profit or 0),
            "cost": float(r.cost or 0),
        }
        for r in result.all()
    ]
    return {"rows": rows}


@router.get("/by-date-range", response_model=list[SambaOrder])
async def list_orders_by_date_range(
    start: str = Query(..., description="시작일 YYYY-MM-DD"),
    end: str = Query(..., description="종료일 YYYY-MM-DD"),
    session: AsyncSession = Depends(get_read_session_dependency),
    tenant_id: Optional[str] = Depends(get_optional_tenant_id),
):
    """기간별 주문 조회 — paid_at(고객결제일) 기준, 제한 없이 전체 반환."""
    from sqlalchemy import select as sa_select, or_
    from backend.utils import kst_date_range_to_utc

    start_dt, end_dt = kst_date_range_to_utc(start, end)

    stmt = (
        sa_select(SambaOrder)
        .where(
            SambaOrder.paid_at != None,  # noqa: E711
            SambaOrder.paid_at >= start_dt,
            SambaOrder.paid_at <= end_dt,
        )
        .order_by(SambaOrder.paid_at.desc())
    )
    if tenant_id is not None:
        stmt = stmt.where(
            or_(
                SambaOrder.tenant_id == tenant_id,
                SambaOrder.tenant_id == None,  # noqa: E711
            )
        )
    result = await session.execute(stmt)
    return list(result.scalars().all())


@router.get("/by-collected-product", response_model=list[SambaOrder])
async def list_orders_by_collected_product(
    collected_product_id: str = Query(..., description="수집상품 ID (cp_ULID)"),
    session: AsyncSession = Depends(get_read_session_dependency),
    tenant_id: Optional[str] = Depends(get_optional_tenant_id),
):
    """수집상품 ID로 해당 상품의 전체 주문 이력 조회."""
    from sqlalchemy import select as sa_select, func as sa_func, or_

    date_col = sa_func.coalesce(SambaOrder.paid_at, SambaOrder.created_at)
    stmt = (
        sa_select(SambaOrder)
        .where(SambaOrder.collected_product_id == collected_product_id)
        .order_by(date_col.desc())
    )
    if tenant_id is not None:
        stmt = stmt.where(
            or_(
                SambaOrder.tenant_id == tenant_id,
                SambaOrder.tenant_id == None,  # noqa: E711
            )
        )
    result = await session.execute(stmt)
    return list(result.scalars().all())


# 한국어 택배사명 → 딜리버리트래커 carrier ID 매핑
SHIPPING_COMPANY_TO_CARRIER_ID: dict[str, str] = {
    "CJ대한통운": "kr.cjlogistics",
    "한진택배": "kr.hanjin",
    "롯데택배": "kr.lotte",
    "로젠택배": "kr.logen",
    "우체국택배": "kr.epost",
    "경동택배": "kr.kdexp",
    "대신택배": "kr.daesin",
    "일양로지스": "kr.ilyanglogis",
    "편의점택배": "kr.cvsnet",
    "합동택배": "kr.hdexp",
    "쿠팡택배": "kr.coupangls",
    "딜리박스": "kr.dilibox",
    "DHL": "de.dhl",
}


@router.get("/tracking")
async def get_tracking(
    carrier: str = Query(..., description="택배사 한국어명 (예: CJ대한통운)"),
    invoice: str = Query(..., description="운송장번호"),
):
    """딜리버리트래커 v1 API를 프록시하여 통합 배송조회 결과를 반환."""
    import httpx

    carrier_id = SHIPPING_COMPANY_TO_CARRIER_ID.get(carrier)
    if not carrier_id:
        raise HTTPException(400, f"지원하지 않는 택배사: {carrier}")

    invoice_clean = re.sub(r"[^0-9A-Za-z]", "", invoice or "")
    if not invoice_clean:
        raise HTTPException(400, "유효하지 않은 송장번호입니다")

    url = f"https://apis.tracker.delivery/carriers/{carrier_id}/tracks/{invoice_clean}"
    try:
        async with httpx.AsyncClient(timeout=10) as hc:
            resp = await hc.get(url)
    except httpx.HTTPError as e:
        logger.warning(
            "[tracking] 외부 API 통신 실패 %s/%s: %s", carrier, invoice_clean, e
        )
        raise HTTPException(502, "택배 조회 서비스에 연결할 수 없습니다")

    if resp.status_code == 404:
        raise HTTPException(
            404, "조회 결과가 없습니다 (송장번호/택배사를 확인해주세요)"
        )
    if resp.status_code >= 400:
        logger.warning(
            "[tracking] 비정상 응답 %s/%s status=%s body=%s",
            carrier,
            invoice_clean,
            resp.status_code,
            resp.text[:200],
        )
        raise HTTPException(502, "택배 조회 결과를 불러오지 못했습니다")

    try:
        data = resp.json()
    except ValueError:
        raise HTTPException(502, "택배 조회 응답 형식 오류")

    progresses = data.get("progresses") or []
    return {
        "carrier_name": carrier,
        "carrier_id": carrier_id,
        "invoice": invoice_clean,
        "from_name": (data.get("from") or {}).get("name"),
        "to_name": (data.get("to") or {}).get("name"),
        "state": (data.get("state") or {}).get("text"),
        "events": [
            {
                "time": p.get("time"),
                "status": (p.get("status") or {}).get("text"),
                "status_code": (p.get("status") or {}).get("id"),
                "location": (p.get("location") or {}).get("name"),
                "description": p.get("description"),
            }
            for p in progresses
        ],
    }


@router.get("/find-by-number")
async def find_by_order_number(
    order_number: str = Query(...),
    session: AsyncSession = Depends(get_read_session_dependency),
    tenant_id: Optional[str] = Depends(get_optional_tenant_id),
):
    """상품주문번호로 주문 조회."""
    svc = _read_service(session)
    order = await svc.repo.find_by_async(order_number=order_number)
    if not order:
        return None
    # 테넌트 소유권 검증
    if tenant_id is not None and order.tenant_id != tenant_id:
        raise HTTPException(403, "해당 주문에 대한 권한이 없습니다")
    return {"id": order.id, "order_number": order.order_number}


@router.post("/{order_id}/sync-tracking")
async def sync_order_tracking(order_id: str, force: bool = False) -> dict:
    """소싱처에서 운송장 추출 잡을 큐에 적재 (단건).

    force=True 면 이미 송장이 있어도 다시 큐잉.
    """
    from backend.domain.samba.tracking_sync.service import enqueue_for_order

    return await enqueue_for_order(order_id, force=force)


@router.post("/sync-tracking/bulk")
async def sync_order_tracking_bulk(
    limit: int = Query(500, ge=1, le=1000),
    days: int = Query(7, ge=1, le=90),
    force: bool = Query(True),
    owner_device: str = Query(
        "",
        description="이 송장수집을 트리거한 PC의 데몬 device_id (전담 송장 PC 지정용)",
    ),
    tenant_id: Optional[str] = Depends(get_optional_tenant_id),
) -> dict:
    """미발송 주문 일괄 송장 추출 큐잉 — 최근 N일 + 소싱처 주문번호 있음 + 송장 미입력.

    owner_device: 전달 시 그 데몬을 '전담 송장 PC'로 저장(samba_settings)하고 잡 owner 로
    지정 → 그 PC만 송장 수신. 여러 PC가 같은 SSG 계정 동시 로그인 → 멀티PC 보안잠금 차단.
    이후 자동 송장수집(스케줄러)도 저장된 전담 PC를 사용.
    """
    from backend.domain.samba.tracking_sync.service import enqueue_pending_orders

    _owner = (owner_device or "").strip()
    # [2026-06-05 송장 확장앱 복구] 송장은 확장앱이 처리하므로 owner = 확장앱 device_id(UUID).
    # 데몬 device_id(samba-daemon-)만 허용하던 구버전 게이트 제거 — 그 게이트가 확장앱 UUID 를
    # ''로 떨궈 owner 미지정 적재 → 아무 PC나 잡 가로채 계정 왔다갔다·WRONG_ACCOUNT 유발.
    # 빈값이면 설정값 사용(enqueue 내부 해석). 값 있으면 그 PC 를 전담으로 저장 + 잡 owner 지정.
    # [#518] 데몬 device_id가 오면 '' 강등 — 데몬은 tracking 잡 dequeue 불가 → 데드존.
    if _owner.startswith("samba-daemon-"):
        _owner = ""
    if _owner:
        try:
            from backend.api.v1.routers.samba.proxy._helpers import _set_setting
            from backend.db.orm import get_write_session

            async with get_write_session() as _s:
                await _set_setting(_s, "tracking_owner_device", _owner)
        except Exception:
            pass

    # [방어] 송장수집이 락/슬로우쿼리로 hang → 프론트 "Failed to fetch"로 원인 숨던 문제.
    # 무한 hang·좀비 방지의 근본 상한은 enqueue_pending_orders 내부 DB 레벨 타임아웃
    # (lock/statement_timeout)에서 보장한다 — 초과 시 Postgres가 abort → 트랜잭션 rollback +
    # 연결 반납(리소스 정리 보장). asyncio.wait_for 는 취소돼도 DB 리소스 해제를 보장하지 못해
    # (좀비 트랜잭션 위험) 사용하지 않는다. 여기서는 그 예외를 잡아 실제 원인을 정상 응답과
    # 동일한 dict 구조로 반환한다(재시도 가능 여부는 메시지로 명시).
    from sqlalchemy.exc import DBAPIError

    # 재시도 가능 SQLSTATE만 좁혀서 처리: 57014=statement_timeout(query canceled),
    # 55P03=lock_timeout(lock not available). 그 외 DB 오류(문법/제약/연결 등)는 재시도 대상 아님.
    _RETRYABLE_SQLSTATES = ("57014", "55P03")
    try:
        return await enqueue_pending_orders(
            tenant_id=tenant_id,
            limit=limit,
            days=days,
            force=force,
            owner_device_id=_owner or None,
        )
    except DBAPIError as _db_exc:
        _orig = getattr(_db_exc, "orig", None)
        _sqlstate = getattr(_orig, "sqlstate", None) or getattr(_orig, "pgcode", None)
        if _sqlstate in _RETRYABLE_SQLSTATES:
            # DB 레벨 타임아웃(statement/lock) 초과 = 락 경합/슬로우쿼리. 컨텍스트 매니저가
            # 이미 rollback + 연결 반납 → 잠시 후 재시도 가능.
            logger.warning(
                f"[송장수집] DB 타임아웃/락(SQLSTATE {_sqlstate}) — 재시도 가능: "
                f"{_orig or _db_exc}"
            )
            return {
                "success": False,
                "queued": 0,
                "skipped": 0,
                "errors": [
                    "송장수집 DB 타임아웃(락 경합/슬로우쿼리) — 잠시 후 다시 시도하세요."
                ],
                "job_ids": [],
            }
        # 재시도 대상 아닌 DB 오류 — 원인 로깅 후 일반 실패로 분리(타임아웃으로 오분류 금지).
        logger.exception("[송장수집] DB 오류(재시도 대상 아님)")
        return {
            "success": False,
            "queued": 0,
            "skipped": 0,
            "errors": [
                f"송장수집 DB 오류(SQLSTATE {_sqlstate or '?'}): {_orig or _db_exc}"
            ],
            "job_ids": [],
        }
    except Exception as _exc:  # noqa: BLE001
        logger.exception("[송장수집] 예기치 못한 실패")
        return {
            "success": False,
            "queued": 0,
            "skipped": 0,
            "errors": [f"송장수집 실패(재시도 전 서버 로그 확인 권장): {_exc}"],
            "job_ids": [],
        }


@router.get("/tracking-sync/owner-device")
async def get_tracking_owner_device() -> dict:
    """현재 전담 송장 PC(데몬 device_id) 조회. ''이면 전담 미지정(모든 PC)."""
    from backend.api.v1.routers.samba.proxy._helpers import _get_setting
    from backend.db.orm import get_write_session

    async with get_write_session() as s:
        v = await _get_setting(s, "tracking_owner_device")
    return {"tracking_owner_device": str(v).strip() if v else ""}


@router.post("/tracking-sync/owner-device")
async def set_tracking_owner_device(device: str = Query("")) -> dict:
    """전담 송장 PC 지정. device='' 면 해제(모든 PC). 확장앱 device_id(UUID)만 허용.

    지정 시 송장 잡 owner_device_id 가 그 PC로 박혀 그 PC만 수신 →
    여러 PC가 같은 SSG 계정 동시 로그인하는 멀티PC 보안잠금 차단.
    [#518] 데몬 device_id(samba-daemon-) 거부 — 데몬은 tracking 잡 dequeue 불가 → 데드존.
    """
    from backend.api.v1.routers.samba.proxy._helpers import _set_setting
    from backend.db.orm import get_write_session

    dev = (device or "").strip()
    if dev.startswith("samba-daemon-"):
        raise HTTPException(
            400, "데몬 device_id는 송장 owner로 지정 불가 (확장앱 전담)"
        )
    async with get_write_session() as s:
        await _set_setting(s, "tracking_owner_device", dev)
    return {"success": True, "tracking_owner_device": dev}


@router.post("/tracking-sync/dispatch/bulk")
async def dispatch_tracking_bulk(dry_run: bool = False) -> dict:
    """SCRAPED + DISPATCH_FAILED 잡 전부 일괄 마켓 전송 (재시도 포함)."""
    from backend.domain.samba.tracking_sync.service import dispatch_pending_to_market

    return await dispatch_pending_to_market(dry_run=dry_run)


@router.post("/tracking-sync/retry-failed")
async def retry_failed_tracking_jobs(
    days: int = 7,
    tenant_id: Optional[str] = Depends(get_optional_tenant_id),
) -> dict:
    """WRONG_ACCOUNT / FAILED / DISPATCH_FAILED 잡들을 자동 재큐잉.

    송장수집이 실패한 주문들만 모아서 다시 자동 로그인 + 송장 추출 시도.
    송장 미입력 주문 전체 재큐잉(sync-tracking/bulk)과 다른 점:
    - 미발송으로 끝난 잡은 제외 (실패한 것만)
    - 한 번에 빠르게 retry 트리거 가능
    """
    from backend.domain.samba.tracking_sync.service import retry_failed_jobs

    return await retry_failed_jobs(tenant_id=tenant_id, days=days)


@router.post("/tracking-sync/{job_id}/dispatch")
async def dispatch_tracking_to_market(job_id: str, dry_run: bool = False) -> dict:
    """추출 완료된(SCRAPED) 잡의 운송장을 마켓으로 push.

    dry_run=True (기본): 페이로드만 로그. False면 실제 마켓 API 호출.
    """
    from backend.domain.samba.tracking_sync.service import dispatch_to_market

    return await dispatch_to_market(job_id, dry_run=dry_run)


@router.get("/tracking-sync/recent")
async def list_recent_tracking_sync_jobs(
    limit: int = Query(50, ge=1, le=200),
    tenant_id: Optional[str] = Depends(get_optional_tenant_id),
) -> dict:
    """최근 송장 자동전송 잡 목록 + 상태 카운트.

    프론트가 일괄 송장수집 후 폴링해서 진행상황 보여주는 용도.
    SambaOrder (상품주문번호/고객명) + SambaSourcingAccount (소싱처 계정 라벨) LEFT JOIN.
    """
    from sqlalchemy import select
    from sqlalchemy.orm import aliased
    from backend.db.orm import get_read_session
    from backend.domain.samba.order.model import (
        EXCLUDED_ORDER_STATUSES,
        SHIPPED_SHIPPING_STATUS_KEYWORDS,
        SambaOrder,
    )
    from backend.domain.samba.sourcing_account.model import SambaSourcingAccount
    from backend.domain.samba.tracking_sync.model import SambaTrackingSyncJob

    def _is_excluded(order_status, shipping_status) -> bool:
        """페이지 필터 '취소/반품/교환 제외 + 배송중/배송완료 제외' 와 동일 기준."""
        if order_status and order_status in EXCLUDED_ORDER_STATUSES:
            return True
        if shipping_status and any(
            kw in shipping_status for kw in SHIPPED_SHIPPING_STATUS_KEYWORDS
        ):
            return True
        return False

    async with get_read_session() as session:
        O = aliased(SambaOrder)
        A = aliased(SambaSourcingAccount)
        # 잡 + 주문 메타를 한 번에 가져와 Python에서 dedup → 카운트/리스트 일관 처리
        # 큐잉 필터(enqueue_pending_orders)와 100% 동일 조건 적용:
        #   2) sourcing_order_number 있음
        #   3) source_site 있음
        #   4) 최근 7일 (created_at >= now-7d)
        #   7) action_tag 에 'kkadaegi' 토큰 없음
        # 1/5/6 (송장 미입력 / 상태 제외 / 배송중·완료 제외) 은 Python loop 에서 처리.
        from datetime import timedelta, timezone
        from sqlalchemy import and_, func, not_, or_

        # KST 캘린더 7일 (오늘 포함 -6일) + paid_at(폴백 created_at) 기준
        _KST = timezone(timedelta(hours=9))
        _today_kst = datetime.now(_KST).replace(
            hour=0, minute=0, second=0, microsecond=0
        )
        _since = (_today_kst - timedelta(days=6)).astimezone(timezone.utc)
        _until = (_today_kst + timedelta(days=1)).astimezone(timezone.utc)
        action_tag_expr = func.concat(",", func.coalesce(O.action_tag, ""), ",")
        date_col = func.coalesce(O.paid_at, O.created_at)
        base_stmt = (
            select(
                SambaTrackingSyncJob,
                O.order_number,
                O.customer_name,
                O.channel_name,
                O.status,
                O.shipping_status,
                A.account_label,
                O.tracking_number,
                O.paid_at,
                O.action_tag,
            )
            .join(O, O.id == SambaTrackingSyncJob.order_id, isouter=True)
            .join(A, A.id == SambaTrackingSyncJob.sourcing_account_id, isouter=True)
            .where(
                and_(
                    O.sourcing_order_number.is_not(None),
                    O.sourcing_order_number != "",
                    # source_site 비어있어도 source_url / collected_product 로 추론 가능하면 포함
                    or_(
                        and_(O.source_site.is_not(None), O.source_site != ""),
                        and_(O.source_url.is_not(None), O.source_url != ""),
                        O.collected_product_id.is_not(None),
                    ),
                    date_col >= _since,
                    date_col < _until,
                    not_(action_tag_expr.like("%,kkadaegi,%")),
                    # 송장 채워졌어도 잡 자체는 표시 (수집 결과 확인용).
                    # 큐 적재 단계에서만 송장 있는 주문 제외 — enqueue_for_order 가 처리.
                )
            )
            .order_by(SambaTrackingSyncJob.updated_at.desc())
            .limit(limit * 10)
        )
        if tenant_id:
            base_stmt = base_stmt.where(SambaTrackingSyncJob.tenant_id == tenant_id)
        raw_rows = (await session.execute(base_stmt)).all()

        # order_id별 최신 1건만 선별 + 페이지 필터와 동일 기준 제외 +
        # 이미 송장 입력된 주문은 처리 대상 아니므로 제외 (모달 = "처리 필요" 잡만 표시)
        seen_order_ids: set[str] = set()
        result_rows = []
        counts: dict[str, int] = {}
        for row in raw_rows:
            j = row[0]
            order_status = row[4]
            shipping_status = row[5]
            order_tracking_number = row[7]
            if j.order_id in seen_order_ids:
                continue
            seen_order_ids.add(j.order_id)
            if _is_excluded(order_status, shipping_status):
                continue
            # 송장 채워진 주문은 모달 대상 아님 — "송장수집 = 송장 미입력건만 처리" 정책.
            # 외부 수동입력/이전 수집완료 무관하게 송장 있으면 숨김.
            if order_tracking_number:
                continue
            counts[j.status] = counts.get(j.status, 0) + 1
            if len(result_rows) < limit:
                result_rows.append(row)

    return {
        "counts": counts,
        "recent": [
            {
                "id": j.id,
                "orderId": j.order_id,
                "orderNumber": order_number or "",
                "customerName": customer_name or "",
                "channelName": channel_name or "",
                "site": j.sourcing_site,
                "sourcingOrderNumber": j.sourcing_order_number,
                "sourcingAccountLabel": account_label or "",
                "status": j.status,
                "courier": j.scraped_courier,
                "tracking": j.scraped_tracking,
                "lastError": j.last_error,
                "attempts": j.attempts,
                "updatedAt": j.updated_at.isoformat() if j.updated_at else None,
                "paidAt": paid_at.isoformat() if paid_at else None,
                "actionTag": action_tag or "",
            }
            for j, order_number, customer_name, channel_name, _os, _ss, account_label, _otn, paid_at, action_tag in result_rows
        ],
    }


@router.post("/tracking-sync/by-ids")
async def list_tracking_sync_jobs_by_ids(body: dict) -> dict:
    """송장수집 배치에 속한 잡들만 조회 — 모달 "이번 배치 고정" 용도.

    프론트가 일괄 송장수집 직후 받은 job_ids 를 그대로 전달.
    송장 채워진 행도 응답에 포함(상태 변화 추적용)하고, 순서는 paid_at ASC 로 고정.
    """
    from sqlalchemy import select
    from sqlalchemy.orm import aliased
    from backend.db.orm import get_read_session
    from backend.domain.samba.order.model import SambaOrder
    from backend.domain.samba.sourcing_account.model import SambaSourcingAccount
    from backend.domain.samba.tracking_sync.model import SambaTrackingSyncJob

    raw_ids = body.get("job_ids") or []
    if not isinstance(raw_ids, list):
        raise HTTPException(400, "job_ids 는 배열이어야 합니다")
    job_ids: list[str] = [str(x) for x in raw_ids if x]
    if not job_ids:
        return {"counts": {}, "recent": []}
    if len(job_ids) > 1000:
        job_ids = job_ids[:1000]

    async with get_read_session() as session:
        from sqlalchemy import func

        O = aliased(SambaOrder)
        A = aliased(SambaSourcingAccount)
        date_col = func.coalesce(O.paid_at, O.created_at)
        stmt = (
            select(
                SambaTrackingSyncJob,
                O.order_number,
                O.customer_name,
                O.channel_name,
                A.account_label,
                O.paid_at,
                O.action_tag,
            )
            .join(O, O.id == SambaTrackingSyncJob.order_id, isouter=True)
            .join(A, A.id == SambaTrackingSyncJob.sourcing_account_id, isouter=True)
            .where(SambaTrackingSyncJob.id.in_(job_ids))
            # 모달 리스트 = 처리 순서. 소싱처(MUSINSA → LOTTEON → SSG...) → 계정(병기 → 성희 → 귀옥...)
            # → 결제일 순으로 정렬. 같은 계정 잡이 연속 표시되고 1번부터 순서대로 처리되어
            # 자동 로그인 swap 횟수 = 계정 수로 최소화.
            .order_by(
                SambaTrackingSyncJob.sourcing_site.asc().nulls_last(),
                SambaTrackingSyncJob.sourcing_account_id.asc().nulls_last(),
                date_col.asc(),
            )
        )
        raw_rows = (await session.execute(stmt)).all()
        # 세션 종료 후 밖에서 컬럼 접근 — detach 방지 (#597)
        session.expunge_all()

    counts: dict[str, int] = {}
    items = []
    for row in raw_rows:
        j = row[0]
        order_number = row[1]
        customer_name = row[2]
        channel_name = row[3]
        account_label = row[4]
        paid_at = row[5]
        action_tag = row[6]
        counts[j.status] = counts.get(j.status, 0) + 1
        items.append(
            {
                "id": j.id,
                "orderId": j.order_id,
                "orderNumber": order_number or "",
                "customerName": customer_name or "",
                "channelName": channel_name or "",
                "site": j.sourcing_site,
                "sourcingOrderNumber": j.sourcing_order_number,
                "sourcingAccountLabel": account_label or "",
                "status": j.status,
                "courier": j.scraped_courier,
                "tracking": j.scraped_tracking,
                "lastError": j.last_error,
                "attempts": j.attempts,
                "updatedAt": j.updated_at.isoformat() if j.updated_at else None,
                "paidAt": paid_at.isoformat() if paid_at else None,
                "actionTag": action_tag or "",
            }
        )

    return {"counts": counts, "recent": items}


@router.post("/tracking-sync/cancel-batch")
async def cancel_tracking_sync_batch(body: dict) -> dict:
    """송장수집 모달 닫기 시 배치 잡 일괄 취소.

    PENDING/DISPATCHED 상태의 잡만 CANCELLED 로 전환. 이미 SCRAPED/SENT 등
    완료된 잡은 변경 안 함 (결과 보존). 확장앱이 in-flight 로 들고 있는 잡은
    apply_tracking_result 진입 시 상태가 CANCELLED 면 결과 폐기.
    """
    from sqlalchemy import update
    from datetime import datetime, timezone
    from backend.db.orm import get_write_session
    from backend.domain.samba.tracking_sync.model import (
        SambaTrackingSyncJob,
        STATUS_PENDING,
        STATUS_DISPATCHED,
        STATUS_CANCELLED,
    )

    raw_ids = body.get("job_ids") or []
    if not isinstance(raw_ids, list):
        raise HTTPException(400, "job_ids 는 배열이어야 합니다")
    job_ids: list[str] = [str(x) for x in raw_ids if x]
    if not job_ids:
        return {"cancelled": 0}
    if len(job_ids) > 1000:
        job_ids = job_ids[:1000]

    async with get_write_session() as session:
        stmt = (
            update(SambaTrackingSyncJob)
            .where(
                SambaTrackingSyncJob.id.in_(job_ids),
                SambaTrackingSyncJob.status.in_([STATUS_PENDING, STATUS_DISPATCHED]),
            )
            .values(
                status=STATUS_CANCELLED,
                last_error="모달 닫기로 배치 취소",
                updated_at=datetime.now(timezone.utc),
            )
        )
        result = await session.execute(stmt)
        await session.commit()
        return {"cancelled": result.rowcount or 0}


@router.get("/cancel-alert-count")
async def get_cancel_alert_count(
    session: AsyncSession = Depends(get_read_session_dependency),
    tenant_id: Optional[str] = Depends(get_optional_tenant_id),
):
    """아직 처리 안 한 취소요청 건수 반환.

    인지 누락 사고 방지가 목적. 조건은 _build_cancel_alert_clause() 와 동일.
    응답에 귀책별 분리 카운트 포함 (#246 PR-6) — 운영자 우선순위 판단용.
    """
    from sqlalchemy import case, select, func
    from backend.domain.samba.order.model import SambaOrder as OrderModel

    base_where = _build_cancel_alert_clause()
    if tenant_id is not None:
        base_where = base_where & (OrderModel.tenant_id == tenant_id)

    customer_expr = func.sum(
        case(
            (
                func.upper(func.coalesce(OrderModel.cancel_fault_by, "")) == "CUSTOMER",
                1,
            ),
            else_=0,
        )
    )
    non_customer_expr = func.sum(
        case(
            (
                func.upper(func.coalesce(OrderModel.cancel_fault_by, "")).in_(
                    ("VENDOR", "COUPANG", "WMS")
                ),
                1,
            ),
            else_=0,
        )
    )
    unknown_expr = func.sum(
        case(
            (
                func.upper(func.coalesce(OrderModel.cancel_fault_by, "")).in_(
                    ("CUSTOMER", "VENDOR", "COUPANG", "WMS")
                ),
                0,
            ),
            else_=1,
        )
    )
    stmt = select(
        func.count(),
        customer_expr,
        non_customer_expr,
        unknown_expr,
    ).where(base_where)
    row = (await session.execute(stmt)).one()
    total, customer, non_customer, unknown = row
    return {
        "count": int(total or 0),
        "by_fault": {
            "customer": int(customer or 0),
            "non_customer": int(non_customer or 0),
            "unknown": int(unknown or 0),
        },
    }


@router.get("/alarm-settings")
async def get_alarm_settings(
    session: AsyncSession = Depends(get_read_session_dependency),
):
    """취소알람 수집 주기 및 영업시간 설정 조회."""
    from backend.api.v1.routers.samba.proxy import _get_setting

    data = await _get_setting(session, "cancel_alarm_settings") or {}
    return {
        "hour": data.get("hour", 0),
        "min": data.get("min", 5),
        "sleep_start": data.get("sleep_start", "23:00"),
        "sleep_end": data.get("sleep_end", "07:00"),
    }


@router.post("/alarm-settings")
async def save_alarm_settings(
    body: dict,
    session: AsyncSession = Depends(get_write_session_dependency),
):
    """취소알람 수집 주기 및 영업시간 설정 저장."""
    from backend.api.v1.routers.samba.proxy import _set_setting

    await _set_setting(
        session,
        "cancel_alarm_settings",
        {
            "hour": int(body.get("hour", 0)),
            "min": int(body.get("min", 5)),
            "sleep_start": body.get("sleep_start", "23:00"),
            "sleep_end": body.get("sleep_end", "07:00"),
        },
    )
    return {"ok": True}


@router.get("/coupang-auto-confirm")
async def get_coupang_auto_confirm(
    session: AsyncSession = Depends(get_read_session_dependency),
) -> dict:
    """쿠팡 자동 발주확인(ACCEPT→INSTRUCT) 토글 조회 (#246 PR-6).

    기본값 True (현재 동작 유지). 운영자가 OFF 시 sync 시 confirm_orders 호출 스킵.
    """
    from backend.api.v1.routers.samba.proxy import _get_setting

    raw = await _get_setting(session, "coupang_auto_confirm_orders")
    enabled = True
    if isinstance(raw, dict):
        v = raw.get("enabled")
        if isinstance(v, bool):
            enabled = v
    elif isinstance(raw, bool):
        enabled = raw
    return {"enabled": enabled}


@router.post("/coupang-auto-confirm")
async def set_coupang_auto_confirm(
    body: dict,
    session: AsyncSession = Depends(get_write_session_dependency),
) -> dict:
    """쿠팡 자동 발주확인 토글 저장 (#246 PR-6)."""
    from backend.api.v1.routers.samba.proxy import _set_setting

    enabled = bool(body.get("enabled", True))
    await _set_setting(session, "coupang_auto_confirm_orders", {"enabled": enabled})
    return {"ok": True, "enabled": enabled}


@router.get("/esm-auto-confirm")
async def get_esm_auto_confirm(
    session: AsyncSession = Depends(get_read_session_dependency),
) -> dict:
    """ESM(G마켓/옥션) 자동 발주확인(OrderCheck) 토글 조회 (#423).

    기본값 True. OFF 시 sync 에서 confirm_order 호출 스킵 → 수동 /confirm 사용.
    """
    from backend.api.v1.routers.samba.proxy import _get_setting

    raw = await _get_setting(session, "esm_auto_confirm_orders")
    enabled = True
    if isinstance(raw, dict):
        v = raw.get("enabled")
        if isinstance(v, bool):
            enabled = v
    elif isinstance(raw, bool):
        enabled = raw
    return {"enabled": enabled}


@router.post("/esm-auto-confirm")
async def set_esm_auto_confirm(
    body: dict,
    session: AsyncSession = Depends(get_write_session_dependency),
) -> dict:
    """ESM 자동 발주확인 토글 저장 (#423)."""
    from backend.api.v1.routers.samba.proxy import _set_setting

    enabled = bool(body.get("enabled", True))
    await _set_setting(session, "esm_auto_confirm_orders", {"enabled": enabled})
    return {"ok": True, "enabled": enabled}


@router.get("/auto-sync-interval")
async def get_auto_sync_interval(
    session: AsyncSession = Depends(get_read_session_dependency),
) -> dict:
    """주문 자동수집 인터벌 설정 조회 (분 단위, 0=OFF)."""
    from backend.api.v1.routers.samba.proxy import _get_setting

    val = await _get_setting(session, "order_auto_sync_interval_minutes")
    try:
        minutes = int(val) if val is not None else 0
    except (TypeError, ValueError):
        minutes = 0
    return {"interval_minutes": minutes}


@router.post("/auto-sync-interval")
async def set_auto_sync_interval(
    body: dict,
    session: AsyncSession = Depends(get_write_session_dependency),
) -> dict:
    """주문 자동수집 인터벌 설정 저장 (분 단위, 0 이하면 OFF)."""
    from backend.api.v1.routers.samba.proxy import _set_setting

    try:
        minutes = int(body.get("interval_minutes", 0))
    except (TypeError, ValueError):
        minutes = 0
    if minutes < 0:
        minutes = 0
    await _set_setting(session, "order_auto_sync_interval_minutes", minutes)
    return {"interval_minutes": minutes}


@router.get("/auto-sync-history")
async def get_auto_sync_history(
    limit: int = 2,
    session: AsyncSession = Depends(get_read_session_dependency),
) -> dict:
    """주문 자동실행(order_sync 잡) 최근 이력 N건 요약.

    프론트 '주문 자동실행' 섹션에서 최근 수집 결과를 표시하기 위함.
    """
    from sqlalchemy import text as _t

    limit = max(1, min(int(limit or 2), 10))
    rows = await session.execute(
        _t(
            "SELECT id, status, created_at, started_at, completed_at, result, error "
            "FROM samba_jobs WHERE job_type = 'order_sync' "
            "ORDER BY created_at DESC LIMIT :lim"
        ),
        {"lim": limit},
    )
    items: list[dict] = []
    for r in rows.fetchall():
        job_id, status, created_at, started_at, completed_at, result, error = r
        result_dict = result if isinstance(result, dict) else {}
        results_list = result_dict.get("results") or []
        per_market: list[dict] = []
        for it in results_list:
            if not isinstance(it, dict):
                continue
            per_market.append(
                {
                    "account": it.get("account", ""),
                    "status": it.get("status", ""),
                    "synced": int(it.get("synced") or 0),
                    "fetched": int(it.get("fetched") or 0),
                    "message": (it.get("message") or "")[:200],
                }
            )
        duration_sec: int | None = None
        if started_at and completed_at:
            duration_sec = int((completed_at - started_at).total_seconds())
        ts = result_dict.get("tracking_sync") or {}
        tracking_summary: dict | None = None
        if isinstance(ts, dict) and ts:
            tracking_summary = {
                "success": bool(ts.get("success")),
                "queued": int(ts.get("queued") or 0),
                "skipped": int(ts.get("skipped") or 0),
                "jobs": int(ts.get("job_ids_count") or 0),
                "errors": [str(e)[:200] for e in (ts.get("errors") or [])][:3],
                "ran_at": ts.get("ran_at"),
            }
        items.append(
            {
                "job_id": job_id,
                "status": status,
                "created_at": created_at.isoformat() if created_at else None,
                "started_at": started_at.isoformat() if started_at else None,
                "completed_at": completed_at.isoformat() if completed_at else None,
                "duration_sec": duration_sec,
                "total_synced": int(result_dict.get("total_synced") or 0),
                "per_market": per_market,
                "tracking_sync": tracking_summary,
                "error": (error or "")[:300] if error else None,
            }
        )
    return {"items": items}


@router.get("/{order_id}", response_model=SambaOrder)
async def get_order(
    order_id: str,
    session: AsyncSession = Depends(get_read_session_dependency),
    tenant_id: Optional[str] = Depends(get_optional_tenant_id),
):
    svc = _read_service(session)
    order = await svc.get_order(order_id)
    if not order:
        raise HTTPException(status_code=404, detail="주문을 찾을 수 없습니다")
    # 테넌트 소유권 검증
    if tenant_id is not None and order.tenant_id != tenant_id:
        raise HTTPException(status_code=403, detail="해당 주문에 대한 권한이 없습니다")
    return order


@router.post("", response_model=SambaOrder, status_code=201)
async def create_order(
    body: OrderCreate,
    session: AsyncSession = Depends(get_write_session_dependency),
):
    svc = _write_service(session)
    return await svc.create_order(body.model_dump(exclude_unset=True))


@router.patch("/{order_id}/link-product")
async def link_order_to_product(
    order_id: str,
    body: dict,
    session: AsyncSession = Depends(get_write_session_dependency),
):
    """주문에 수집상품 ID 연결 (지연 채움)."""
    cpid = body.get("collected_product_id", "")
    if not cpid:
        raise HTTPException(400, "collected_product_id 필수")
    from sqlalchemy import text as _t

    await session.execute(
        _t(
            "UPDATE samba_order SET collected_product_id = :cpid WHERE id = :oid AND collected_product_id IS NULL"
        ),
        {"cpid": cpid, "oid": order_id},
    )
    await session.commit()
    return {"ok": True}


@router.post("/backfill-product-links")
async def backfill_product_links(
    session: AsyncSession = Depends(get_write_session_dependency),
):
    """기존 주문의 collected_product_id 일괄 백필."""
    from sqlalchemy import text as _t

    # market_product_nos에서 역매핑 캐시 빌드
    cp_rows = await session.execute(
        _t(
            "SELECT id, market_product_nos FROM samba_collected_product "
            "WHERE market_product_nos IS NOT NULL"
        )
    )
    mpn_map: dict[str, str] = {}
    # #534 — 같은 상품번호를 복수 cp가 점유하면 오연결 위험. 충돌 키는 매핑서 제외.
    _mpn_conflicts: set[str] = set()

    def _put(_key: str, _cpid: str) -> None:
        # __claiming__<epoch> 등록중 임시 마커 — 실제 상품번호 아님 (이슈 #579)
        if _key.startswith("__claiming__"):
            return
        _prev = mpn_map.get(_key)
        if _prev is not None and _prev != _cpid:
            _mpn_conflicts.add(_key)
        else:
            mpn_map[_key] = _cpid

    for cpid, mpnos in cp_rows.fetchall():
        if not mpnos or not isinstance(mpnos, dict):
            continue
        for _v in mpnos.values():
            if not _v:
                continue
            if isinstance(_v, dict):
                for sv in [
                    _v.get("smartstoreChannelProductNo"),
                    _v.get("originProductNo"),
                    _v.get("channelProductNo"),
                ]:
                    if sv:
                        _put(str(sv), cpid)
            else:
                _put(str(_v), cpid)
    # 충돌 키 제거 — 오연결 방지(#534). 관리자 확인용 로그.
    for _ck in _mpn_conflicts:
        mpn_map.pop(_ck, None)
    if _mpn_conflicts:
        logger.warning(
            "[주문링크] #534 identity 충돌 %d건 매핑 제외: %s",
            len(_mpn_conflicts),
            ", ".join(sorted(_mpn_conflicts)[:20]),
        )

    # collected_product_id가 없는 주문 조회
    null_orders = await session.execute(
        _t(
            "SELECT id, product_id FROM samba_order "
            "WHERE collected_product_id IS NULL AND product_id IS NOT NULL"
        )
    )
    linked = 0
    for oid, pid in null_orders.fetchall():
        cpid = mpn_map.get(str(pid))
        if cpid:
            await session.execute(
                _t(
                    "UPDATE samba_order SET collected_product_id = :cpid WHERE id = :oid"
                ),
                {"cpid": cpid, "oid": oid},
            )
            linked += 1
    await session.commit()
    return {"linked": linked, "total_cache": len(mpn_map)}


@router.post("/backfill-playauto-style-code")
async def backfill_playauto_style_code(
    session: AsyncSession = Depends(get_write_session_dependency),
):
    """플레이오토 미등록 주문을 상품명 style_code로 일괄 백필.

    - collected_product_id IS NULL + source='playauto' + product_name 있는 주문 대상
    - _lh_style_tokens 로 토큰 추출 후 samba_collected_product.style_code 단일후보 매칭
    - 글로벌 단일후보: 해당 토큰 style_code 가진 CP가 정확히 1개일 때만 연결 (다중 skip)
    - 배치 CP 조회(토큰셋 1회) → N × DB 호출 없이 효율적 처리
    """
    from sqlalchemy import text as _t

    # 미등록 플레이오토 주문 조회
    null_rows = (
        await session.execute(
            _t(
                "SELECT id, product_name FROM samba_order "
                "WHERE source = 'playauto' "
                "AND collected_product_id IS NULL "
                "AND product_name IS NOT NULL AND product_name != ''"
            )
        )
    ).fetchall()

    if not null_rows:
        return {"linked": 0, "skipped_ambiguous": 0, "no_cp": 0, "total": 0}

    # 전체 토큰 수집
    all_tokens: set[str] = set()
    order_tokens: list[tuple[str, list[str]]] = []
    for oid, pname in null_rows:
        tokens = _lh_style_tokens(str(pname or ""))
        order_tokens.append((str(oid), tokens))
        all_tokens.update(tokens)

    if not all_tokens:
        return {
            "linked": 0,
            "skipped_ambiguous": 0,
            "no_cp": 0,
            "total": len(null_rows),
        }

    # 토큰 → CP 배치 조회 (1회)
    _cols = (
        "id, style_code, source_site, source_url, (images->>0) AS thumb, category, cost"
    )
    cp_rows = (
        await session.execute(
            _t(
                f"SELECT {_cols} FROM samba_collected_product "
                "WHERE style_code = ANY(:t)"
            ),
            {"t": list(all_tokens)},
        )
    ).fetchall()

    # 토큰 → [cp_id, ...] 인덱스 구성
    token_to_cp: dict[str, list[str]] = {}
    for row in cp_rows:
        sc = str(row[1] or "")
        if sc:
            token_to_cp.setdefault(sc, []).append(str(row[0]))

    # 주문별 매칭
    linked = skipped_ambiguous = no_cp = 0
    for oid, tokens in order_tokens:
        if not tokens:
            no_cp += 1
            continue
        # 가장 긴 토큰부터 단독 고유 매칭 시도
        matched_cpid: str | None = None
        for tok in sorted(tokens, key=len, reverse=True):
            cands = token_to_cp.get(tok, [])
            if len(cands) == 1:
                matched_cpid = cands[0]
                break
            elif len(cands) > 1:
                skipped_ambiguous += 1
                break
        if matched_cpid:
            await session.execute(
                _t(
                    "UPDATE samba_order SET collected_product_id = :cpid "
                    "WHERE id = :oid AND collected_product_id IS NULL"
                ),
                {"cpid": matched_cpid, "oid": oid},
            )
            linked += 1
        else:
            no_cp += 1

    await session.commit()
    logger.info(
        f"[백필/플레이오토-style] linked={linked} ambiguous={skipped_ambiguous} no_cp={no_cp}"
    )
    return {
        "linked": linked,
        "skipped_ambiguous": skipped_ambiguous,
        "no_cp": no_cp,
        "total": len(null_rows),
    }


@router.post("/fix-musinsa-fashionplus-mismatch")
async def fix_musinsa_fashionplus_mismatch(
    session: AsyncSession = Depends(get_write_session_dependency),
):
    """FashionPlus로 잘못 매칭된 무신사 주문 일괄 수정.

    상품명 끝 숫자가 MUSINSA site_product_id와 일치하는데
    collected_product_id가 FashionPlus 상품을 가리키는 주문을 찾아 수정한다.
    """
    import re as _re
    from sqlalchemy import text as _t

    # FashionPlus로 매칭된 주문 중 상품명 끝에 숫자가 있는 건 조회
    bad_orders = await session.execute(
        _t(
            "SELECT o.id, o.product_name, o.collected_product_id "
            "FROM samba_order o "
            "JOIN samba_collected_product cp ON cp.id = o.collected_product_id "
            "WHERE cp.source_site = 'FashionPlus' "
            "AND o.product_name ~ E'\\\\d{7,}\\\\s*$'"
        )
    )
    rows = bad_orders.fetchall()

    fixed = 0
    skipped = 0
    for oid, pname, old_cpid in rows:
        m = _re.search(r"(\d{7,})\s*$", pname or "")
        if not m:
            skipped += 1
            continue
        sid = m.group(1)

        # 동일 site_product_id를 가진 MUSINSA 상품 조회
        cp_row = await session.execute(
            _t(
                "SELECT id FROM samba_collected_product "
                "WHERE site_product_id = :sid AND source_site = 'MUSINSA' "
                "ORDER BY (market_product_nos IS NOT NULL) DESC, created_at ASC "
                "LIMIT 1"
            ),
            {"sid": sid},
        )
        correct_cp = cp_row.fetchone()
        if not correct_cp:
            skipped += 1
            continue

        await session.execute(
            _t(
                "UPDATE samba_order "
                "SET collected_product_id = :cpid, source_site = 'MUSINSA' "
                "WHERE id = :oid"
            ),
            {"cpid": correct_cp[0], "oid": oid},
        )
        fixed += 1

    await session.commit()
    return {"fixed": fixed, "skipped": skipped, "total_checked": len(rows)}


@router.put("/{order_id}", response_model=SambaOrder)
async def update_order(
    order_id: str,
    body: OrderUpdate,
    session: AsyncSession = Depends(get_write_session_dependency),
):
    from sqlalchemy import text as _t

    svc = _write_service(session)
    data = body.model_dump(exclude_unset=True)
    order = await svc.update_order(order_id, data)
    if not order:
        raise HTTPException(status_code=404, detail="주문을 찾을 수 없습니다")

    # source_url/product_image 변경 시 동일 product_id+channel_name 주문 일괄 업데이트
    batch_fields = {
        k: v for k, v in data.items() if k in ("source_url", "product_image")
    }
    if batch_fields and order.product_id and order.channel_name:
        set_clauses = ", ".join(f"{k} = :{k}" for k in batch_fields)
        params = {
            **batch_fields,
            "pid": order.product_id,
            "cname": order.channel_name,
            "oid": order_id,
        }
        await session.execute(
            _t(
                f"UPDATE samba_order SET {set_clauses} "
                "WHERE product_id = :pid AND channel_name = :cname AND id != :oid"
            ),
            params,
        )
        await session.commit()

    return order


# ─────────────────────────────────────────────────────────────────────────
# SNKRDUNK 해외송장 자동수집 (크림 해외매입 — 사무국→구매자 발송송장)
# MFA(SMS OTP)라 백엔드 id/pw 자동로그인 불가 → 확장앱이 SNKRDUNK 로그인 세션쿠키를
# 캡처해 아래로 전송. 백엔드가 이 쿠키로 /v1/orders/{취引ID} + get-delivery-company 호출.
# sourcing_order_number = SNKRDUNK 취引ID (③에서 채움).
# ─────────────────────────────────────────────────────────────────────────
_SNKR_COOKIE_KEY = "snkrdunk_session_cookie"


class SnkrCookieBody(BaseModel):
    cookie: str


async def _get_snkr_session_cookie(session: AsyncSession) -> str:
    """저장된 SNKRDUNK 세션쿠키 조회 (samba_settings)."""
    from backend.domain.samba.forbidden.model import SambaSettings

    r = await session.execute(
        select(SambaSettings).where(SambaSettings.key == _SNKR_COOKIE_KEY)
    )
    row = r.scalars().first()
    val = row.value if row else None
    if isinstance(val, dict):
        return str(val.get("cookie") or "").strip()
    return str(val).strip() if isinstance(val, str) else ""


async def _apply_snkr_overseas_tracking(
    session: AsyncSession, order: SambaOrder, cookie: str
) -> dict:
    """주문 1건에 대해 SNKRDUNK 해외송장 조회 → 발송됐으면 DB 저장."""
    from datetime import timezone as _tz

    from backend.domain.samba.proxy.snkrdunk import fetch_order_overseas_tracking

    ord_no = (order.sourcing_order_number or "").strip()
    if not ord_no:
        return {"success": False, "error": "소싱주문번호(취引ID) 없음"}
    r = await fetch_order_overseas_tracking(cookie, ord_no)
    if r.get("error"):
        return {"success": False, "error": r["error"]}
    if not r.get("shipped"):
        # 아직 사무국→구매자 발송 전 — 송장 미존재
        return {
            "success": True,
            "shipped": False,
            "order_status": r.get("order_status"),
        }
    order.overseas_shipping_company = r["delivery_company"]
    order.overseas_tracking_number = r["tracking_number"]
    # 해외송장 수집 완료 → 상태 '국내배송중'(shipping) — 배송완료/확정/취소/반품은 유지
    if order.status not in ("delivered", "confirmed", "cancelled", "returned"):
        order.status = "shipping"
    order.updated_at = datetime.now(_tz.utc)
    await session.commit()
    return {
        "success": True,
        "shipped": True,
        "delivery_company": r["delivery_company"],
        "tracking_number": r["tracking_number"],
        "order_status": r.get("order_status"),
    }


@router.post("/snkrdunk/session-cookie")
async def save_snkrdunk_session_cookie(
    body: SnkrCookieBody,
    session: AsyncSession = Depends(get_write_session_dependency),
):
    """확장앱이 캡처한 SNKRDUNK 로그인 세션쿠키 저장 (samba_settings upsert)."""
    from datetime import UTC, datetime as _dt
    from sqlalchemy import func as _func
    from sqlalchemy.dialects.postgresql import insert as pg_insert

    from backend.core.tenant_context import current_tenant_id
    from backend.domain.samba.forbidden.model import SambaSettings

    cookie = (body.cookie or "").strip()
    if cookie.lower().startswith("session="):
        cookie = cookie.split("=", 1)[1]
    if not cookie:
        raise HTTPException(status_code=400, detail="cookie 비어있음")

    now = _dt.now(UTC)
    tid = current_tenant_id.get()
    value = {"cookie": cookie, "updated_at": now.isoformat()}
    ins = pg_insert(SambaSettings).values(
        key=_SNKR_COOKIE_KEY, value=value, updated_at=now, tenant_id=tid
    )
    stmt = ins.on_conflict_do_update(
        index_elements=["key"],
        set_={
            "value": value,
            "updated_at": now,
            "tenant_id": _func.coalesce(
                ins.excluded.tenant_id, SambaSettings.__table__.c.tenant_id
            ),
        },
    )
    await session.execute(stmt)
    await session.commit()
    return {"success": True}


@router.post("/{order_id}/fetch-snkrdunk-tracking")
async def fetch_snkrdunk_tracking(
    order_id: str,
    session: AsyncSession = Depends(get_write_session_dependency),
):
    """주문 1건 해외송장(사무국→구매자 발송) 조회 + 저장."""
    order = await session.get(SambaOrder, order_id)
    if not order:
        raise HTTPException(status_code=404, detail="주문을 찾을 수 없습니다")
    cookie = await _get_snkr_session_cookie(session)
    if not cookie:
        return {
            "success": False,
            "error": "SNKRDUNK 세션쿠키 없음 — 확장앱으로 SNKRDUNK 로그인 필요",
        }
    return await _apply_snkr_overseas_tracking(session, order, cookie)


@router.post("/snkrdunk/sync-overseas-tracking")
async def sync_snkrdunk_overseas_tracking(
    limit: int = 200,
    session: AsyncSession = Depends(get_write_session_dependency),
):
    """미수집 KREAM 주문 일괄 해외송장 조회 (소싱주문번호 有 & 해외송장 空)."""
    from sqlalchemy import func as _func, or_ as _or

    cookie = await _get_snkr_session_cookie(session)
    if not cookie:
        return {
            "success": False,
            "error": "SNKRDUNK 세션쿠키 없음 — 확장앱으로 SNKRDUNK 로그인 필요",
        }
    stmt = (
        select(SambaOrder)
        .where(
            SambaOrder.sourcing_order_number.is_not(None),
            SambaOrder.sourcing_order_number != "",
            (SambaOrder.overseas_tracking_number.is_(None))
            | (SambaOrder.overseas_tracking_number == ""),
            _or(
                _func.upper(_func.coalesce(SambaOrder.source_site, "")) == "KREAM",
                _func.upper(_func.coalesce(SambaOrder.sales_channel_alias, "")).like(
                    "%KREAM%"
                ),
            ),
        )
        .limit(max(1, min(int(limit or 200), 500)))
    )
    rows = (await session.execute(stmt)).scalars().all()
    checked = 0
    shipped = 0
    for o in rows:
        checked += 1
        res = await _apply_snkr_overseas_tracking(session, o, cookie)
        if res.get("shipped"):
            shipped += 1
        await asyncio.sleep(0.3)  # SNKRDUNK 레이트리밋 보수값
    logger.info(f"[SNKRDUNK해외송장] 일괄조회 checked={checked} shipped={shipped}")
    return {"success": True, "checked": checked, "shipped": shipped}


# ─────────────────────────────────────────────────────────────────────────
# 허브넷(kpartner.ehub24.net) 택배번호 자동입력
# 크림 해외판매 배대지 — 스니덩크 해외송장을 허브넷 행에 기입해야 국내 재발송됨.
# 서버사이드 로그인(auth) → search_kream 으로 (A-LI주문번호 → 행PK) 매핑 →
# bulk_tracking_update 로 일괄 기입. 크레덴셜은 samba_settings.hubnet_credentials.
# ─────────────────────────────────────────────────────────────────────────
_HUBNET_BASE = "https://kpartner.ehub24.net"


async def _push_hubnet_tracking(session: AsyncSession) -> dict:
    """해외송장 보유 크림주문 → 허브넷 택배번호 일괄 기입. 실패해도 예외 안 던짐."""
    import json  # noqa: F811 — 로컬 import (모듈 최상위에 없음)

    import httpx as _httpx

    from backend.domain.samba.forbidden.model import SambaSettings

    r = await session.execute(
        select(SambaSettings).where(SambaSettings.key == "hubnet_credentials")
    )
    row = r.scalars().first()
    creds = row.value if row and isinstance(row.value, dict) else None
    if not creds or not creds.get("email"):
        return {"updated": 0, "error": "hubnet_credentials 없음"}

    orders = (
        await session.execute(
            select(SambaOrder.order_number, SambaOrder.overseas_tracking_number).where(
                SambaOrder.order_number.like("A-LI%"),
                SambaOrder.overseas_tracking_number.is_not(None),
                SambaOrder.overseas_tracking_number != "",
            )
        )
    ).all()
    trk = {o[0]: o[1] for o in orders}
    if not trk:
        return {"updated": 0, "error": None}

    from datetime import date as _date, timedelta as _td

    start = (_date.today() - _td(days=30)).isoformat()
    end = _date.today().isoformat()
    ua = (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"
    )
    try:
        async with _httpx.AsyncClient(
            headers={
                "User-Agent": ua,
                "X-Requested-With": "XMLHttpRequest",
                "Referer": f"{_HUBNET_BASE}/list",
            },
            timeout=30,
            follow_redirects=True,
        ) as client:
            login = await client.post(
                f"{_HUBNET_BASE}/auth",
                data={
                    "action": "login",
                    "email": creds["email"],
                    "password": creds.get("password", ""),
                },
            )
            if '"success":true' not in login.text:
                return {"updated": 0, "error": f"허브넷 로그인 실패: {login.text[:80]}"}

            search = await client.post(
                f"{_HUBNET_BASE}/list_ajax",
                data={
                    "mode": "search_kream",
                    "start_date": start,
                    "end_date": end,
                    "date_type": "order",
                    "search_type": "hbl",
                    "numbers": "",
                    "work_status": "",
                    "origin": "",
                },
            )
            data = search.json()
            if not data.get("success"):
                return {"updated": 0, "error": "허브넷 조회 실패"}
            payload = []
            for hrow in data.get("data", []):
                onum = str(hrow.get("add1") or "").strip()
                if onum in trk:
                    payload.append({"no": hrow.get("no"), "tracking_no": trk[onum]})
            if not payload:
                return {"updated": 0, "error": None}
            upd = await client.post(
                f"{_HUBNET_BASE}/list_ajax",
                data={
                    "mode": "bulk_tracking_update",
                    "update_data": json.dumps(payload, ensure_ascii=False),
                },
            )
            ok = '"success":true' in upd.text
            logger.info(f"[허브넷] 택배번호 기입 {len(payload)}건 ok={ok}")
            return {
                "updated": len(payload) if ok else 0,
                "error": None if ok else upd.text[:80],
            }
    except Exception as e:
        logger.warning(f"[허브넷] 자동기입 실패(무시): {e}")
        return {"updated": 0, "error": str(e)[:80]}


@router.put("/{order_id}/status", response_model=SambaOrder)
async def update_order_status(
    order_id: str,
    body: OrderStatusUpdate,
    session: AsyncSession = Depends(get_write_session_dependency),
):
    svc = _write_service(session)
    order = await svc.update_order_status(order_id, body.status)
    if not order:
        raise HTTPException(status_code=404, detail="주문을 찾을 수 없습니다")
    return order


@router.delete("/{order_id}")
async def delete_order(
    order_id: str,
    session: AsyncSession = Depends(get_write_session_dependency),
):
    svc = _write_service(session)
    deleted = await svc.delete_order(order_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="주문을 찾을 수 없습니다")
    return {"ok": True}


# ══════════════════════════════════════════════
# 취소승인
# ══════════════════════════════════════════════


@router.post("/{order_id}/approve-cancel")
async def approve_cancel(
    order_id: str,
    session: AsyncSession = Depends(get_write_session_dependency),
):
    """취소요청 주문에 대해 마켓 취소승인 실행."""
    from backend.domain.samba.account.repository import SambaMarketAccountRepository
    from backend.domain.samba.forbidden.repository import SambaSettingsRepository

    svc = _write_service(session)
    order = await svc.get_order(order_id)
    if not order:
        raise HTTPException(status_code=404, detail="주문을 찾을 수 없습니다")

    if not order.order_number:
        raise HTTPException(status_code=400, detail="상품주문번호가 없습니다")

    # 마켓 계정 조회
    if not order.channel_id:
        raise HTTPException(status_code=400, detail="마켓 계정 정보가 없습니다")

    account_repo = SambaMarketAccountRepository(session)
    account = await account_repo.get_async(order.channel_id)
    if not account:
        raise HTTPException(status_code=400, detail="마켓 계정을 찾을 수 없습니다")

    if account.market_type == "smartstore":
        from backend.domain.samba.proxy.smartstore import SmartStoreClient

        extras = account.additional_fields or {}
        client_id = extras.get("clientId", "") or account.api_key or ""
        client_secret = extras.get("clientSecret", "") or account.api_secret or ""
        if not client_id or not client_secret:
            settings_repo = SambaSettingsRepository(session)
            row = await settings_repo.find_by_async(key="store_smartstore")
            if row and isinstance(row.value, dict):
                client_id = client_id or row.value.get("clientId", "")
                client_secret = client_secret or row.value.get("clientSecret", "")
        if not client_id or not client_secret:
            raise HTTPException(status_code=400, detail="스마트스토어 인증정보 없음")

        client = SmartStoreClient(client_id, client_secret)
        try:
            await client.approve_cancel(order.order_number)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"취소승인 실패: {e}")

        # DB 상태 업데이트 — status='cancelled' 도 같이 변경 (쿠팡/롯데ON/eBay 분기와 일관).
        # status 누락 시 OrdersTable 의 isCancelRequested(=status==='cancel_requested')
        # 가 true 로 유지돼 빨간 '취소요청' 배지·승인/거부 버튼이 안 사라지는 UX 버그
        # 발생 (2026-06-08 사용자 보고).
        await svc.update_order(
            order_id,
            {
                "shipping_status": "취소완료",
                "status": "cancelled",
            },
        )
        logger.info(f"[취소승인] {order.order_number} 취소승인 완료")
        return {"ok": True, "message": "취소승인 완료"}

    elif account.market_type == "11st":
        from backend.domain.samba.proxy.elevenst import ElevenstClient
        from backend.domain.samba.returns.repository import SambaReturnRepository

        api_key = (
            (account.additional_fields or {}).get("apiKey", "") or account.api_key or ""
        )
        if not api_key:
            raise HTTPException(status_code=400, detail="11번가 API 키 없음")

        client = ElevenstClient(api_key)
        return_repo = SambaReturnRepository(session)
        existing_returns = await return_repo.filter_by_async(order_id=order_id)
        ret = existing_returns[0] if existing_returns else None
        clm_req_seq = (ret.clm_req_seq if ret else None) or ""
        ord_prd_seq = (ret.ord_prd_seq if ret else None) or ""
        confirm_ord_no = order.order_number

        # 미수집(반품동기화 잡 미실행 또는 방금 취소요청 접수) 시 → 11번가 취소요청 목록을
        # 라이브 조회해 이 주문을 매칭, 클레임번호(ordPrdCnSeq)·주문순번을 즉시 확보한다.
        # 동기화를 기다리지 않고 취소요청 들어오자마자 바로 승인 가능하게 함.
        if not clm_req_seq or not ord_prd_seq:
            from datetime import datetime as _dt, timedelta as _td

            _fmt = "%Y%m%d%H%M"
            _now = _dt.now()
            try:
                _cancel_items = await client.get_cancel_requests(
                    (_now - _td(days=30)).strftime(_fmt), _now.strftime(_fmt)
                )
            except Exception as _le:  # noqa: BLE001
                _cancel_items = []
                logger.warning(f"[취소승인][11번가] 라이브 취소목록 조회 실패: {_le}")
            for _it in _cancel_items:
                _onum = _it.get("ordPrdNo", "") or _it.get("ordNo", "")
                if _onum and str(_onum) == str(order.order_number):
                    clm_req_seq = _it.get("ordPrdCnSeq", "") or clm_req_seq
                    ord_prd_seq = _it.get("ordPrdSeq", "") or ord_prd_seq
                    confirm_ord_no = _it.get("ordNo", "") or order.order_number
                    logger.info(
                        "[취소승인][11번가] 라이브 조회로 클레임정보 확보: "
                        f"ordPrdCnSeq={clm_req_seq} ordPrdSeq={ord_prd_seq}"
                    )
                    break

        if not clm_req_seq or not ord_prd_seq:
            raise HTTPException(
                status_code=400,
                detail="11번가 취소 클레임 정보 없음 — 라이브 조회에도 취소요청이 "
                "없습니다 (이미 처리됐거나 취소요청 미접수)",
            )

        try:
            await client.confirm_cancel(clm_req_seq, confirm_ord_no, ord_prd_seq)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"취소승인 실패: {e}")

        # status='cancelled' 도 같이 변경 — 쿠팡/롯데ON/eBay 분기와 일관.
        # status 누락 시 빨간 '취소요청' 배지가 처리 후에도 안 사라지는 UX 버그
        # (2026-06-08 사용자 보고).
        await svc.update_order(
            order_id, {"shipping_status": "취소완료", "status": "cancelled"}
        )
        if ret:
            await return_repo.update_async(
                ret.id, status="cancelled", market_order_status="취소완료"
            )

        logger.info(f"[취소승인][11번가] {order.order_number} 취소승인 완료")
        return {"ok": True, "message": "취소승인 완료"}

    elif account.market_type == "ebay":
        # eBay는 seller_cancel_order로 이미 취소 처리됨 → DB 상태만 동기화
        await svc.update_order(
            order_id,
            {"shipping_status": "취소완료", "status": "cancelled"},
        )
        # samba_return 상태도 업데이트
        from backend.domain.samba.returns.repository import SambaReturnRepository

        ret_repo = SambaReturnRepository(session)
        rets = await ret_repo.filter_by_async(order_id=order_id)
        for ret in rets:
            await ret_repo.update_async(
                ret.id,
                status="completed",
                market_order_status="취소완료",
            )
        logger.info(f"[취소승인] eBay {order.order_number} 취소완료 동기화")
        return {"ok": True, "message": "eBay 취소완료 처리"}

    elif account.market_type == "coupang":
        # 쿠팡 취소승인 — returnRequests v6 stoppedShipment (미출고 케이스만 자동) (#246 PR-4)
        # - release_status='N' (미출고)   → stopped_shipment 호출 → 출고중지완료 처리
        # - release_status='A' (이미출고) → 별도 엔드포인트 /approve-cancel-with-shipment
        #                                    (운영자 송장 정보 입력 필요)
        # - release_status='Y'/'S'/None  → 처리 불가/이미 처리됨 → 400
        from backend.domain.samba.proxy.coupang import CoupangApiError, CoupangClient

        extras = account.additional_fields or {}
        access_key = extras.get("accessKey", "") or account.api_key or ""
        secret_key = extras.get("secretKey", "") or account.api_secret or ""
        vendor_id = extras.get("vendorId", "") or account.seller_id or ""
        if not all([access_key, secret_key, vendor_id]):
            raise HTTPException(
                status_code=400,
                detail="쿠팡 인증정보 없음 (accessKey/secretKey/vendorId)",
            )

        client = CoupangClient(access_key, secret_key, vendor_id)

        # 0) 결제완료 단계 즉시취소 분기 — 쿠팡이 자동 취소하므로 승인 호출 불필요.
        #    발주서 단건 재조회로 쿠팡 현재 상태(진실의 원천)를 확인해 이미 CANCEL이면
        #    상태만 '취소완료'로 동기화하고 종료. (receiptId 없는 즉시취소 케이스 구제)
        #
        # ⚠️ 진실의 원천 = 쿠팡. DB 자체 자료(order.status)만 보고 이 분기에 빠지면
        # 마켓 API 한 번도 호출 안 한 채 DB만 cancelled로 박는 false-success 발생
        # (사례: 동기화 잡이 status를 cancelled로 선점 → 사용자가 취소승인 누름 →
        #  마켓엔 '출고중지요청' 그대로인데 삼바만 '취소완료'로 보이는 사고).
        # 따라서 ① 항상 단건 조회로 live_status 확보 시도 ② live_status == 'CANCEL' 일
        # 때만 즉시취소 분기로 통과. live_status 확보 실패면 명시적 에러 → 운영자가
        # 동기화 후 재시도하도록 유도.
        live_status = ""
        _box_id = None
        try:
            # 쿠팡 order_number = shipmentBoxId (parse 규칙)
            _box_id = int(order.order_number)
        except (TypeError, ValueError):
            _box_id = None
        if _box_id:
            try:
                _sheet = await client.get_ordersheet_by_box_id(_box_id)
                _data = _sheet.get("data") if isinstance(_sheet, dict) else None
                if isinstance(_data, list):
                    _data = _data[0] if _data else None
                if isinstance(_data, dict):
                    live_status = (_data.get("status") or "").upper()
            except Exception as _le:
                logger.warning(f"[취소승인] 쿠팡 단건 조회 실패: {_le}")

        if live_status == "CANCEL":
            await svc.update_order(
                order_id,
                {"shipping_status": "취소완료", "status": "cancelled"},
            )
            logger.info(
                f"[취소승인] 쿠팡 {order.order_number} 즉시취소 확인 — 상태 동기화 "
                f"(live_status=CANCEL)"
            )
            return {"ok": True, "message": "쿠팡 즉시취소 완료 (상태 동기화)"}

        # 1) 상품준비중 단계 출고중지 승인 경로 — receiptId 필수.
        #    미수집 시 라이브로 취소·반품 요청 목록을 조회해 이 주문(orderId=shipment_id)의
        #    receiptId·releaseStatus를 즉시 확보한다(CANCEL=배송전취소 우선). 동기화를 기다리지
        #    않고 취소요청 들어오자마자 바로 승인 가능하게 함.
        _receipt_id = order.cancel_receipt_id
        _release_status = order.cancel_release_status
        if not _receipt_id:
            _target_oid = str(order.shipment_id or order.ext_order_number or "")
            try:
                _cr_items = await client.get_cancel_and_return_requests(days=30)
            except Exception as _le:  # noqa: BLE001
                _cr_items = []
                logger.warning(
                    f"[취소승인][쿠팡] 라이브 취소·반품 목록 조회 실패: {_le}"
                )
            _matched = None
            for _cr in _cr_items or []:
                if not isinstance(_cr, dict):
                    continue
                if _target_oid and str(_cr.get("orderId", "") or "") == _target_oid:
                    # CANCEL(배송전 취소) 우선 — 이미 매칭된 게 CANCEL이 아니면 교체
                    if _matched is None or (
                        (_cr.get("receiptType") or "").upper() == "CANCEL"
                        and (_matched.get("receiptType") or "").upper() != "CANCEL"
                    ):
                        _matched = _cr
            if _matched:
                try:
                    _receipt_id = int(_matched.get("receiptId"))
                except (TypeError, ValueError):
                    _receipt_id = None
                _rit = _matched.get("returnItems") or []
                if isinstance(_rit, list) and _rit and isinstance(_rit[0], dict):
                    _release_status = _rit[0].get("releaseStatus") or _release_status
                logger.info(
                    f"[취소승인][쿠팡] 라이브 조회로 receiptId={_receipt_id} "
                    f"releaseStatus={_release_status} 확보 (orderId={_target_oid})"
                )

        if not _receipt_id:
            raise HTTPException(
                status_code=400,
                detail="쿠팡 취소 receiptId 미수집 — 라이브 조회에도 취소요청이 "
                "없습니다 (이미 처리됐거나 취소요청 미접수)",
            )

        rls = (_release_status or "").upper()
        if rls == "A":
            raise HTTPException(
                status_code=400,
                detail=(
                    "이미출고 상태 — /orders/{id}/approve-cancel-with-shipment 로 "
                    "택배사·송장번호 함께 호출 필요"
                ),
            )
        if rls and rls not in ("N",):
            raise HTTPException(
                status_code=400,
                detail=f"쿠팡 release_status={rls} — 처리 불가 또는 이미 처리됨",
            )

        cancel_count = int(order.quantity or 1)
        try:
            await client.stopped_shipment(
                receipt_id=int(_receipt_id),
                cancel_count=cancel_count,
            )
        except CoupangApiError as e:
            raise HTTPException(status_code=500, detail=f"쿠팡 취소승인 실패: {e}")
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"쿠팡 취소승인 실패: {e}")

        await svc.update_order(
            order_id,
            {"shipping_status": "취소완료", "status": "cancelled"},
        )
        logger.info(
            f"[취소승인] 쿠팡 {order.order_number} stoppedShipment 완료 "
            f"(receiptId={_receipt_id}, count={cancel_count})"
        )
        return {"ok": True, "message": "쿠팡 취소승인 완료 (출고중지)"}

    elif account.market_type == "lotteon":
        # 롯데ON 취소요청 승인 — 라이브 취소클레임 조회로 승인 대상 판별 후 cnclRequestApproval.
        # 클레임 없으면 판매자직접취소. 자동취소와 동일 로직(_lotteon_approve_or_direct_cancel).
        from backend.api.v1.routers.samba.proxy.sourcing import (
            _lotteon_approve_or_direct_cancel,
        )
        from backend.domain.samba.proxy.lotteon import LotteonClient

        extras = account.additional_fields or {}
        api_key = extras.get("apiKey", "") or account.api_key or ""
        if not api_key:
            raise HTTPException(status_code=400, detail="롯데ON API Key 없음")

        client = LotteonClient(api_key)
        try:
            ok, message = await _lotteon_approve_or_direct_cancel(client, order)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"롯데ON 취소승인 실패: {e}")
        if not ok:
            raise HTTPException(status_code=500, detail=f"롯데ON 취소 실패: {message}")

        await svc.update_order(
            order_id,
            {"shipping_status": "취소완료", "status": "cancelled"},
        )
        logger.info(f"[취소승인][롯데ON] {order.order_number} {message}")
        return {"ok": True, "message": f"롯데ON {message}"}

    elif account.market_type == "lottehome":
        # 롯데홈쇼핑 취소 → registDeliver(proc_gubun=imps) 발송불가 처리
        # order_number 형식: "ord_no:ord_dtl_sn"
        from backend.domain.samba.proxy.lottehome import LotteHomeClient

        raw = order.order_number or ""
        parts = raw.split(":")
        if len(parts) < 2 or not parts[0] or not parts[1]:
            raise HTTPException(
                status_code=400,
                detail=f"롯데홈쇼핑 주문번호 형식 오류 (ord_no:ord_dtl_sn 필요, 현재값={raw!r})",
            )
        ord_no, ord_dtl_sn = parts[0], parts[1]

        extras = account.additional_fields or {}
        user_id = extras.get("userId", "") or account.seller_id or ""
        password = extras.get("password", "")
        agnc_no = extras.get("agncNo", "")
        env = extras.get("env", "prod")
        if not user_id:
            raise HTTPException(status_code=400, detail="롯데홈쇼핑 로그인 정보 없음")

        client = LotteHomeClient(user_id, password, agnc_no, env)
        try:
            res = await client.register_deliver(ord_no, ord_dtl_sn, "imps")
        except Exception as e:
            raise HTTPException(
                status_code=500, detail=f"롯데홈쇼핑 발송불가 처리 실패: {e}"
            )
        if not res.get("ok"):
            raise HTTPException(
                status_code=500,
                detail=f"롯데홈쇼핑 발송불가 처리 실패: result={res.get('result')}",
            )

        await svc.update_order(
            order_id,
            {"shipping_status": "취소완료", "status": "cancelled"},
        )
        logger.info(
            f"[취소승인][롯데홈쇼핑] {ord_no}:{ord_dtl_sn} 발송불가(imps) 처리 완료"
        )
        return {"ok": True, "message": "롯데홈쇼핑 발송불가 처리 완료"}

    elif account.market_type == "ssg":
        # SSG 셀러 API — POST /api/claim/v2/cancel/request/approve
        # 파라미터: ordNo (=order.order_number), ordItemSeq (=order.ord_prd_seq).
        # ord_prd_seq 는 주문동기화 시 SSG 응답의 ordItemSeq 를 그대로 저장 (ssg.py:2400 참고).
        # SSGClient.approve_cancel 은 resultCode 00·91 모두 성공으로 처리.
        from backend.domain.samba.proxy.ssg import SSGApiError, SSGClient

        extras = account.additional_fields or {}
        api_key = (
            extras.get("apiKey", "")
            or extras.get("api_key", "")
            or account.api_key
            or ""
        )
        if not api_key:
            raise HTTPException(status_code=400, detail="SSG API 키 없음")

        if not order.ord_prd_seq:
            raise HTTPException(
                status_code=400,
                detail="SSG ordItemSeq 미수집 — 주문 동기화 후 다시 시도해주세요",
            )

        site_no = extras.get("siteNo", "") or extras.get("site_no", "") or "6004"
        client = SSGClient(api_key, site_no=site_no)
        try:
            await client.approve_cancel(order.order_number, str(order.ord_prd_seq))
        except SSGApiError as e:
            raise HTTPException(status_code=500, detail=f"SSG 취소승인 실패: {e}")
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"SSG 취소승인 실패: {e}")

        # status='cancelled' 도 같이 update — PR #376 일관성.
        # 누락 시 OrdersTable.isCancelRequested 가 true 로 남아 빨간 '취소요청'
        # 배지·승인/거부 버튼이 안 사라지는 UX 사고가 발생함.
        await svc.update_order(
            order_id,
            {"shipping_status": "취소완료", "status": "cancelled"},
        )
        logger.info(f"[취소승인][SSG] {order.order_number} 취소승인 완료")
        return {"ok": True, "message": "SSG 취소승인 완료"}

    elif account.market_type in ("gmarket", "auction"):
        # ESM(옥션/G마켓) 취소승인 — PUT /claim/v1/sa/Cancel/{OrderNo}
        # site_type: 옥션=1, G마켓=2 (PUT 엔드포인트 기준, search_cancels 의 1/3 과 다름)
        from backend.domain.samba.proxy.esmplus import (
            ESMPlusClient,
            resolve_esm_credentials,
        )
        from backend.domain.samba.returns.repository import SambaReturnRepository

        hosting_id, secret_key = await resolve_esm_credentials(session, account)
        seller_id = (account.seller_id or "").strip()
        if not hosting_id or not secret_key:
            raise HTTPException(status_code=400, detail="ESM 인증정보 없음")
        if not seller_id:
            raise HTTPException(status_code=400, detail="ESM seller_id 없음")

        site_type = 2 if account.market_type == "gmarket" else 1
        client = ESMPlusClient(
            hosting_id, secret_key, seller_id, site=account.market_type
        )
        try:
            await client.approve_cancel_by_orderno(order.order_number, site_type)
        except Exception as e:
            # 옥션 resultCode=8668 (BizRuleCode W8-2) = 이미 취소승인된 건 → 멱등 성공 처리
            if "8668" in str(e):
                logger.info(
                    f"[취소승인][ESM] {order.order_number} 이미 취소승인됨(멱등 처리)"
                )
            else:
                raise HTTPException(status_code=500, detail=f"취소승인 실패: {e}")
        finally:
            try:
                await client.aclose()
            except Exception:
                pass

        # status='cancelled' 도 같이 변경 — 다른 마켓 분기와 일관(빨간 '취소요청' 배지 제거)
        await svc.update_order(
            order_id, {"shipping_status": "취소완료", "status": "cancelled"}
        )
        ret_repo = SambaReturnRepository(session)
        for ret in await ret_repo.filter_by_async(order_id=order_id):
            await ret_repo.update_async(
                ret.id, status="cancelled", market_order_status="취소완료"
            )
        logger.info(f"[취소승인][ESM] {order.order_number} 취소승인 완료")
        return {"ok": True, "message": "ESM 취소승인 완료"}

    else:
        raise HTTPException(
            status_code=400, detail=f"{account.market_type} 취소승인 미지원"
        )


# ══════════════════════════════════════════════
# 소싱처 발주 자동취소 (헤드리스 데몬)
# ══════════════════════════════════════════════


@router.post("/{order_id}/sourcing-cancel")
async def sourcing_cancel_order(
    order_id: str,
    session: AsyncSession = Depends(get_write_session_dependency),
):
    """소싱처 발주 헤드리스 자동취소 — 운영자 수동 트리거.

    가드:
      - sourcing_order_number 있어야 함 (실제 발주 완료된 주문)
      - shipping_status가 '배송중'/'배송완료' 면 차단 (이미 발송)
    동작:
      - SourcingQueue 에 cancel_order 잡 enqueue → 데몬 처리 → cancel-result 콜백
      - 결과는 비동기. 즉시 {jobId, accepted: True} 반환.
    """
    from backend.domain.samba.order.service import SambaOrderService
    from backend.domain.samba.proxy.sourcing_queue import SourcingQueue

    svc = SambaOrderService(session)
    order = await svc.get_order(order_id)
    if not order:
        raise HTTPException(status_code=404, detail="주문 없음")
    if not (order.sourcing_order_number or "").strip():
        raise HTTPException(
            status_code=400, detail="소싱처 발주번호 없음 — 발주 안 된 주문"
        )
    blocked_shipping = ("배송중", "배송완료", "출고완료", "구매확정")
    if (order.shipping_status or "").strip() in blocked_shipping:
        raise HTTPException(
            status_code=400,
            detail=f"이미 발송 단계({order.shipping_status}) — 소싱처 자동취소 불가",
        )

    site = (order.source_site or "").strip()
    if not site:
        raise HTTPException(status_code=400, detail="source_site 미상")

    request_id, _future = await SourcingQueue.add_cancel_order_job(
        site=site,
        sourcing_order_number=order.sourcing_order_number,
        order_id=order_id,
        sourcing_account_id=order.sourcing_account_id or "",
    )
    logger.info(
        f"[소싱취소] 잡 enqueue order={order_id} site={site} "
        f"ord={order.sourcing_order_number} req={request_id}"
    )
    return {"accepted": True, "jobId": request_id, "site": site}


# ══════════════════════════════════════════════
# 판매자 주도 취소 (재고부족, 가격변동 등)
# ══════════════════════════════════════════════


class SellerCancelBody(BaseModel):
    # 기본값 135=판매자취소(고객변심)=구매자귀책(셀러 무페널티). 사유 미지정 시 안전값.
    # 롯데ON clmRsnCd표: 111=판매자취소(판매자)=판매자귀책 페널티 → 기본값으로 위험(#592).
    # (111=품절 132=가격오등록 133=리셀러 135=고객변심 137=택배불가)
    reason_code: str = "135"
    reason_text: Optional[str] = None


@router.post("/{order_id}/seller-cancel")
async def seller_cancel(
    order_id: str,
    body: SellerCancelBody,
    session: AsyncSession = Depends(get_write_session_dependency),
):
    """판매자 주도 주문 취소 (재고부족/가격변동 등)."""
    from backend.domain.samba.account.repository import SambaMarketAccountRepository

    svc = _write_service(session)
    order = await svc.get_order(order_id)
    if not order:
        raise HTTPException(status_code=404, detail="주문을 찾을 수 없습니다")
    if not order.order_number:
        raise HTTPException(status_code=400, detail="상품주문번호가 없습니다")
    if not order.channel_id:
        raise HTTPException(status_code=400, detail="마켓 계정 정보가 없습니다")

    account_repo = SambaMarketAccountRepository(session)
    account = await account_repo.get_async(order.channel_id)
    if not account:
        raise HTTPException(status_code=400, detail="마켓 계정을 찾을 수 없습니다")

    if account.market_type == "lotteon":
        from backend.domain.samba.proxy.lotteon import LotteonClient

        extras = account.additional_fields or {}
        api_key = extras.get("apiKey", "") or account.api_key or ""
        if not api_key:
            raise HTTPException(status_code=400, detail="롯데ON API Key 없음")

        client = LotteonClient(api_key)
        try:
            await client.test_auth()
            success, message = await client.seller_cancel_order(
                od_no=order.od_no or order.order_number,
                reason_code=body.reason_code,
                reason_text=body.reason_text or "고객변심",
                od_seq=int(order.od_seq or 1),
                proc_seq=int(order.proc_seq or 1),
            )
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"판매자 취소 실패: {e}")

        if not success:
            raise HTTPException(status_code=500, detail=f"판매자 취소 실패: {message}")

        await svc.update_order(
            order_id,
            {"shipping_status": "취소완료", "status": "cancelled"},
        )
        # 롯데ON은 단일 itemList 요청으로 같은 odNo의 모든 옵션이 함께 취소됨.
        # 삼바 DB도 같은 odNo의 다른 옵션 레코드를 일괄 cancelled 처리해 UI 정합성 유지.
        od_no_val = order.od_no
        sibling_count = 0
        if od_no_val:
            from sqlmodel import select

            sibling_stmt = (
                select(SambaOrder)
                .where(SambaOrder.od_no == od_no_val)
                .where(SambaOrder.channel_id == order.channel_id)
                .where(SambaOrder.id != order_id)
                .where(SambaOrder.status != "cancelled")
            )
            sibling_rows = (await session.execute(sibling_stmt)).scalars().all()
            for sib in sibling_rows:
                await svc.update_order(
                    sib.id,
                    {"shipping_status": "취소완료", "status": "cancelled"},
                )
            sibling_count = len(sibling_rows)
        if sibling_count:
            logger.info(
                f"[판매자취소] 롯데ON {order.order_number} 동일 주문 옵션 {sibling_count}건 동반 취소"
            )
        logger.info(
            f"[판매자취소] 롯데ON {order.order_number} 완료 ({body.reason_code})"
        )
        user_msg = (
            "이미 취소된 주문 — DB 상태 갱신 완료"
            if message == "이미 취소된 주문"
            else "판매자 취소 완료"
        )
        return {"ok": True, "message": user_msg, "detail": message}

    elif account.market_type == "smartstore":
        from backend.domain.samba.proxy.smartstore import SmartStoreClient
        from backend.domain.samba.forbidden.repository import SambaSettingsRepository

        extras = account.additional_fields or {}
        client_id = extras.get("clientId", "") or account.api_key or ""
        client_secret = extras.get("clientSecret", "") or account.api_secret or ""
        if not client_id or not client_secret:
            settings_repo = SambaSettingsRepository(session)
            row = await settings_repo.find_by_async(key="store_smartstore")
            if row and isinstance(row.value, dict):
                client_id = client_id or row.value.get("clientId", "")
                client_secret = client_secret or row.value.get("clientSecret", "")
        if not client_id or not client_secret:
            raise HTTPException(status_code=400, detail="스마트스토어 인증정보 없음")

        client = SmartStoreClient(client_id, client_secret)
        try:
            await client.request_cancel(
                product_order_id=order.order_number,
                cancel_reason="INTENT_CHANGED",
            )
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"판매자 취소 실패: {e}")

        await svc.update_order(
            order_id,
            {"shipping_status": "취소완료", "status": "cancelled"},
        )
        logger.info(
            f"[판매자취소] 스마트스토어 {order.order_number} 완료 (INTENT_CHANGED)"
        )
        return {"ok": True, "message": "판매자 취소 완료"}

    elif account.market_type == "playauto":
        # 플레이오토 EMP API는 주문확인 상태변경 미지원 (송장입력만 가능)
        # DB 상태만 변경하여 이행 불가 건 구분용으로 사용
        await svc.update_order(
            order_id,
            {"shipping_status": "주문확인"},
        )
        logger.info(f"[주문확인] 플레이오토 {order.order_number} 주문확인 완료 (DB)")
        return {"ok": True, "message": "주문확인 완료"}

    elif account.market_type == "ebay":
        from backend.domain.samba.proxy.ebay import EbayApiError, EbayClient

        extras = account.additional_fields or {}
        app_id = extras.get("clientId") or extras.get("appId") or account.api_key or ""
        cert_id = (
            extras.get("clientSecret")
            or extras.get("certId")
            or account.api_secret
            or ""
        )
        refresh_token = extras.get("oauthToken") or extras.get("authToken", "") or ""
        if not (app_id and cert_id and refresh_token):
            raise HTTPException(status_code=400, detail="eBay 인증정보 없음")

        client = EbayClient(
            app_id=app_id,
            dev_id="",
            cert_id=cert_id,
            refresh_token=refresh_token,
            sandbox=bool(extras.get("sandbox", False)),
        )
        # order_number에 legacyOrderId 저장되어 있음
        try:
            reason_map = {
                "111": "OUT_OF_STOCK_OR_CANNOT_FULFILL",
                "SOLD_OUT": "OUT_OF_STOCK_OR_CANNOT_FULFILL",
                "112": "BUYER_CANCEL_OR_ADDRESS_ISSUE",
                "113": "BUYER_ASKED_CANCEL",
            }
            ebay_reason = reason_map.get(
                body.reason_code, "OUT_OF_STOCK_OR_CANNOT_FULFILL"
            )
            await client.seller_cancel_order(
                legacy_order_id=order.order_number,
                reason=ebay_reason,
            )
        except EbayApiError as e:
            raise HTTPException(status_code=500, detail=f"eBay 취소 실패: {e}")

        await svc.update_order(
            order_id,
            {"shipping_status": "취소요청", "status": "cancel_requested"},
        )
        logger.info(f"[판매자취소] eBay {order.order_number} 취소 요청 완료")
        return {"ok": True, "message": "eBay 판매자 취소 요청 완료"}

    elif account.market_type == "11st":
        # 11번가 판매불가처리 (재고부족 등 판매자 주도 취소)
        # 사유코드 20(구매의사 없어짐, 구매자 귀책) 고정 — 신용점수 차감 회피.
        # 기존 코드 10은 셀러오피스 사유 목록에 없는 폐기 코드로, 11번가 화면에
        # "타사이트 상품주문"으로 표기되던 버그 (2026-07-09 셀러오피스 실측 교정).
        # 운영 가이드: 고객 동의 후 진행
        from backend.domain.samba.proxy.elevenst import (
            ElevenstApiError,
            ElevenstClient,
        )

        api_key = (
            (account.additional_fields or {}).get("apiKey", "") or account.api_key or ""
        )
        if not api_key:
            raise HTTPException(status_code=400, detail="11번가 API Key 없음")

        if not order.ord_prd_seq:
            raise HTTPException(
                status_code=400,
                detail="11번가 ordPrdSeq 미수집 — 주문 동기화 후 다시 시도해주세요",
            )

        client = ElevenstClient(api_key)
        try:
            await client.reject_order(
                ord_no=order.order_number,
                ord_prd_seq=order.ord_prd_seq,
                ord_cn_rsn_cd="20",  # 구매의사 없어짐 (구매자 귀책)
                ord_cn_dtls_rsn="구매자 요청으로 취소 처리",
            )
        except ElevenstApiError as e:
            raise HTTPException(
                status_code=500, detail=f"11번가 판매불가처리 실패: {e}"
            )
        except Exception as e:
            raise HTTPException(
                status_code=500, detail=f"11번가 판매불가처리 실패: {e}"
            )

        await svc.update_order(
            order_id,
            {"shipping_status": "취소완료", "status": "cancelled"},
        )
        logger.info(
            f"[판매자취소] 11번가 {order.order_number} 판매불가처리 완료 (사유=20/구매의사 없어짐)"
        )
        return {"ok": True, "message": "11번가 판매불가처리 완료"}

    elif account.market_type == "coupang":
        # 쿠팡 판매자 능동 취소 — POST .../orders/{orderId}/cancel (#246 PR-4)
        # 판매자 귀책 사유코드만 사용 가능 → 판매자 점수 하락 주의
        from backend.domain.samba.proxy.coupang import CoupangApiError, CoupangClient

        # SambaOrder reason_code → 쿠팡 middleCancelCode 매핑
        #   111(품절)     → CCTTER 재고 연동 오류
        #   132(가격오등록)→ CCPRER 가격등재오류
        #   133(리셀러)   → CCPNER 제휴사이트 오류
        #   135(고객변심) → CCTTER (판매자 능동에는 고객 귀책 코드 불가 → 재고 fallback)
        #   137(택배불가) → CCPNER 배송지 문제
        coupang_reason_map = {
            "111": "CCTTER",
            "132": "CCPRER",
            "133": "CCPNER",
            "135": "CCTTER",
            "137": "CCPNER",
        }
        middle_code = coupang_reason_map.get(body.reason_code, "CCTTER")

        extras = account.additional_fields or {}
        access_key = extras.get("accessKey", "") or account.api_key or ""
        secret_key = extras.get("secretKey", "") or account.api_secret or ""
        vendor_id = extras.get("vendorId", "") or account.seller_id or ""
        if not all([access_key, secret_key, vendor_id]):
            raise HTTPException(
                status_code=400,
                detail="쿠팡 인증정보 없음 (accessKey/secretKey/vendorId)",
            )
        if not order.shipment_id:
            raise HTTPException(
                status_code=400,
                detail="쿠팡 orderId(shipment_id) 미수집 — 동기화 후 재시도",
            )
        if not order.vendor_item_id:
            raise HTTPException(
                status_code=400,
                detail="쿠팡 vendorItemId 미수집 — 동기화 후 재시도",
            )

        # userId: 쿠팡 공식 명세상 의미 미확정 (wing 로그인 ID 추정).
        # extras["coupangUserId"] 우선 → 없으면 vendor_id fallback. 운영 실측 후 보정.
        coupang_user_id = extras.get("coupangUserId") or vendor_id

        client = CoupangClient(access_key, secret_key, vendor_id)
        try:
            await client.seller_cancel_order(
                order_id=int(order.shipment_id),
                vendor_item_ids=[int(order.vendor_item_id)],
                receipt_counts=[int(order.quantity or 1)],
                middle_cancel_code=middle_code,
                user_id=str(coupang_user_id),
            )
        except CoupangApiError as e:
            raise HTTPException(status_code=500, detail=f"쿠팡 판매자 취소 실패: {e}")
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"쿠팡 판매자 취소 실패: {e}")

        await svc.update_order(
            order_id,
            {"shipping_status": "취소완료", "status": "cancelled"},
        )
        logger.info(
            f"[판매자취소] 쿠팡 {order.order_number} cancel 완료 "
            f"(reason={body.reason_code}/{middle_code})"
        )
        return {"ok": True, "message": "쿠팡 판매자 취소 완료"}

    elif account.market_type == "lottehome":
        # 롯데홈쇼핑은 판매자 취소 = registDeliver(proc_gubun=imps) 발송불가 처리
        # order_number 형식: "ord_no:ord_dtl_sn"
        from backend.domain.samba.proxy.lottehome import LotteHomeClient

        raw = order.order_number or ""
        parts = raw.split(":")
        if len(parts) < 2 or not parts[0] or not parts[1]:
            raise HTTPException(
                status_code=400,
                detail=f"롯데홈쇼핑 주문번호 형식 오류 (ord_no:ord_dtl_sn 필요, 현재값={raw!r})",
            )
        ord_no, ord_dtl_sn = parts[0], parts[1]

        extras = account.additional_fields or {}
        user_id = extras.get("userId", "") or account.seller_id or ""
        password = extras.get("password", "")
        agnc_no = extras.get("agncNo", "")
        env = extras.get("env", "prod")
        if not user_id:
            raise HTTPException(status_code=400, detail="롯데홈쇼핑 로그인 정보 없음")

        client = LotteHomeClient(user_id, password, agnc_no, env)
        try:
            res = await client.register_deliver(ord_no, ord_dtl_sn, "imps")
        except Exception as e:
            raise HTTPException(
                status_code=500, detail=f"롯데홈쇼핑 발송불가 처리 실패: {e}"
            )
        if not res.get("ok"):
            raise HTTPException(
                status_code=500,
                detail=f"롯데홈쇼핑 발송불가 처리 실패: result={res.get('result')}",
            )

        await svc.update_order(
            order_id,
            {"shipping_status": "취소완료", "status": "cancelled"},
        )
        logger.info(
            f"[판매자취소][롯데홈쇼핑] {ord_no}:{ord_dtl_sn} 발송불가(imps) 처리 완료"
        )
        return {"ok": True, "message": "롯데홈쇼핑 발송불가 처리 완료"}

    elif account.market_type == "ssg":
        # SSG 판매자취소 = 결품등록(saveNoSellRequestRegist, scEvnt=I).
        # 등록 후 익일 17시 SSG가 자동으로 취소/환불처리 + 고객에게 품절안내 발송.
        # 배송지시/피킹완료 상태에서만 가능.
        # shipment_id 형식: "shppNo|shppSeq" — parse_order에서 저장.
        from backend.domain.samba.proxy.ssg import SSGClient
        from backend.domain.samba.forbidden.repository import SambaSettingsRepository

        raw_shipment = order.shipment_id or ""
        parts = raw_shipment.split("|")
        if len(parts) < 2 or not parts[0] or not parts[1]:
            raise HTTPException(
                status_code=400,
                detail=f"SSG 배송번호 형식 오류 (shppNo|shppSeq 필요, 현재값={raw_shipment!r})",
            )
        shpp_no, shpp_seq = parts[0], parts[1]

        item_id = order.product_id or ""
        if not item_id:
            raise HTTPException(status_code=400, detail="SSG 상품코드(itemId) 없음")

        extras = account.additional_fields or {}
        ssg_api_key = extras.get("apiKey", "") or account.api_key or ""
        if not ssg_api_key:
            settings_repo = SambaSettingsRepository(session)
            row = await settings_repo.find_by_async(key="store_ssg")
            if row and isinstance(row.value, dict):
                ssg_api_key = row.value.get("apiKey", "") or ""
        if not ssg_api_key:
            raise HTTPException(status_code=400, detail="SSG API Key 없음")

        client = SSGClient(ssg_api_key)
        try:
            await client.register_no_sell(
                shpp_no=shpp_no,
                shpp_seq=shpp_seq,
                item_id=item_id,
                reason_code="08",  # 08 = 상품정보오류 → SSG 화면에 "품절(상품정보/가격오류)" 표시
                reason_text="품절",
            )
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"SSG 결품 등록 실패: {e}")

        # 결품 등록은 신청 상태 — SSG가 익일 17시에 자동으로 취소완료 처리하므로
        # 즉시 "취소완료"가 아닌 "취소요청"으로 마킹. 자동 동기화 시 취소완료로 갱신됨.
        await svc.update_order(
            order_id,
            {"shipping_status": "취소요청", "status": "cancel_requested"},
        )
        logger.info(
            f"[판매자취소][SSG] {order.order_number} 결품 등록 완료 "
            f"(shppNo={shpp_no}, shppSeq={shpp_seq}, itemId={item_id})"
        )
        return {
            "ok": True,
            "message": "SSG 결품 등록 완료 — 익일 17시 자동 취소/환불 예정",
        }

    raise HTTPException(
        status_code=400, detail=f"{account.market_type} 판매자 취소 미지원"
    )


class ApproveCancelWithShipmentBody(BaseModel):
    """쿠팡 이미출고 케이스 — 운영자 송장 정보 입력 (#246 PR-4)."""

    delivery_company_code: str
    invoice_number: str


@router.post("/{order_id}/approve-cancel-with-shipment")
async def approve_cancel_with_shipment(
    order_id: str,
    body: ApproveCancelWithShipmentBody,
    session: AsyncSession = Depends(get_write_session_dependency),
):
    """쿠팡 이미출고 취소승인 — completedShipment 처리 (#246 PR-4).

    조건: order.cancel_release_status == 'A'.
    운영자가 실제 발송한 택배사·송장번호를 입력해야 호출 가능.
    주의: 왕복 배송비 판매자 부담.
    """
    from backend.domain.samba.account.repository import SambaMarketAccountRepository
    from backend.domain.samba.proxy.coupang import CoupangApiError, CoupangClient

    svc = _write_service(session)
    order = await svc.get_order(order_id)
    if not order:
        raise HTTPException(status_code=404, detail="주문을 찾을 수 없습니다")
    if not order.channel_id:
        raise HTTPException(status_code=400, detail="마켓 계정 정보가 없습니다")
    if not order.cancel_receipt_id:
        raise HTTPException(status_code=400, detail="쿠팡 취소 receiptId 미수집")
    if (order.cancel_release_status or "").upper() != "A":
        raise HTTPException(
            status_code=400,
            detail=(
                f"release_status={order.cancel_release_status or 'None'} — "
                "이미출고(A) 케이스만 이 엔드포인트 사용. 미출고는 /approve-cancel 호출"
            ),
        )

    account_repo = SambaMarketAccountRepository(session)
    account = await account_repo.get_async(order.channel_id)
    if not account or account.market_type != "coupang":
        raise HTTPException(status_code=400, detail="쿠팡 계정에만 사용 가능")

    extras = account.additional_fields or {}
    access_key = extras.get("accessKey", "") or account.api_key or ""
    secret_key = extras.get("secretKey", "") or account.api_secret or ""
    vendor_id = extras.get("vendorId", "") or account.seller_id or ""
    if not all([access_key, secret_key, vendor_id]):
        raise HTTPException(status_code=400, detail="쿠팡 인증정보 없음")

    client = CoupangClient(access_key, secret_key, vendor_id)
    try:
        await client.confirm_completed_shipment(
            receipt_id=int(order.cancel_receipt_id),
            delivery_company_code=body.delivery_company_code,
            invoice_number=body.invoice_number,
        )
    except CoupangApiError as e:
        raise HTTPException(status_code=500, detail=f"쿠팡 이미출고 처리 실패: {e}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"쿠팡 이미출고 처리 실패: {e}")

    await svc.update_order(
        order_id,
        {"shipping_status": "취소완료(이미출고)", "status": "cancelled"},
    )
    logger.info(
        f"[취소승인] 쿠팡 {order.order_number} completedShipment 완료 "
        f"(receiptId={order.cancel_receipt_id}, company={body.delivery_company_code}, "
        f"invoice={body.invoice_number})"
    )
    return {"ok": True, "message": "쿠팡 이미출고 취소승인 완료"}


@router.post("/{order_id}/reject-cancel")
async def reject_cancel(
    order_id: str,
    session: AsyncSession = Depends(get_write_session_dependency),
):
    """취소 거부 — 내부 상태만 cancel_reject_pending 으로 표시 (#246 PR-4).

    쿠팡: 거부 전용 공식 API 없음 → 운영자에게 Wing 화면에서 수동 처리 안내.
    프론트는 응답 후 토스트로 안내 표시.
    """
    from backend.domain.samba.account.repository import SambaMarketAccountRepository

    svc = _write_service(session)
    order = await svc.get_order(order_id)
    if not order:
        raise HTTPException(status_code=404, detail="주문을 찾을 수 없습니다")
    if not order.channel_id:
        raise HTTPException(status_code=400, detail="마켓 계정 정보가 없습니다")

    account_repo = SambaMarketAccountRepository(session)
    account = await account_repo.get_async(order.channel_id)
    if not account:
        raise HTTPException(status_code=400, detail="마켓 계정을 찾을 수 없습니다")

    if account.market_type == "coupang":
        await svc.update_order(
            order_id,
            {"status": "cancel_reject_pending"},
        )
        logger.info(
            f"[취소거부] 쿠팡 {order.order_number} 내부 pending 처리 "
            "(Wing 수동 처리 필요)"
        )
        return {
            "ok": True,
            "message": "쿠팡 취소거부 — Wing 화면에서 수동 처리해주세요",
            "manual_required": True,
        }

    raise HTTPException(
        status_code=400, detail=f"{account.market_type} 취소 거부 미지원"
    )


@router.post("/{order_id}/confirm")
async def confirm_order(
    order_id: str,
    session: AsyncSession = Depends(get_write_session_dependency),
):
    """주문확인(발주확인) 수동 처리 — 원소싱처 재고/가격 확인 후 사용자가 실행."""
    from backend.domain.samba.account.repository import SambaMarketAccountRepository
    from backend.domain.samba.order.model import is_order_cancelled

    svc = _write_service(session)
    order = await svc.get_order(order_id)
    if not order:
        raise HTTPException(status_code=404, detail="주문을 찾을 수 없습니다")
    # 취소 가드 — 발주확인(주문확인) 직전 차단. 마켓 인지 후 잘못 발주되는 사고 방지.
    if is_order_cancelled(order):
        raise HTTPException(
            status_code=409,
            detail=(
                f"취소요청 상태(주문={order.status}/마켓={order.shipping_status})라 "
                "발주확인을 진행할 수 없습니다"
            ),
        )
    if not order.order_number:
        raise HTTPException(status_code=400, detail="상품주문번호가 없습니다")
    if not order.channel_id:
        raise HTTPException(status_code=400, detail="마켓 계정 정보가 없습니다")

    account_repo = SambaMarketAccountRepository(session)
    account = await account_repo.get_async(order.channel_id)
    if not account:
        raise HTTPException(status_code=400, detail="마켓 계정을 찾을 수 없습니다")

    if account.market_type == "lotteon":
        from backend.domain.samba.proxy.lotteon import LotteonClient

        extras = account.additional_fields or {}
        api_key = extras.get("apiKey", "") or account.api_key or ""
        if not api_key:
            raise HTTPException(status_code=400, detail="롯데ON API Key 없음")

        # SellerIfCompleteInform은 odNo/odSeq/procSeq만 필요 (비클레임은 기본 1/1)
        client = LotteonClient(api_key)
        try:
            await client.test_auth()
            ok = await client.confirm_orders(
                [
                    {
                        "odNo": order.od_no or order.order_number,
                        "odSeq": int(order.od_seq or 1),
                        "procSeq": int(order.proc_seq or 1),
                    }
                ]
            )
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"주문확인 실패: {e}")

        if not ok:
            raise HTTPException(
                status_code=500,
                detail="롯데ON 주문확인 실패 — SellerIfCompleteInform 응답 rsltCd≠0000 (서버 로그 확인)",
            )

        await svc.update_order(order_id, {"shipping_status": "출고지시"})
        logger.info(f"[주문확인] 롯데ON {order.order_number} 완료")
        return {"ok": True, "message": "주문확인 완료"}

    if account.market_type in ("gmarket", "auction"):
        from backend.domain.samba.proxy.esmplus import (
            ESMPlusClient,
            resolve_esm_credentials,
        )

        extras = account.additional_fields or {}
        hosting_id, secret_key = await resolve_esm_credentials(session, account)
        seller_id = (
            extras.get("apiKey") or extras.get("sellerId") or (account.seller_id or "")
        ).strip()
        if not (hosting_id and secret_key and seller_id):
            raise HTTPException(status_code=400, detail="ESM 인증정보 없음")
        client = ESMPlusClient(
            hosting_id, secret_key, seller_id, site=account.market_type
        )
        try:
            await client.confirm_order(order.order_number)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"주문확인 실패: {e}")
        finally:
            await client.aclose()
        await svc.update_order(order_id, {"shipping_status": "배송준비중"})
        logger.info(f"[주문확인] ESM({account.market_type}) {order.order_number} 완료")
        return {"ok": True, "message": "주문확인 완료"}

    raise HTTPException(
        status_code=400, detail=f"{account.market_type} 주문확인 미지원"
    )


@router.post("/{order_id}/market-delete")
async def market_delete_order_product(
    order_id: str,
    session: AsyncSession = Depends(get_write_session_dependency),
):
    """주문 카드의 '마켓상품삭제' — 해당 주문 상품을 마켓에서 완전 삭제(판매종료가 아닌 삭제)."""
    from backend.domain.samba.account.repository import SambaMarketAccountRepository

    svc = _write_service(session)
    order = await svc.get_order(order_id)
    if not order:
        raise HTTPException(status_code=404, detail="주문을 찾을 수 없습니다")
    if not order.product_id:
        raise HTTPException(status_code=400, detail="마켓 상품번호가 없습니다")
    if not order.channel_id:
        raise HTTPException(status_code=400, detail="마켓 계정 정보가 없습니다")

    account_repo = SambaMarketAccountRepository(session)
    account = await account_repo.get_async(order.channel_id)
    if not account:
        raise HTTPException(status_code=400, detail="마켓 계정을 찾을 수 없습니다")

    if account.market_type == "lotteon":
        from backend.domain.samba.proxy.lotteon import LotteonClient

        extras = account.additional_fields or {}
        api_key = extras.get("apiKey", "") or account.api_key or ""
        if not api_key:
            raise HTTPException(status_code=400, detail="롯데ON API Key 없음")

        spd_no = order.product_id
        client = LotteonClient(api_key)
        try:
            await client.test_auth()
            result = await client.delete_product(spd_no)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"마켓상품삭제 실패: {e}")

        logger.info(
            f"[마켓상품삭제] 롯데ON spdNo={spd_no} order={order.order_number} result={result}"
        )
        return {"ok": True, "message": "마켓 상품 삭제 완료", "detail": result}

    if account.market_type == "smartstore":
        from backend.domain.samba.proxy.smartstore import SmartStoreClient

        extras = account.additional_fields or {}
        client_id = extras.get("clientId", "") or account.api_key or ""
        client_secret = extras.get("clientSecret", "") or account.api_secret or ""
        if not client_id or not client_secret:
            raise HTTPException(status_code=400, detail="스마트스토어 인증 정보 없음")

        # originProductNo: collected_product의 market_product_nos에서 우선 조회
        origin_product_no = ""
        if order.collected_product_id:
            from backend.domain.samba.collector.repository import (
                SambaCollectorRepository,
            )

            cp_repo = SambaCollectorRepository(session)
            cp = await cp_repo.get_async(order.collected_product_id)
            if cp and cp.market_product_nos:
                origin_product_no = (cp.market_product_nos or {}).get(
                    order.channel_id, ""
                )

        # fallback: channelProductNo (order.product_id)
        if not origin_product_no:
            origin_product_no = order.product_id or ""

        if not origin_product_no:
            raise HTTPException(
                status_code=400, detail="스마트스토어 상품번호를 찾을 수 없습니다"
            )

        client = SmartStoreClient(client_id, client_secret)
        try:
            result = await client.delete_product(origin_product_no)
            logger.info(
                f"[마켓상품삭제] 스마트스토어 삭제 성공 productNo={origin_product_no} "
                f"order={order.order_number}"
            )
            return {"ok": True, "message": "마켓 상품 삭제 완료", "detail": result}
        except Exception as del_err:
            # 진행중 주문 등으로 삭제 불가 시 → 전 옵션 재고 0 (품절) 폴백
            logger.warning(
                f"[마켓상품삭제] 스마트스토어 삭제 실패({del_err}), 품절 폴백 시도: {origin_product_no}"
            )

        try:
            existing = await client.get_product(origin_product_no)
            origin = existing.get("originProduct", {})
            for k in ["productNo", "channelProducts", "regDate", "modifiedDate"]:
                origin.pop(k, None)

            # 전 옵션 재고 0 + usable=False
            origin["stockQuantity"] = 0
            opt_info = origin.get("detailAttribute", {}).get("optionInfo") or {}
            combos = opt_info.get("optionCombinations") or opt_info.get(
                "combinations", []
            )
            for combo in combos:
                combo["stockQuantity"] = 0
                combo["usable"] = False

            put_data: dict[str, Any] = {"originProduct": origin}
            if "smartstoreChannelProduct" in existing:
                put_data["smartstoreChannelProduct"] = existing[
                    "smartstoreChannelProduct"
                ]

            await client.update_product(origin_product_no, put_data)
            logger.info(
                f"[마켓상품삭제] 스마트스토어 품절 폴백 완료 productNo={origin_product_no}"
            )
            return {
                "ok": True,
                "message": "마켓 삭제 불가 — 전 옵션 품절처리 완료",
                "fallback": True,
            }
        except Exception as fb_err:
            raise HTTPException(
                status_code=500,
                detail=f"마켓상품삭제 및 품절처리 모두 실패: {fb_err}",
            )

    raise HTTPException(
        status_code=400, detail=f"{account.market_type} 마켓상품삭제 미지원"
    )


class CancelSourceOrderRequest(BaseModel):
    order_number: str
    reason: str = "단순변심"


@router.post("/cancel-source-order")
async def cancel_source_order(
    req: CancelSourceOrderRequest,
    session: AsyncSession = Depends(get_write_session_dependency),
):
    """소싱처 원주문 취소 (무신사 등 소비자 주문취소)."""
    from backend.domain.samba.forbidden.repository import SambaSettingsRepository

    settings_repo = SambaSettingsRepository(session)

    # 현재는 무신사만 지원
    cookie_row = await settings_repo.find_by_async(key="musinsa_cookie")
    musinsa_cookie = cookie_row.value if cookie_row else ""
    if not musinsa_cookie:
        raise HTTPException(status_code=400, detail="무신사 쿠키가 설정되지 않았습니다")

    from backend.domain.samba.proxy.musinsa import MusinsaClient

    client = MusinsaClient(cookie=musinsa_cookie)

    try:
        result = await client.cancel_order(req.order_number, req.reason)
        return result
    except Exception as e:
        logger.error(f"[원주문취소] 실패: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ══════════════════════════════════════════════
# 교환 처리 (재배송 / 거부 / 반품변경)
# ══════════════════════════════════════════════


class ExchangeActionBody(BaseModel):
    action: str  # "reship" | "reject" | "convert_return"
    reason: Optional[str] = None
    clm_no: Optional[str] = None  # 롯데ON 교환 클레임번호
    tracking_number: Optional[str] = None  # 롯데ON 교환 재배송 송장번호
    shipping_company: Optional[str] = None  # 롯데ON 교환 재배송 택배사


@router.post("/{order_id}/exchange-action")
async def exchange_action(
    order_id: str,
    body: ExchangeActionBody,
    session: AsyncSession = Depends(get_write_session_dependency),
):
    """교환요청에 대한 처리 (재배송/거부/반품변경)."""
    from backend.domain.samba.account.repository import SambaMarketAccountRepository
    from backend.domain.samba.forbidden.repository import SambaSettingsRepository

    svc = _write_service(session)
    order = await svc.get_order(order_id)
    if not order:
        raise HTTPException(status_code=404, detail="주문을 찾을 수 없습니다")
    if not order.order_number:
        raise HTTPException(status_code=400, detail="상품주문번호가 없습니다")
    if not order.channel_id:
        raise HTTPException(status_code=400, detail="마켓 계정 정보가 없습니다")

    account_repo = SambaMarketAccountRepository(session)
    account = await account_repo.get_async(order.channel_id)
    if not account:
        raise HTTPException(status_code=400, detail="마켓 계정을 찾을 수 없습니다")

    if account.market_type == "smartstore":
        from backend.domain.samba.proxy.smartstore import SmartStoreClient

        extras = account.additional_fields or {}
        client_id = extras.get("clientId", "") or account.api_key or ""
        client_secret = extras.get("clientSecret", "") or account.api_secret or ""
        if not client_id or not client_secret:
            settings_repo = SambaSettingsRepository(session)
            row = await settings_repo.find_by_async(key="store_smartstore")
            if row and isinstance(row.value, dict):
                client_id = client_id or row.value.get("clientId", "")
                client_secret = client_secret or row.value.get("clientSecret", "")
        if not client_id or not client_secret:
            raise HTTPException(status_code=400, detail="스마트스토어 인증정보 없음")

        client = SmartStoreClient(client_id, client_secret)
        action_labels = {
            "reship": "교환재배송",
            "reject": "교환거부",
            "convert_return": "반품변경",
        }
        label = action_labels.get(body.action, body.action)

        try:
            if body.action == "reship":
                await client.approve_exchange(order.order_number)
                new_status = "교환완료"
            elif body.action == "reject":
                await client.reject_exchange(
                    order.order_number, body.reason or "판매자 교환 거부"
                )
                new_status = "교환거부"
            elif body.action == "convert_return":
                await client.convert_exchange_to_return(order.order_number)
                new_status = "반품변경"
            else:
                raise HTTPException(
                    status_code=400, detail=f"알 수 없는 액션: {body.action}"
                )
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"{label} 실패: {e}")

        await svc.update_order(order_id, {"shipping_status": new_status})
        logger.info(f"[교환처리] {order.order_number} {label} 완료")
        return {"ok": True, "message": f"{label} 완료"}

    elif account.market_type == "lotteon":
        from backend.domain.samba.proxy.lotteon import LotteonClient

        extras = account.additional_fields or {}
        api_key = extras.get("apiKey", "") or account.api_key or ""
        if not api_key:
            raise HTTPException(status_code=400, detail="롯데ON API 키 없음")

        client = LotteonClient(api_key=api_key)
        await client.test_auth()

        # 교환 클레임 정보 자동 탐색 (clmNo, procSeq, orglProcSeq)
        clm_no = body.clm_no or ""
        found_claim: dict = {}
        try:
            exchange_claims = await client.get_exchanges(days=30)
            for claim in exchange_claims:
                if str(claim.get("odNo", "")) == str(order.od_no or order.order_number):
                    if not clm_no:
                        clm_no = claim.get("clmNo", "")
                    found_claim = claim
                    logger.info(
                        f"[교환처리] clmNo 탐색 성공: {clm_no} stepCd={claim.get('odPrgsStepCd', '')}"
                    )
                    break
        except Exception as ce:
            logger.warning(f"[교환처리] 클레임 탐색 실패: {ce}")

        if body.action == "reship":
            # 교환 재배송: 승인 → 발송 처리
            tracking_number = body.tracking_number or ""
            shipping_company = body.shipping_company or ""
            sitm_no = order.shipment_id or ""
            spd_no = order.product_id or ""
            quantity = order.quantity or 1

            if not tracking_number:
                raise HTTPException(
                    status_code=400, detail="교환 재배송 송장번호가 필요합니다"
                )

            # 교환 승인 (회수 지시) — 접수(03) 상태인 경우 먼저 승인
            step_cd = str(found_claim.get("odPrgsStepCd", "") or "")
            if step_cd == "03" and clm_no:
                proc_seq = str(found_claim.get("procSeq", 1))
                orgl_proc_seq = str(found_claim.get("orglProcSeq", 1))
                clm_rsn_cd = str(found_claim.get("clmRsnCd", "204"))
                try:
                    approved = await client.approve_exchange(
                        od_no=order.od_no or order.order_number,
                        clm_no=clm_no,
                        items=[
                            {
                                "odSeq": int(order.od_seq or 1),
                                "procSeq": int(proc_seq),
                                "orglProcSeq": int(orgl_proc_seq),
                                "slrRsnCd": clm_rsn_cd,
                            }
                        ],
                    )
                    if approved:
                        logger.info(f"[교환처리] {order.order_number} 교환 승인 완료")
                except Exception as ae:
                    logger.warning(f"[교환처리] 교환 승인 실패 (계속 진행): {ae}")

            try:
                sent = await client.ship_order_exchange(
                    od_no=order.od_no or order.order_number,
                    od_seq=order.od_seq or "1",
                    proc_seq=order.proc_seq or "1",
                    sitm_no=sitm_no,
                    spd_no=spd_no,
                    clm_no=clm_no,
                    quantity=quantity,
                    shipping_company=shipping_company,
                    tracking_number=tracking_number,
                )
                if not sent:
                    raise HTTPException(
                        status_code=500, detail="롯데ON 교환 재배송 전송 실패"
                    )
            except HTTPException:
                raise
            except Exception as e:
                raise HTTPException(status_code=500, detail=f"교환 재배송 실패: {e}")

            await svc.update_order(
                order_id,
                {
                    "shipping_status": "교환재배송",
                    "tracking_number": tracking_number,
                    "shipping_company": shipping_company,
                },
            )
            logger.info(f"[교환처리] {order.order_number} 롯데ON 교환재배송 완료")
            return {"ok": True, "message": "교환 재배송 처리 완료"}

        elif body.action == "convert_return":
            # 교환→반품 변경: 롯데ON API 미지원 → 삼바 내부 처리만
            # 반품교환 레코드 타입을 exchange→return으로 변경
            from backend.domain.samba.returns.repository import SambaReturnRepository

            return_repo = SambaReturnRepository(session)
            ret = await return_repo.find_by_async(order_id=order_id)
            if ret:
                await return_repo.update_async(
                    ret.id,
                    type="return",
                    market_order_status="반품요청",
                    status="pending",
                )
            await svc.update_order(
                order_id, {"shipping_status": "반품요청", "status": "return_requested"}
            )
            logger.info(
                f"[교환처리] {order.order_number} 교환→반품 변경 완료 (삼바 내부)"
            )
            return {
                "ok": True,
                "message": "교환→반품 변경 완료 (롯데ON 판매자센터에서도 별도 처리 필요)",
            }

        elif body.action == "reject":
            # 교환 거부: 삼바 내부 상태 업데이트 (롯데ON 교환 거부 API 스펙 확인 후 연동 필요)
            from backend.domain.samba.returns.repository import SambaReturnRepository

            return_repo = SambaReturnRepository(session)
            ret = await return_repo.find_by_async(order_id=order_id)
            if ret:
                await return_repo.update_async(
                    ret.id,
                    status="rejected",
                    market_order_status="교환거부",
                )
            await svc.update_order(order_id, {"shipping_status": "교환거부"})
            logger.info(f"[교환처리] {order.order_number} 교환거부 완료 (삼바 내부)")
            return {
                "ok": True,
                "message": "교환거부 완료 (롯데ON 판매자센터에서도 별도 처리 필요)",
            }

        else:
            raise HTTPException(
                status_code=400, detail=f"롯데ON 교환처리 미지원 액션: {body.action}"
            )

    elif account.market_type == "11st":
        from backend.domain.samba.forbidden.repository import SambaSettingsRepository
        from backend.domain.samba.proxy.elevenst_exchange import (
            ElevenstApiError,
            ElevenstExchangeClient,
        )
        from backend.domain.samba.returns.repository import SambaReturnRepository

        api_key = account.api_key or ""
        if not api_key:
            # account.api_key 미설정 시 settings 테이블의 store_11st.apiKey fallback
            settings_repo = SambaSettingsRepository(session)
            st_row = await settings_repo.find_by_async(key="store_11st")
            if st_row and isinstance(st_row.value, dict):
                api_key = st_row.value.get("apiKey", "") or ""
        if not api_key:
            raise HTTPException(status_code=400, detail="11번가 API 키가 없습니다")

        return_repo = SambaReturnRepository(session)
        ret_records = await return_repo.list_by_order(order_id)
        ret = next((r for r in ret_records if r.type == "exchange"), None)

        if body.action in ("reject", "approve", "reship"):
            clm_req_seq = (ret.clm_req_seq or "") if ret else ""
            ord_prd_seq = (ret.ord_prd_seq or "") if ret else ""
            ord_no = order.order_number or ""

            if not clm_req_seq or not ord_no or not ord_prd_seq:
                raise HTTPException(
                    status_code=400,
                    detail="교환 처리에 필요한 클레임 식별자(clm_req_seq, ord_no, ord_prd_seq)가 없습니다",
                )

            client = ElevenstExchangeClient(api_key)
            action_labels = {
                "reship": "교환승인(재배송)",
                "approve": "교환승인(재배송)",
                "reject": "교환거부",
            }
            label = action_labels.get(body.action, body.action)

            try:
                if body.action in ("reship", "approve"):
                    await client.confirm_exchange(clm_req_seq, ord_no, ord_prd_seq)
                    new_status = "교환승인"
                else:
                    await client.reject_exchange(
                        clm_req_seq,
                        ord_no,
                        ord_prd_seq,
                        refs_rsn_cd="204",
                        refs_rsn=body.reason or "기타",
                    )
                    new_status = "교환거부"
            except HTTPException:
                raise
            except ElevenstApiError as e:
                raise HTTPException(status_code=502, detail=f"{label} API 오류: {e}")
            except Exception as e:
                raise HTTPException(status_code=500, detail=f"{label} 실패: {e}")

            await svc.update_order(order_id, {"shipping_status": new_status})
            if ret:
                await return_repo.update_async(
                    ret.id,
                    status="approved" if new_status == "교환승인" else "rejected",
                    market_order_status=new_status,
                )
            logger.info(f"[교환처리] {order.order_number} 11번가 {label} 완료")
            return {"ok": True, "message": f"{label} 완료"}

        elif body.action == "convert_return":
            if ret:
                await return_repo.update_async(
                    ret.id,
                    type="return",
                    market_order_status="반품요청",
                    status="pending",
                )
            await svc.update_order(
                order_id, {"shipping_status": "반품요청", "status": "return_requested"}
            )
            logger.info(f"[교환처리] {order.order_number} 11번가 교환→반품 변경 완료")
            return {
                "ok": True,
                "message": "교환→반품 변경 완료 (11번가 판매자센터에서도 별도 처리 필요)",
            }

        else:
            raise HTTPException(
                status_code=400, detail=f"11번가 교환처리 미지원 액션: {body.action}"
            )

    else:
        raise HTTPException(
            status_code=400, detail=f"{account.market_type} 교환처리 미지원"
        )


# ══════════════════════════════════════════════
# 반품 처리 (승인 / 거부)
# ══════════════════════════════════════════════


class ReturnActionBody(BaseModel):
    action: str  # "approve" | "reject"
    reason: Optional[str] = None


@router.post("/{order_id}/return-action")
async def return_action(
    order_id: str,
    body: ReturnActionBody,
    session: AsyncSession = Depends(get_write_session_dependency),
):
    """반품요청에 대한 처리 (승인/거부)."""
    from backend.domain.samba.account.repository import SambaMarketAccountRepository
    from backend.domain.samba.forbidden.repository import SambaSettingsRepository

    svc = _write_service(session)
    order = await svc.get_order(order_id)
    if not order:
        raise HTTPException(status_code=404, detail="주문을 찾을 수 없습니다")
    if not order.order_number:
        raise HTTPException(status_code=400, detail="상품주문번호가 없습니다")
    if not order.channel_id:
        raise HTTPException(status_code=400, detail="마켓 계정 정보가 없습니다")

    account_repo = SambaMarketAccountRepository(session)
    account = await account_repo.get_async(order.channel_id)
    if not account:
        raise HTTPException(status_code=400, detail="마켓 계정을 찾을 수 없습니다")

    if account.market_type == "smartstore":
        from backend.domain.samba.proxy.smartstore import SmartStoreClient

        extras = account.additional_fields or {}
        client_id = extras.get("clientId", "") or account.api_key or ""
        client_secret = extras.get("clientSecret", "") or account.api_secret or ""
        if not client_id or not client_secret:
            settings_repo = SambaSettingsRepository(session)
            row = await settings_repo.find_by_async(key="store_smartstore")
            if row and isinstance(row.value, dict):
                client_id = client_id or row.value.get("clientId", "")
                client_secret = client_secret or row.value.get("clientSecret", "")
        if not client_id or not client_secret:
            raise HTTPException(status_code=400, detail="스마트스토어 인증정보 없음")

        client = SmartStoreClient(client_id, client_secret)
        label = "반품승인" if body.action == "approve" else "반품거부"

        try:
            if body.action == "approve":
                try:
                    await client.approve_return(order.order_number)
                except Exception as first_err:
                    if "환불보류" in str(first_err):
                        # 환불보류 해제 후 재시도
                        logger.info(
                            f"[반품처리] {order.order_number} 환불보류 감지 → 보류해제 후 재시도"
                        )
                        await client.release_return_hold(order.order_number)
                        await client.approve_return(order.order_number)
                    else:
                        raise
                new_status = "반품승인"
            elif body.action == "reject":
                await client.reject_return(
                    order.order_number, body.reason or "판매자 반품 거부"
                )
                new_status = "반품거부"
            else:
                raise HTTPException(
                    status_code=400, detail=f"알 수 없는 액션: {body.action}"
                )
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"{label} 실패: {e}")

        await svc.update_order(order_id, {"shipping_status": new_status})

        # 반품교환(samba_return) 레코드도 상태 업데이트
        from backend.domain.samba.returns.repository import SambaReturnRepository
        from datetime import UTC, datetime

        return_repo = SambaReturnRepository(session)
        existing_returns = await return_repo.filter_by_async(order_id=order_id)
        if existing_returns:
            ret = existing_returns[0]
            if body.action == "approve":
                await return_repo.update_async(
                    ret.id,
                    status="completed",
                    market_order_status="반품완료",
                    completion_date=datetime.now(UTC),
                )
            elif body.action == "reject":
                await return_repo.update_async(
                    ret.id,
                    status="rejected",
                    market_order_status="반품거부",
                )

        logger.info(f"[반품처리] {order.order_number} {label} 완료")
        return {"ok": True, "message": f"{label} 완료"}

    elif account.market_type == "lotteon":
        from backend.domain.samba.proxy.lotteon import LotteonClient

        api_key = (
            (account.additional_fields or {}).get("apiKey", "") or account.api_key or ""
        )
        if not api_key:
            raise HTTPException(status_code=400, detail="롯데ON API 키 없음")

        client = LotteonClient(api_key=api_key)
        label = "반품승인" if body.action == "approve" else "반품거부"

        try:
            if body.action == "approve":
                # 반품 클레임 목록에서 해당 주문 item 조회
                raw_returns = await client.get_returns(days=30)
                _lo_od_no = order.od_no or order.order_number
                claim_items = [i for i in raw_returns if i.get("odNo") == _lo_od_no]
                if not claim_items:
                    raise HTTPException(
                        status_code=400,
                        detail="롯데ON 반품 클레임 정보 없음 (최근 30일 내 조회되지 않음)",
                    )
                ci = claim_items[0]
                clm_no = ci.get("clmNo", "")
                od_seq = int(ci.get("odSeq") or 1)
                proc_seq = int(ci.get("procSeq") or od_seq)
                orgl_proc_seq = int(ci.get("orglProcSeq") or proc_seq)
                items_payload = [
                    {
                        "odSeq": od_seq,
                        "procSeq": proc_seq,
                        "orglProcSeq": orgl_proc_seq,
                        "spdNo": ci.get("spdNo", ""),
                        "spdNm": ci.get("spdNm", ""),
                        "sitmNo": ci.get("sitmNo", ""),
                        "sitmNm": ci.get("sitmNm", ""),
                    }
                ]
                await client.approve_return(_lo_od_no, clm_no, items_payload)
                new_status = "반품승인"
            elif body.action == "reject":
                await client.reject_return(_lo_od_no, body.reason or "")
                new_status = "반품거부"
            else:
                raise HTTPException(
                    status_code=400, detail=f"알 수 없는 액션: {body.action}"
                )
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"{label} 실패: {e}")

        await svc.update_order(order_id, {"shipping_status": new_status})

        # samba_return 상태 업데이트
        from backend.domain.samba.returns.repository import SambaReturnRepository
        from datetime import UTC, datetime

        return_repo = SambaReturnRepository(session)
        existing_returns = await return_repo.filter_by_async(order_id=order_id)
        if existing_returns:
            ret = existing_returns[0]
            if body.action == "approve":
                await return_repo.update_async(
                    ret.id,
                    status="completed",
                    market_order_status="반품완료",
                    completion_date=datetime.now(UTC),
                )
            elif body.action == "reject":
                await return_repo.update_async(
                    ret.id,
                    status="rejected",
                    market_order_status="반품거부",
                )

        logger.info(f"[반품처리][롯데ON] {order.order_number} {label} 완료")
        return {"ok": True, "message": f"{label} 완료"}

    elif account.market_type == "11st":
        from datetime import UTC, datetime

        from backend.domain.samba.proxy.elevenst import ElevenstClient
        from backend.domain.samba.returns.repository import SambaReturnRepository

        api_key = (
            (account.additional_fields or {}).get("apiKey", "") or account.api_key or ""
        )
        if not api_key:
            raise HTTPException(status_code=400, detail="11번가 API 키 없음")

        return_repo = SambaReturnRepository(session)
        existing_returns = await return_repo.filter_by_async(order_id=order_id)
        ret = existing_returns[0] if existing_returns else None
        clm_req_seq = (ret.clm_req_seq if ret else None) or ""
        ord_prd_seq = (ret.ord_prd_seq if ret else None) or ""

        if not clm_req_seq or not ord_prd_seq:
            raise HTTPException(
                status_code=400,
                detail="11번가 반품 클레임 정보 없음 (clm_req_seq 또는 ord_prd_seq 미수집)",
            )

        client = ElevenstClient(api_key)
        label = "반품승인" if body.action == "approve" else "반품거부"

        try:
            if body.action == "approve":
                await client.confirm_return(
                    clm_req_seq, order.order_number, ord_prd_seq
                )
                new_status = "반품승인"
            elif body.action == "reject":
                await client.reject_return(clm_req_seq, order.order_number, ord_prd_seq)
                new_status = "반품거부"
            else:
                raise HTTPException(
                    status_code=400, detail=f"알 수 없는 액션: {body.action}"
                )
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"{label} 실패: {e}")

        await svc.update_order(order_id, {"shipping_status": new_status})

        if ret:
            if body.action == "approve":
                await return_repo.update_async(
                    ret.id,
                    status="completed",
                    market_order_status="반품완료",
                    completion_date=datetime.now(UTC),
                )
            elif body.action == "reject":
                await return_repo.update_async(
                    ret.id,
                    status="rejected",
                    market_order_status="반품거부",
                )

        logger.info(f"[반품처리][11번가] {order.order_number} {label} 완료")
        return {"ok": True, "message": f"{label} 완료"}

    elif account.market_type == "ebay":
        # eBay 반품은 SambaReturn.market_order_status 에 저장된 returnId 필요
        from backend.domain.samba.proxy.ebay import EbayApiError, EbayClient
        from backend.domain.samba.returns.repository import SambaReturnRepository

        extras = account.additional_fields or {}
        app_id = extras.get("clientId") or extras.get("appId") or account.api_key or ""
        cert_id = (
            extras.get("clientSecret")
            or extras.get("certId")
            or account.api_secret
            or ""
        )
        refresh_token = extras.get("oauthToken") or extras.get("authToken", "") or ""
        if not (app_id and cert_id and refresh_token):
            raise HTTPException(status_code=400, detail="eBay 인증정보 없음")

        # returnId 는 samba_return.notes 또는 market_order_status에 저장 권장
        ret_repo = SambaReturnRepository(session)
        existing = await ret_repo.filter_by_async(order_id=order_id)
        if not existing:
            raise HTTPException(
                status_code=400, detail="해당 주문에 반품 데이터가 없습니다"
            )
        return_id = existing[0].memo or existing[0].market_order_status or ""
        # memo/market_order_status 에 returnId 저장 관례. 비어있으면 사용자 입력 필요
        if not return_id:
            raise HTTPException(
                status_code=400,
                detail="eBay returnId 없음 (samba_return.memo에 저장 필요)",
            )

        client = EbayClient(
            app_id=app_id,
            dev_id="",
            cert_id=cert_id,
            refresh_token=refresh_token,
            sandbox=bool(extras.get("sandbox", False)),
        )
        try:
            if body.action == "approve":
                await client.approve_return(return_id)
                new_status = "반품승인"
                ret_update = {"status": "completed", "market_order_status": "반품승인"}
            elif body.action == "reject":
                await client.reject_return(return_id, body.reason or "Seller decline")
                new_status = "반품거부"
                ret_update = {"status": "rejected", "market_order_status": "반품거부"}
            else:
                raise HTTPException(
                    status_code=400, detail=f"eBay 반품 액션 미지원: {body.action}"
                )
        except EbayApiError as e:
            raise HTTPException(status_code=500, detail=f"eBay 반품처리 실패: {e}")

        await svc.update_order(order_id, {"shipping_status": new_status})
        await ret_repo.update_async(existing[0].id, **ret_update)
        logger.info(f"[반품처리][eBay] {order.order_number} {body.action} 완료")
        return {"ok": True, "message": f"eBay 반품 {body.action} 완료"}

    else:
        raise HTTPException(
            status_code=400, detail=f"{account.market_type} 반품처리 미지원"
        )


# ══════════════════════════════════════════════
# 송장번호 전송 (발송처리)
# ══════════════════════════════════════════════


class ShipRequest(BaseModel):
    shipping_company: str
    tracking_number: str


@router.post("/{order_id}/ship")
async def ship_order(
    order_id: str,
    body: ShipRequest,
    session: AsyncSession = Depends(get_write_session_dependency),
):
    """송장번호 저장 + 마켓 발송처리."""
    svc = _write_service(session)
    order = await svc.get_order(order_id)
    if not order:
        raise HTTPException(404, "주문을 찾을 수 없습니다")

    # DB 저장 (마켓 전송 성공 여부와 무관하게 항상 저장)
    await svc.update_order(
        order_id,
        {
            "shipping_company": body.shipping_company,
            "tracking_number": body.tracking_number,
        },
    )

    # 마켓 송장 전송 — 통일 service (자동 dispatch_to_market 도 같은 함수 호출).
    # [통일 2026-05-16] 이전엔 이곳과 dispatch_to_market 가 마켓별 분기를 중복 구현 →
    # 자동 dispatch 가 자격증명 누락/필드 차이로 실패하던 회귀 차단. 단일 진실의 출처.
    from backend.domain.samba.order.dispatch_service import send_invoice_to_market

    market_sent, market_msg = await send_invoice_to_market(
        order, body.shipping_company, body.tracking_number, session
    )

    # 마켓 송장 전송 성공 시 status를 '국내배송중'으로 일괄 변경
    if market_sent:
        await svc.update_order(
            order_id,
            {"shipping_status": "송장전송완료", "status": "shipping"},
        )

    return {
        "ok": True,
        "market_sent": market_sent,
        "message": market_msg or "송장번호 저장 완료",
    }


# ══════════════════════════════════════════════
# URL에서 상품 대표이미지 추출
# ══════════════════════════════════════════════


@router.post("/fetch-product-image")
async def fetch_product_image(
    body: FetchProductImageRequest,
    session: AsyncSession = Depends(get_read_session_dependency),
):
    """URL에서 상품 대표이미지를 추출해 반환."""
    from urllib.parse import urlparse

    import httpx

    url = body.url.strip()
    if not url.startswith("http"):
        raise HTTPException(400, "올바른 URL을 입력해주세요")

    parsed = urlparse(url)
    host = parsed.hostname or ""

    try:
        # ── 무신사 ──
        if "musinsa.com" in host:
            # URL에서 상품번호 추출: /products/1234 또는 /app/goods/1234
            m = re.search(r"(?:/products/|/app/goods/|/goods/)(\d+)", url)
            if not m:
                raise HTTPException(400, "무신사 상품번호를 URL에서 추출할 수 없습니다")
            goods_no = m.group(1)

            from backend.domain.samba.proxy.musinsa import MusinsaClient

            # 쿠키 로드
            from backend.domain.samba.forbidden.repository import (
                SambaSettingsRepository,
            )

            settings_repo = SambaSettingsRepository(session)
            row = await settings_repo.find_by_async(key="musinsa_cookie")
            cookie = ""
            if row and row.value:
                cookie = str(row.value)
            client = MusinsaClient(cookie=cookie)
            detail = await client.get_goods_detail(goods_no)
            images = detail.get("images", [])
            if images:
                return {"image_url": images[0]}
            raise HTTPException(404, "무신사 상품에서 이미지를 찾을 수 없습니다")

        # ── KREAM ──
        elif "kream.co.kr" in host:
            async with httpx.AsyncClient(timeout=15, follow_redirects=True) as hc:
                resp = await hc.get(
                    url,
                    headers={
                        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
                    },
                )
                text = resp.text
            m = re.search(r'<meta[^>]+property="og:image"[^>]+content="([^"]*)"', text)
            if m:
                return {"image_url": m.group(1).split("?")[0]}
            raise HTTPException(404, "KREAM 상품에서 이미지를 찾을 수 없습니다")

        # ── 범용 fallback (og:image) ──
        else:
            async with httpx.AsyncClient(timeout=15, follow_redirects=True) as hc:
                resp = await hc.get(
                    url,
                    headers={
                        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
                    },
                )
                text = resp.text
            # og:image 추출
            m = re.search(r'<meta[^>]+property="og:image"[^>]+content="([^"]*)"', text)
            if not m:
                # content가 앞에 오는 경우도 처리
                m = re.search(
                    r'<meta[^>]+content="([^"]*)"[^>]+property="og:image"', text
                )
            if m:
                return {"image_url": m.group(1)}
            raise HTTPException(404, "해당 페이지에서 대표이미지를 찾을 수 없습니다")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[fetch-product-image] 이미지 추출 실패: {e}")
        raise HTTPException(500, f"이미지 추출 중 오류: {str(e)}")


# ══════════════════════════════════════════════
# 마켓 주문 동기화
# ══════════════════════════════════════════════


class SyncOrdersRequest(BaseModel):
    days: int = 7
    account_id: Optional[str] = None  # 특정 계정만 동기화
    # 명시적 날짜 범위 — 지정 시 days 무시. KST 기준 YYYY-MM-DD 또는 YYYYMMDD.
    # PlayAuto/스마트스토어 등 start_date 지원 마켓에 그대로 전달, 그 외 마켓은
    # (end - start + 1) 일수를 days 로 환산해 사용.
    start_date: Optional[str] = None
    end_date: Optional[str] = None


@router.post("/sync-from-markets")
async def sync_orders_from_markets(
    body: SyncOrdersRequest,
    session: AsyncSession = Depends(get_write_session_dependency),
    tenant_id: Optional[str] = Depends(get_optional_tenant_id),
):
    """활성 마켓 계정에서 주문 데이터를 가져와 DB에 저장."""
    from backend.domain.samba.account.repository import SambaMarketAccountRepository
    from backend.domain.samba.forbidden.repository import SambaSettingsRepository

    # 명시적 start_date/end_date 가 들어오면 days 환산.
    # 프론트의 날짜 input 이 daysMap 프리셋만 보고 days=1 로 박히던 버그 보완.
    # YYYY-MM-DD / YYYYMMDD 모두 허용.
    if body.start_date and body.end_date:
        from datetime import date as _bd_date

        def _parse_ymd(s: str) -> _bd_date | None:
            s = (s or "").strip().replace("-", "").replace(".", "").replace("/", "")
            if len(s) == 8 and s.isdigit():
                try:
                    return _bd_date(int(s[:4]), int(s[4:6]), int(s[6:8]))
                except ValueError:
                    return None
            return None

        _sd_dt = _parse_ymd(body.start_date)
        _ed_dt = _parse_ymd(body.end_date)
        if _sd_dt and _ed_dt and _ed_dt >= _sd_dt:
            body.days = max(1, (_ed_dt - _sd_dt).days + 1)

    account_repo = SambaMarketAccountRepository(session)

    # 특정 계정 또는 전체 활성 계정
    if body.account_id:
        target = await account_repo.get_async(body.account_id)
        if not target:
            active_accounts = []
        else:
            # 테넌트 소유권 검증
            if tenant_id is not None and target.tenant_id != tenant_id:
                raise HTTPException(403, "해당 계정에 대한 권한이 없습니다")
            active_accounts = [target]
    else:
        # 테넌트 필터링: tenant_id가 있으면 해당 테넌트 계정만 조회
        if tenant_id is not None:
            active_accounts = await account_repo.filter_by_async(
                is_active=True, order_by="created_at", order_by_desc=True
            )
            # in-memory 필터링으로 tenant_id 또는 None(공용) 계정만 유지
            active_accounts = [
                a
                for a in active_accounts
                if a.tenant_id == tenant_id or a.tenant_id is None
            ]
        else:
            active_accounts = await account_repo.filter_by_async(
                is_active=True, order_by="created_at", order_by_desc=True
            )

    svc = _write_service(session)
    results: list[dict[str, Any]] = []
    total_synced = 0

    # ORM 객체를 딕셔너리로 미리 추출 — rollback 후 lazy loading MissingGreenlet 방지
    account_snapshots = [
        {
            "id": a.id,
            "market_type": a.market_type,
            "market_name": a.market_name,
            "seller_id": a.seller_id or "",
            "api_key": a.api_key,
            "api_secret": a.api_secret,
            "additional_fields": a.additional_fields or {},
            "tenant_id": a.tenant_id,
        }
        for a in active_accounts
    ]

    # 소싱처별 원문 URL 템플릿 (상수)
    _sourcing_urls = {
        "MUSINSA": "https://www.musinsa.com/products/{}",
        "KREAM": "https://kream.co.kr/products/{}",
        "FashionPlus": "https://www.fashionplus.co.kr/goods/detail/{}",
        "ABCmart": "https://www.a-rt.com/product?prdtNo={}",
        "GrandStage": "https://www.a-rt.com/product?prdtNo={}",
        "REXMONDE": "https://www.okmall.com/products/detail/{}",
        "LOTTEON": "https://www.lotteon.com/p/product/{}",
        "GSShop": "https://www.gsshop.com/prd/prd.gs?prdid={}",
        "ElandMall": "https://www.elandmall.com/goods/goods.action?goodsNo={}",
        "SSF": "https://www.ssfshop.com/goods/{}",
        "SSG": "https://www.ssg.com/item/itemView.ssg?itemId={}",
        "Nike": "https://www.nike.com/kr/t/{}",
        "Adidas": "https://www.adidas.co.kr/{}.html",
    }

    # ── 병렬 사전조회: 각 마켓 get_orders() HTTP 호출을 동시에 실행 ──────────
    # 세션 없이 순수 HTTP만 병렬화 — DB 작업/파싱/발주확인은 기존 루프에서 수행
    _pre_settings_keys: set[str] = set()
    for _pacc in account_snapshots:
        _pmt = _pacc["market_type"]
        _pex = _pacc["additional_fields"]
        if _pmt == "smartstore" and not (
            (_pex.get("clientId") or _pacc["api_key"])
            and (_pex.get("clientSecret") or _pacc["api_secret"])
        ):
            _pre_settings_keys.add("store_smartstore")
        elif _pmt == "11st" and not (_pex.get("apiKey") or _pacc["api_key"]):
            _pre_settings_keys.add("store_11st")
        elif _pmt == "ebay" and not (
            (_pex.get("clientId") or _pex.get("appId") or _pacc["api_key"])
            and (_pex.get("clientSecret") or _pex.get("certId") or _pacc["api_secret"])
            and (_pex.get("oauthToken") or _pex.get("authToken"))
        ):
            _pre_settings_keys.add("store_ebay")
        elif _pmt == "ssg" and not (_pex.get("apiKey") or _pacc["api_key"]):
            _pre_settings_keys.add("store_ssg")

    _pre_settings: dict[str, dict] = {}
    if _pre_settings_keys:
        _pre_svc_repo = SambaSettingsRepository(session)
        for _psk in _pre_settings_keys:
            _prow = await _pre_svc_repo.find_by_async(key=_psk)
            if _prow and isinstance(_prow.value, dict):
                _pre_settings[_psk] = _prow.value

    async def _pre_fetch_orders(
        acc: dict[str, Any], days: int
    ) -> tuple[str, list | None]:
        """마켓 API에서 초기 주문 목록 조회 (세션 없음, HTTP만)"""
        _aid = acc["id"]
        _mtype = acc["market_type"]
        _extr = acc["additional_fields"]
        _sid = acc["seller_id"]
        try:
            if _mtype == "smartstore":
                _cid = _extr.get("clientId", "") or acc["api_key"] or ""
                _csec = _extr.get("clientSecret", "") or acc["api_secret"] or ""
                if not _cid or not _csec:
                    _sv = _pre_settings.get("store_smartstore", {})
                    _cid = _cid or _sv.get("clientId", "")
                    _csec = _csec or _sv.get("clientSecret", "")
                if not _cid or not _csec:
                    return _aid, None
                from backend.domain.samba.proxy.smartstore import SmartStoreClient

                _c = SmartStoreClient(_cid, _csec)
                return _aid, await _c.get_orders(days=days)

            elif _mtype == "lotteon":
                _ak = _extr.get("apiKey", "") or acc["api_key"] or ""
                if not _ak:
                    return _aid, None
                from backend.domain.samba.proxy.lotteon import LotteonClient

                _c = LotteonClient(_ak)
                await _c.test_auth()
                return _aid, await _c.get_orders(days=days)

            elif _mtype == "poison":
                _app_key = (
                    _extr.get("app_key", "")
                    or _extr.get("appKey", "")
                    or acc["api_key"]
                    or ""
                )
                _app_secret = (
                    _extr.get("app_secret", "")
                    or _extr.get("appSecret", "")
                    or acc["api_secret"]
                    or ""
                )
                if not _app_key or not _app_secret:
                    return _aid, None
                from backend.domain.samba.proxy.poison import PoisonClient

                _c = PoisonClient(_app_key, _app_secret)
                return _aid, await _c.get_orders(days=days)

            elif _mtype == "playauto":
                _ak = _extr.get("apiKey", "") or acc["api_key"] or ""
                if not _ak:
                    return _aid, None
                from datetime import UTC as _paut, datetime as _padt, timedelta as _patd

                from backend.domain.samba.proxy.playauto import PlayAutoClient

                _c = PlayAutoClient(_ak)
                try:
                    _sd = (_padt.now(_paut) - _patd(days=days)).strftime("%Y%m%d")
                    return _aid, await _c.get_orders(start_date=_sd, count=500)
                finally:
                    await _c.close()

            elif _mtype == "coupang":
                _ack = _extr.get("accessKey", "") or acc.get("api_key", "") or ""
                _sck = _extr.get("secretKey", "") or acc.get("api_secret", "") or ""
                _vid = _extr.get("vendorId", "") or _sid or ""
                if not all([_ack, _sck, _vid]):
                    return _aid, None
                from backend.domain.samba.proxy.coupang import CoupangClient

                _c = CoupangClient(_ack, _sck, _vid)
                return _aid, await _c.get_orders(days=days)

            elif _mtype == "11st":
                _ak = _extr.get("apiKey", "") or acc["api_key"] or ""
                if not _ak:
                    _sv = _pre_settings.get("store_11st", {})
                    _ak = _sv.get("apiKey", "") or ""
                if not _ak:
                    return _aid, None
                from datetime import datetime as _11dt, timedelta as _11td
                from zoneinfo import ZoneInfo as _11zi

                from backend.domain.samba.proxy.elevenst import ElevenstClient

                _KST11 = _11zi("Asia/Seoul")
                _fmt11 = "%Y%m%d%H%M"
                _st11 = (_11dt.now(_KST11) - _11td(days=days)).strftime(_fmt11)
                _et11 = _11dt.now(_KST11).strftime(_fmt11)
                _c = ElevenstClient(_ak)
                return _aid, await _c.get_orders(_st11, _et11)

            elif _mtype == "ebay":
                _appid = _extr.get("clientId") or _extr.get("appId") or acc["api_key"]
                _certid = (
                    _extr.get("clientSecret")
                    or _extr.get("certId")
                    or acc["api_secret"]
                )
                _rtok = _extr.get("oauthToken") or _extr.get("authToken", "")
                if not (_appid and _certid and _rtok):
                    _sv = _pre_settings.get("store_ebay", {})
                    _appid = _appid or _sv.get("clientId", "") or _sv.get("appId", "")
                    _certid = (
                        _certid or _sv.get("clientSecret", "") or _sv.get("certId", "")
                    )
                    _rtok = (
                        _rtok or _sv.get("oauthToken", "") or _sv.get("authToken", "")
                    )
                if not (_appid and _certid and _rtok):
                    return _aid, None
                from backend.domain.samba.proxy.ebay import EbayClient

                _c = EbayClient(
                    app_id=_appid,
                    dev_id="",
                    cert_id=_certid,
                    refresh_token=_rtok,
                    sandbox=bool(_extr.get("sandbox", False)),
                )
                return _aid, await _c.get_orders(days=days)

            elif _mtype == "ssg":
                _ak = _extr.get("apiKey", "") or acc["api_key"] or ""
                if not _ak:
                    _sv = _pre_settings.get("store_ssg", {})
                    _ak = _sv.get("apiKey", "") or ""
                if not _ak:
                    return _aid, None
                from backend.domain.samba.proxy.ssg import SSGClient

                _c = SSGClient(_ak)
                return _aid, await _c.get_orders(days=days)

            elif _mtype == "gsshop":
                # GS샵은 본 루프에서 직접 수집 (processType 분기 필요)
                # 병렬 사전조회 생략 — None 반환 시 본 루프가 처리
                return _aid, None

        except Exception as _pfe:
            logger.warning(f"[주문동기화] 병렬 사전조회 실패 ({_mtype}): {_pfe}")
        return _aid, None

    _prefetch_raw = await asyncio.gather(
        *[_pre_fetch_orders(acc, body.days) for acc in account_snapshots],
        return_exceptions=True,
    )
    _raw_cache: dict[str, list] = {}
    for _pr in _prefetch_raw:
        if isinstance(_pr, Exception):
            continue
        _paid, _praw = _pr
        if _praw is not None:
            _raw_cache[_paid] = _praw
    logger.info(
        f"[주문동기화] 병렬 사전조회 완료: {len(_raw_cache)}/{len(account_snapshots)}개 계정"
    )
    # ── 병렬 사전조회 끝 ──────────────────────────────────────────────────────

    for account in account_snapshots:
        market_type = account["market_type"]
        extras = account["additional_fields"]
        seller_id = account["seller_id"]
        label = f"{account['market_name']}({seller_id})"

        # 마켓 클라이언트들의 httpx keepalive 좀비 차단 — 매 계정 처리 후 명시적 aclose.
        # 미회수 시 hang 한 번에 다음 계정·다른 마켓 호출까지 영향(2026-05-15 사고).
        _clients_to_close: list[Any] = []

        try:
            orders_data: list[dict[str, Any]] = []
            unconfirmed_ids: list[str] = []
            _lh_replaced_old_keys: list[
                str
            ] = []  # deliver_list가 교체한 index-format order_numbers

            if market_type == "smartstore":
                from backend.domain.samba.proxy.smartstore import SmartStoreClient

                client_id = extras.get("clientId", "") or account["api_key"] or ""
                client_secret = (
                    extras.get("clientSecret", "") or account["api_secret"] or ""
                )
                if not client_id or not client_secret:
                    # fallback: 공유 설정
                    settings_repo = SambaSettingsRepository(session)
                    row = await settings_repo.find_by_async(key="store_smartstore")
                    if row and isinstance(row.value, dict):
                        client_id = client_id or row.value.get("clientId", "")
                        client_secret = client_secret or row.value.get(
                            "clientSecret", ""
                        )
                if not client_id or not client_secret:
                    results.append(
                        {"account": label, "status": "skip", "message": "인증정보 없음"}
                    )
                    continue
                client = SmartStoreClient(client_id, client_secret)
                _clients_to_close.append(client)
                raw_orders = _raw_cache.get(account["id"])
                if raw_orders is None:
                    raw_orders = await client.get_orders(days=body.days)
                # 발주 미확인(PAYED) 주문 자동 발주확인
                unconfirmed_ids = []
                for ro in raw_orders:
                    po = ro.get("productOrder", ro)
                    order_info = ro.get("order", {})
                    # 클레임 정보: claim / cancel / currentClaim 순으로 확인
                    # 취소요청 시 응답 최상위에 'cancel' 키로 오는 경우 처리
                    claim_info = (
                        ro.get("claim")
                        or ro.get("cancel")
                        or ro.get("currentClaim")
                        or po.get("claim")
                        or {}
                    )
                    orders_data.append(
                        _parse_smartstore_order(
                            po, order_info, account["id"], label, claim_info=claim_info
                        )
                    )
                    if (
                        po.get("placeOrderStatus") == "NOT_YET"
                        and po.get("productOrderStatus") == "PAYED"
                    ):
                        unconfirmed_ids.append(po.get("productOrderId", ""))
                # 발주확인 실행
                if unconfirmed_ids:
                    try:
                        await client.confirm_product_orders(unconfirmed_ids)
                        logger.info(
                            f"[주문동기화] {label}: {len(unconfirmed_ids)}건 발주확인 완료"
                        )
                    except Exception as ce:
                        logger.warning(f"[주문동기화] {label}: 발주확인 실패 — {ce}")

                # last-changed API 권한 제한 보완:
                # DB에 있는 미완결 주문을 직접 재조회하여 배송완료/취소요청 등 최신 상태 반영
                # '취소요청' 포함 이유 — 고객이 취소를 철회하면 Naver API는
                # claimStatus=null + productOrderStatus=PURCHASE_DECIDED 로 응답하지만
                # last-changed 윈도우(body.days)를 벗어난 주문은 본 쿼리에 포함되지
                # 않아 영영 '취소요청' 으로 남던 사고 방지 (issue #192)
                _pending_statuses = {
                    "발주미확인",
                    "발송대기",
                    "결제완료",
                    "배송대기중",
                    "송장전송완료",
                    "국내배송중",
                    "취소요청",
                    "취소처리중",
                }
                _already_fetched = {
                    d["order_number"] for d in orders_data if d.get("order_number")
                }
                from sqlalchemy import and_ as _and_, or_ as _or_, select as _sa_select
                from backend.domain.samba.order.model import SambaOrder as _SambaOrder
                from datetime import datetime as _dt, timedelta, timezone as _tz

                # 취소요청/취소처리중은 철회가 30일 이후에도 발생 가능 → 시간 cap 제거.
                # 그 외 미완결은 stuck 누적 방지를 위해 기존 30일 cap 유지.
                _cutoff = _dt.now(_tz.utc) - timedelta(days=max(body.days, 30))
                _cancel_pending = {"취소요청", "취소처리중"}
                _other_pending = _pending_statuses - _cancel_pending
                _stmt = (
                    _sa_select(_SambaOrder.order_number)
                    .where(
                        _SambaOrder.channel_id == account["id"],
                        _or_(
                            _SambaOrder.shipping_status.in_(_cancel_pending),
                            _and_(
                                _SambaOrder.shipping_status.in_(_other_pending),
                                _SambaOrder.updated_at >= _cutoff,
                            ),
                        ),
                    )
                    .order_by(_SambaOrder.updated_at.desc())
                    .limit(300)
                )
                _res = await session.execute(_stmt)
                _pending_numbers = [
                    r[0]
                    for r in _res.fetchall()
                    if r[0] and r[0] not in _already_fetched
                ]
                if _pending_numbers:
                    logger.info(
                        f"[주문동기화] {label}: 미완결 주문 {len(_pending_numbers)}건 직접 재조회"
                    )
                    try:
                        _extra_raws = await client.get_product_orders_by_ids(
                            _pending_numbers
                        )
                        for ro2 in _extra_raws:
                            po2 = ro2.get("productOrder", ro2)
                            order_info2 = ro2.get("order", {})
                            claim_info2 = (
                                ro2.get("claim")
                                or ro2.get("cancel")
                                or ro2.get("currentClaim")
                                or po2.get("claim")
                                or {}
                            )
                            orders_data.append(
                                _parse_smartstore_order(
                                    po2,
                                    order_info2,
                                    account["id"],
                                    label,
                                    claim_info=claim_info2,
                                )
                            )
                    except Exception as _ex:
                        logger.warning(
                            f"[주문동기화] {label}: 미완결 주문 직접 재조회 실패 — {_ex}"
                        )

            elif market_type == "lotteon":
                from backend.domain.samba.proxy.lotteon import LotteonClient

                api_key = extras.get("apiKey", "") or account["api_key"] or ""
                if not api_key:
                    results.append(
                        {
                            "account": label,
                            "status": "skip",
                            "message": "롯데ON API Key 없음",
                        }
                    )
                    continue
                lotteon_client = LotteonClient(api_key)
                _clients_to_close.append(lotteon_client)
                await lotteon_client.test_auth()
                raw_orders = _raw_cache.get(account["id"])
                if raw_orders is None:
                    raw_orders = await lotteon_client.get_orders(days=body.days)
                logger.info(
                    f"[주문동기화] {label}: 롯데ON 주문 {len(raw_orders)}건 조회"
                )
                # 신규주문(odPrgsStepCd=11=출고지시) 자동 연동완료 통보 대상 수집
                # SellerDeliveryOrdersSearch는 11(출고지시)/23(회수지시)만 반환 — "10"은 영원히 안 잡힘(공식 문서 기준)
                # SellerIfCompleteInform(ifCplYN=Y) 호출 시 롯데ON에서 자동으로 11→12(상품준비)로 전이됨
                lotteon_confirmed_count = 0
                unconfirmed_items: list[dict] = []
                for ro in raw_orders:
                    orders_data.append(_parse_lotteon_order(ro, account["id"], label))
                    step_cd = str(ro.get("odPrgsStepCd", "") or "")
                    if step_cd == "11":
                        unconfirmed_items.append(
                            {
                                "odNo": ro.get("odNo", ""),
                                "odSeq": ro.get("odSeq", 1) or 1,
                                "procSeq": ro.get("procSeq", 1) or 1,
                            }
                        )

                # 주문확인(SellerIfCompleteInform, ifCplYN=Y) 일괄 실행 — 호출 후 셀러센터에서 상품준비중 자동 전이
                if unconfirmed_items:
                    try:
                        ok = await lotteon_client.confirm_orders(unconfirmed_items)
                        if ok:
                            lotteon_confirmed_count = len(unconfirmed_items)
                            logger.info(
                                f"[주문동기화] {label}: {len(unconfirmed_items)}건 주문확인 완료 (출고지시→상품준비중 자동 전이)"
                            )
                            # 로컬 표시도 즉시 상품준비중으로 갱신 (다음 sync까지 기다리지 않음)
                            _confirmed_keys = {
                                f"{it['odNo']}_{it['odSeq']}_{it['procSeq']}"
                                for it in unconfirmed_items
                            }
                            for od in orders_data:
                                if (
                                    od.get("source") == "lotteon"
                                    and od.get("order_number") in _confirmed_keys
                                    and od.get("shipping_status")
                                    in ("발주확인대기", "출고지시")
                                ):
                                    od["shipping_status"] = "상품준비"
                                    od["status"] = "preparing"
                        else:
                            logger.warning(
                                f"[주문동기화] {label}: 주문확인 API 응답 실패(rsltCd != 0000)"
                            )
                    except Exception as ce:
                        logger.warning(f"[주문동기화] {label}: 주문확인 실패 — {ce}")

                # ── 정산예상 계산용 raw 필드 매핑 (롯데ON 공식 정산공식, 2026-04-30 재확인) ─
                # SellerDeliveryOrdersSearch 실제 응답 키:
                #   slAmt            = 총판매금액 (= 판매단가 × 수량)
                #   actualAmt        = 고객결제금액 (= 슬amt − 전체할인)
                #   prSfcoShrAmtSum  = 당사(롯데/이커머스) 부담 할인 합 (= ajstDcAmt 역할)
                #   prEntpShrAmtSum  = 제휴몰 부담 할인 합
                #   sptDcPgmCmsnSum  = 셀러 부담 할인 합 (지원할인 PGM)
                #   fvrAmtSum        = 전체 할인합 (= prSfco + prEntp + sptDcPgm)
                # → bseCmsn/pcsCmsn/dvCmsn/ajstDcAmt 필드는 이 API에 존재하지 않음.
                #   기본수수료는 카테고리 fee_rate × slAmt, PCS는 가격비교 채널만 부과,
                #   조정(당사부담환급)은 prSfcoShrAmtSum 으로 대체.
                # 정산공식: pymtAmt = actualAmt − (bseCmsn + pcsCmsn + dvCmsn − ajstDcAmt)
                # 키: (odNo, odSeq) — 같은 odNo에 여러 옵션/수량이 묶인 멀티라인 주문에서
                # odNo만 사용하면 한 라인의 값이 다른 라인을 덮어써 모든 라인의 결제/정산 금액이
                # 동일해지는 버그가 발생함(2026-05-15 수정).
                sl_amt_map: dict[tuple[str, str], int] = {}  # 총판매금액 (slAmt)
                fvr_amt_map: dict[tuple[str, str], int] = {}  # 전체 할인합
                actual_amt_map: dict[
                    tuple[str, str], int
                ] = {}  # 고객결제금액 (actualAmt)
                lotte_dc_map: dict[
                    tuple[str, str], int
                ] = {}  # 당사부담할인 (prSfcoShrAmtSum)
                slr_dc_map: dict[
                    tuple[str, str], int
                ] = {}  # 셀러부담할인 (sptDcPgmCmsnSum + 셀러즉시) — 2026-05-20 추가
                ch_no_map: dict[
                    str, str
                ] = {}  # 채널번호 (chNo) — 주문 단위라 odNo 키 유지

                def _pick(d: dict, *keys: str) -> int:
                    for k in keys:
                        v = d.get(k)
                        if v not in (None, "", 0, "0"):
                            try:
                                return int(float(v))
                            except (TypeError, ValueError):
                                continue
                    return 0

                for ro in raw_orders:
                    _od_no = str(ro.get("odNo") or "")
                    if not _od_no:
                        continue
                    _od_seq = str(ro.get("odSeq", "1") or "1")
                    _line_key = (_od_no, _od_seq)
                    _slamt = _pick(ro, "slAmt", "slPrc")
                    _fvr = _pick(ro, "fvrAmtSum")
                    _actual = _pick(ro, "actualAmt")
                    _lotte_dc = _pick(ro, "prSfcoShrAmtSum")
                    # 셀러 부담 할인 — 정산 화면 "상품할인(셀러부담)" 5,922원이 누락되던 사고(2026-05-20)
                    # sptDcPgmCmsnSum(지원할인 PGM 셀러부담) + 셀러즉시할인(slrDcAmt 계열)
                    _slr_dc = _pick(ro, "sptDcPgmCmsnSum") + _pick(
                        ro, "slrDcAmt", "slrDcSptAmt", "slrImdDcAmt"
                    )
                    _ch_no = str(ro.get("chNo") or "")
                    # 라인(odSeq) 단위 저장 — 같은 odNo의 다른 옵션/수량이 서로 덮어쓰지 않도록.
                    if _slamt > sl_amt_map.get(_line_key, 0):
                        sl_amt_map[_line_key] = _slamt
                    if _fvr > fvr_amt_map.get(_line_key, 0):
                        fvr_amt_map[_line_key] = _fvr
                    if _actual > actual_amt_map.get(_line_key, 0):
                        actual_amt_map[_line_key] = _actual
                    if _lotte_dc > lotte_dc_map.get(_line_key, 0):
                        lotte_dc_map[_line_key] = _lotte_dc
                    if _slr_dc > slr_dc_map.get(_line_key, 0):
                        slr_dc_map[_line_key] = _slr_dc
                    if _ch_no:
                        ch_no_map[_od_no] = _ch_no
                logger.info(
                    f"[주문동기화] {label}: 정산필드 매핑 {len(sl_amt_map)}건 "
                    f"(raw_orders {len(raw_orders)}건)"
                )

                # ── 정산금액 매칭 (SettleItmdSales) ─────────────────────────
                # 정산 데이터는 배송완료 → 구매확정 후 수일 지나서 생성되므로
                # 주문 조회 기간(body.days)보다 넓게(최대 30일) 조회해야 매칭률 ↑.
                # 최대값 30은 api_client.get_settlement_items 내부에서 cap.
                try:
                    settle_items = await lotteon_client.get_settlement_items(days=30)
                    # (odNo, odSeq, procSeq) → 정산 데이터 매핑
                    settle_map: dict[tuple[str, str, str], dict] = {}
                    for si in settle_items:
                        key = (
                            str(si.get("odNo", "")),
                            str(si.get("odSeq", "")),
                            str(si.get("procSeq", "")),
                        )
                        settle_map[key] = si
                    # 매출 주문에 매칭 → revenue/fee_rate 갱신
                    matched = 0
                    for i, ro in enumerate(raw_orders):
                        key = (
                            str(ro.get("odNo", "")),
                            str(ro.get("odSeq", "1")),
                            str(ro.get("procSeq", "1")),
                        )
                        si = settle_map.get(key)
                        if not si:
                            continue
                        pymt_amt = float(si.get("pymtAmt", 0) or 0)
                        sl_amt = float(si.get("slAmt", 0) or 0)
                        sl_qty = float(si.get("slQty", 1) or 1)
                        gross = sl_amt * sl_qty
                        # 고객결제금액 = 총판매 - 셀러부담할인 - 상품할인(셀러+이커머스)
                        slr_dc = float(si.get("slrDcAmt", 0) or 0)
                        pd_dc_slr = float(si.get("pdDcSlrAmt", 0) or 0)
                        pd_dc_oco = float(si.get("pdDcOcoAmt", 0) or 0)
                        customer_paid = max(0.0, gross - slr_dc - pd_dc_slr - pd_dc_oco)
                        if pymt_amt > 0 and customer_paid > 0:
                            fee_rate = round((1 - pymt_amt / customer_paid) * 100, 2)
                            orders_data[i]["revenue"] = pymt_amt
                            orders_data[i]["fee_rate"] = fee_rate
                            orders_data[i]["total_payment_amount"] = customer_paid
                            matched += 1
                        elif pymt_amt > 0 and gross > 0:
                            # 할인 필드가 비어 있으면 기존 방식(총판매 기준)으로 폴백
                            fee_rate = round((1 - pymt_amt / gross) * 100, 2)
                            orders_data[i]["revenue"] = pymt_amt
                            orders_data[i]["fee_rate"] = fee_rate
                            matched += 1
                    logger.info(
                        f"[주문동기화] {label}: 정산 매칭 {matched}/{len(raw_orders)}건 "
                        f"(정산 API {len(settle_items)}건)"
                    )

                    # ── 기존 DB 주문 보정 (구매확정 후 정산 데이터로 정확값 덮어쓰기) ─
                    # raw_orders는 odPrgsStepCd=11/23만 반환하므로,
                    # 이미 발주확인되어 raw에서 빠진 주문은 위 in-memory 매칭으로 보정 안 됨.
                    # 정산 API에 있는 모든 키에 대해 DB를 직접 UPDATE 한다.
                    db_updated = 0
                    from sqlalchemy import text as _sa_text

                    for (od_no_k, od_seq_k, proc_seq_k), si in settle_map.items():
                        if not od_no_k:
                            continue
                        pymt_amt = float(si.get("pymtAmt", 0) or 0)
                        if pymt_amt <= 0:
                            continue
                        sl_amt = float(si.get("slAmt", 0) or 0)
                        sl_qty = float(si.get("slQty", 1) or 1)
                        gross = sl_amt * sl_qty
                        slr_dc = float(si.get("slrDcAmt", 0) or 0)
                        pd_dc_slr = float(si.get("pdDcSlrAmt", 0) or 0)
                        pd_dc_oco = float(si.get("pdDcOcoAmt", 0) or 0)
                        customer_paid = max(0.0, gross - slr_dc - pd_dc_slr - pd_dc_oco)
                        base = customer_paid if customer_paid > 0 else gross
                        if base <= 0:
                            continue
                        new_fee_rate = round((1 - pymt_amt / base) * 100, 2)
                        # od_seq/proc_seq는 SambaOrder에 Text로 저장되어 있음
                        # 동일 odNo + odSeq + procSeq 매칭 (account 무관 — odNo는 전역 유일)
                        try:
                            res = await session.execute(
                                _sa_text(
                                    "UPDATE samba_order "
                                    "SET revenue = :rev, fee_rate = :fr, "
                                    "    total_payment_amount = COALESCE(NULLIF(:cp, 0), total_payment_amount), "
                                    "    updated_at = now() "
                                    "WHERE source = 'lotteon' "
                                    "  AND od_no = :od "
                                    "  AND COALESCE(od_seq, '1') = :os "
                                    "  AND COALESCE(proc_seq, '1') = :ps "
                                    "  AND (revenue IS NULL OR revenue <> :rev)"
                                ),
                                {
                                    "rev": pymt_amt,
                                    "fr": new_fee_rate,
                                    "cp": customer_paid,
                                    "od": od_no_k,
                                    "os": od_seq_k or "1",
                                    "ps": proc_seq_k or "1",
                                },
                            )
                            db_updated += res.rowcount or 0
                        except Exception as ue:
                            logger.warning(
                                f"[주문동기화] {label}: 정산 DB UPDATE 실패 odNo={od_no_k} — {ue}"
                            )
                    if db_updated:
                        logger.info(
                            f"[주문동기화] {label}: 정산 API → DB 보정 {db_updated}건 "
                            "(구매확정된 기존 주문 revenue/fee_rate 갱신)"
                        )
                except Exception as se:
                    logger.warning(f"[주문동기화] {label}: 정산 조회 실패 — {se}")

                # 발주확인은 수동 처리 (원소싱처 재고/가격 확인 후 사용자가 결정)
                # 교환 클레임 조회 → 기존 주문 shipping_status 업데이트
                try:
                    exchange_claims = await lotteon_client.get_exchanges(days=body.days)
                    logger.info(f"[롯데ON] 교환 클레임 조회: {len(exchange_claims)}건")
                    if exchange_claims:
                        exchange_step_map = {
                            "21": "교환요청",
                            "22": "교환회수완료",
                            "23": "교환회수완료",
                            "24": "교환재배송",
                            "25": "교환완료",
                        }
                        exchange_priority = {
                            "교환요청": 1,
                            "교환회수완료": 2,
                            "교환재배송": 3,
                            "교환완료": 4,
                        }
                        for claim in exchange_claims:
                            ex_od_no = claim.get("odNo", "")
                            clm_no = claim.get("clmNo", "")
                            step_cd = str(claim.get("odPrgsStepCd", "") or "")
                            ex_status = exchange_step_map.get(step_cd, "교환요청")
                            logger.info(
                                f"[롯데ON][교환클레임] odNo={ex_od_no} clmNo={clm_no} stepCd={step_cd} → {ex_status}"
                            )
                            found_in_data = False
                            for od in orders_data:
                                # order_number는 합성키(odNo_odSeq_procSeq)이므로 od_no로 비교
                                if od.get("od_no") == ex_od_no:
                                    cur_status = od.get("shipping_status", "")
                                    cur_p = exchange_priority.get(cur_status, 0)
                                    new_p = exchange_priority.get(ex_status, 0)
                                    if cur_p == 0 or new_p >= cur_p:
                                        od["shipping_status"] = ex_status
                                        if step_cd in ("21", "22", "23", "24"):
                                            od["status"] = "exchanging"
                                        elif step_cd == "25":
                                            od["status"] = "exchanged"
                                    found_in_data = True
                                    break
                            if not found_in_data and ex_od_no:
                                from sqlalchemy import text as _sa_text_ex

                                _ex_row = await session.execute(
                                    _sa_text_ex(
                                        "SELECT id FROM samba_order "
                                        "WHERE source = 'lotteon' AND od_no = :od_no LIMIT 1"
                                    ),
                                    {"od_no": ex_od_no},
                                )
                                _ex_id = (_ex_row.fetchone() or [None])[0]
                                existing = (
                                    await svc.repo.get_async(_ex_id) if _ex_id else None
                                )
                                if existing:
                                    cur_p = exchange_priority.get(
                                        existing.shipping_status, 0
                                    )
                                    new_p = exchange_priority.get(ex_status, 0)
                                    if cur_p == 0 or new_p >= cur_p:
                                        await svc.update_order(
                                            existing.id,
                                            {"shipping_status": ex_status},
                                        )
                                        logger.info(
                                            f"[롯데ON][교환클레임] DB 직접 업데이트: {ex_od_no} → {ex_status}"
                                        )
                except Exception as ex_err:
                    logger.warning(f"[롯데ON] 교환 클레임 조회 실패: {ex_err}")

                # 취소 클레임 조회 → samba_order.status 갱신
                # odPrgsStepCd 실측(2026-06-01 getCancellationRequestAndComplateList):
                #   02=요청 / 21=취소완료 / 22=철회.
                # (구 매핑 11/12/13 은 실제값과 안 맞아 한 번도 매칭 안 되던 죽은 코드였음.
                #  claim 기반 status 갱신이 전혀 안 돼 cancel_requested 는 메인목록 ordPrdStat
                #  파싱만 세팅해 왔음.)
                # 22 철회(고객이 취소요청 회수)는 취소 진행 아님 → 매핑 안 함, status 유지.
                # 21 취소완료는 의도적으로 매핑 안 함 — 종결(cancelled) 상태는 메인 주문목록
                #   파싱이 권위 소스. claim-sync 로 21 을 일괄 flip 하면 배송완료/구매확정 주문이
                #   cancelled 로 뒤집히고 profit/cost 가 0으로 정리 안 돼 정산 불일치 발생.
                #   여기선 actionable 한 02 요청만 반영(auto-cancel 트리거).
                try:
                    cancel_claims = await lotteon_client.get_cancel_orders(
                        days=body.days
                    )
                    logger.info(f"[롯데ON] 취소 클레임 조회: {len(cancel_claims)}건")
                    cancel_step_map = {
                        "02": ("취소요청", "cancel_requested"),
                        # 21=취소완료 (롯데ON 공식문서 odPrgsStepCd: 02 요청/21 취소완료/22 철회).
                        # #326: od_seq 정밀매칭 + 전체취소(rmdrQty=0) + shipped_guard 로 안전.
                        # 배송완료/구매확정 등 종결·정산 주문은 shipped_guard 가 차단.
                        "21": ("취소완료", "cancelled"),
                    }
                    cancel_priority = {
                        "취소요청": 1,
                        "취소완료": 3,  # 종결 — 반품/교환요청보다 우선
                    }
                    # 배송 진행 단계 보호 — 송장출력 이후로 진행한 주문은 좀비/지연
                    # cancel claim 으로 '취소요청'으로 되돌리지 않음 ('취소처리중'/'취소완료'
                    # 는 실제 종결 상태이므로 그대로 반영)
                    _lo_shipped_guard = {
                        "송장전송완료",
                        "국내배송중",
                        "배송완료",
                        "구매확정",
                        "발송완료",
                    }
                    for claim in cancel_claims:
                        cn_od_no = claim.get("odNo", "")
                        # od_seq 정밀매칭 — 다중 품목 주문에서 취소된 od_seq 만 정확히 갱신.
                        # (od_no-only 매칭 시 같은 주문의 배송완료 다른 품목을 오취소할 위험, #326)
                        cn_od_seq = str(claim.get("odSeq", "") or "")
                        step_cd_c = str(claim.get("odPrgsStepCd", "") or "")
                        mapped = cancel_step_map.get(step_cd_c)
                        if not mapped or not cn_od_no:
                            continue
                        # 부분취소(잔여수량>0)는 전체취소 아님 → status 전이 스킵.
                        # 수량 처리는 메인 주문목록 파싱(ordPrdStat/quantity)에 위임.
                        # 안 막으면 21 취소완료 클레임이 부분취소 주문을 전체 cancelled 로 오염.
                        try:
                            _cn_rmdr = int(claim.get("rmdrQty", 0) or 0)
                        except (TypeError, ValueError):
                            _cn_rmdr = 0
                        if _cn_rmdr > 0:
                            logger.info(
                                f"[롯데ON][취소클레임] 부분취소 스킵: {cn_od_no} "
                                f"잔여수량={_cn_rmdr}"
                            )
                            continue
                        cn_ship_status, cn_status = mapped
                        found_in_data_c = False
                        for od in orders_data:
                            if od.get("od_no") == cn_od_no and (
                                not cn_od_seq or str(od.get("od_seq", "")) == cn_od_seq
                            ):
                                cur_ss = od.get("shipping_status", "")
                                # 취소요청·취소완료 모두 배송 진행/종결 상태는 보호 (정산 주문 오취소 차단)
                                if (
                                    cn_ship_status in ("취소요청", "취소완료")
                                    and cur_ss in _lo_shipped_guard
                                ):
                                    logger.info(
                                        f"[롯데ON][취소클레임] 배송 진행 상태 보호: {cn_od_no} "
                                        f"{cur_ss} → {cn_ship_status} 차단"
                                    )
                                    found_in_data_c = True
                                    break
                                cur_p = cancel_priority.get(cur_ss, 0)
                                new_p = cancel_priority.get(cn_ship_status, 0)
                                if cur_p == 0 or new_p >= cur_p:
                                    od["shipping_status"] = cn_ship_status
                                    od["status"] = cn_status
                                    if cn_status == "cancelled":
                                        # 정산 finalize — _finalize_cancelled 관례와 동일
                                        # (cost/shipping_fee/profit 0, revenue 는 건드리지 않음, #326)
                                        od["cost"] = 0
                                        od["shipping_fee"] = 0
                                        od["profit"] = 0
                                found_in_data_c = True
                                break
                        if not found_in_data_c:
                            from sqlalchemy import text as _sa_text_cn

                            # od_seq 정밀매칭 — claim.odSeq 있으면 정확한 품목 row 만 조회
                            if cn_od_seq:
                                _cn_row = await session.execute(
                                    _sa_text_cn(
                                        "SELECT id FROM samba_order "
                                        "WHERE source = 'lotteon' AND od_no = :od_no "
                                        "AND od_seq = :od_seq LIMIT 1"
                                    ),
                                    {"od_no": cn_od_no, "od_seq": cn_od_seq},
                                )
                            else:
                                _cn_row = await session.execute(
                                    _sa_text_cn(
                                        "SELECT id FROM samba_order "
                                        "WHERE source = 'lotteon' AND od_no = :od_no LIMIT 1"
                                    ),
                                    {"od_no": cn_od_no},
                                )
                            _cn_id = (_cn_row.fetchone() or [None])[0]
                            existing_c = (
                                await svc.repo.get_async(_cn_id) if _cn_id else None
                            )
                            if existing_c:
                                if (
                                    cn_ship_status in ("취소요청", "취소완료")
                                    and existing_c.shipping_status in _lo_shipped_guard
                                ):
                                    logger.info(
                                        f"[롯데ON][취소클레임] 배송 진행 상태 보호(DB): {cn_od_no} "
                                        f"{existing_c.shipping_status} → {cn_ship_status} 차단"
                                    )
                                    continue
                                cur_p = cancel_priority.get(
                                    existing_c.shipping_status, 0
                                )
                                new_p = cancel_priority.get(cn_ship_status, 0)
                                if cur_p == 0 or new_p >= cur_p:
                                    # status도 함께 갱신 — orders_data 분기와 일치 (2026-05-20)
                                    # 누락 시 status=cancelled인데 ship=교환요청/반품요청 잔존 사고
                                    _cn_upd = {
                                        "shipping_status": cn_ship_status,
                                        "status": cn_status,
                                    }
                                    if cn_status == "cancelled":
                                        # 정산 finalize — _finalize_cancelled 관례와 동일
                                        # (cost/shipping_fee/profit 0, revenue 유지, #326)
                                        _cn_upd.update(
                                            {
                                                "cost": 0,
                                                "shipping_fee": 0,
                                                "profit": 0,
                                            }
                                        )
                                    await svc.update_order(existing_c.id, _cn_upd)
                                    logger.info(
                                        f"[롯데ON][취소클레임] DB 직접 업데이트: {cn_od_no} → "
                                        f"{cn_status}/{cn_ship_status}"
                                    )
                except Exception as cn_err:
                    logger.warning(f"[롯데ON] 취소 클레임 조회 실패: {cn_err}")

                # 반품 클레임 조회 → samba_order.status 갱신
                # step_cd: 11=반품요청, 12=반품수거중, 13=반품완료, 14=반품거부
                try:
                    return_claims = await lotteon_client.get_returns(days=body.days)
                    logger.info(f"[롯데ON] 반품 클레임 조회: {len(return_claims)}건")
                    return_step_map = {
                        "11": ("반품요청", "return_requested"),
                        "12": ("반품요청", "returning"),
                        "13": ("반품완료", "returned"),
                        "14": ("반품거부", "return_requested"),
                    }
                    return_priority = {
                        "반품요청": 1,
                        "반품거부": 1,
                        "반품완료": 2,
                    }
                    for claim in return_claims:
                        rt_od_no = claim.get("odNo", "")
                        step_cd_r = str(claim.get("odPrgsStepCd", "") or "")
                        mapped_r = return_step_map.get(step_cd_r)
                        if not mapped_r or not rt_od_no:
                            continue
                        rt_ship_status, rt_status = mapped_r
                        found_in_data_r = False
                        for od in orders_data:
                            if od.get("od_no") == rt_od_no:
                                # 취소완료(종결)는 반품클레임이 덮어쓰지 않음 (#326 — 취소가 권위)
                                if od.get("status") == "cancelled":
                                    found_in_data_r = True
                                    break
                                cur_p = return_priority.get(
                                    od.get("shipping_status", ""), 0
                                )
                                new_p = return_priority.get(rt_ship_status, 0)
                                if cur_p == 0 or new_p >= cur_p:
                                    od["shipping_status"] = rt_ship_status
                                    od["status"] = rt_status
                                found_in_data_r = True
                                break
                        if not found_in_data_r:
                            from sqlalchemy import text as _sa_text_rt

                            _rt_row = await session.execute(
                                _sa_text_rt(
                                    "SELECT id FROM samba_order "
                                    "WHERE source = 'lotteon' AND od_no = :od_no LIMIT 1"
                                ),
                                {"od_no": rt_od_no},
                            )
                            _rt_id = (_rt_row.fetchone() or [None])[0]
                            existing_r = (
                                await svc.repo.get_async(_rt_id) if _rt_id else None
                            )
                            if existing_r:
                                # 취소완료(종결)는 반품클레임이 덮어쓰지 않음 (#326)
                                if existing_r.status == "cancelled":
                                    continue
                                cur_p = return_priority.get(
                                    existing_r.shipping_status, 0
                                )
                                new_p = return_priority.get(rt_ship_status, 0)
                                if cur_p == 0 or new_p >= cur_p:
                                    await svc.update_order(
                                        existing_r.id,
                                        {"shipping_status": rt_ship_status},
                                    )
                                    logger.info(
                                        f"[롯데ON][반품클레임] DB 직접 업데이트: {rt_od_no} → {rt_ship_status}"
                                    )
                except Exception as rt_err:
                    logger.warning(f"[롯데ON] 반품 클레임 조회 실패: {rt_err}")

                # 배송 진행 상태 갱신 (SellerDeliveryProgressStateSearch)
                # 이미 수집된 주문(상품준비→발송완료→배송완료→구매확정) 상태 업데이트
                _lo_delivery_status_map = {
                    "11": ("출고지시", "preparing"),
                    "12": ("상품준비", "preparing"),
                    "13": ("발송완료", "shipping"),
                    "14": ("배송완료", "delivered"),
                    "15": ("수취완료", "delivered"),
                    "21": ("취소완료", "cancelled"),
                    "22": ("철회", "cancelled"),
                    "23": ("회수지시", "return_requested"),
                    "24": ("회수진행", "return_requested"),
                    "25": ("회수완료", "return_requested"),
                    "26": ("회수확정", "return_requested"),
                    "27": ("반품완료", "return_requested"),
                }
                # 이미 orders_data에서 처리한 주문은 중복 갱신 불필요
                _already_in_data = {
                    od.get("order_number")
                    for od in orders_data
                    if od.get("order_number")
                }
                try:
                    progress_states = await lotteon_client.get_delivery_progress_states(
                        days=body.days
                    )
                    _ps_updated = 0
                    for ps in progress_states:
                        od_no = str(ps.get("odNo", "") or "")
                        od_seq = str(ps.get("odSeq", 1) or 1)
                        if not od_no:
                            continue
                        # 저장 시 키와 동일하게 (odNo, odSeq) 2부분만 사용
                        # procSeq는 처리 단계마다 바뀌므로 키에서 제외
                        order_number = f"{od_no}_{od_seq}"
                        if order_number in _already_in_data:
                            continue
                        step_cd = str(ps.get("odPrgsStepCd", "") or "")
                        mapped = _lo_delivery_status_map.get(step_cd)
                        if not mapped:
                            continue
                        new_ship_status, new_status = mapped
                        invc_no = str(ps.get("invcNo", "") or "")
                        dv_co_cd = str(ps.get("dvCoCd", "") or "")
                        from sqlalchemy import text as _sa_text_ps

                        _set_parts = [
                            "shipping_status = :ship_status",
                            "updated_at = now()",
                        ]
                        _ps_params: dict[str, Any] = {
                            "order_number": order_number,
                            "ship_status": new_ship_status,
                        }
                        if invc_no:
                            _set_parts.append("tracking_number = :invc_no")
                            _ps_params["invc_no"] = invc_no
                        if dv_co_cd:
                            _set_parts.append("shipping_company = :dv_co_cd")
                            _ps_params["dv_co_cd"] = dv_co_cd
                        _ps_result = await session.execute(
                            _sa_text_ps(
                                f"UPDATE samba_order SET {', '.join(_set_parts)} "
                                "WHERE source = 'lotteon' AND order_number = :order_number "
                                "AND status NOT IN ('cancelled', 'confirmed', 'return_requested')"
                            ),
                            _ps_params,
                        )
                        if _ps_result.rowcount:
                            _ps_updated += 1
                    if _ps_updated:
                        logger.info(
                            f"[주문동기화] {label}: 배송상태 갱신 {_ps_updated}건"
                        )
                except Exception as ps_err:
                    logger.warning(
                        f"[주문동기화] {label}: 롯데ON 배송상태 갱신 실패 — {ps_err}"
                    )

            elif market_type == "poison":
                from backend.domain.samba.proxy.poison import PoisonClient

                app_key = (
                    extras.get("app_key", "")
                    or extras.get("appKey", "")
                    or account["api_key"]
                    or ""
                )
                app_secret = (
                    extras.get("app_secret", "")
                    or extras.get("appSecret", "")
                    or account["api_secret"]
                    or ""
                )
                if not app_key or not app_secret:
                    results.append(
                        {
                            "account": label,
                            "status": "skip",
                            "message": "POIZON app_key/app_secret 없음",
                        }
                    )
                    continue
                poison_client = PoisonClient(app_key, app_secret)
                raw_orders = _raw_cache.get(account["id"])
                if raw_orders is None:
                    raw_orders = await poison_client.get_orders(days=body.days)
                logger.info(
                    f"[주문동기화] {label}: POIZON 주문 {len(raw_orders)}건 조회"
                )
                for ro in raw_orders:
                    orders_data.append(_parse_poison_order(ro, account["id"], label))

            elif market_type == "playauto":
                from datetime import UTC, datetime, timedelta

                from backend.domain.samba.proxy.playauto import PlayAutoClient

                api_key = extras.get("apiKey", "") or account["api_key"] or ""
                if not api_key:
                    results.append(
                        {"account": label, "status": "skip", "message": "API Key 없음"}
                    )
                    continue
                # 별칭 매핑 로드 — 2026-05-25 store_* samba_settings 폐기 후
                # samba_market_account.additional_fields 가 단일 진실 출처.
                # 현재 처리 중인 playauto 계정의 extras(=additional_fields)에서 alias1~5 추출.
                alias_map: dict[str, str] = {}
                try:
                    for ak in ("alias1", "alias2", "alias3", "alias4", "alias5"):
                        av = str(extras.get(ak, "") or "")
                        code, nick = parse_playauto_alias_entry(av)
                        if code and nick:
                            alias_map[code] = nick
                except Exception:
                    pass
                pa_client = PlayAutoClient(api_key)
                _clients_to_close.append(pa_client)
                try:
                    start_date = (
                        datetime.now(UTC) - timedelta(days=body.days)
                    ).strftime("%Y%m%d")
                    # 전체 상태 한번에 조회 (상태 필터 없이)
                    raw_orders = _raw_cache.get(account["id"])
                    if raw_orders is None:
                        raw_orders = await pa_client.get_orders(
                            start_date=start_date,
                            count=500,
                        )
                    logger.info(f"[주문동기화] 플레이오토: {len(raw_orders)}건 조회")

                    # 롯데홈쇼핑(롯데아이몰)은 삼바 직수집(lottehome) 전용 — EMP(플토)
                    # 주문은 계정 활성 여부와 무관하게 무조건 스킵한다. EMP의 롯데아이몰
                    # 연동은 플토 화면에서 주문 확인용으로 유지되므로(해제 불가) 삼바
                    # 유입만 차단. 과거 "활성 lottehome 계정 존재" 조건부 게이트는 계정이
                    # 잠시 꺼진 사이 EMP 주문이 유입돼 신규주문 중복(2026-07-14, 68행)을
                    # 만들었다.
                    # 이미 들어온 중복행 정리 — 몰 이름은 sales_channel_alias
                    # ("롯데아이몰(계정)")에 담긴다(source_site 는 소싱처 전용이라 과거
                    # source_site 대조는 한 번도 매칭된 적 없는 죽은 조건이었음).
                    from sqlalchemy import text as _pa_text

                    _del_result = await session.execute(
                        _pa_text(
                            "DELETE FROM samba_order "
                            "WHERE source = 'playauto' "
                            "AND channel_id = :cid "
                            "AND (source_site LIKE '%롯데아이몰%' OR source_site LIKE '%롯데홈쇼핑%' "
                            "     OR sales_channel_alias LIKE '%롯데아이몰%' OR sales_channel_alias LIKE '%롯데홈쇼핑%')"
                        ),
                        {"cid": account["id"]},
                    )
                    if _del_result.rowcount:
                        logger.info(
                            f"[주문동기화] 플레이오토 롯데홈쇼핑 중복 주문 {_del_result.rowcount}건 삭제"
                        )

                    for ro in raw_orders:
                        # 파생 주문 스킵 (사본-취소마감, ★교환주문 — 원주문에 이미 정보 포함)
                        _pname = ro.get("ProdName", "")
                        if _pname.startswith("[사본-") or "★교환주문" in _pname:
                            continue
                        # 롯데아이몰 주문은 직수집 전용 — 플레이오토 유입 차단
                        _ro_site = str(ro.get("SiteName", "") or "")
                        if "롯데아이몰" in _ro_site or "롯데홈쇼핑" in _ro_site:
                            continue
                        orders_data.append(
                            _parse_playauto_order(ro, account["id"], label, alias_map)
                        )
                except Exception as e:
                    logger.warning(f"[주문동기화] {label}: 플레이오토 조회 실패 — {e}")
                    results.append(
                        {"account": label, "status": "error", "message": str(e)[:100]}
                    )
                    continue
                finally:
                    await pa_client.close()
            elif market_type == "coupang":
                from backend.domain.samba.proxy.coupang import CoupangClient

                access_key = (
                    extras.get("accessKey", "") or account.get("api_key", "") or ""
                )
                secret_key = (
                    extras.get("secretKey", "") or account.get("api_secret", "") or ""
                )
                vendor_id = extras.get("vendorId", "") or seller_id or ""

                if not all([access_key, secret_key, vendor_id]):
                    results.append(
                        {
                            "account": label,
                            "status": "skip",
                            "message": "쿠팡 인증정보 없음 (accessKey/secretKey/vendorId)",
                        }
                    )
                    continue

                client = CoupangClient(access_key, secret_key, vendor_id)
                _clients_to_close.append(client)
                try:
                    raw_orders = _raw_cache.get(account["id"])
                    if raw_orders is None:
                        raw_orders = await client.get_orders(days=body.days)
                    logger.info(f"[주문동기화] 쿠팡({label}): {len(raw_orders)}건 조회")

                    # 취소·반품 요청 통합 조회 (#246) — ordersheets v5에 cancelRequests/
                    # returnRequests 필드가 없으므로 returnRequests v6 API를 별도 호출해 머지.
                    # 키: (orderId, vendorItemId) — 멀티옵션 주문에서 옵션별 정확한 매칭 (#296)
                    # vendorItemId가 없는 receipt는 (orderId, None) fallback
                    cancel_map: dict[tuple[int, int | None], dict] = {}
                    try:
                        cr_list = await client.get_cancel_and_return_requests(
                            days=max(body.days, 30)
                        )
                        for cr in cr_list or []:
                            if not isinstance(cr, dict):
                                continue
                            oid_raw = cr.get("orderId")
                            try:
                                oid = int(oid_raw) if oid_raw is not None else None
                            except (TypeError, ValueError):
                                oid = None
                            if oid is None:
                                continue
                            # cancelItems / returnItems에서 vendorItemId 추출
                            items = cr.get("cancelItems") or cr.get("returnItems") or []
                            vids: list[int | None] = []
                            for itm in items if isinstance(items, list) else []:
                                if isinstance(itm, dict):
                                    vid_raw = itm.get("vendorItemId")
                                    try:
                                        vids.append(
                                            int(vid_raw)
                                            if vid_raw is not None
                                            else None
                                        )
                                    except (TypeError, ValueError):
                                        vids.append(None)
                            if not vids:
                                vids = [
                                    None
                                ]  # vendorItemId 없는 receipt → orderId만 fallback

                            def _register(_key: tuple[int, int | None]) -> None:
                                """CANCEL 우선 정책으로 cancel_map 등록.

                                같은 키에 이미 entry 있으면 CANCEL(배송 전 취소) 을
                                RETURN(반품) 보다 우선 매핑한다.
                                """
                                _prev = cancel_map.get(_key)
                                if _prev is None:
                                    cancel_map[_key] = cr
                                elif (
                                    cr.get("receiptType") or ""
                                ).upper() == "CANCEL" and (
                                    _prev.get("receiptType") or ""
                                ).upper() != "CANCEL":
                                    cancel_map[_key] = cr

                            for vid in vids:
                                _register((oid, vid))
                            # ── (oid, None) fallback 항상 등록 ────────────────────
                            # 2026-06-09 사용자 보고: 휠라 1010099522 쿠팡 출고중지요청이
                            # 마켓엔 들어왔지만 우리 DB 에 cancel_requested 로 매핑 안 됨.
                            # 진앞: cancelItems[].vendorItemId 와 raw_orders[].orderItems[0]
                            # .vendorItemId 가 옵션 차이로 어긋나면 (oid, vid) 정확매칭 실패.
                            # 매칭측(아래 line ~5963)은 (oid, None) fallback 을 시도하는데
                            # 등록측에서 items 가 비어있을 때만 (oid, None) 키를 만들어 fallback
                            # 자체가 비어있었음. 항상 등록하도록 보강. CANCEL 우선 정책 동일.
                            _register((oid, None))
                        logger.info(
                            f"[주문동기화] 쿠팡({label}): "
                            f"취소·반품 요청 {len(cr_list or [])}건 머지 "
                            f"({len(cancel_map)} orderId×vendorItemId)"
                        )
                    except Exception as cre:
                        logger.warning(
                            f"[주문동기화] {label}: 쿠팡 취소·반품 조회 실패 — {cre}"
                        )

                    # [#599] orphan 취소·반품 receipt 되살리기 — 배송완료 고착 해소.
                    #   receipt 는 orderId 만 있고 shipmentBoxId 가 없다. 배송완료로 종결돼
                    #   get_orders 기간창(raw_orders)에서 빠진 주문은 매칭 대상이 없어 취소·반품
                    #   신호가 버려졌다(#599 증상2). cancel_map 의 orderId 중 raw_orders 에 없는
                    #   것을 orderId 발주서 단건 조회로 되살려 raw_orders 에 추가 → 아래 매칭 루프가
                    #   자동으로 cancel_info 를 붙여 파싱한다.
                    try:
                        _raw_oids: set[int] = set()
                        for _ro in raw_orders:
                            _o = _ro.get("orderId")
                            if _o is not None:
                                try:
                                    _raw_oids.add(int(_o))
                                except (TypeError, ValueError):
                                    pass
                        _orphan_oids = {k[0] for k in cancel_map.keys()} - _raw_oids
                        _ORPHAN_CAP = 50  # 폭주 가드 (쿠팡 API rate limit 보호)
                        _orphan_list = list(_orphan_oids)
                        if len(_orphan_list) > _ORPHAN_CAP:
                            logger.warning(
                                f"[주문동기화] 쿠팡({label}): orphan 취소·반품 "
                                f"{len(_orphan_list)}건 중 {_ORPHAN_CAP}건만 재조회 "
                                f"(나머지 {len(_orphan_list) - _ORPHAN_CAP}건은 다음 sync)"
                            )
                        _recovered = 0
                        for _oid in _orphan_list[:_ORPHAN_CAP]:
                            try:
                                _sheets = await client.get_ordersheets_by_order_id(_oid)
                            except Exception as _re:
                                logger.warning(
                                    f"[주문동기화] 쿠팡({label}): "
                                    f"orphan orderId={_oid} 재조회 실패 — {_re}"
                                )
                                continue
                            for _sheet in _sheets:
                                raw_orders.append(_sheet)
                                _recovered += 1
                        if _recovered:
                            logger.info(
                                f"[주문동기화] 쿠팡({label}): "
                                f"orphan 취소·반품 주문 {_recovered}건 재조회 복원"
                            )
                    except Exception as orphan_err:
                        logger.warning(
                            f"[주문동기화] {label}: orphan 재조회 단계 실패 — {orphan_err}"
                        )

                    # ACCEPT(결제완료) + 취소·반품 머지 없음 → 자동 발주확인 대상
                    unconfirmed_box_ids: list[int] = []
                    for ro in raw_orders:
                        oid_raw = ro.get("orderId")
                        try:
                            oid = int(oid_raw) if oid_raw is not None else None
                        except (TypeError, ValueError):
                            oid = None
                        # 해당 라인 vendorItemId로 정확한 매칭, 없으면 (oid, None) fallback
                        ci = None
                        if oid is not None:
                            first_item_tmp = (
                                (ro.get("orderItems") or [{}])[0]
                                if ro.get("orderItems")
                                else {}
                            )
                            vid_raw_tmp = (
                                first_item_tmp.get("vendorItemId")
                                if isinstance(first_item_tmp, dict)
                                else None
                            )
                            try:
                                vid_tmp = (
                                    int(vid_raw_tmp)
                                    if vid_raw_tmp is not None
                                    else None
                                )
                            except (TypeError, ValueError):
                                vid_tmp = None
                            ci = cancel_map.get((oid, vid_tmp)) or cancel_map.get(
                                (oid, None)
                            )
                            # 운영 추적용 — cancel_map 에 그 oid 의 다른 키 entry 가 있는데
                            # 정확매칭+fallback 모두 실패한 케이스 로깅. 정상 흐름에서는 발생
                            # 안 해야 하지만, 쿠팡 응답 스키마 변경 등 회귀 빠른 감지용.
                            if ci is None and cancel_map:
                                _other_keys = [
                                    k for k in cancel_map.keys() if k[0] == oid
                                ]
                                if _other_keys:
                                    logger.warning(
                                        f"[주문동기화] 쿠팡({label}): "
                                        f"orderId={oid} cancel/return receipt 있는데 "
                                        f"vendorItemId({vid_tmp}) 매칭 실패. "
                                        f"cancel_map 키들 for this oid={_other_keys}"
                                    )
                        try:
                            orders_data.append(
                                _parse_coupang_order(
                                    ro, account["id"], label, cancel_info=ci
                                )
                            )
                        except Exception as parse_err:
                            logger.warning(f"[주문동기화] 쿠팡 파싱 실패: {parse_err}")
                            continue
                        if (ro.get("status") or "").upper() == "ACCEPT" and ci is None:
                            box_id_raw = ro.get("shipmentBoxId")
                            try:
                                if box_id_raw is not None:
                                    unconfirmed_box_ids.append(int(box_id_raw))
                            except (TypeError, ValueError):
                                pass

                    # 발주확인 호출 (ACCEPT → INSTRUCT, 상품준비중)
                    # 자동 발주확인 토글 (#246 PR-6) — samba_settings.coupang_auto_confirm_orders
                    # 기본값 True (현재 동작 유지). 운영자가 OFF 시 box_id만 모으고 호출 스킵 → 운영자가 /confirm 수동 실행.
                    from backend.api.v1.routers.samba.proxy import _get_setting

                    _auto_setting = await _get_setting(
                        session, "coupang_auto_confirm_orders"
                    )
                    auto_confirm = True
                    if isinstance(_auto_setting, dict):
                        v = _auto_setting.get("enabled")
                        if isinstance(v, bool):
                            auto_confirm = v
                    elif isinstance(_auto_setting, bool):
                        auto_confirm = _auto_setting
                    if unconfirmed_box_ids and not auto_confirm:
                        logger.info(
                            f"[주문동기화] 쿠팡({label}): "
                            f"자동 발주확인 OFF — {len(unconfirmed_box_ids)}건 스킵"
                        )
                    elif unconfirmed_box_ids:
                        try:
                            ack_results = await client.confirm_orders(
                                unconfirmed_box_ids
                            )
                            success_box_strs = {
                                str(r["shipmentBoxId"])
                                for r in ack_results
                                if r.get("success")
                            }
                            if success_box_strs:
                                # 로컬 표시도 즉시 상품준비중으로 갱신 (다음 sync 까지 대기 X)
                                for od in orders_data:
                                    if (
                                        od.get("source") == "coupang"
                                        and od.get("order_number") in success_box_strs
                                        and od.get("shipping_status") == "결제완료"
                                    ):
                                        od["shipping_status"] = "상품준비중"

                                # 공식 가이드: 발주확인 후 단건 조회로 배송지 변경 여부 재확인 (#246).
                                # 옵션 동작이라 실패해도 동기화 자체는 진행. 변경 감지 시 로그만.
                                for box_str in success_box_strs:
                                    try:
                                        box_id_int = int(box_str)
                                        ord_sheet = (
                                            await client.get_ordersheet_by_box_id(
                                                box_id_int
                                            )
                                        )
                                        if isinstance(ord_sheet, dict):
                                            new_addr = (
                                                (ord_sheet.get("receiver") or {}).get(
                                                    "addr1"
                                                )
                                                or ord_sheet.get("receiverAddr1")
                                                or ""
                                            )
                                            if new_addr:
                                                for od in orders_data:
                                                    if (
                                                        od.get("source") == "coupang"
                                                        and od.get("order_number")
                                                        == box_str
                                                        and od.get("customer_address")
                                                        and od.get("customer_address")
                                                        != new_addr.strip()
                                                    ):
                                                        logger.warning(
                                                            f"[주문동기화] 쿠팡({label}): "
                                                            f"발주확인 후 배송지 변경 감지 boxId={box_str} "
                                                            f"old='{od.get('customer_address')}' new='{new_addr.strip()}'"
                                                        )
                                                        od["customer_address"] = (
                                                            new_addr.strip()
                                                        )
                                    except Exception as _re:
                                        logger.warning(
                                            f"[주문동기화] 쿠팡 단건 재조회 실패 boxId={box_str}: {_re}"
                                        )
                            logger.info(
                                f"[주문동기화] 쿠팡({label}): "
                                f"{len(success_box_strs)}/{len(unconfirmed_box_ids)}건 발주확인 완료"
                            )
                        except Exception as ce:
                            logger.warning(
                                f"[주문동기화] {label}: 쿠팡 발주확인 실패 — {ce}"
                            )
                except Exception as e:
                    logger.warning(f"[주문동기화] {label}: 쿠팡 조회 실패 — {e}")
                    results.append(
                        {"account": label, "status": "error", "message": str(e)[:100]}
                    )
                    continue
            elif market_type == "11st":
                from datetime import UTC, datetime, timedelta

                from backend.domain.samba.proxy.elevenst import ElevenstClient

                api_key = extras.get("apiKey", "") or account["api_key"] or ""
                if not api_key:
                    # SambaSettings의 store_11st에서 fallback
                    settings_repo = SambaSettingsRepository(session)
                    _11st_setting = await settings_repo.find_by_async(key="store_11st")
                    if _11st_setting and isinstance(_11st_setting.value, dict):
                        api_key = _11st_setting.value.get("apiKey", "") or ""
                if not api_key:
                    results.append(
                        {
                            "account": label,
                            "status": "skip",
                            "message": "11번가 API Key 없음",
                        }
                    )
                    continue

                _11st_client = ElevenstClient(api_key)
                _clients_to_close.append(_11st_client)
                _confirm_targets: list[dict[str, str]] = []
                _confirmed = 0
                _fmt = "%Y%m%d%H%M"
                # 11번가 API는 KST 기준 시간을 요구 (UTC+9)
                from zoneinfo import ZoneInfo

                _KST = ZoneInfo("Asia/Seoul")
                _start_dt = datetime.now(_KST) - timedelta(days=body.days)
                _end_dt = datetime.now(_KST)
                _start_time = _start_dt.strftime(_fmt)
                _end_time = _end_dt.strftime(_fmt)

                try:
                    # 결제완료 주문 조회
                    _raw_orders = _raw_cache.get(account["id"])
                    if _raw_orders is None:
                        _raw_orders = await _11st_client.get_orders(
                            _start_time, _end_time
                        )
                    logger.info(
                        f"[주문동기화] {label}: 11번가 주문 {len(_raw_orders)}건 조회"
                    )
                    # 결제완료(ordPrdStat=200) 주문 자동 발주확인
                    for _ro in _raw_orders:
                        # ordPrdStat=900(취소완료)은 orders_data에서 제외
                        # 취소 상태는 get_cancel_requests(취소클레임)에서만 처리
                        # → 이렇게 하지 않으면 취소요청 선제 업데이트 이후 upsert가 취소완료로 덮어씀
                        if str(_ro.get("ordPrdStat", "")) == "900":
                            continue
                        orders_data.append(
                            _parse_elevenst_order(_ro, account["id"], label)
                        )
                        # 결제완료(200) 및 처리중(202) 모두 발주확인 대상
                        if str(_ro.get("ordPrdStat", "")) in ("200", "202"):
                            _ord_no = str(_ro.get("ordNo", "") or "")
                            _ord_prd_seq = str(_ro.get("ordPrdSeq", "") or "")
                            _dlv_no = str(_ro.get("dlvNo", "") or "")
                            if _ord_no and _ord_prd_seq and _dlv_no:
                                _confirm_targets.append(
                                    {
                                        "ord_no": _ord_no,
                                        "ord_prd_seq": _ord_prd_seq,
                                        "dlv_no": _dlv_no,
                                    }
                                )
                            else:
                                logger.warning(
                                    "[주문동기화] %s: 발주확인 스킵 (dlvNo 없음) ordNo=%s ordPrdSeq=%s dlvNo=%r",
                                    label,
                                    _ord_no,
                                    _ord_prd_seq,
                                    _dlv_no,
                                )

                    if _confirm_targets:
                        _confirmed = 0
                        _confirmed_ord_nos: set[str] = set()
                        for _ct in _confirm_targets:
                            try:
                                await _11st_client.confirm_order(
                                    _ct["ord_no"], _ct["ord_prd_seq"], _ct["dlv_no"]
                                )
                                _confirmed += 1
                                _confirmed_ord_nos.add(_ct["ord_no"])
                            except Exception as _ce:
                                logger.warning(
                                    f"[주문동기화] {label}: 11번가 발주확인 실패 "
                                    f"ordNo={_ct['ord_no']} — {_ce}"
                                )
                        # 발주확인 성공한 주문의 status/shipping_status를 배송대기중으로 업데이트
                        for _od in orders_data:
                            if _od.get("order_number") in _confirmed_ord_nos:
                                _od["status"] = "wait_ship"
                                _od["shipping_status"] = "배송대기중"
                        # 이미 DB에 저장된 주문도 즉시 배송대기중으로 갱신
                        for _ord_no in _confirmed_ord_nos:
                            _ex = await svc.repo.find_by_async(order_number=_ord_no)
                            if _ex:
                                await svc.update_order(
                                    _ex.id,
                                    {"shipping_status": "배송대기중"},
                                )
                        logger.info(
                            f"[주문동기화] {label}: 11번가 발주확인 {_confirmed}/{len(_confirm_targets)}건 완료"
                        )

                    # 배송준비중 주문 추가 수집 (결제완료 목록에 없는 건만)
                    _raw_packaging = await _11st_client.get_packaging_orders(
                        _start_time, _end_time
                    )
                    logger.info(
                        f"[주문동기화] {label}: 11번가 배송준비중 {len(_raw_packaging)}건 조회"
                    )
                    # dedup 키 = (ordNo, ordPrdSeq) — 한 주문 다중옵션(여러 ordPrdSeq)
                    # 2번째+ 라인이 ordNo 단독 dedup 으로 탈락하던 누락 수정(#422, #208 회귀).
                    _fetched_keys = {
                        (d["order_number"], str(d.get("ord_prd_seq") or ""))
                        for d in orders_data
                    }
                    for _ro in _raw_packaging:
                        _ord_no = _ro.get("ordNo", "")
                        _seq = str(_ro.get("ordPrdSeq", "") or "").strip()
                        if _ord_no and (_ord_no, _seq) not in _fetched_keys:
                            orders_data.append(
                                _parse_elevenst_order(_ro, account["id"], label)
                            )
                            _fetched_keys.add((_ord_no, _seq))

                except Exception as _e:
                    logger.warning(
                        f"[주문동기화] {label}: 11번가 주문 조회 실패 — {_e}"
                    )
                    results.append(
                        {"account": label, "status": "error", "message": str(_e)[:100]}
                    )
                    continue

                # 취소/반품/교환 클레임 → 주문 상태 업데이트 (3종 병렬 조회)
                try:
                    import asyncio as _asyncio

                    from backend.domain.samba.proxy.elevenst_exchange import (
                        ElevenstExchangeClient,
                    )

                    _exchange_client = ElevenstExchangeClient(api_key)
                    _clients_to_close.append(_exchange_client)
                    # return_exceptions=True — 한 종류(취소/반품/교환) 조회가 실패해도
                    # 나머지는 처리되도록 격리. 과거 교환 빈결과(-1) 예외가 gather 전체를
                    # 무너뜨려 취소/반품 클레임 처리가 통째로 누락되던 회귀 방지 (issue #316)
                    (
                        _cancel_claims,
                        _return_claims,
                        _exchange_claims,
                    ) = await _asyncio.gather(
                        _11st_client.get_cancel_requests(_start_time, _end_time),
                        _11st_client.get_return_requests(_start_time, _end_time),
                        _exchange_client.get_exchange_requests(_start_time, _end_time),
                        return_exceptions=True,
                    )
                    if isinstance(_cancel_claims, BaseException):
                        logger.warning(
                            f"[주문동기화] {label}: 11번가 취소 조회 실패(무시) — {_cancel_claims}"
                        )
                        _cancel_claims = []
                    if isinstance(_return_claims, BaseException):
                        logger.warning(
                            f"[주문동기화] {label}: 11번가 반품 조회 실패(무시) — {_return_claims}"
                        )
                        _return_claims = []
                    if isinstance(_exchange_claims, BaseException):
                        logger.warning(
                            f"[주문동기화] {label}: 11번가 교환 조회 실패(무시) — {_exchange_claims}"
                        )
                        _exchange_claims = []
                    logger.info(
                        f"[주문동기화] {label}: 취소 {len(_cancel_claims)}건, "
                        f"반품 {len(_return_claims)}건, "
                        f"교환 {len(_exchange_claims)}건"
                    )

                    # 배송 진행 단계 보호 — 송장출력 이후로 마켓이 진행한 주문은
                    # 좀비/지연 cancel claim 으로 '취소요청'으로 되돌리지 않음
                    _shipped_guard = {
                        "송장전송완료",
                        "국내배송중",
                        "배송완료",
                        "구매확정",
                    }
                    for _claim in _cancel_claims:
                        # 상품주문번호(ordPrdNo) 우선, 없으면 주문번호(ordNo) 폴백 —
                        # returns.py:1808 검증된 패턴. silent continue 금지(issue #316)
                        _c_ord_no = _claim.get("ordNo", "")
                        _c_prd_no = _claim.get("ordPrdNo", "")
                        _match_no = _c_prd_no or _c_ord_no
                        if not _match_no:
                            logger.warning(
                                f"[주문동기화][11번가] 취소 클레임에 주문번호 없음 — 스킵: {_claim}"
                            )
                            continue
                        _found = False
                        for _od in orders_data:
                            if _od.get("order_number") in (_match_no, _c_ord_no):
                                if _od.get("shipping_status") in _shipped_guard:
                                    logger.info(
                                        f"[주문동기화][11번가] 배송 진행 상태 보호: {_match_no} "
                                        f"{_od.get('shipping_status')} → 취소요청 차단"
                                    )
                                else:
                                    _od["shipping_status"] = "취소요청"
                                    _od["status"] = "cancelled"
                                _found = True
                                break
                        # _found 여부와 관계없이 DB에 즉시 반영
                        # (upsert 단계에서 ordPrdStat=900 → 취소완료로 덮어씌워질 수 있으므로 선제 업데이트)
                        # ordPrdNo → ordNo 양방향 조회로 매칭 누락 방지
                        _ex_cancel = await svc.repo.find_by_async(
                            order_number=_match_no
                        )
                        if not _ex_cancel and _c_ord_no and _c_ord_no != _match_no:
                            _ex_cancel = await svc.repo.find_by_async(
                                order_number=_c_ord_no
                            )
                        if _ex_cancel:
                            if _ex_cancel.shipping_status in _shipped_guard:
                                logger.info(
                                    f"[주문동기화][11번가] 배송 진행 상태 보호(DB): {_match_no} "
                                    f"{_ex_cancel.shipping_status} → 취소요청 차단"
                                )
                            else:
                                await svc.update_order(
                                    _ex_cancel.id,
                                    {"shipping_status": "취소요청"},
                                )

                    for _claim in _return_claims:
                        _r_ord_no = _claim.get("ordNo", "")
                        if not _r_ord_no:
                            continue
                        _found = False
                        for _od in orders_data:
                            if _od.get("order_number") == _r_ord_no:
                                _od["shipping_status"] = "반품요청"
                                _od["status"] = "return_requested"
                                _found = True
                                break
                        if not _found:
                            _ex_return = await svc.repo.find_by_async(
                                order_number=_r_ord_no
                            )
                            if _ex_return:
                                await svc.update_order(
                                    _ex_return.id,
                                    {"shipping_status": "반품요청"},
                                )

                    for _claim in _exchange_claims:
                        _e_ord_no = _claim.get("ordNo", "")
                        if not _e_ord_no:
                            continue
                        _found = False
                        for _od in orders_data:
                            if _od.get("order_number") == _e_ord_no:
                                _od["shipping_status"] = "교환요청"
                                _od["status"] = "exchange_requested"
                                _found = True
                                break
                        # orders_data에 없어도 DB에 즉시 반영
                        # (반품거부 후 교환요청 시 orders_data에 해당 주문이 없을 수 있음)
                        _ex_exchange = await svc.repo.find_by_async(
                            order_number=_e_ord_no
                        )
                        if _ex_exchange:
                            logger.info(
                                f"[주문동기화] {label}: 교환요청 DB 반영 "
                                f"{_e_ord_no} {_ex_exchange.shipping_status} → 교환요청"
                            )
                            await svc.update_order(
                                _ex_exchange.id,
                                {"shipping_status": "교환요청"},
                            )

                except Exception as _ce:
                    logger.warning(
                        f"[주문동기화] {label}: 11번가 클레임 조회 실패 — {_ce}"
                    )
            elif market_type == "ebay":
                from backend.domain.samba.proxy.ebay import (
                    EbayApiError,
                    EbayClient,
                )

                app_id = (
                    extras.get("clientId") or extras.get("appId") or account["api_key"]
                )
                cert_id = (
                    extras.get("clientSecret")
                    or extras.get("certId")
                    or account["api_secret"]
                )
                refresh_token = extras.get("oauthToken") or extras.get("authToken", "")
                # SambaSettings 폴백
                if not (app_id and cert_id and refresh_token):
                    settings_repo = SambaSettingsRepository(session)
                    row = await settings_repo.find_by_async(key="store_ebay")
                    if row and isinstance(row.value, dict):
                        app_id = (
                            app_id
                            or row.value.get("clientId", "")
                            or row.value.get("appId", "")
                        )
                        cert_id = (
                            cert_id
                            or row.value.get("clientSecret", "")
                            or row.value.get("certId", "")
                        )
                        refresh_token = (
                            refresh_token
                            or row.value.get("oauthToken", "")
                            or row.value.get("authToken", "")
                        )
                if not (app_id and cert_id and refresh_token):
                    results.append(
                        {
                            "account": label,
                            "status": "skip",
                            "message": "eBay 인증정보 없음",
                        }
                    )
                    continue

                ebay_client = EbayClient(
                    app_id=app_id,
                    dev_id="",
                    cert_id=cert_id,
                    refresh_token=refresh_token,
                    sandbox=bool(extras.get("sandbox", False)),
                )
                _clients_to_close.append(ebay_client)
                raw_orders = _raw_cache.get(account["id"])
                if raw_orders is None:
                    try:
                        raw_orders = await ebay_client.get_orders(days=body.days)
                    except EbayApiError as e:
                        err = str(e)
                        if (
                            "scope" in err.lower()
                            or "invalid_scope" in err.lower()
                            or "insufficient" in err.lower()
                        ):
                            results.append(
                                {
                                    "account": label,
                                    "status": "error",
                                    "message": "sell.fulfillment scope 누락 — eBay 재인증 필요",
                                }
                            )
                        else:
                            results.append(
                                {
                                    "account": label,
                                    "status": "error",
                                    "message": err[:150],
                                }
                            )
                        continue

                logger.info(f"[주문동기화] {label}: eBay 주문 {len(raw_orders)}건 조회")

                # USD → KRW 환율 (exchange_rate_service의 USD effectiveRate 우선)
                ebay_exchange_rate = 1400.0
                try:
                    from backend.domain.samba.exchange_rate_service import (
                        build_exchange_rate_response,
                        get_exchange_rate_settings,
                        get_latest_exchange_rates,
                    )

                    _er_settings = await get_exchange_rate_settings(
                        session, account["tenant_id"] or tenant_id
                    )
                    _er_latest = await get_latest_exchange_rates()
                    _er_resp = build_exchange_rate_response(_er_settings, _er_latest)
                    _usd_info = _er_resp.get("currencies", {}).get("USD", {}) or {}
                    _eff_rate = float(_usd_info.get("effectiveRate") or 0)
                    if _eff_rate > 0:
                        ebay_exchange_rate = _eff_rate
                except Exception as e:
                    logger.warning(
                        f"[주문동기화] {label}: 환율 조회 실패, 폴백 1400 사용 — {e}"
                    )

                _ebay_new_this_acc: list[dict[str, Any]] = []
                for ro in raw_orders:
                    _parsed = _parse_ebay_order(
                        ro, account["id"], label, ebay_exchange_rate
                    )
                    orders_data.append(_parsed)
                    _ebay_new_this_acc.append(_parsed)

                # Finance API 실제 정산액 조회 — orderId → (net_usd, fee_usd) 매핑
                # sell.finances scope 필요. 방금 들어온 주문은 거래 미확정 상태라 매핑 없을 수 있음
                try:
                    tx_list = await ebay_client.get_transactions(days=body.days)
                    # Finance API 응답 필드:
                    #   amount                = net (이미 수수료 차감된 값)
                    #   totalFeeBasisAmount   = gross (판매가)
                    #   totalFeeAmount        = 실제 수수료
                    # 같은 orderId에 여러 거래(SALE, SHIPPING_LABEL 등) 있을 수 있음 → 누적
                    tx_map: dict[str, dict[str, float]] = {}
                    for tx in tx_list:
                        oid = tx.get("orderId", "") or ""
                        if not oid:
                            continue
                        net = float((tx.get("amount") or {}).get("value", 0) or 0)
                        gross = float(
                            (tx.get("totalFeeBasisAmount") or {}).get("value", 0) or 0
                        )
                        fee = float(
                            (tx.get("totalFeeAmount") or {}).get("value", 0) or 0
                        )
                        booking = tx.get("bookingEntry", "CREDIT")
                        tx_type = tx.get("transactionType", "")
                        tx_id = tx.get("transactionId", "")
                        tx_status = tx.get("transactionStatus", "")
                        logger.info(
                            "[eBay Finance tx] order=%s type=%s book=%s status=%s "
                            "gross=%.2f fee=%.2f net=%.2f id=%s",
                            oid,
                            tx_type,
                            booking,
                            tx_status,
                            gross,
                            fee,
                            net,
                            tx_id,
                        )
                        # DEBIT = 판매자 잔액 차감 (환불, 배송라벨 등)
                        if booking == "DEBIT":
                            net = -net
                            gross = -gross
                            fee = -fee
                        cur = tx_map.setdefault(
                            oid, {"net": 0.0, "gross": 0.0, "fee": 0.0}
                        )
                        cur["net"] += net
                        cur["gross"] += gross
                        cur["fee"] += fee

                    matched = 0
                    for od in orders_data:
                        oid = od.get("ext_order_number") or ""
                        if oid in tx_map:
                            net_usd = tx_map[oid]["net"]
                            gross_usd = tx_map[oid]["gross"]
                            fee_usd = tx_map[oid]["fee"]
                            od["revenue"] = int(round(net_usd * ebay_exchange_rate))
                            if gross_usd > 0:
                                od["fee_rate"] = round(fee_usd / gross_usd * 100, 2)
                            od["notes"] = (
                                f"gross ${gross_usd:.2f} - fee ${fee_usd:.2f} "
                                f"= net ${net_usd:.2f} @ {ebay_exchange_rate:.2f}원/USD "
                                f"(Finance API)"
                            )
                            matched += 1
                    logger.info(
                        f"[주문동기화] {label}: Finance 실제 정산 매칭 "
                        f"{matched}/{len(orders_data)}건"
                    )
                except Exception as e:
                    logger.warning(
                        f"[주문동기화] {label}: Finance API 조회 실패 "
                        f"(예상 수수료 유지) — {e}"
                    )

                # 반품/취소 수집 (최근 90일 고정)
                try:
                    returns_raw = await ebay_client.get_returns(days=90)
                    cancellations_raw = await ebay_client.get_cancellations(days=90)
                    _apply_ebay_claims_to_orders(
                        orders_data, returns_raw, cancellations_raw
                    )
                    logger.info(
                        f"[주문동기화] {label}: eBay 반품 {len(returns_raw)}건 "
                        f"+ 취소 {len(cancellations_raw)}건 매칭 (90일)"
                    )
                except Exception as e:
                    logger.warning(
                        f"[주문동기화] {label}: eBay 반품/취소 조회 실패 — {e}"
                    )

                # 오버셀 방지 — 번장(C2C 단일재고) 소싱상품이 이번 주기에 실제로 판매됐으면
                # 번장 원 판매자가 자기 글을 "판매완료"로 바꾸는 걸 기다리지 않고 즉시
                # 우리 쪽에서 sold_out 처리 + eBay 재고 0으로 내림. 2026-07-13 같은 카드가
                # 재등록(오토튠 refresh)으로 재고 1로 리셋돼 중복판매된 사고 재발 방지.
                for _od in _ebay_new_this_acc:
                    _sku = _od.get("shipment_id") or ""
                    if not _sku:
                        continue
                    try:
                        from sqlalchemy import text as _sa_text

                        _prow = await session.execute(
                            _sa_text(
                                "SELECT source_site, sale_status FROM samba_collected_product "
                                "WHERE id = :pid"
                            ),
                            {"pid": _sku},
                        )
                        _prec = _prow.first()
                        if not _prec or _prec[0] != "BUNJANG" or _prec[1] == "sold_out":
                            continue
                        await session.execute(
                            _sa_text(
                                "UPDATE samba_collected_product SET sale_status='sold_out' "
                                "WHERE id = :pid"
                            ),
                            {"pid": _sku},
                        )
                        try:
                            _offers = await ebay_client.get_offers_by_sku(_sku)
                            for _off in _offers:
                                await ebay_client.withdraw_offer(_off["offerId"])
                            logger.info(
                                f"[오버셀방지] {_sku} 판매 감지 → sold_out 처리 + eBay 재고 내림"
                            )
                        except Exception as _e:
                            logger.warning(
                                f"[오버셀방지] {_sku} eBay 재고 내림 실패: {_e}"
                            )
                    except Exception as _e:
                        logger.warning(f"[오버셀방지] {_sku} 처리 실패: {_e}")
            # (dead code 제거: 두 번째 롯데ON 블록 → 첫 번째에 병합 완료)
            elif market_type == "ssg":
                from backend.domain.samba.proxy.ssg import SSGClient

                # 계정 설정(additional_fields.feeRate)에서 수수료율 조회
                _ssg_fee_rate = float(extras.get("feeRate", 0) or 0)

                _ssg_api_key = extras.get("apiKey", "") or account["api_key"] or ""
                if not _ssg_api_key:
                    settings_repo = SambaSettingsRepository(session)
                    _ssg_setting = await settings_repo.find_by_async(key="store_ssg")
                    if _ssg_setting and isinstance(_ssg_setting.value, dict):
                        _ssg_api_key = _ssg_setting.value.get("apiKey", "") or ""
                if not _ssg_api_key:
                    results.append(
                        {
                            "account": label,
                            "status": "skip",
                            "message": "SSG API Key 없음",
                        }
                    )
                    continue

                _ssg_client = SSGClient(_ssg_api_key)
                _clients_to_close.append(_ssg_client)
                try:
                    _ssg_raw_orders = _raw_cache.get(account["id"])
                    if _ssg_raw_orders is None:
                        _ssg_raw_orders = await _ssg_client.get_orders(days=body.days)
                    # 출고대기(피킹완료) 주문 추가 조회 — listShppDirection은 배송지시(11)만 반환
                    # 캐시 여부와 무관하게 항상 호출
                    _ssg_wo_orders = await _ssg_client.get_warehouse_out_orders(
                        days=body.days
                    )
                    if _ssg_wo_orders:
                        _ssg_raw_orders = list(_ssg_raw_orders) + _ssg_wo_orders
                    logger.info(
                        f"[주문동기화] {label}: SSG 주문 {len(_ssg_raw_orders)}건 조회"
                    )
                    _ssg_unconfirmed: list[tuple[str, str]] = []
                    for _ssg_ro in _ssg_raw_orders:
                        _ord = _ssg_client.parse_order(
                            _ssg_ro, account["id"], label, fee_rate=_ssg_fee_rate
                        )
                        orders_data.append(_ord)
                        # 상품준비중(11) = 발주확인 미처리 신규주문 → 출고대기로 변경 대상
                        if str(_ssg_ro.get("shppProgStatDtlCd", "")) == "11":
                            _ssg_unconfirmed.append(
                                (
                                    str(_ssg_ro.get("shppNo", "")),
                                    str(_ssg_ro.get("shppSeq", "")),
                                )
                            )

                    # 자동 발주확인 (출고대기로 변경)
                    if _ssg_unconfirmed:
                        _ssg_confirm_ok = 0
                        for _shpp_no, _shpp_seq in _ssg_unconfirmed:
                            try:
                                await _ssg_client.confirm_order(_shpp_no, _shpp_seq)
                                _ssg_confirm_ok += 1
                                # 발주확인 성공 → orders_data의 해당 주문 상태를 출고대기로 업데이트
                                # (listShppDirection API는 발주확인 후 출고대기 주문을 반환하지 않으므로
                                #  confirm 성공 시 직접 DB 반영)
                                _confirmed_sid = f"{_shpp_no}|{_shpp_seq}"
                                for _od in orders_data:
                                    if _od.get("shipment_id") == _confirmed_sid:
                                        _od["shipping_status"] = "출고대기"
                                        break
                            except Exception as _ce:
                                logger.warning(
                                    f"[주문동기화] {label}: 발주확인 실패 "
                                    f"shppNo={_shpp_no} — {_ce}"
                                )
                        logger.info(
                            f"[주문동기화] {label}: {_ssg_confirm_ok}/{len(_ssg_unconfirmed)}건 발주확인 완료"
                        )

                    # 취소신청 주문 조회 → 상태 업데이트
                    _ssg_cancels: list[dict] = []
                    try:
                        _ssg_cancels = await _ssg_client.get_cancel_requests(
                            days=body.days
                        )
                        for _ssg_cr in _ssg_cancels:
                            orders_data.append(
                                _ssg_client.parse_cancel_request(
                                    _ssg_cr,
                                    account["id"],
                                    label,
                                    fee_rate=_ssg_fee_rate,
                                )
                            )
                        if _ssg_cancels:
                            logger.info(
                                f"[주문동기화] {label}: 취소신청 {len(_ssg_cancels)}건 조회"
                            )
                    except Exception as _ssg_ce:
                        logger.warning(
                            f"[주문동기화] {label}: SSG 취소신청 조회 실패 — {_ssg_ce}"
                        )

                    # 반품/교환 회수 대상 조회 → 상태 업데이트
                    try:
                        _ssg_returns = await _ssg_client.get_return_requests(
                            days=body.days
                        )
                        for _ret in _ssg_returns:
                            _ret_ord_no = str(
                                _ret.get("orordNo") or _ret.get("ordNo") or ""
                            )
                            if not _ret_ord_no:
                                continue
                            _div_cd = str(_ret.get("shppDivDtlCd") or "")
                            _status = "return_requested"
                            _shipping_status = (
                                "교환요청" if _div_cd == "22" else "반품요청"
                            )
                            orders_data.append(
                                {
                                    "order_number": _ret_ord_no,
                                    "channel_id": account["id"],
                                    "channel_name": label,
                                    "status": _status,
                                    "shipping_status": _shipping_status,
                                    "source": "ssg",
                                    "sale_price": 0.0,
                                    "revenue": 0.0,
                                    "fee_rate": _ssg_fee_rate,
                                    "cost": 0,
                                }
                            )
                        if _ssg_returns:
                            logger.info(
                                f"[주문동기화] {label}: 반품/교환 {len(_ssg_returns)}건 조회"
                            )
                    except Exception as _ssg_re:
                        logger.warning(
                            f"[주문동기화] {label}: SSG 반품조회 실패 — {_ssg_re}"
                        )

                    # SSG 취소 상태 전환 감지
                    # 1) 활성 주문 중 listShppDirection에 없는 것 → 취소요청 여부 단건 확인
                    # 2) 취소요청 주문 중 get_cancel_requests+listShppDirection 모두에 없는 것 → 취소완료
                    # 3) 취소요청 주문이 listShppDirection에 다시 나타나면 → parse_order가 이미 처리
                    try:
                        from sqlalchemy import text as _sa_text_cdet
                        from datetime import (
                            datetime as _cdet_dt,
                            timezone as _ctz,
                            timedelta as _ctd,
                        )

                        _ssg_seen_ord_nos = {
                            str(_ro.get("ordNo") or "")
                            for _ro in _ssg_raw_orders
                            if _ro.get("ordNo")
                        }
                        # get_cancel_requests 결과에서 아직 취소신청 중인 주문번호 집합
                        _ssg_cancel_req_nos = {
                            str(_cr.get("ordNo") or "")
                            for _cr in _ssg_cancels
                            if _cr.get("ordNo")
                        }
                        _cdet_cutoff = _cdet_dt.now(_ctz(_ctd(hours=9))) - _ctd(
                            days=body.days
                        )
                        async with get_read_session() as _cdet_sess:
                            _cdet_q = await _cdet_sess.execute(
                                _sa_text_cdet(
                                    "SELECT order_number, shipping_status FROM samba_order "
                                    "WHERE source = 'ssg' "
                                    "AND channel_id = :cid "
                                    "AND shipping_status NOT IN ("
                                    "  '취소완료','반품완료','구매확정'"
                                    ") "
                                    "AND (paid_at IS NULL OR paid_at >= :cutoff) "
                                    "AND order_number IS NOT NULL AND order_number != ''"
                                ),
                                {"cid": account["id"], "cutoff": _cdet_cutoff},
                            )
                            _cdet_rows = _cdet_q.fetchall()
                        _db_active_nos = {
                            r[0]
                            for r in _cdet_rows
                            if r[1] not in ("취소요청", "취소처리중")
                        }
                        _db_cancel_req_nos = {
                            r[0]
                            for r in _cdet_rows
                            if r[1] in ("취소요청", "취소처리중")
                        }

                        # 활성 주문 중 listShppDirection에 없는 것 → 단건 조회로 취소요청 확인
                        _ssg_need_check = _db_active_nos - _ssg_seen_ord_nos
                        if _ssg_need_check:
                            logger.info(
                                f"[주문동기화] {label}: SSG 취소 확인 대상 "
                                f"{len(_ssg_need_check)}건"
                            )
                            _ssg_cancel_found = 0
                            # API 호출 과다 방지 — 최대 30건
                            for _chk_ord_no in list(_ssg_need_check)[:30]:
                                try:
                                    _detail_items = await _ssg_client.get_order_detail(
                                        _chk_ord_no
                                    )
                                    _divs = {
                                        str(it.get("ordItemDiv", ""))
                                        for it in _detail_items
                                    }
                                    if "021" in _divs:
                                        orders_data.append(
                                            {
                                                "order_number": _chk_ord_no,
                                                "channel_id": account["id"],
                                                "channel_name": label,
                                                "status": "cancel_requested",
                                                "shipping_status": "취소요청",
                                                "source": "ssg",
                                                "sale_price": 0.0,
                                                "revenue": 0.0,
                                                "fee_rate": _ssg_fee_rate,
                                                "cost": 0,
                                            }
                                        )
                                        _ssg_cancel_found += 1
                                        logger.info(
                                            f"[주문동기화] {label}: SSG 취소 감지 "
                                            f"— {_chk_ord_no}"
                                        )
                                    elif _divs & {"031", "041"}:
                                        orders_data.append(
                                            {
                                                "order_number": _chk_ord_no,
                                                "channel_id": account["id"],
                                                "channel_name": label,
                                                "status": "return_requested",
                                                "shipping_status": "반품요청",
                                                "source": "ssg",
                                                "sale_price": 0.0,
                                                "revenue": 0.0,
                                                "fee_rate": _ssg_fee_rate,
                                                "cost": 0,
                                            }
                                        )
                                        logger.info(
                                            f"[주문동기화] {label}: SSG 반품 감지 "
                                            f"— {_chk_ord_no}"
                                        )
                                except Exception as _chk_e:
                                    logger.warning(
                                        f"[주문동기화] {label}: SSG 단건 조회 실패 "
                                        f"{_chk_ord_no} — {_chk_e}"
                                    )
                            if _ssg_cancel_found:
                                logger.info(
                                    f"[주문동기화] {label}: SSG 취소 감지 "
                                    f"{_ssg_cancel_found}건 취소요청 처리"
                                )

                        # 취소요청 주문 중 cancel_requests·listShppDirection 모두에 없는 것.
                        # #531 — 과거엔 이 집합을 무조건 '취소완료'로 flip 했으나(음성추론),
                        # 배송완료 종결주문이 listShppDirection(배송지시 only) 조회창에서
                        # 빠지면서 취소 철회 후 배송된 정상주문이 오취소됐다.
                        # get_order_detail 단건 양성확인으로 전환 — 롯데ON '21 미매핑' 가드와 동일 취지.
                        _ssg_completed = (
                            _db_cancel_req_nos - _ssg_cancel_req_nos - _ssg_seen_ord_nos
                        )
                        if _ssg_completed:

                            def _ssg_iqty(v) -> int:
                                try:
                                    return int(float(str(v or "0")))
                                except (TypeError, ValueError):
                                    return 0

                            _cpno_list = list(_ssg_completed)
                            if len(_cpno_list) > 30:
                                logger.info(
                                    f"[주문동기화] {label}: SSG 취소완료 후보 "
                                    f"{len(_cpno_list)}건 중 30건만 확인 — "
                                    f"{len(_cpno_list) - 30}건 다음 싱크로 이월"
                                )
                            logger.info(
                                f"[주문동기화] {label}: SSG 취소완료 후보 "
                                f"{min(len(_cpno_list), 30)}건 단건 양성확인"
                            )
                            _ssg_cmpl_cancel = 0
                            _ssg_cmpl_fix = 0
                            # API 호출 과다 방지 — 최대 30건
                            for _cpno in _cpno_list[:30]:
                                try:
                                    _cd_items = await _ssg_client.get_order_detail(
                                        _cpno
                                    )
                                except Exception as _cd_e:
                                    # 조회 실패 — 판단 불가, 보수적 스킵(오flip 방지)
                                    logger.warning(
                                        f"[주문동기화] {label}: SSG 취소완료 확인 조회 실패 "
                                        f"{_cpno} — {_cd_e} (스킵)"
                                    )
                                    continue
                                if not _cd_items:
                                    # 빈 응답 — 판단 불가, 보수적 스킵
                                    continue
                                _divs2 = {
                                    str(it.get("ordItemDiv", "")) for it in _cd_items
                                }
                                _cncl_qty = sum(
                                    _ssg_iqty(it.get("cnclQty")) for it in _cd_items
                                )
                                _shpmt_qty = sum(
                                    _ssg_iqty(it.get("shpmtQty")) for it in _cd_items
                                )
                                if "021" in _divs2 or _cncl_qty > 0:
                                    # 실제 취소 확인(취소구분 또는 취소수량>0) → 취소완료
                                    orders_data.append(
                                        {
                                            "order_number": _cpno,
                                            "channel_id": account["id"],
                                            "channel_name": label,
                                            "status": "cancelled",
                                            "shipping_status": "취소완료",
                                            "source": "ssg",
                                            "sale_price": 0.0,
                                            "revenue": 0.0,
                                            "fee_rate": _ssg_fee_rate,
                                            "cost": 0,
                                        }
                                    )
                                    _ssg_cmpl_cancel += 1
                                    logger.info(
                                        f"[주문동기화] {label}: SSG 취소완료 확인 — {_cpno}"
                                    )
                                elif (
                                    "011" in _divs2
                                    and _cncl_qty == 0
                                    and _shpmt_qty > 0
                                ):
                                    # 취소 철회 후 출고/배송 — 오취소 방지, 배송완료로 정정.
                                    # (financial 미포함 dict → upsert 좀비해제 분기가
                                    #  status=delivered + cancel_requested_at 해제, 금액 보존)
                                    orders_data.append(
                                        {
                                            "order_number": _cpno,
                                            "channel_id": account["id"],
                                            "channel_name": label,
                                            "status": "delivered",
                                            "shipping_status": "배송완료",
                                            "source": "ssg",
                                        }
                                    )
                                    _ssg_cmpl_fix += 1
                                    logger.info(
                                        f"[주문동기화] {label}: SSG 취소철회·출고 → "
                                        f"배송완료 정정 — {_cpno}"
                                    )
                                # 그 외(미출고·불명) → 보수적 스킵
                            if _ssg_cmpl_cancel or _ssg_cmpl_fix:
                                logger.info(
                                    f"[주문동기화] {label}: SSG 취소완료 확인 "
                                    f"{_ssg_cmpl_cancel}건 / 배송완료 정정 {_ssg_cmpl_fix}건"
                                )
                    except Exception as _cdet_e:
                        logger.warning(
                            f"[주문동기화] {label}: SSG 취소 감지 실패 — {_cdet_e}"
                        )

                except Exception as _ssg_e:
                    logger.warning(
                        f"[주문동기화] {label}: SSG 주문 조회 실패 — {_ssg_e}"
                    )
                    results.append(
                        {
                            "account": label,
                            "status": "error",
                            "message": f"SSG 주문 조회 실패: {_ssg_e}",
                        }
                    )
                    continue
            elif market_type == "gsshop":
                from backend.domain.samba.proxy.gsshop import GsShopClient
                from backend.domain.samba.account.resolver import resolve_market_creds

                _gs_creds: dict = dict(extras) if extras else {}
                if not (_gs_creds.get("supCd") or _gs_creds.get("apiKeyProd")):
                    _gs_creds = (
                        await resolve_market_creds(
                            session,
                            account["tenant_id"],
                            market_type="gsshop",
                            store_key="store_gsshop",
                        )
                        or {}
                    )
                _gs_sup_cd = (
                    _gs_creds.get("supCd", "")
                    or _gs_creds.get("storeId", "")
                    or extras.get("storeId", "")
                    or account[
                        "seller_id"
                    ]  # GS supCd는 seller_id 컬럼(gsshop_creds 빌더와 동일)
                )
                _gs_aes_key = _gs_creds.get("apiKeyProd", "") or extras.get(
                    "apiKeyProd", ""
                )
                if not _gs_sup_cd or not _gs_aes_key:
                    results.append(
                        {
                            "account": label,
                            "status": "skip",
                            "message": "GS샵 협력사코드 또는 인증키 없음",
                        }
                    )
                    continue

                _gs_client = GsShopClient(
                    sup_cd=_gs_sup_cd,
                    aes_key=_gs_aes_key,
                    sub_sup_cd=_gs_sup_cd,
                    env="prod",
                )
                from datetime import datetime as _gsdt, timedelta as _gstd
                from zoneinfo import ZoneInfo as _gszi

                _gs_KST = _gszi("Asia/Seoul")
                _gs_today = _gsdt.now(_gs_KST)
                _gs_days = int(body.days or 7)
                # GS ORD01은 sdDt '하루치'만 반환 — 기간 내 날짜별로 반복 조회해야 함
                _gs_date_list = [
                    (_gs_today - _gstd(days=_i)).strftime("%Y%m%d")
                    for _i in range(_gs_days + 1)
                ]

                # 신규주문(S), 취소(C) 각각 수집 — 날짜×processType 순회 + order_number 중복제거
                _gs_raw_orders: list[dict] = []
                _gs_seen_keys: set[str] = set()
                for _gs_d in _gs_date_list:
                    for _gs_pt in ("S", "C"):
                        try:
                            _gs_rows = await _gs_client.get_orders(
                                sd_dt=_gs_d, process_type=_gs_pt
                            )
                        except Exception as _gs_e:
                            logger.warning(
                                f"[주문동기화] {label}: GS샵 {_gs_d}/{_gs_pt} 실패 — {_gs_e}"
                            )
                            continue
                        for _gr in _gs_rows:
                            _gk = f"{_gr.get('ordNo', '')}:{_gr.get('ordItemNo', '')}"
                            if _gk in _gs_seen_keys:
                                continue
                            _gs_seen_keys.add(_gk)
                            _gs_raw_orders.append(_gr)

                # 상품명·원가 보강 — GS 주문 API의 prdNm은 '송장명'(30byte 절단)이고
                # 원가는 안 온다. 우리가 등록한 상품은 supPrdCd(=style_code/
                # site_product_id)로 samba_collected_product를 찾아 풀네임 + 소싱 원가(cost)
                # 를 가져온다. 미보유분(PlayAuto 등록 등)은 송장명/원가0 폴백.
                _gs_info_map: dict[str, dict] = {}
                _gs_sup_codes = {
                    str(_r.get("supPrdCd", "") or "")
                    for _r in _gs_raw_orders
                    if _r.get("supPrdCd")
                }
                if _gs_sup_codes:
                    from sqlalchemy import text as _gs_text

                    _gs_info_rows = await session.execute(
                        _gs_text(
                            "SELECT style_code, site_product_id, name, cost "
                            "FROM samba_collected_product "
                            "WHERE (style_code = ANY(:codes) "
                            "OR site_product_id = ANY(:codes)) "
                            "AND name IS NOT NULL AND name <> ''"
                        ),
                        {"codes": list(_gs_sup_codes)},
                    )
                    for _sc, _sp, _nm, _cst in _gs_info_rows:
                        _info = {"name": _nm, "cost": _cst}
                        if _sc:
                            _gs_info_map[str(_sc)] = _info
                        if _sp:
                            _gs_info_map[str(_sp)] = _info

                # GS 주문 수집은 읽기 전용 — 발주확인(ORD02)은 수집에서 분리.
                # GS는 주문 수집 단계에 '발주확인' 개념이 없고(배송 워크플로우는
                # 출하지시→출고완료→배송완료), 해당 주문은 플레이오토가 관리하므로
                # 수집 중 발주확인을 쏘면 충돌 위험.
                for ro in _gs_raw_orders:
                    _gs_ord_no = str(ro.get("ordNo", "") or "")
                    _gs_ord_item_no = str(ro.get("ordItemNo", "") or "")
                    _gs_ord_type = str(ro.get("ordTypeCd", "") or "")
                    _gs_ord_st = str(ro.get("ordStCd", "") or "")
                    _gs_sup_prd_cd = str(ro.get("supPrdCd", "") or "")
                    _gs_dtl_prd_cd = str(ro.get("dtlPrdCd", "") or "")
                    _gs_qty = int(ro.get("ordQty", 1) or 1)
                    _gs_ord_dt = str(ro.get("ordDt", "") or "")
                    _gs_buyer = str(ro.get("rlOrdPrsnNm", "") or "")
                    _gs_receiver = str(ro.get("custPrsnNm", "") or "")
                    _gs_phone = str(ro.get("custPrsnCelTel", "") or "")
                    _gs_zip = str(ro.get("delivZip", "") or "")
                    _gs_addr1 = str(
                        ro.get("roadNmDelivAddr1", "") or ro.get("delivAddr1", "") or ""
                    )
                    _gs_addr2 = str(
                        ro.get("roadNmDelivAddr2", "") or ro.get("delivAddr2", "") or ""
                    )
                    _gs_msg = str(ro.get("delivMsg", "") or "")
                    # 결제가=stdUprc(GS 화면 '결제'), 정산/공급가=supGivRtamt('정산')
                    # salePrc는 stdUprc-할인이라 결제가가 아님 → stdUprc 우선
                    _gs_std_uprc = int(ro.get("stdUprc", 0) or 0)
                    _gs_sup_give = int(ro.get("supGivRtamt", 0) or 0)
                    _gs_sale_prc = _gs_std_uprc or int(ro.get("salePrc", 0) or 0)
                    # 수수료율 = (결제 − 정산)/결제 — 정책 수수료와 동일(마놀25%/캐논13%)
                    _gs_fee_rate = (
                        round((_gs_std_uprc - _gs_sup_give) / _gs_std_uprc * 100, 1)
                        if _gs_std_uprc > 0 and _gs_sup_give > 0
                        else 0.0
                    )
                    _gs_opt1 = str(ro.get("attrTypNm1", "") or "")
                    _gs_opt2 = str(ro.get("attrTypNm2", "") or "")
                    _gs_opt3 = str(ro.get("attrTypNm3", "") or "")
                    _gs_prd_nm = str(ro.get("prdNm", "") or "")
                    _gs_prd_cd = str(ro.get("prdCd", "") or "")

                    if not _gs_ord_no or not _gs_ord_item_no:
                        continue

                    # 주문번호: ordNo:ordItemNo 조합.
                    # 반품(R)/교환(X)은 GS가 새 주문번호(ordNo)를 부여하고 원주문번호를
                    # orgOrdNo/orgOrdItemNo에 담아 보낸다 → 원주문번호로 매칭해야 원주문
                    # (배송완료)이 반품요청/교환요청으로 전환되고, 반품이 별개 주문으로 잡혀
                    # 정산 이중계산되는 것을 막는다. orgOrdNo 없으면 기존대로 ordNo 사용.
                    _gs_org_no = str(ro.get("orgOrdNo", "") or "")
                    _gs_org_item = str(ro.get("orgOrdItemNo", "") or "")
                    _gs_claim_order_number = None
                    if _gs_ord_type in ("R", "X") and _gs_org_no and _gs_org_item:
                        _gs_order_number = f"{_gs_org_no}:{_gs_org_item}"
                        # 반품이 부여받은 새 주문번호 — 주문 화면 표시·반품 처리용
                        _gs_claim_order_number = f"{_gs_ord_no}:{_gs_ord_item_no}"
                    else:
                        _gs_order_number = f"{_gs_ord_no}:{_gs_ord_item_no}"

                    # 상태 매핑
                    # ordTypeCd: O=주문, C=취소, R=반품, X=교환주문
                    # ordStCd: 21/22=결제완료, 31=발주완료, 44=출고지시완료
                    if _gs_ord_type == "C":
                        _gs_status = "취소완료"
                    elif _gs_ord_type == "R":
                        _gs_status = "반품요청"
                    elif _gs_ord_type == "X":
                        _gs_status = "교환요청"
                    elif _gs_ord_st in ("21", "22"):
                        _gs_status = "결제완료"
                    elif _gs_ord_st == "31":
                        _gs_status = "발주완료"
                    elif _gs_ord_st == "44":
                        _gs_status = "배송준비"
                    else:
                        _gs_status = "결제완료"

                    # 옵션 조합 ('None'/'null' 문자열 제외)
                    _gs_opt_parts = [
                        o
                        for o in [_gs_opt1, _gs_opt2, _gs_opt3]
                        if o and o.lower() not in ("none", "null")
                    ]
                    _gs_option_str = " / ".join(_gs_opt_parts) if _gs_opt_parts else ""

                    # ordDt('YYYY-MM-DD' 문자열) → timestamptz 컬럼용 datetime 변환
                    # (문자열을 그대로 넘기면 asyncpg DataError 발생)
                    _gs_paid_at = None
                    if _gs_ord_dt:
                        try:
                            _gs_paid_at = _gsdt.strptime(
                                _gs_ord_dt[:10], "%Y-%m-%d"
                            ).replace(tzinfo=_gs_KST)
                        except Exception:
                            _gs_paid_at = None

                    # 우리 등록상품이면 DB 풀네임 + 소싱 원가, 없으면 송장명/원가0 폴백
                    _gs_info = (
                        _gs_info_map.get(_gs_sup_prd_cd)
                        or _gs_info_map.get(_gs_prd_cd)
                        or {}
                    )
                    _gs_full_nm = _gs_info.get("name")
                    _gs_src_cost = int(_gs_info.get("cost") or 0)

                    orders_data.append(
                        {
                            "order_number": _gs_order_number,
                            "claim_order_number": _gs_claim_order_number,
                            "source": "gsshop",
                            "channel_id": account["id"],
                            "channel_name": label,
                            "product_name": _gs_full_nm or _gs_prd_nm or _gs_sup_prd_cd,
                            "product_id": _gs_prd_cd or _gs_sup_prd_cd,
                            "product_option": _gs_option_str,
                            "quantity": _gs_qty,
                            "paid_at": _gs_paid_at,
                            "orderer_name": _gs_buyer,
                            "customer_name": _gs_receiver,
                            "customer_phone": _gs_phone,
                            "customer_postal_code": _gs_zip,
                            "customer_address": f"{_gs_addr1} {_gs_addr2}".strip(),
                            "customer_note": _gs_msg,
                            "sale_price": _gs_sale_prc,
                            "revenue": _gs_sup_give,
                            "fee_rate": _gs_fee_rate,
                            "cost": _gs_src_cost,
                            "shipping_status": _gs_status,
                            "tenant_id": account["tenant_id"],
                        }
                    )

            elif market_type == "lottehome":
                from backend.domain.samba.proxy.lottehome import LotteHomeClient
                from backend.domain.samba.forbidden.model import SambaSettings
                from sqlalchemy import text as _sa_text  # noqa: F811 — pre-sync 고아 정리 블록에서 사용 (8504 지역 import보다 먼저 필요)
                from sqlmodel import select as _select_lh

                _lh_creds_result = await session.exec(
                    _select_lh(SambaSettings).where(
                        SambaSettings.key == "lottehome_credentials"
                    )
                )
                _lh_creds_row = _lh_creds_result.first()
                lh_creds = _lh_creds_row.value if _lh_creds_row else {}

                lh_user_id = (
                    lh_creds.get("userId", "")
                    or extras.get("userId", "")
                    or account["seller_id"]
                    or ""
                )
                lh_password = (
                    lh_creds.get("password", "") or extras.get("password", "") or ""
                )
                lh_agnc_no = lh_creds.get("agncNo", "") or extras.get("agncNo", "")
                lh_env = lh_creds.get("env", "prod")

                if not lh_user_id or not lh_password:
                    results.append(
                        {
                            "account": label,
                            "status": "skip",
                            "message": "롯데홈쇼핑 인증정보 없음",
                        }
                    )
                    continue

                await session.commit()

                # sync 시작 전: 구형식(order_number에 콜론 없음) 고아 레코드 정리.
                # 신형식(OrdNo:OrdDtlSn) 레코드가 이미 있는 경우에만 삭제하므로
                # 데이터 손실 없음. sync 후가 아닌 시작 전 실행으로 신규 생성 레코드와
                # 충돌 없음.
                try:
                    await session.execute(
                        _sa_text(
                            "DELETE FROM samba_order "
                            "WHERE source = 'lottehome' "
                            "AND channel_id = :cid "
                            "AND order_number NOT LIKE '%:%' "
                            "AND ("
                            "  (ext_order_number LIKE '%:%' AND EXISTS ("
                            "    SELECT 1 FROM samba_order s2 "
                            "    WHERE s2.channel_id = :cid "
                            "    AND s2.order_number = samba_order.ext_order_number"
                            "  ))"
                            "  OR EXISTS ("
                            "    SELECT 1 FROM samba_order s2 "
                            "    WHERE s2.channel_id = :cid "
                            "    AND s2.order_number LIKE samba_order.order_number || ':%'"
                            "  )"
                            ")"
                        ),
                        {"cid": account["id"]},
                    )
                    # 인덱스 형식(K72118:0) 레코드가 DlvUnitSn 형식(K72118:1001)과 공존하면
                    # 인덱스 형식 삭제 — 같은 OrdNo에 더 긴 suffix 레코드가 있을 때만
                    await session.execute(
                        _sa_text(
                            "DELETE FROM samba_order "
                            "WHERE source = 'lottehome' "
                            "AND channel_id = :cid "
                            "AND order_number ~ ':[0-9]$' "
                            "AND EXISTS ("
                            "  SELECT 1 FROM samba_order s2 "
                            "  WHERE s2.source = 'lottehome' "
                            "  AND s2.channel_id = :cid "
                            "  AND SPLIT_PART(s2.order_number, ':', 1) = SPLIT_PART(samba_order.order_number, ':', 1) "
                            "  AND s2.order_number != samba_order.order_number "
                            "  AND s2.order_number !~ ':[0-9]$'"
                            ")"
                        ),
                        {"cid": account["id"]},
                    )
                    # 상품명이 비어있는 레코드는 깨진 sync 결과 → 삭제 후 재생성
                    await session.execute(
                        _sa_text(
                            "DELETE FROM samba_order "
                            "WHERE source = 'lottehome' "
                            "AND channel_id = :cid "
                            "AND (product_name IS NULL OR product_name = '')"
                        ),
                        {"cid": account["id"]},
                    )
                    await session.commit()
                except Exception as _pre_clean_e:
                    await session.rollback()
                    logger.warning(
                        f"[주문동기화] {label}: pre-sync 고아 레코드 정리 실패(무시): {_pre_clean_e}"
                    )

                lh_client = LotteHomeClient(lh_user_id, lh_password, lh_agnc_no, lh_env)
                _clients_to_close.append(lh_client)

                from datetime import datetime as _dt, timedelta as _td, UTC as _UTC

                lh_end = _dt.now(_UTC)
                lh_start = lh_end - _td(days=body.days)
                lh_start_str = lh_start.strftime("%Y%m%d")
                lh_end_str = lh_end.strftime("%Y%m%d")

                _lh_seen: set[str] = set()
                _lh_seen_ordno: set[str] = set()  # list-ProdInfo로 처리된 OrdNo

                def _lh_order_key(ro: dict) -> str:
                    prod = (
                        ro.get("ProdInfo", {})
                        if isinstance(ro.get("ProdInfo"), dict)
                        else {}
                    )
                    _ord_no = str(ro.get("OrdNo", "") or "")
                    _dtl_sn = str(
                        prod.get("OrdDtlSn")
                        or prod.get("DlvUnitSn")
                        or prod.get("OrgOrdDtlSn")
                        or prod.get("ProdSeq")
                        or prod.get("ProdCode")
                        or ""
                    )
                    if _ord_no and _dtl_sn:
                        return f"{_ord_no}:{_dtl_sn}"
                    return str(
                        ro.get("SubOrdNo")
                        or prod.get("DlvUnitSn")
                        or prod.get("OrdDtlSn")
                        or _ord_no
                        or ""
                    )

                # deliver_list를 먼저 수집: OrdNo → DlvUnitSn 목록 매핑 확보.
                # new_orders에서 인덱스 대신 DlvUnitSn으로 키를 통일해
                # 다음 sync에서도 동일 키로 upsert되도록 한다.
                _dlv_status_map = {
                    "15": ("shipping", "출고지시"),
                    "16": ("shipping", "배송대기중"),
                    "17": ("delivered", "배송완료"),
                    "18": ("confirmed", "구매확정"),
                }
                _lh_dlv_cache: dict[str, list[dict]] = {}
                _lh_dlvsn_map: dict[str, list[str]] = {}  # OrdNo → [DlvUnitSn, ...]
                for _lh_stat in ["15", "16", "17", "18"]:
                    try:
                        _cached = await lh_client.search_deliver_list(
                            lh_start_str, lh_end_str, ord_dtl_stat_cd=_lh_stat
                        )
                        _lh_dlv_cache[_lh_stat] = _cached
                        for _ro in _cached:
                            _ono = str(_ro.get("OrdNo", "") or "")
                            _pi = _ro.get("ProdInfo")
                            if not _ono:
                                continue
                            if isinstance(_pi, list):
                                for _pitem in _pi:
                                    if isinstance(_pitem, dict):
                                        _dsn = str(
                                            _pitem.get("DlvUnitSn")
                                            or _pitem.get("OrdDtlSn")
                                            or ""
                                        )
                                        if _dsn and _dsn not in _lh_dlvsn_map.get(
                                            _ono, []
                                        ):
                                            _lh_dlvsn_map.setdefault(_ono, []).append(
                                                _dsn
                                            )
                            elif isinstance(_pi, dict):
                                _dsn = str(
                                    _pi.get("DlvUnitSn") or _pi.get("OrdDtlSn") or ""
                                )
                                if _dsn and _dsn not in _lh_dlvsn_map.get(_ono, []):
                                    _lh_dlvsn_map.setdefault(_ono, []).append(_dsn)
                    except Exception as _dlv_pre_e:
                        logger.warning(
                            f"[주문동기화] {label}: 배송조회(stat={_lh_stat}) 수집 실패: {_dlv_pre_e}"
                        )
                        _lh_dlv_cache[_lh_stat] = []

                _new_ord_status_map = {
                    "01": ("pending", "주문접수"),
                    "02": ("pending", "출하지시"),
                    "03": ("pending", "발송약정"),
                }
                for _lh_sel in ["01", "02", "03"]:
                    try:
                        _lh_orders = await lh_client.search_new_orders(
                            lh_start_str, lh_end_str, sel_option=_lh_sel
                        )
                    except Exception as _lh_ne:
                        # 0001=데이터없음 포함 — 한 sel_option 실패가 전체 롯데홈 블록을 크래시시키지 않도록
                        logger.warning(
                            f"[주문동기화] {label}: search_new_orders sel={_lh_sel} 실패(계속): {_lh_ne}"
                        )
                        _lh_orders = []
                    _fs, _fss = _new_ord_status_map[_lh_sel]
                    for ro in _lh_orders:
                        _prod_info_raw = ro.get("ProdInfo")
                        if isinstance(_prod_info_raw, list):
                            _no_key = str(ro.get("OrdNo", "") or "")
                            # deliver_list에서 수집한 DlvUnitSn 사용 → 키 일관성 유지
                            _dlvsn_list = _lh_dlvsn_map.get(_no_key, [])
                            for _i, _prod in enumerate(_prod_info_raw):
                                _flat = dict(ro)
                                _flat["ProdInfo"] = (
                                    _prod if isinstance(_prod, dict) else {}
                                )
                                if _dlvsn_list and _i < len(_dlvsn_list):
                                    _flat["_lh_prod_idx"] = _dlvsn_list[_i]
                                    # DlvUnitSn 키로 교체 시 이전 index-format 레코드 삭제 대상 등록
                                    _lh_replaced_old_keys.append(f"{_no_key}:{_i}")
                                else:
                                    _flat["_lh_prod_idx"] = _i
                                _p = _parse_lottehome_order(
                                    _flat, account["id"], label, _fs, _fss
                                )
                                _p["shipping_status"] = _fss
                                _dedup_key = _p.get("ext_order_number") or _p.get(
                                    "order_number", ""
                                )
                                if _dedup_key and _dedup_key not in _lh_seen:
                                    _lh_seen.add(_dedup_key)
                                    orders_data.append(_p)
                            if _no_key:
                                _lh_seen_ordno.add(_no_key)
                        else:
                            _oid = _lh_order_key(ro)
                            if _oid and _oid not in _lh_seen:
                                _lh_seen.add(_oid)
                                orders_data.append(
                                    _parse_lottehome_order(
                                        ro, account["id"], label, _fs, _fss
                                    )
                                )

                # deliver_list 처리: 캐시 재사용 (API 이중 호출 없음).
                # new_orders에서 이미 처리된 합배송 주문은 상태만 업데이트.
                _lh_dlv_replaced: set[str] = set()
                for _lh_stat in ["15", "16", "17", "18"]:
                    _fs, _fss = _dlv_status_map[_lh_stat]
                    for ro in _lh_dlv_cache.get(_lh_stat, []):
                        _dlv_ord_no = str(ro.get("OrdNo", "") or "")
                        _prod_info_raw = ro.get("ProdInfo")
                        # new_orders에서 list-ProdInfo로 이미 처리된 합배송 주문 →
                        # 상품 데이터 교체 없이 상태만 업데이트
                        if _dlv_ord_no and _dlv_ord_no in _lh_seen_ordno:
                            if _dlv_ord_no not in _lh_dlv_replaced:
                                for _o in orders_data:
                                    if _o.get("source") == "lottehome" and str(
                                        _o.get("order_number", "")
                                    ).startswith(f"{_dlv_ord_no}:"):
                                        _o["status"] = _fs
                                        _o["shipping_status"] = _fss
                                _lh_dlv_replaced.add(_dlv_ord_no)
                            continue

                        if isinstance(_prod_info_raw, list):
                            for _p in _parse_lottehome_order_multi(
                                ro, account["id"], label, _fs
                            ):
                                _p["shipping_status"] = _fss
                                _dedup_key = _p.get("ext_order_number") or _p.get(
                                    "order_number", ""
                                )
                                if _dedup_key and _dedup_key not in _lh_seen:
                                    _lh_seen.add(_dedup_key)
                                    orders_data.append(_p)
                        else:
                            _oid = _lh_order_key(ro)
                            if _oid and _oid not in _lh_seen:
                                _lh_seen.add(_oid)
                                orders_data.append(
                                    _parse_lottehome_order(
                                        ro, account["id"], label, _fs, _fss
                                    )
                                )

                def _lh_override(parsed: dict) -> None:
                    _oid = parsed.get("order_number", "")
                    if not _oid:
                        return
                    # 1) exact match → 기존 레코드 교체
                    if any(o.get("order_number") == _oid for o in orders_data):
                        orders_data[:] = [
                            o for o in orders_data if o.get("order_number") != _oid
                        ]
                        orders_data.append(parsed)
                        _lh_seen.add(_oid)
                        return
                    # 2) exact match 없음 → OrdNo prefix로 탐색 후 배송완료 건 하나를 교체
                    _ord_no = _oid.split(":")[0]
                    _prefix_matches = [
                        o
                        for o in orders_data
                        if o.get("order_number", "").split(":")[0] == _ord_no
                    ]
                    if _prefix_matches:
                        # 반품/취소 아닌 건(배송완료 등) 우선 제거, 없으면 첫 번째 제거
                        _to_remove = next(
                            (
                                o
                                for o in _prefix_matches
                                if o.get("status")
                                not in (
                                    "cancelled",
                                    "return_requested",
                                    "return_completed",
                                )
                            ),
                            _prefix_matches[0],
                        )
                        orders_data.remove(_to_remove)
                    orders_data.append(parsed)
                    _lh_seen.add(_oid)

                try:
                    _lh_cncl = await lh_client.search_cancel_orders(
                        lh_start_str, lh_end_str
                    )
                    for ro in _lh_cncl:
                        # #528 — 취소조회 OrdDtlSn 은 재발급 클레임 라인번호라
                        # 원주문(OrgOrdDtlSn)과 어긋남 → prefer_org_dtl_sn=True 로
                        # 원주문 매칭(반품 #393 과 동일). 유령 취소행 방지.
                        for parsed in _parse_lottehome_order_multi(
                            ro,
                            account["id"],
                            label,
                            "cancelled",
                            prefer_org_dtl_sn=True,
                        ):
                            _lh_override(parsed)
                except Exception as _e:
                    logger.warning(f"[주문동기화] {label}: 취소주문 실패: {_e}")

                for _ret_stat in ["20", "21"]:
                    try:
                        _lh_ret = await lh_client.search_return_orders(
                            lh_start_str, lh_end_str, ord_dtl_stat_cd=_ret_stat
                        )
                        ret_status = (
                            "return_requested"
                            if _ret_stat == "20"
                            else "return_completed"
                        )
                        for ro in _lh_ret:
                            _ret_ord_no = str(ro.get("OrdNo", "") or "")
                            _ret_prod_raw = ro.get("ProdInfo", [])
                            if isinstance(_ret_prod_raw, dict):
                                _ret_prod_raw = [_ret_prod_raw]
                            if not _ret_prod_raw:
                                _ret_prod_raw = [{}]
                            _ret_dlvsn_list = _lh_dlvsn_map.get(_ret_ord_no, [])
                            for _ri, _ret_prod in enumerate(_ret_prod_raw):
                                _ret_flat = dict(ro)
                                _ret_flat["ProdInfo"] = (
                                    _ret_prod if isinstance(_ret_prod, dict) else {}
                                )
                                # DlvUnitSn 없으면 deliver_list에서 수집한 값으로 보완
                                _has_dlvsn = bool(
                                    _ret_flat["ProdInfo"].get("OrdDtlSn")
                                    or _ret_flat["ProdInfo"].get("DlvUnitSn")
                                    or _ret_flat["ProdInfo"].get("OrgOrdDtlSn")
                                )
                                if (
                                    not _has_dlvsn
                                    and _ret_dlvsn_list
                                    and _ri < len(_ret_dlvsn_list)
                                ):
                                    _ret_flat["_lh_prod_idx"] = _ret_dlvsn_list[_ri]
                                parsed = _parse_lottehome_order(
                                    _ret_flat,
                                    account["id"],
                                    label,
                                    prefer_org_dtl_sn=True,
                                )
                                parsed["status"] = ret_status
                                parsed["shipping_status"] = (
                                    "반품요청"
                                    if ret_status == "return_requested"
                                    else "회수확정"
                                )
                                _lh_override(parsed)
                    except Exception as _e:
                        logger.warning(
                            f"[주문동기화] {label}: 반품조회(stat={_ret_stat}) 실패: {_e}"
                        )

                logger.info(
                    f"[주문동기화] {label}: 롯데홈쇼핑 주문 {len(orders_data)}건 조회"
                )

                # ② 유령상품 goods_no 자동복구(#434) — 방금 수집한 주문(product_id=
                # SiteGoodsNo, product_name=GoodsName)으로 번호 잃은 등록상품을 이름
                # 단일매칭으로 재연결. 롯데홈 API 무호출(IP리스크 0). 등록은 됐는데 번호
                # 유실된 케이스를 다음 주문 동기화 때 스스로 복구.
                try:
                    import json as _bf_json
                    import re as _bf_re

                    from sqlalchemy import text as _bf_text

                    def _bf_norm(s: object) -> str:
                        return _bf_re.sub(r"[^0-9a-z가-힣]", "", str(s or "").lower())

                    _bf_acc = str(account["id"])
                    _bf_map: dict[str, set] = {}
                    for _bf_od in orders_data:
                        if _bf_od.get("source") != "lottehome":
                            continue
                        _bf_pid = str(_bf_od.get("product_id") or "")
                        _bf_pn = _bf_norm(_bf_od.get("product_name"))
                        if _bf_pn and _bf_pid.isdigit():
                            _bf_map.setdefault(_bf_pn, set()).add(_bf_pid)
                    if _bf_map:
                        _bf_ghosts = (
                            await session.execute(
                                _bf_text(
                                    "SELECT id, name, COALESCE(market_names->>:k,'') "
                                    "FROM samba_collected_product "
                                    "WHERE registered_accounts @> CAST(:a AS jsonb) "
                                    "AND NOT jsonb_exists("
                                    "COALESCE(market_product_nos,'{}'::jsonb), :k)"
                                ),
                                {"k": _bf_acc, "a": _bf_json.dumps([_bf_acc])},
                            )
                        ).fetchall()
                        _bf_n = 0
                        for _bf_gid, _bf_gn, _bf_gm in _bf_ghosts:
                            _bf_cand = _bf_map.get(_bf_norm(_bf_gm)) or _bf_map.get(
                                _bf_norm(_bf_gn)
                            )
                            if _bf_cand and len(_bf_cand) == 1:
                                await session.execute(
                                    _bf_text(
                                        "UPDATE samba_collected_product SET "
                                        "market_product_nos = "
                                        "COALESCE(market_product_nos,'{}'::jsonb) || "
                                        "jsonb_build_object(CAST(:k AS text), "
                                        "to_jsonb(CAST(:v AS text))) WHERE id = :i"
                                    ),
                                    {
                                        "k": _bf_acc,
                                        "v": next(iter(_bf_cand)),
                                        "i": _bf_gid,
                                    },
                                )
                                _bf_n += 1
                        if _bf_n:
                            await session.commit()
                            logger.info(
                                f"[주문동기화] {label}: 롯데홈 유령 goods_no "
                                f"자동복구 {_bf_n}건(#434)"
                            )
                except Exception as _bf_e:
                    logger.warning(
                        f"[주문동기화] {label}: 롯데홈 유령 복구 스킵(무시): {_bf_e}"
                    )

            elif market_type in ("gmarket", "auction"):
                from backend.domain.samba.proxy.esmplus import (
                    ESMPlusClient,
                    resolve_esm_credentials,
                )

                # 인증 정보 — account 모델 직접 조회 (sync는 dict 스냅샷이라 model 어댑터 작성)
                class _AccountAdapter:
                    def __init__(self, fields: dict[str, Any]) -> None:
                        self.additional_fields = fields

                _esm_account = _AccountAdapter(extras)
                esm_hosting_id, esm_secret_key = await resolve_esm_credentials(
                    session, _esm_account
                )
                if not esm_hosting_id or not esm_secret_key:
                    results.append(
                        {
                            "account": label,
                            "status": "skip",
                            "message": "ESM 인증정보 없음",
                        }
                    )
                    continue
                if not seller_id:
                    results.append(
                        {
                            "account": label,
                            "status": "skip",
                            "message": "ESM seller_id 없음",
                        }
                    )
                    continue

                _esm_site = market_type  # "gmarket" or "auction"
                _esm_site_type = 2 if market_type == "gmarket" else 1
                # 기간 클램프: G마켓 31일 / 옥션 180일. to=내일(+1) 여유 위해 -1.
                _esm_max_days = 31 if market_type == "gmarket" else 180
                _esm_days = min(int(body.days or 1), _esm_max_days - 1)

                from datetime import (
                    datetime as _esm_dt,
                    timedelta as _esm_td,
                    timezone as _esm_tz,
                )

                # KST 기준 + requestDateTo=내일 — ESM은 to=오늘이면 당일/전날밤 경계
                # 주문을 date-only 경계로 제외(필터 아닌 조회조건이라 영구 누락). #369
                _esm_KST = _esm_tz(_esm_td(hours=9))
                _esm_now = _esm_dt.now(_esm_KST)
                _esm_from = (_esm_now - _esm_td(days=_esm_days)).strftime("%Y-%m-%d")
                _esm_to = (_esm_now + _esm_td(days=1)).strftime("%Y-%m-%d")

                esm_client = ESMPlusClient(
                    esm_hosting_id, esm_secret_key, seller_id, site=_esm_site
                )
                _clients_to_close.append(esm_client)

                _esm_seen: set[str] = set()
                _esm_total = 0
                # 신규주문(결제완료, OrderStatus=1) — 자동 발주확인(OrderCheck) 대상(#423)
                _esm_confirm_nos: list[str] = []
                # OrderStatus 루프 — 1=결제완료, 2=배송준비, 3=배송중, 4=배송완료, 5=구매결정
                # search_orders 내부 _esm_order_throttle()로 5.2초 인터벌 보장
                for _esm_status in (1, 2, 3, 4, 5):
                    _esm_page_index = 1
                    while True:
                        try:
                            _esm_resp = await esm_client.search_orders(
                                {
                                    "siteType": _esm_site_type,
                                    "orderStatus": _esm_status,
                                    "requestDateType": 1,
                                    "requestDateFrom": _esm_from,
                                    "requestDateTo": _esm_to,
                                    "pageIndex": _esm_page_index,
                                    "pageSize": 500,
                                }
                            )
                        except Exception as _esm_e:
                            logger.warning(
                                f"[주문동기화] {label}: ESM 주문 조회 실패 "
                                f"status={_esm_status} page={_esm_page_index} — {_esm_e}"
                            )
                            break
                        _esm_data = (
                            _esm_resp.get("Data")
                            if isinstance(_esm_resp, dict)
                            else None
                        ) or {}
                        _esm_items = _esm_data.get("RequestOrders") or []
                        if not _esm_items:
                            break
                        for _esm_it in _esm_items:
                            if not isinstance(_esm_it, dict):
                                continue
                            _oid = str(_esm_it.get("OrderNo") or "")
                            if not _oid or _oid in _esm_seen:
                                continue
                            _esm_seen.add(_oid)
                            orders_data.append(
                                _parse_esmplus_order(
                                    _esm_it, account["id"], label, market_type
                                )
                            )
                            _esm_total += 1
                            # 결제완료(신규) → 발주확인 대상 적재
                            if _esm_status == 1:
                                _esm_confirm_nos.append(_oid)
                        # 다음 페이지 종료 조건 — 500 미만이면 끝
                        if len(_esm_items) < 500:
                            break
                        _esm_page_index += 1
                logger.info(
                    f"[주문동기화] {label}: ESM({market_type}) 주문 {_esm_total}건 조회"
                )

                # 자동 발주확인(OrderCheck) — 신규주문(결제완료)을 배송준비중으로 전이(#423).
                # 토글 esm_auto_confirm_orders 기본 True. 멱등(이미확인 무시)·_call_api
                # rate-limit·try/except 로 sync 비중단. 쿠팡/11번가 패턴 미러.
                if _esm_confirm_nos:
                    from backend.api.v1.routers.samba.proxy import _get_setting

                    _esm_auto = await _get_setting(session, "esm_auto_confirm_orders")
                    _esm_auto_on = True
                    if isinstance(_esm_auto, dict):
                        _v = _esm_auto.get("enabled")
                        if isinstance(_v, bool):
                            _esm_auto_on = _v
                    elif isinstance(_esm_auto, bool):
                        _esm_auto_on = _esm_auto
                    if not _esm_auto_on:
                        logger.info(
                            f"[주문동기화] {label}: ESM 자동 발주확인 OFF — "
                            f"{len(_esm_confirm_nos)}건 스킵"
                        )
                    else:
                        _esm_conf_ok = 0
                        for _cno in _esm_confirm_nos:
                            try:
                                await esm_client.confirm_order(_cno)
                                _esm_conf_ok += 1
                                # 로컬 표시도 즉시 배송준비중으로 갱신
                                for od in orders_data:
                                    if (
                                        od.get("order_number") == _cno
                                        and od.get("shipping_status") == "결제완료"
                                    ):
                                        od["shipping_status"] = "배송준비중"
                            except Exception as _ce:
                                # 이미 확인된 주문 등 — 멱등 처리(경고만, sync 비중단)
                                logger.warning(
                                    f"[주문동기화] {label}: ESM 발주확인 실패 "
                                    f"ord={_cno} — {str(_ce)[:120]}"
                                )
                        logger.info(
                            f"[주문동기화] {label}: ESM({market_type}) 발주확인 "
                            f"{_esm_conf_ok}/{len(_esm_confirm_nos)}건 완료"
                        )

                # ── ESM 판매대금 정산 reconcile (#532) ──────────────────────
                # 주문 조회의 revenue/fee_rate 는 ServiceFee 기반 추정이라 실수수료
                # (해외채널 +5% 등)를 반영 못 함. 정산 API(getsettleorder)의
                # SettlementPrice(실 정산금)로 덮어쓴다. 정산은 배송완료·구매확정
                # 후 수일 지나 생성되므로 60일 창으로 넓게 조회(롯데온 패턴 미러).
                try:
                    _esm_settle_site = "G" if market_type == "gmarket" else "A"
                    _esm_settle_from = (_esm_now - _esm_td(days=60)).strftime(
                        "%Y-%m-%d"
                    )
                    _esm_settle_to = _esm_now.strftime("%Y-%m-%d")

                    def _esm_find_settle_rows(obj):
                        # 응답 컨테이너 키가 불확실 — ContrNo 를 가진 dict 리스트를
                        # 재귀 탐색. 못 찾으면 [] → 0매칭 no-op(오염 없음).
                        if isinstance(obj, list):
                            if (
                                obj
                                and isinstance(obj[0], dict)
                                and any(k in obj[0] for k in ("ContrNo", "contrNo"))
                            ):
                                return obj
                            for _e in obj:
                                _r = _esm_find_settle_rows(_e)
                                if _r:
                                    return _r
                        elif isinstance(obj, dict):
                            for _v in obj.values():
                                _r = _esm_find_settle_rows(_v)
                                if _r:
                                    return _r
                        return None

                    def _esm_settle_f(d, *keys):
                        for k in keys:
                            if k in d and d[k] not in (None, ""):
                                try:
                                    return float(str(d[k]))
                                except (TypeError, ValueError):
                                    return 0.0
                        return 0.0

                    # {ContrNo: [net_settlement, gross_sell]} — 환불은 반대부호로 합산
                    _esm_settle_map: dict[str, list[float]] = {}
                    _esm_settle_page = 1
                    _esm_settle_rows_total = 0
                    while _esm_settle_page <= 20:
                        try:
                            _st_resp = await esm_client.search_settle_orders(
                                {
                                    "SiteType": _esm_settle_site,
                                    "SrchType": "D1",
                                    "SrchStartDate": _esm_settle_from,
                                    "SrchEndDate": _esm_settle_to,
                                    "PageNo": _esm_settle_page,
                                    "PageRowCnt": 500,
                                }
                            )
                        except Exception as _st_e:
                            logger.warning(
                                f"[주문동기화] {label}: ESM 정산 조회 실패 "
                                f"page={_esm_settle_page} — {_st_e}"
                            )
                            break
                        _st_rows = _esm_find_settle_rows(_st_resp) or []
                        if not _st_rows:
                            break
                        for _sr in _st_rows:
                            if not isinstance(_sr, dict):
                                continue
                            _cn = str(
                                _sr.get("ContrNo") or _sr.get("contrNo") or ""
                            ).strip()
                            if not _cn:
                                continue
                            _settle = _esm_settle_f(
                                _sr, "SettlementPrice", "settlementPrice"
                            )
                            _sell = _esm_settle_f(
                                _sr, "SellOrderPrice", "sellOrderPrice"
                            )
                            _qty = _esm_settle_f(_sr, "OrderQty", "orderQty") or 1.0
                            _acc = _esm_settle_map.setdefault(_cn, [0.0, 0.0])
                            _acc[0] += _settle
                            _acc[1] += _sell * _qty
                        _esm_settle_rows_total += len(_st_rows)
                        if len(_st_rows) < 500:
                            break
                        _esm_settle_page += 1

                    # in-memory 매칭 — 이번 sync 로 들어온 주문에 실 정산값 반영
                    _esm_settle_mem = 0
                    for _od in orders_data:
                        if _od.get("source") != market_type:
                            continue
                        _acc = _esm_settle_map.get(str(_od.get("order_number") or ""))
                        if not _acc:
                            continue
                        _net, _gross = _acc
                        if _net == 0 or _gross <= 0:
                            continue
                        _od["revenue"] = _net
                        _od["fee_rate"] = round((1 - _net / _gross) * 100, 2)
                        _esm_settle_mem += 1

                    # DB 보정 — 조회창에서 빠진 구매확정 주문(롯데온 db_updated 패턴).
                    # 8093 의 방어적 rollback 이 uncommitted 를 날리므로 여기서 commit.
                    _esm_settle_db = 0
                    if _esm_settle_map:
                        from sqlalchemy import text as _sa_text_esm

                        for _cn, (_net, _gross) in _esm_settle_map.items():
                            if _net == 0 or _gross <= 0:
                                continue
                            _fr = round((1 - _net / _gross) * 100, 2)
                            try:
                                _res = await session.execute(
                                    _sa_text_esm(
                                        "UPDATE samba_order "
                                        "SET revenue = :rev, fee_rate = :fr, "
                                        "    updated_at = now() "
                                        "WHERE source = :src "
                                        "  AND order_number = :cn "
                                        "  AND (revenue IS NULL OR revenue <> :rev)"
                                    ),
                                    {
                                        "rev": _net,
                                        "fr": _fr,
                                        "src": market_type,
                                        "cn": _cn,
                                    },
                                )
                                _esm_settle_db += _res.rowcount or 0
                            except Exception as _ue:
                                logger.warning(
                                    f"[주문동기화] {label}: ESM 정산 DB UPDATE 실패 "
                                    f"ContrNo={_cn} — {_ue}"
                                )
                        try:
                            await session.commit()
                        except Exception as _ce:
                            logger.warning(
                                f"[주문동기화] {label}: ESM 정산 commit 실패 — {_ce}"
                            )
                    logger.info(
                        f"[주문동기화] {label}: ESM({market_type}) 정산 reconcile — "
                        f"정산행 {_esm_settle_rows_total}건 / in-memory "
                        f"{_esm_settle_mem}건 / DB보정 {_esm_settle_db}건"
                    )
                except Exception as _esm_settle_e:
                    logger.warning(
                        f"[주문동기화] {label}: ESM 정산 reconcile 실패 — "
                        f"{_esm_settle_e}"
                    )

                # 클레임 상태 배선 — 취소/교환/반품 중인 주문 shipping_status 업데이트
                try:
                    _esm_claim_site_type = 3 if market_type == "gmarket" else 1
                    _esm_claim_days = min(_esm_days, 6)
                    _esm_from_claim = (
                        _esm_now - _esm_td(days=_esm_claim_days)
                    ).strftime("%Y-%m-%d")
                    _esm_to_claim = _esm_now.strftime("%Y-%m-%d")
                    _claim_base = {
                        "SiteType": _esm_claim_site_type,
                        "Type": 2,
                        "StartDate": _esm_from_claim,
                        "EndDate": _esm_to_claim,
                    }
                    _cancel_ss_map = {
                        1: "취소요청",
                        2: "취소완료",
                        3: "취소완료",
                        4: "취소거부",
                    }
                    _exchange_ss_map = {
                        1: "교환요청",
                        2: "교환진행",
                        3: "교환진행",
                        4: "교환완료",
                        5: "교환거부",
                    }
                    _return_ss_map = {
                        1: "반품요청",
                        2: "반품진행",
                        3: "반품진행",
                        4: "반품완료",
                        5: "반품거부",
                    }
                    try:
                        _oc_cancels_resp = await esm_client.search_cancels(
                            {**_claim_base, "CancelStatus": 0}
                        )
                    except Exception:
                        _oc_cancels_resp = {}
                    try:
                        _oc_exchanges_resp = await esm_client.search_exchanges(
                            {**_claim_base, "ExchangeStatus": 0}
                        )
                    except Exception:
                        _oc_exchanges_resp = {}
                    try:
                        _oc_returns_resp = await esm_client.search_returns(
                            {**_claim_base, "ReturnStatus": 0}
                        )
                    except Exception:
                        _oc_returns_resp = {}

                    _oc_cancels = (
                        _oc_cancels_resp.get("Data")
                        if isinstance(_oc_cancels_resp, dict)
                        else []
                    ) or []
                    _oc_exchanges = (
                        _oc_exchanges_resp.get("Data")
                        if isinstance(_oc_exchanges_resp, dict)
                        else []
                    ) or []
                    _oc_returns = (
                        _oc_returns_resp.get("Data")
                        if isinstance(_oc_returns_resp, dict)
                        else []
                    ) or []

                    from sqlalchemy import text as _claim_text

                    _claim_updates = 0
                    for _items, _status_key, _ss_map in [
                        (_oc_cancels, "CancelStatus", _cancel_ss_map),
                        (_oc_exchanges, "ExchangeStatus", _exchange_ss_map),
                        (_oc_returns, "ReturnStatus", _return_ss_map),
                    ]:
                        for _cl in _items:
                            if not isinstance(_cl, dict):
                                continue
                            _cl_ord_no = str(_cl.get("OrderNo") or "")
                            if not _cl_ord_no:
                                continue
                            _cl_status = int(_cl.get(_status_key) or 1)
                            _cl_ss = _ss_map.get(_cl_status)
                            if not _cl_ss:
                                continue
                            try:
                                _cl_res = await session.execute(
                                    _claim_text(
                                        "UPDATE samba_order "
                                        "SET shipping_status = :ss, updated_at = now() "
                                        "WHERE source = :src AND order_number = :ono "
                                        "AND (shipping_status IS NULL OR shipping_status != :ss)"
                                    ),
                                    {
                                        "ss": _cl_ss,
                                        "src": market_type,
                                        "ono": _cl_ord_no,
                                    },
                                )
                                _claim_updates += _cl_res.rowcount or 0
                            except Exception as _cl_e:
                                logger.warning(
                                    f"[주문동기화][ESM] {label}: 클레임 상태 업데이트 실패 "
                                    f"OrderNo={_cl_ord_no} — {_cl_e}"
                                )
                    if _claim_updates:
                        await session.commit()
                    logger.info(
                        f"[주문동기화] {label}: ESM 클레임 배선 — "
                        f"cancels={len(_oc_cancels)}, exchanges={len(_oc_exchanges)}, "
                        f"returns={len(_oc_returns)}, updated={_claim_updates}"
                    )
                except Exception as _esm_claim_e:
                    logger.warning(
                        f"[주문동기화] {label}: ESM 클레임 배선 실패 — {_esm_claim_e}"
                    )

            else:
                results.append(
                    {
                        "account": label,
                        "status": "skip",
                        "message": f"{market_type} 주문 조회 미지원",
                    }
                )
                continue

            # 수집상품 매칭 캐시 — 모듈 전역 60초 TTL 캐시 사용 (sync마다 재빌드 X)
            from sqlalchemy import text as _sa_text

            # 외부 마켓 API 호출이 길어 write session이 idle in transaction
            # timeout으로 끊겼을 수 있음. 이후 INSERT/UPDATE 전 rollback으로
            # 죽은 connection을 invalidate하고 풀에서 새 connection을 받는다.
            try:
                await session.rollback()
            except BaseException as _rb_e:
                logger.warning(
                    f"[주문동기화] write session rollback 실패(무시): {_rb_e}"
                )

            _mpn_global, _mpn_by_account = await _get_mpn_cache(session, _sourcing_urls)

            # 소싱계정 캐시 — (tenant_id, site_name) → sourcing_account_id (#299)
            # 주문 동기화 시 sourcing_account_id 누락/"etc" 잔존 방지용
            # is_login_default=True 우선, 없으면 단일 계정만, 여러 개면 None(모호)
            _sa_map: dict[tuple[str, str], str | None] = {}
            try:
                async with get_read_session() as _sa_sess:
                    _sa_rows = (
                        await _sa_sess.execute(
                            _sa_text(
                                "SELECT id, tenant_id, site_name, is_login_default "
                                "FROM samba_sourcing_account WHERE is_active = true"
                            )
                        )
                    ).fetchall()
                _sa_by_key: dict[tuple[str, str], dict] = {}
                for _sa_id, _sa_tid, _sa_site, _sa_default in _sa_rows:
                    _k = (str(_sa_tid or ""), str(_sa_site or ""))
                    if _k not in _sa_by_key:
                        _sa_by_key[_k] = {"default": None, "count": 0, "first": None}
                    info = _sa_by_key[_k]
                    info["count"] += 1
                    if _sa_default:
                        info["default"] = str(_sa_id)
                    elif info["count"] == 1:
                        info["first"] = str(_sa_id)
                for _k, _info in _sa_by_key.items():
                    if _info["default"]:
                        _sa_map[_k] = _info["default"]
                    elif _info["count"] == 1:
                        _sa_map[_k] = _info["first"]
                    else:
                        _sa_map[_k] = None  # 모호 — 보정 불가
            except Exception as _sa_e:
                logger.warning(f"[주문동기화] _sa_map 빌드 실패(무시): {_sa_e}")
                _sa_map = {}

            # 미등록 입력 캐시 — 정확 키 매칭만 허용(2026-05-11 보완).
            # 과거 사고: 동일 (product_id, channel_name) 키 헐거움 → 시계 cp 800건 오염.
            # 보완:
            #   - 키: (channel_id, product_id) — 마켓×상품 정확 식별
            #   - playauto: (channel_id, product_id, _pa_site_id) — 1채널 5별칭 분리
            #   - 소스: 수동 입력본(collected_product_id IS NULL + source_url 존재)만
            #     자동매칭으로 채워진 행은 _matched 경로가 이미 처리하므로 캐시 미포함.
            _unreg_cache: dict[str, dict[str, str]] = {}
            try:
                async with get_read_session() as _unreg_sess:
                    _unreg_result = await _unreg_sess.execute(
                        _sa_text(
                            "SELECT channel_id, product_id, source, product_name, source_url, product_image "
                            "FROM samba_order "
                            "WHERE source_url IS NOT NULL AND source_url <> '' "
                            "AND collected_product_id IS NULL "
                            "AND channel_id IS NOT NULL "
                            "AND product_id IS NOT NULL"
                        )
                    )
                    _unreg_rows = _unreg_result.fetchall()
                for _ur in _unreg_rows:
                    _u_ch = str(_ur[0] or "")
                    _u_pid = str(_ur[1] or "")
                    _u_src = str(_ur[2] or "")
                    if not _u_ch or not _u_pid:
                        continue
                    if _u_src == "playauto":
                        # playauto는 _pa_site_id 차원이 필요하지만 DB엔 별도 컬럼 없음.
                        # 별칭 cross-매칭 사고 방지 위해 playauto 수동입력 전파는 보류.
                        continue
                    _ukey_build = f"{_u_ch}|{_u_pid}"
                    _unreg_cache[_ukey_build] = {
                        "source_url": _ur[4],
                        "product_image": _ur[5] or "",
                    }
            except Exception as _unreg_e:
                logger.warning(f"[주문동기화] _unreg_cache 빌드 실패(무시): {_unreg_e}")
                _unreg_cache = {}

            # 비-롯데ON 주문: order_number 배치 조회로 N+1 SELECT 제거
            _non_lotteon_nos = list(
                {
                    str(od.get("order_number", ""))
                    for od in orders_data
                    if od.get("source") != "lotteon" and od.get("order_number")
                }
            )
            # 키 = order_number(타 마켓) 또는 (order_number, ord_prd_seq)(11번가).
            # 11번가 한 주문 다중옵션(여러 ord_prd_seq)이 order_number 단독 키로
            # seq1 행에 매칭돼 seq2 가 UPDATE 경로로 조용히 소실되던 회귀 수정(#422).
            _existing_id_map: dict[Any, int] = {}

            def _existing_key(_onum: str, _src: str, _seq) -> Any:
                # 한 주문에 여러 라인(ord_prd_seq)을 분리해 내려주는 마켓은
                # (order_number, ord_prd_seq) 복합키로 매칭 — 2번째+ 라인 누락/덮어쓰기
                # 방지. 11번가(#422)·SSG(#424). 타 마켓은 order_number 단독(무회귀).
                # SSG 클레임 합성 레코드는 ord_prd_seq 없음(#521) — seq 없으면 단독 키로
                # fallback해 원본 주문을 덮어쓰기(status/shipping_status 갱신)할 수 있게 함.
                if _src == "ssg":
                    return (_onum, str(_seq)) if _seq else _onum
                if _src == "11st":
                    return (_onum, str(_seq or ""))
                return _onum

            if _non_lotteon_nos:
                _batch_tid = account["tenant_id"] or tenant_id
                _batch_cid = next(
                    (
                        od.get("channel_id")
                        for od in orders_data
                        if od.get("channel_id")
                    ),
                    None,
                )
                # asyncpg text()에서 list 파라미터 타입 오류 방지 — IN (...)으로 처리
                _ph = ", ".join(f":no_{i}" for i in range(len(_non_lotteon_nos)))
                _bulk_params: dict = {
                    f"no_{i}": v for i, v in enumerate(_non_lotteon_nos)
                }
                _bulk_params["tid"] = _batch_tid
                _bulk_params["cid"] = _batch_cid
                _bulk_q = await session.execute(
                    _sa_text(
                        f"SELECT id, order_number, source, ord_prd_seq FROM samba_order "
                        f"WHERE order_number IN ({_ph}) "
                        f"AND tenant_id IS NOT DISTINCT FROM :tid "
                        f"AND channel_id IS NOT DISTINCT FROM :cid "
                        f"ORDER BY created_at DESC"
                    ),
                    _bulk_params,
                )
                for _br in _bulk_q.fetchall():
                    _k = _existing_key(_br[1], _br[2], _br[3])
                    if _k not in _existing_id_map:
                        _existing_id_map[_k] = _br[0]
                    # SSG 클레임 합성 레코드(#521): ord_prd_seq 없어서 order_number 단독
                    # 키로 조회 → 원본 행(복합키로 저장) 못 찾아 유령 행 생성 방지.
                    # 원본 행을 order_number 단독 키로도 등록해 클레임 레코드가 찾을 수 있게 함.
                    if _br[2] == "ssg" and _br[1] not in _existing_id_map:
                        _existing_id_map[_br[1]] = _br[0]
                logger.info(
                    f"[주문동기화] {label}: 배치 중복 조회 완료 "
                    f"{len(_existing_id_map)}/{len(_non_lotteon_nos)}건 기존"
                )

            # 롯데홈쇼핑 역방향 조회: order_number가 "ord_no:ord_dtl_sn" 형식으로 바뀌었지만
            # 기존 DB 레코드는 구형식(ord_no만) order_number로 저장돼 있어 위 조회에서 못 찾음.
            # ext_order_number 필드는 구버전에도 "ord_no:ord_dtl_sn"으로 저장됐으므로 역조회.
            _lh_unfound = [
                str(od.get("order_number", ""))
                for od in orders_data
                if od.get("source") == "lottehome"
                and str(od.get("order_number", "")) not in _existing_id_map
                and od.get("order_number")
            ]
            if _lh_unfound:
                _lh_ph = ", ".join(f":lh_{i}" for i in range(len(_lh_unfound)))
                _lh_prm: dict = {f"lh_{i}": v for i, v in enumerate(_lh_unfound)}
                _lh_prm["tid"] = account["tenant_id"] or tenant_id
                _lh_q = await session.execute(
                    _sa_text(
                        f"SELECT id, ext_order_number FROM samba_order "
                        f"WHERE ext_order_number IN ({_lh_ph}) "
                        f"AND tenant_id IS NOT DISTINCT FROM :tid "
                        f"ORDER BY created_at DESC"
                    ),
                    _lh_prm,
                )
                for _lhr in _lh_q.fetchall():
                    if _lhr[1] and _lhr[1] not in _existing_id_map:
                        _existing_id_map[_lhr[1]] = _lhr[0]

            # 중복 확인 후 저장 (기존 주문은 금액/상태 업데이트)
            synced = 0
            _processed = 0
            _total = len(orders_data)
            # 청크 commit (issue #401): 건당 commit → 100건마다 + 루프 끝 일괄 commit.
            # 롯데홈쇼핑처럼 cancel 대량(대부분 update) 계정의 per-account 300초 timeout 방지.
            # _pending 은 create/update 양쪽에서 증가 — synced(create만 증가)에 묶으면
            # update 대량 계정이 중간 commit을 못 타 전체가 한 번에 몰림.
            _pending = 0
            _PERSIST_CHUNK = 100
            # 롯데홈쇼핑 style_code 보강 캐시 (issue #365) — account 단위.
            # (ch, 토큰셋) → _matched entry. 같은 토큰 조합 DB 재조회 차단.
            _lh_style_cache: dict = {}
            for order_data in orders_data:
                _processed += 1
                if _processed % 50 == 0:
                    logger.info(
                        f"[주문동기화] {label}: 주문 처리 중 {_processed}/{_total}건"
                    )
                # tenant_id 주입 (멀티테넌트 격리 — account 우선, JWT fallback)
                _tid = account["tenant_id"] or tenant_id
                if _tid:
                    order_data["tenant_id"] = _tid
                # 수집상품 매칭 — collected_product_id, product_image, source_site, source_url 보충
                # 매칭 우선순위 (오염 방지):
                #   1) (channel_id, product_id) 정확 매칭 (by_account)
                #   2) playauto master_code 글로벌 매칭 (충돌 시 거부)
                #   3) product_id 글로벌 매칭 (충돌 시 거부)
                _pid = str(order_data.get("product_id", ""))
                _pa_mc = str(order_data.get("_pa_master_code") or "")
                _ch_id = str(order_data.get("channel_id") or "")
                _matched = None
                # 1) 정확 매칭 — (channel_id, product_id)
                #    #534 — 같은 (account_id, product_no)를 다른 cp가 점유(ambiguous)면
                #    자동매칭 보류(엉뚱한 cp 오연결 방지). 관리자 확인용 경고 로그.
                if _ch_id and _pid:
                    _cand = _mpn_by_account.get(f"{_ch_id}:{_pid}")
                    if _cand and not _cand.get("ambiguous"):
                        _matched = _cand
                    elif _cand and _cand.get("ambiguous"):
                        logger.warning(
                            "[주문동기화] #534 identity 충돌 — (%s:%s) 복수 CP 점유, 자동매칭 보류",
                            _ch_id,
                            _pid,
                        )
                # 2) playauto master_code 글로벌 (master_code는 통상 unique)
                if not _matched and order_data.get("source") == "playauto" and _pa_mc:
                    _cand = _mpn_global.get(_pa_mc)
                    if _cand and not _cand.get("ambiguous"):
                        _matched = _cand
                # 3) product_id 글로벌 — 충돌(ambiguous)이면 거부
                if not _matched and _pid:
                    _cand = _mpn_global.get(_pid)
                    if _cand and not _cand.get("ambiguous"):
                        _matched = _cand
                # 3.5) 롯데홈쇼핑 중복 goods_no 보강 (issue #365) — 위 정확/글로벌 모두 실패 시.
                # 주문 goods_no(구·판매중)와 cp 저장 goods_no(신·품절)가 중복등록으로
                # 불일치 → product_name의 제조사 style_code로 cp를 매칭(순수 DB, 외부 API 無).
                # product_id 가드 없음 — 취소/배송 주문(ProdCode/GoodsNo 미제공으로 _pid 빈값,
                # product_id 없는 미등록 1,600여건)도 product_name만 있으면 매칭(issue #365 P4).
                if (
                    not _matched
                    and order_data.get("source") == "lottehome"
                    and order_data.get("product_name")
                ):
                    _matched = await _lh_resolve_by_style_code(
                        str(order_data.get("product_name", "")),
                        _ch_id,
                        _lh_style_cache,
                    )
                # 3.6) 쿠팡 sellerProductId 글로벌 폴백 (#408) — 다중옵션 리스팅의
                # 비대표 옵션 주문 미등록 대비. sellerProductId는 상품당 1개·옵션무관·
                # 안정키라 productId/vendorItemId가 인덱스에 없어도 매칭됨(백필 불필요).
                # 충돌(ambiguous) 거부 + product_id attempt 뒤에만 발동 → 회귀 없음.
                if not _matched:
                    _spid = str(order_data.get("seller_product_id") or "")
                    if _spid:
                        _cand = _mpn_global.get(_spid)
                        if _cand and not _cand.get("ambiguous"):
                            _matched = _cand
                # 3.7) 쿠팡 vendor_item_id 글로벌 폴백 — _pid(productId) 오저장/노후화 대비 (#398).
                # 등록 직후 임시 productId가 승인 후 바뀌어도 _vid(옵션ID)는 안정적.
                if not _matched:
                    _vid = str(order_data.get("vendor_item_id") or "")
                    if _vid:
                        _cand = _mpn_global.get(_vid)
                        if _cand and not _cand.get("ambiguous"):
                            _matched = _cand
                # 3.8) PlayAuto product_name style_code 보강 — MasterCode 미확보 후속.
                # lottehome 동일 style_code DB 조회 재활용 (_lh_resolve_by_style_code).
                # registered_accounts @> [channel_id] 가드로 타 채널 cp 유입 차단;
                # multi-candidate skip으로 별칭 교차오염(1채널 내 복수별칭 동일style_code) 방어.
                if (
                    not _matched
                    and order_data.get("source") == "playauto"
                    and _ch_id
                    and order_data.get("product_name")
                ):
                    _matched = await _lh_resolve_by_style_code(
                        str(order_data.get("product_name", "")),
                        _ch_id,
                        _lh_style_cache,
                    )
                # 플레이오토 별칭(site_id) 단위 매칭 검증 — 1 channel_id에 5개 별칭이
                # 묶인 구조에서 사용자가 특정 별칭에만 등록한 cp가 다른 별칭 주문에
                # 잘못 매칭되는 것을 차단. cp.market_product_nos에 `{account_id}_sites`
                # 키가 있을 때만 엄격 매칭, 없으면 호환 모드(기존 동작).
                if _matched and order_data.get("source") == "playauto":
                    _order_site_id = str(order_data.get("_pa_site_id") or "").strip()
                    _account_id = str(order_data.get("channel_id") or "")
                    _allowed_sites = _matched.get("site_ids_by_account", {}).get(
                        _account_id
                    )
                    if (
                        _allowed_sites
                        and _order_site_id
                        and _order_site_id not in _allowed_sites
                    ):
                        # 등록된 site_id에 해당 주문의 별칭이 없음 → 매칭 거부
                        _matched = None
                if _matched:
                    if not order_data.get("collected_product_id"):
                        order_data["collected_product_id"] = _matched[
                            "collected_product_id"
                        ]
                    if not order_data.get("product_image"):
                        order_data["product_image"] = _matched["product_image"]
                    if not order_data.get(
                        "source_site"
                    ) and _can_override_source_site_from_sourcing(order_data):
                        order_data["source_site"] = _matched["source_site"]
                    if not order_data.get("source_url") and _matched.get(
                        "original_link"
                    ):
                        order_data["source_url"] = _matched["original_link"]
                elif _pid and _ch_id and not order_data.get("collected_product_id"):
                    # 매칭 실패 → 삼바에서 등록했다가 삭제된 상품 케이스.
                    # 같은 (channel_id, product_id) 과거 주문에서 이미지/소싱처 백필
                    # + collected_product_id='DELETED' 표시.
                    try:
                        async with get_read_session() as _ghost_sess:
                            _ghost_row = (
                                await _ghost_sess.execute(
                                    _sa_text(
                                        "SELECT product_image, source_url, source_site "
                                        "FROM samba_order "
                                        "WHERE channel_id = :ch AND product_id = :pid "
                                        "  AND (product_image IS NOT NULL OR source_url IS NOT NULL) "
                                        "ORDER BY created_at DESC LIMIT 1"
                                    ),
                                    {"ch": _ch_id, "pid": _pid},
                                )
                            ).fetchone()
                        if _ghost_row and any(_ghost_row):
                            if _ghost_row[0] and not order_data.get("product_image"):
                                order_data["product_image"] = _ghost_row[0]
                            if _ghost_row[1] and not order_data.get("source_url"):
                                order_data["source_url"] = _ghost_row[1]
                            if _ghost_row[2] and not order_data.get("source_site"):
                                order_data["source_site"] = _ghost_row[2]
                            order_data["collected_product_id"] = "DELETED"
                    except Exception as _ge:
                        logger.warning(
                            "[주문동기화] 삭제상품 백필 실패(무시): %s", str(_ge)[:80]
                        )
                # sourcing_account_id 보충 — source_site 확정됐고 계정이 비어있으면 (#299)
                # LOTTEON 등 source_site 매칭 성공 후 sourcing_account_id="etc"/NULL 잔존 방지
                _cur_said = order_data.get("sourcing_account_id") or ""
                if not _cur_said or _cur_said == "etc":
                    _ss = order_data.get("source_site") or ""
                    _sa_key = (_tid or "", _ss)
                    if _ss and _sa_key in _sa_map and _sa_map[_sa_key]:
                        order_data["sourcing_account_id"] = _sa_map[_sa_key]
                # 매칭 검증용 임시 키 제거 (DB 저장 직전, 모델에 없는 필드)
                order_data.pop("_pa_site_id", None)
                order_data.pop("_pa_master_code", None)
                # 롯데ON 예상 정산금액 계산 (롯데ON 공식 정산공식, 2026-05-20 셀러부담 할인 반영)
                # 공식(SettleItmdSales):
                #   pymtAmt = slAmt - (셀러즉시 + 셀러부담 + 롯데부담)            # 고객결제 → actualAmt
                #             + 배송비정산 - 배송비할인
                #             - (기본수수료 + PCS수수료 + 배송비수수료 - 조정할인)   # 조정 = 롯데부담
                # 정리하면(배송비 0 가정): pymtAmt = slAmt − 셀러부담할인 − 기본수수료 − PCS수수료
                #   (당사부담할인은 고객결제 차감과 수수료 환급으로 상쇄됨)
                # 정산 API(SettleItmdSales) 매칭으로 이미 revenue가 세팅됐으면 확정값이므로 건드리지 않음.
                if order_data.get("source") == "lotteon":
                    _od_no = str(order_data.get("od_no") or "")
                    _od_seq = str(order_data.get("od_seq", "1") or "1")
                    _line_key = (_od_no, _od_seq)
                    _slamt = int(sl_amt_map.get(_line_key, 0))
                    _actual = int(actual_amt_map.get(_line_key, 0))
                    _lotte_dc = int(lotte_dc_map.get(_line_key, 0))
                    _slr_dc = int(slr_dc_map.get(_line_key, 0))
                    _ch_no = ch_no_map.get(_od_no, "")

                    # 가격비교 채널 = PCS 수수료 부과 대상
                    # account.additional_fields.lotteon_price_compare=True 면 PCS 부과.
                    # 폴백: 운영 표본 chNo (가디 100065).
                    _af = account.get("additional_fields") or {}
                    _pcs_on = bool(_af.get("lotteon_price_compare"))
                    _pcs_rate = 2.0 if (_pcs_on or _ch_no in {"100065"}) else 0.0

                    # 수수료율 결정 (우선순위)
                    # 1) account.additional_fields.lotteon_fee_rate (운영자 수동 지정, %)
                    # 2) _matched.category 1뎁스가 LotteON 한국어 1뎁스와 일치할 때만 채택
                    #    (소싱 카탈로그 카테고리는 영문/소싱 path라 거의 미매칭 — 임의 매칭 차단)
                    # 3) DEFAULT_LOTTEON_FEE_RATE (13%)
                    # 정산 확정 후 SettleItmdSales.pymtAmt 매칭으로 덮어씀.
                    from backend.domain.samba.proxy.lotteon.category_fees import (
                        DEFAULT_LOTTEON_FEE_RATE,
                        LOTTEON_CATEGORY_FEE_RATES,
                    )

                    _fee: float
                    _override_fee = _af.get("lotteon_fee_rate")
                    if _override_fee is not None:
                        try:
                            _fee = float(_override_fee)
                        except (TypeError, ValueError):
                            _fee = DEFAULT_LOTTEON_FEE_RATE
                    else:
                        _cat_for_fee = _matched.get("category", "") if _matched else ""
                        _first = (
                            _cat_for_fee.split(">")[0].strip() if _cat_for_fee else ""
                        )
                        _fee = LOTTEON_CATEGORY_FEE_RATES.get(
                            _first, DEFAULT_LOTTEON_FEE_RATE
                        )

                    if _slamt > 0:
                        # 고객결제금액 = actualAmt 우선, 없으면 slAmt − fvrAmtSum 폴백
                        # actualAmt가 slAmt와 같게 들어오는 케이스(=할인 미반영) 방지 위해
                        # slr_dc 있으면 fallback 강제: slAmt − fvr (할인 반영된 실결제)
                        _fvr = int(fvr_amt_map.get(_line_key, 0))
                        if _actual > 0 and _actual < _slamt:
                            _customer_paid = _actual
                        elif _fvr > 0:
                            _customer_paid = max(0, _slamt - _fvr)
                        else:
                            # raw에 할인합도 없음 — 셀러부담+롯데부담만으로 계산
                            _customer_paid = max(0, _slamt - _slr_dc - _lotte_dc)
                        order_data["total_payment_amount"] = _customer_paid

                        # revenue=0(손실 주문 등)을 unset으로 오인하지 않도록 sentinel(키 존재) 사용.
                        if (
                            "revenue" not in order_data
                            or order_data.get("revenue") is None
                        ):
                            _bse_cmsn = int(_slamt * _fee / 100)
                            _pcs_cmsn = int(_slamt * _pcs_rate / 100)
                            # ─────────────────────────────────────────────────────────────
                            # 롯데ON 정산예상 공식 [2026-06-02 실측 검증 · 이슈 #313]
                            #
                            # SellerDeliveryOrdersSearch raw 응답 실측 확인 (2026-06-02):
                            #   - prSfcoShrAmtSum (롯데ON부담 환급) = 전 주문 present, non-zero
                            #   - prEntpShrAmtSum (제휴몰부담) = 일부 present
                            #   - slrDcAmt / bseCmsn / pcsCmsn / pymtAmt = MISSING (SettleItmdSales에만)
                            #   → 2026-05-23 가드의 "raw에 환급 필드 없음" 가정 = 틀림
                            #
                            # 공식 (롯데ON 정산예정금액 엑셀과 일치):
                            #   pymtAmt = actualAmt − (bseCmsn + pcsCmsn − prSfcoShrAmtSum)
                            #           = _customer_paid − _bse_cmsn − _pcs_cmsn + _lotte_dc
                            #
                            # ⛔ 회귀 방지 — 다음 패턴 절대 추가 금지:
                            #   1. `_slamt − _slr_dc` 또는 `_slamt − _slr_dc − _lotte_dc`
                            #      → actualAmt가 이미 전체할인(셀러+롯데+제휴몰) 반영했는데
                            #        다시 일부 할인만 차감 = 항상 한쪽이 깨짐 (a401c15e 사고)
                            #   2. `_slamt − _slr_dc − fvrAmtSum` (66fc0837 이중차감 사고)
                            #   3. `+ _entp_dc` (prEntpShrAmtSum 제휴몰부담 환급) 추가
                            #      → 롯데ON 정산 공식 비포함 — 제휴몰이 별도 정산하는 구조
                            #
                            # 핵심 원칙: 할인은 _customer_paid 계산에서 한 번만 반영.
                            #          revenue = _customer_paid − 수수료 + 롯데부담환급(_lotte_dc)
                            # 확정값: SettleItmdSales.pymtAmt 매칭 성공 시 덮어씀.
                            # ─────────────────────────────────────────────────────────────
                            _revenue = max(
                                0,
                                _customer_paid - _bse_cmsn - _pcs_cmsn + _lotte_dc,
                            )
                            order_data["revenue"] = _revenue
                            # 화면 수수료율 — 마켓수수료/실결제 기준 (롯데ON 정산내역 "실수수료율" 정의)
                            order_data["fee_rate"] = (
                                round((_bse_cmsn + _pcs_cmsn) / _customer_paid * 100, 2)
                                if _customer_paid > 0
                                else 0
                            )
                    elif (
                        "revenue" not in order_data or order_data.get("revenue") is None
                    ):
                        # raw 매핑 실패 폴백 — 위에서 결정된 _fee 재사용, PCS도 동일 적용
                        _sp = int(order_data.get("sale_price", 0) or 0)
                        _bse_cmsn = int(_sp * _fee / 100)
                        _pcs_cmsn = int(_sp * _pcs_rate / 100)
                        order_data["total_payment_amount"] = _sp
                        # 실효율 통일 — 정상경로와 동일하게 마켓수수료/실결제 기준
                        order_data["fee_rate"] = (
                            round((_bse_cmsn + _pcs_cmsn) / _sp * 100, 2)
                            if _sp > 0
                            else _fee
                        )
                        # fallback: raw 없으면 _lotte_dc=0 — 공식 일관성 유지
                        order_data["revenue"] = max(
                            0, _sp - _bse_cmsn - _pcs_cmsn + _lotte_dc
                        )
                # 롯데홈쇼핑 정산금액 계산 — account.additional_fields.commission_rate 우선, 폴백 25%
                if order_data.get("source") == "lottehome":
                    _lh_fee = float(
                        (account.get("additional_fields") or {}).get("commission_rate")
                        or 25.0
                    )
                    _lh_total = int(order_data.get("total_payment_amount") or 0)
                    order_data["fee_rate"] = _lh_fee
                    if not order_data.get("revenue") and _lh_total > 0:
                        order_data["revenue"] = max(
                            0, int(_lh_total * (1 - _lh_fee / 100))
                        )
                # 미등록 입력 자동 적용 — 정확 키 매칭만 허용(2026-05-11 보완).
                # 과거 (product_id, channel_name) 키는 헐거워서 시계 cp 800건 오염 사고 발생.
                # 보완: (channel_id, product_id) 정확 매칭 + playauto는 site_id 추가.
                # _matched(수집상품 자동매칭)가 이미 채운 경우 그쪽 우선이므로 건드리지 않음.
                if not _matched and _ch_id and _pid:
                    if order_data.get("source") == "playauto":
                        _pa_sid = str(order_data.get("_pa_site_id") or "")
                        _ukey = f"{_ch_id}|{_pid}|{_pa_sid}"
                    else:
                        _ukey = f"{_ch_id}|{_pid}"
                    _unreg_matched = _unreg_cache.get(_ukey)
                    if _unreg_matched:
                        if not order_data.get("source_url"):
                            order_data["source_url"] = _unreg_matched["source_url"]
                        if (
                            not order_data.get("product_image")
                            and _unreg_matched["product_image"]
                        ):
                            order_data["product_image"] = _unreg_matched[
                                "product_image"
                            ]
                # status는 사용자가 직접 관리 — shipping_status 따라 자동변경 금지
                # 상품명에서 소싱처 상품번호 추출 → source_site/source_url 보충
                # 플레이오토는 1 channel에 5 별칭이 묶인 구조라 product_name 끝 공통 무신사
                # goods_no가 별칭 무관하게 cross-매칭됨 (예: 캐논 주문이 고경 등록 cp에 매칭).
                # → 플레이오토 주문은 본 분기 비활성화. master_code 직접 매칭만 신뢰.
                if (
                    not order_data.get("source_url")
                    and order_data.get("source") != "playauto"
                ):
                    import re as _re

                    _pname = order_data.get("product_name", "")
                    _id_match = _re.search(r"\b(\d{6,})\s*$", _pname)
                    if _id_match:
                        _sid = _id_match.group(1)
                        # 1차-A: site_product_id 정확 매칭
                        # cp.source_url을 직접 끌어와 sourcing_urls 템플릿 기반 추정보다 우선 사용한다
                        # (2026-05-20: 상품명 끝 숫자로 추정한 URL이 옵션/스타일코드와 충돌해
                        # 엉뚱한 상품을 열어주던 사고 — 푸마↔스파이더 — 재발 방지).
                        _cp_check = await session.execute(
                            _sa_text(
                                "SELECT id, source_site, images, site_product_id, source_url "
                                "FROM samba_collected_product "
                                "WHERE site_product_id = :sid "
                                "ORDER BY (market_product_nos IS NOT NULL) DESC, created_at ASC "
                                "LIMIT 1"
                            ),
                            {"sid": _sid},
                        )
                        _cp_row = _cp_check.fetchone()
                        # 1차-B prefix 매칭 영구 제거 (2026-05-20).
                        # 상품명 끝 6자리(_sid='403372')가 무관한 다른 cp의 7자리
                        # site_product_id(예: '4033721' 스파이더)와 prefix LIKE로 우연
                        # 일치하여 엉뚱한 상품으로 매칭되는 사고 발생.
                        # SSG itemId 끝자리 잘림은 정확 매칭만으로 처리하거나 별도 정규화 필요.
                        if _cp_row:
                            _matched_spid = _cp_row[3] or _sid
                            _cp_source_url = _cp_row[4] if len(_cp_row) > 4 else None
                            if not order_data.get("collected_product_id"):
                                order_data["collected_product_id"] = _cp_row[0]
                            if _can_override_source_site_from_sourcing(order_data):
                                order_data["source_site"] = _cp_row[1]
                            # cp.source_url 우선, 없으면 sourcing_urls 템플릿 fallback
                            order_data["source_url"] = (
                                _cp_source_url
                                or _sourcing_urls.get(_cp_row[1], "").format(
                                    _matched_spid
                                )
                            )
                            if (
                                not order_data.get("product_image")
                                and _cp_row[2]
                                and isinstance(_cp_row[2], list)
                            ):
                                order_data["product_image"] = _cp_row[2][0]
                        # 매칭 실패 시 무신사 단정하지 않음 — source_site/url 오염 방지
                        # (과거 자릿수만으로 MUSINSA로 추론하던 fallback 제거: 2026-05-10)
                # 중복 체크: 롯데ON은 od_no+od_seq 기반, 기타는 order_number 기반
                # proc_seq는 주문 상태 변경 시 바뀌므로 중복 체크에서 제외
                _normalize_synced_order_status(order_data)
                if order_data.get("source") == "lotteon" and order_data.get("od_no"):
                    # 중복 차단 — channel_id 제외하고 (tenant_id, od_no, od_seq)로만 매칭.
                    # 동일 API key를 공유한 2개 마켓계정이 같은 주문을 양쪽 channel에 중복
                    # 저장하던 사고 방지(2026-05-25).
                    _lo_row = await session.execute(
                        _sa_text(
                            "SELECT id FROM samba_order "
                            "WHERE source = 'lotteon' "
                            "AND tenant_id IS NOT DISTINCT FROM :tid "
                            "AND od_no = :od_no "
                            "AND od_seq = :od_seq "
                            "LIMIT 1"
                        ),
                        {
                            "tid": order_data.get("tenant_id"),
                            "od_no": order_data["od_no"],
                            "od_seq": order_data.get("od_seq", "1"),
                        },
                    )
                    _lo_id = (_lo_row.fetchone() or [None])[0]
                    existing = await svc.repo.get_async(_lo_id) if _lo_id else None
                else:
                    _existing_id = _existing_id_map.get(
                        _existing_key(
                            str(order_data.get("order_number", "")),
                            order_data.get("source", ""),
                            order_data.get("ord_prd_seq"),
                        )
                    )
                    existing = (
                        await svc.repo.get_async(_existing_id) if _existing_id else None
                    )
                if (
                    not existing
                    and order_data.get("shipment_id")
                    and order_data.get("product_id")
                    # 롯데ON 제외: 같은 sitmNo(shipment_id)에 서로 다른 odNo의 주문이 다수 존재
                    # 가능 — fallback 매칭이 다른 사람 행을 잘못 매칭해 한 행에 두 주문 데이터를
                    # 짬뽕시키는 사고 원인 (2026-05-19 임재광/최호선 사례).
                    # 롯데ON은 (channel_id, od_no, od_seq) 매칭만 신뢰.
                    and order_data.get("source") != "lotteon"
                ):
                    # 같은 orderId + 상품번호로 이미 있는 주문 검색
                    _dup_candidates = await svc.repo.filter_by_async(
                        shipment_id=order_data["shipment_id"], limit=10
                    )
                    existing = next(
                        (
                            d
                            for d in _dup_candidates
                            if d.product_id == order_data["product_id"]
                            and (d.product_option or "")
                            == (order_data.get("product_option") or "")
                            # 11번가·SSG는 ord_prd_seq 일치까지 요구 — 같은 배송번호·
                            # 상품의 동일옵션 다중라인 오합치 차단(#422, #424)
                            and (
                                order_data.get("source") not in ("11st", "ssg")
                                or (d.ord_prd_seq or "")
                                == (order_data.get("ord_prd_seq") or "")
                            )
                        ),
                        None,
                    )
                    if existing:
                        # order_number 갱신 (발주확인 후 변경된 productOrderId)
                        await svc.repo.update_async(
                            existing.id, order_number=order_data["order_number"]
                        )
                if existing:
                    # 기존 주문: sale_price, 이미지, 상태, 마켓주문상태 업데이트
                    update_fields: dict[str, Any] = {}
                    # tenant_id 보충 (기존 NULL 데이터 대응)
                    if order_data.get("tenant_id") and not existing.tenant_id:
                        update_fields["tenant_id"] = order_data["tenant_id"]
                    if (
                        order_data.get("sale_price")
                        and order_data["sale_price"] != existing.sale_price
                    ):
                        update_fields["sale_price"] = order_data["sale_price"]
                        if order_data.get("revenue") is not None:
                            update_fields["revenue"] = order_data["revenue"]
                        if order_data.get("fee_rate") is not None:
                            update_fields["fee_rate"] = order_data["fee_rate"]
                    # 고객결제금액 갱신: 변경됐거나 기존 NULL이면 채움
                    new_total_paid = order_data.get("total_payment_amount")
                    if new_total_paid is not None:
                        existing_total = (
                            existing.total_payment_amount
                            if existing.total_payment_amount is not None
                            else None
                        )
                        if existing_total is None or float(new_total_paid) != float(
                            existing_total
                        ):
                            update_fields["total_payment_amount"] = float(
                                new_total_paid
                            )
                    # 결제금액 1,000원 이하인 기존 주문이 아직 pending 이면 배송완료로 전환
                    if existing.status == "pending":
                        _ex_pamt = float(
                            update_fields.get("total_payment_amount")
                            or existing.total_payment_amount
                            or order_data.get("sale_price")
                            or 0
                        )
                        if 0 < _ex_pamt <= 1000:
                            update_fields["status"] = "delivered"
                    if order_data.get("product_image") and not existing.product_image:
                        update_fields["product_image"] = order_data["product_image"]
                    # 상품명/옵션명이 빈 경우 새 데이터로 복구
                    if order_data.get("product_name") and not existing.product_name:
                        update_fields["product_name"] = order_data["product_name"]
                    if order_data.get("product_option") and not existing.product_option:
                        update_fields["product_option"] = order_data["product_option"]
                    new_source_site = str(order_data.get("source_site") or "").strip()
                    existing_source_site = str(existing.source_site or "").strip()
                    if new_source_site and not existing_source_site:
                        update_fields["source_site"] = new_source_site
                    elif (
                        order_data.get("source") == "playauto"
                        and new_source_site
                        and new_source_site != existing_source_site
                        and "(" in new_source_site
                    ):
                        update_fields["source_site"] = new_source_site
                    if order_data.get("source_url") and not existing.source_url:
                        update_fields["source_url"] = order_data["source_url"]
                    # collected_product_id 백필 — 과거 매칭 캐시 LIMIT 컷오프로 끊긴
                    # 기존 주문이 다음 sync 때 자동 재연결되도록.
                    if (
                        order_data.get("collected_product_id")
                        and not existing.collected_product_id
                    ):
                        update_fields["collected_product_id"] = order_data[
                            "collected_product_id"
                        ]
                    if order_data.get("customer_note") and order_data[
                        "customer_note"
                    ] != str(existing.customer_note or ""):
                        update_fields["customer_note"] = order_data["customer_note"]
                    # 반품/교환 클레임 주문번호 — 원주문에 반품 새 번호 보관(GS 등)
                    if order_data.get("claim_order_number") and order_data[
                        "claim_order_number"
                    ] != str(existing.claim_order_number or ""):
                        update_fields["claim_order_number"] = order_data[
                            "claim_order_number"
                        ]
                    # SSG 취소신청 동기화는 shppNo 없는 "|seq" 형식 shipment_id를 만든다.
                    # 같은 주문이 출고대기(shppNo 있음)와 취소신청에 동시 존재하면 정상
                    # "shppNo|seq"를 "|seq"가 덮어써 송장 전송이 shppNo 누락으로 실패한다.
                    # 기존에 유효 shppNo가 있으면 빈-shppNo 값으로 덮어쓰지 않도록 가드.
                    # (타 마켓 shipment_id는 "|"로 시작하지 않아 무영향)
                    _new_sid = str(order_data.get("shipment_id") or "")
                    _old_sid = str(existing.shipment_id or "")
                    if (
                        _new_sid
                        and _new_sid != _old_sid
                        and not (_new_sid.startswith("|") and _old_sid.split("|")[0])
                    ):
                        update_fields["shipment_id"] = _new_sid
                    if order_data.get("ord_prd_seq") and not existing.ord_prd_seq:
                        update_fields["ord_prd_seq"] = order_data["ord_prd_seq"]
                    # 쿠팡 vendor_item_id 백필 — 컬럼 추가(2026-05-26) 이전 수집된 기존 주문은
                    # NULL 이므로 재수집 시 보충해줘야 송장업로드 가능
                    if (
                        order_data.get("source") == "coupang"
                        and order_data.get("vendor_item_id")
                        and not (existing.vendor_item_id or "")
                    ):
                        update_fields["vendor_item_id"] = order_data["vendor_item_id"]
                    # 결제일 갱신: 기존이 NULL이거나 더 이른 값일 때만 채택
                    # (고객 결제시각은 변하지 않음 — 더 늦은 값은 마켓이 sync/처리시각을 결제칸으로 돌려준 케이스로 간주하고 무시)
                    # tz-aware/naive 혼재 방지: 비교 직전 양쪽을 UTC tz-aware로 normalize
                    new_paid = order_data.get("paid_at")
                    if new_paid:
                        if existing.paid_at is None:
                            update_fields["paid_at"] = new_paid
                        else:
                            from datetime import timezone as _tz

                            _np = (
                                new_paid.replace(tzinfo=_tz.utc)
                                if new_paid.tzinfo is None
                                else new_paid
                            )
                            _ep = (
                                existing.paid_at.replace(tzinfo=_tz.utc)
                                if existing.paid_at.tzinfo is None
                                else existing.paid_at
                            )
                            if _np < _ep:
                                update_fields["paid_at"] = new_paid
                    # 수령인 정보 갱신 — 선물하기 주문 등에서 보내는 사람으로 잘못 저장된
                    # customer_name/phone을 다시 가져오기로 수령인 기준으로 교정.
                    # 마켓 응답에 값이 있고 기존과 다르면 덮어쓴다.
                    new_cust_name = order_data.get("customer_name")
                    if new_cust_name and new_cust_name != str(
                        existing.customer_name or ""
                    ):
                        update_fields["customer_name"] = new_cust_name
                    new_orderer_name = order_data.get("orderer_name")
                    if new_orderer_name and new_orderer_name != str(
                        existing.orderer_name or ""
                    ):
                        update_fields["orderer_name"] = new_orderer_name
                    new_cust_phone = order_data.get("customer_phone")
                    # #536 — 기존이 실번호인데 새 값이 안심번호(050x)면 덮지 않음(실번호 보존).
                    if (
                        new_cust_phone
                        and new_cust_phone != str(existing.customer_phone or "")
                        and not (
                            _is_safe_phone(new_cust_phone)
                            and existing.customer_phone
                            and not _is_safe_phone(existing.customer_phone)
                        )
                    ):
                        update_fields["customer_phone"] = new_cust_phone
                    new_cust_addr = order_data.get("customer_address")
                    if new_cust_addr and new_cust_addr != str(
                        existing.customer_address or ""
                    ):
                        update_fields["customer_address"] = new_cust_addr
                    new_cust_addr_dtl = order_data.get("customer_address_detail")
                    if new_cust_addr_dtl is not None and new_cust_addr_dtl != str(
                        existing.customer_address_detail or ""
                    ):
                        update_fields["customer_address_detail"] = new_cust_addr_dtl
                    # 우편번호 — UPDATE path 에서도 채움 (신규 INSERT 만 채워지던 버그 fix)
                    new_postal = order_data.get("customer_postal_code")
                    if new_postal and new_postal != (
                        existing.customer_postal_code or ""
                    ):
                        update_fields["customer_postal_code"] = new_postal
                    # 마켓 상품번호 보충 (기존 주문에 없으면 채움)
                    if order_data.get("product_id") and not existing.product_id:
                        update_fields["product_id"] = order_data["product_id"]
                    # quantity 자기치유 (issue #213 롯데ON → 전 소싱처 확대):
                    # 재동기화 수량 > 1 이고 기존이 known-bad(=1) 일 때만 교정.
                    # 쿠팡 orderQuantity→shippingCount 키 교정(4a7ccda2) 이전에 들어와
                    # quantity=1 로 박힌 멀티수량 주문을 재동기화로 자동 복구하기 위함.
                    # 조건이 보수적(>1 & 기존=1)이라 정상 단품 주문은 영향 없음.
                    try:
                        _new_qty = int(order_data.get("quantity") or 0)
                    except (TypeError, ValueError):
                        _new_qty = 0
                    if _new_qty > 1 and (existing.quantity or 1) == 1:
                        update_fields["quantity"] = _new_qty
                    # 송장전송완료/배송중 이상 상태는 덮어쓰지 않음
                    # 단, 롯데ON은 발송완료/배송중/배송완료로 진행된 경우 갱신 허용
                    new_ship_status = order_data.get("shipping_status")
                    # Recovery — 마켓이 '배송완료'/'구매확정' 같은 종결 상태를 보낸 경우
                    # 좀비 '취소요청' 잔존을 자동 해제 (PlayAuto 5/19 수취확인됐는데
                    # DB는 13일째 취소요청으로 박혀있던 사고 방지).
                    # 마켓의 종결 신호가 진실의 원천 — 취소가 실제로 진행됐다면 마켓이
                    # '취소완료'를 보냈을 것.
                    if (
                        new_ship_status in ("배송완료", "구매확정")
                        and existing.shipping_status == "취소요청"
                    ):
                        update_fields["shipping_status"] = new_ship_status
                        update_fields["status"] = "delivered"
                        if existing.cancel_requested_at is not None:
                            update_fields["cancel_requested_at"] = None
                        logger.info(
                            f"[주문동기화] 취소요청 좀비 해제: "
                            f"{order_data.get('order_number')} "
                            f"취소요청 → {new_ship_status} (마켓 종결 신호)"
                        )
                        new_ship_status = None  # 아래 분기 스킵
                    if new_ship_status:
                        cancel_statuses = {"취소요청", "취소처리중", "취소완료"}
                        exchange_statuses = {
                            "교환요청",
                            "교환회수완료",
                            "교환재배송",
                            "교환완료",
                        }
                        advanced = {"발송완료", "국내배송중", "배송완료", "구매확정"}
                        if new_ship_status in cancel_statuses:
                            # 취소 상태 갱신 규칙:
                            #  - 이미 반품 진행 중인 주문은 취소로 되돌리지 않음
                            #  - 새로 들어오는 값이 '취소요청'인데 마켓이 송장출력 이상으로
                            #    진행한 주문(송장전송완료/국내배송중/배송완료/구매확정)은
                            #    덮어쓰지 않음 — 스마트스토어/쿠팡/롯데ON/11번가/eBay 공통,
                            #    좀비 claim 으로 배송 진행 주문이 '취소요청'으로 표시되던
                            #    사고 방지 (참조: 419d42d4 플레이오토 동일 버그)
                            #  - 단, 마켓이 '취소처리중'/'취소완료'를 보낸 경우는 실제 종결
                            #    상태이므로 그대로 반영
                            if existing.shipping_status in (
                                "반품요청",
                                "반품완료",
                                "반품거부",
                            ):
                                logger.info(
                                    f"[주문동기화] 반품 상태 보호: {order_data.get('order_number')} "
                                    f"{existing.shipping_status} → {new_ship_status} 차단"
                                )
                            elif (
                                new_ship_status == "취소요청"
                                and existing.shipping_status
                                in (
                                    "송장전송완료",
                                    "국내배송중",
                                    "배송완료",
                                    "구매확정",
                                )
                            ):
                                logger.info(
                                    f"[주문동기화] 배송 진행 상태 보호: {order_data.get('order_number')} "
                                    f"{existing.shipping_status} → 취소요청 차단"
                                )
                            else:
                                update_fields["shipping_status"] = new_ship_status
                        elif new_ship_status in exchange_statuses:
                            # 교환 상태는 항상 갱신 (배송완료 → 교환요청 등 역행 허용)
                            # 단, 이미 반품/취소 상태인 주문은 교환으로 되돌리지 않음
                            # 취소 상태 보호 — samba_return 활성 stale 레코드(type=exchange)로
                            # 인해 status=cancelled 주문이 매 sync마다 '교환요청'으로 덮어쓰여
                            # inconsistent state 되는 사고 방지 (issue #224, 롯데ON 6건 사례)
                            if existing.shipping_status in (
                                "반품요청",
                                "반품완료",
                                "반품거부",
                                "취소요청",
                                "취소처리중",
                                "취소완료",
                            ):
                                logger.info(
                                    f"[주문동기화] 반품/취소 상태 보호: {order_data.get('order_number')} "
                                    f"{existing.shipping_status} → {new_ship_status} 차단"
                                )
                            else:
                                update_fields["shipping_status"] = new_ship_status
                        elif (
                            existing.shipping_status == "송장전송완료"
                            and new_ship_status in advanced
                        ):
                            update_fields["shipping_status"] = new_ship_status
                        elif (
                            existing.shipping_status == "국내배송중"
                            and new_ship_status in ("배송완료", "구매확정")
                        ):
                            # #524 — 국내배송중→배송완료/구매확정 종결 전이
                            # 쿠팡 DEPARTURE/DELIVERING 둘 다 국내배송중으로 매핑되어
                            # FINAL_DELIVERY 도달 시 종결이 영구 차단되던 문제 해소.
                            update_fields["shipping_status"] = new_ship_status
                        elif new_ship_status in (
                            "반품요청",
                            "반품완료",
                            "반품거부",
                        ) and existing.shipping_status in (
                            "취소요청",
                            "취소처리중",
                            "취소완료",
                        ):
                            # 취소 종결/진행 상태는 마켓 진실의 원천 — 반품으로 덮지 않음
                            # samba_return 활성 stale 레코드(type=return)로 인한
                            # 매 sync 덮어쓰기 차단 (issue #224)
                            logger.info(
                                f"[주문동기화] 취소 상태 보호: {order_data.get('order_number')} "
                                f"{existing.shipping_status} → {new_ship_status} 차단"
                            )
                        elif new_ship_status in (
                            "반품요청",
                            "반품완료",
                            "반품거부",
                        ) and existing.shipping_status in (
                            "송장전송완료",
                            "국내배송중",
                            "배송완료",
                            "구매확정",
                        ):
                            # [#599] 배송완료 종결 주문에 반품 접수 — 반품이 최신 진실.
                            #   배송 후 반품(쿠팡 releaseStatus=Y/A)은 정상 흐름인데, 기존엔
                            #   배송완료→반품 허용 분기가 없어 반품 신호가 무시돼 '배송완료'로
                            #   고착됐다(증상2). 종결 상태를 반품으로 갱신 허용.
                            #   (취소요청→배송완료 차단은 9657 분기가, 취소 종결→반품 차단은
                            #    9706 분기가 각각 그대로 담당 — 여기는 배송 진행/완료→반품만)
                            update_fields["shipping_status"] = new_ship_status
                            logger.info(
                                f"[주문동기화] 배송완료→반품 전이: {order_data.get('order_number')} "
                                f"{existing.shipping_status} → {new_ship_status} (반품 접수)"
                            )
                        elif (
                            new_ship_status in ("반품요청", "반품완료", "반품거부")
                            and existing.shipping_status in exchange_statuses
                        ):
                            # 반품 상태는 교환 상태를 덮어씀 (교환→반품 재접수 케이스)
                            update_fields["shipping_status"] = new_ship_status
                            logger.info(
                                f"[주문동기화] 교환→반품 상태 전환: {order_data.get('order_number')} "
                                f"{existing.shipping_status} → {new_ship_status}"
                            )
                        elif new_ship_status in (
                            "반품요청",
                            "반품완료",
                            "반품거부",
                        ) and existing.shipping_status in (
                            "국내배송중",
                            "배송완료",
                            "구매확정",
                            "송장전송완료",
                        ):
                            # 배송 진행 후 반품 접수 허용 (국내배송중/배송완료 → 반품요청)
                            update_fields["shipping_status"] = new_ship_status
                            logger.info(
                                f"[주문동기화] 배송→반품 상태 전환: {order_data.get('order_number')} "
                                f"{existing.shipping_status} → {new_ship_status}"
                            )
                        elif existing.shipping_status not in (
                            "송장전송완료",
                            "국내배송중",
                            "배송완료",
                            "교환재배송",
                            "교환요청",
                            "교환회수완료",
                            "교환완료",
                            "교환거부",
                            "반품요청",
                            "반품완료",
                            "반품거부",
                            "회수확정",
                            "취소요청",
                            "취소처리중",
                            "취소완료",
                        ):
                            update_fields["shipping_status"] = new_ship_status
                    # shipping_status 가 "국내배송중"으로 진입 시 status 드롭다운도 함께 동기화.
                    # 라벨/드롭다운이 어긋난 채 wait_ship 으로 남아 페이지 필터를 통과해 노출되던 사고 방지.
                    _new_ss_final = update_fields.get(
                        "shipping_status", existing.shipping_status
                    )
                    if _new_ss_final == "국내배송중" and existing.status in (
                        "pending",
                        "preparing",
                        "wait_ship",
                        "arrived",
                        "processing",
                        "shipped",
                        "ship_failed",
                    ):
                        update_fields["status"] = "shipping"
                    # 마켓이 발송완료/배송완료/구매확정 신호를 주면 ship_failed 잔존 해제.
                    # 우리 송장전송이 false-negative 로 실패했지만 마켓 측엔 실제 송장이 들어간 케이스 보정.
                    elif (
                        _new_ss_final in ("송장전송완료",)
                        and existing.status == "ship_failed"
                    ):
                        update_fields["status"] = "shipping"
                    elif _new_ss_final in (
                        "배송완료",
                        "구매확정",
                    ) and existing.status in ("ship_failed", "wait_ship", "shipping"):
                        update_fields["status"] = "delivered"
                    # issue #393 — 반품 케이스 부재로 배송상태만 '반품요청'되고 주문상태는
                    # '배송완료'에 잔존하던 버그. 터미널 상태(반품완료/취소)는 회귀 금지.
                    elif _new_ss_final == "반품요청" and existing.status not in (
                        "return_requested",
                        "returned",
                        "return_completed",
                        "cancelled",
                    ):
                        update_fields["status"] = "return_requested"
                    elif _new_ss_final in (
                        "회수확정",
                        "반품완료",
                    ) and existing.status not in (
                        "returned",
                        "return_completed",
                        "cancelled",
                    ):
                        update_fields["status"] = "return_completed"
                    elif _new_ss_final == "취소완료" and existing.status != "cancelled":
                        update_fields["status"] = "cancelled"
                    elif (
                        _new_ss_final == "취소요청"
                        and existing.status != "cancel_requested"
                    ):
                        update_fields["status"] = "cancel_requested"
                        # 자동 발주취소 트리거는 SambaOrder after_flush event listener 가 단일 진입점.
                        # 여기 별도 호출 추가 시 중복 잡 발행(dedup race) 발생 — 절대 금지.
                    # 플레이오토 미등록 주문의 취소요청/취소완료는 status 드롭다운도 동기화.
                    # status는 사용자가 직접 관리 — shipping_status 따라 자동변경 금지
                    # 정산금액(revenue) / 수수료율 갱신
                    new_revenue = order_data.get("revenue")
                    new_fee_rate = order_data.get("fee_rate")
                    sp = float(
                        update_fields.get("sale_price", existing.sale_price) or 0
                    )
                    if new_revenue and float(new_revenue) != float(
                        existing.revenue or 0
                    ):
                        rev = float(new_revenue)
                        update_fields["revenue"] = rev
                        update_fields["fee_rate"] = (
                            new_fee_rate
                            if new_fee_rate is not None
                            else (existing.fee_rate or 0)
                        )
                        cost = float(existing.cost or 0)
                        ship_fee = float(existing.shipping_fee or 0)
                        update_fields["profit"] = rev - cost - ship_fee
                        update_fields["profit_rate"] = (
                            f"{((rev - cost - ship_fee) / rev * 100):.2f}"
                            if rev > 0
                            else "0.00"
                        )
                    elif "sale_price" in update_fields:
                        fr = float(
                            new_fee_rate
                            if new_fee_rate is not None
                            else (existing.fee_rate or 0)
                        )
                        rev = sp * (1 - fr / 100)
                        cost = float(existing.cost or 0)
                        ship_fee = float(existing.shipping_fee or 0)
                        update_fields["revenue"] = rev
                        update_fields["profit"] = rev - cost - ship_fee
                        update_fields["profit_rate"] = (
                            f"{((rev - cost - ship_fee) / rev * 100):.2f}"
                            if rev > 0
                            else "0.00"
                        )
                    # 취소·반품 클레임 필드 백필 — 기존주문이 나중에 취소요청/반품요청으로
                    # 전환될 때 receiptId·release_status·사유가 update_fields에서 누락돼
                    # 영영 저장 안 되던 버그 수정 (쿠팡 취소승인 시 receiptId 미수집 차단).
                    # parse가 채워준 값만, 기존값과 다를 때만 반영 (NULL 덮어쓰기 금지).
                    for _cf in (
                        "cancel_receipt_id",
                        "cancel_release_status",
                        "cancel_release_stop_status",
                        "cancel_reason_code",
                        "cancel_reason_text",
                        "cancel_reason_category1",
                        "cancel_reason_category2",
                        "cancel_fault_by",
                        "cancel_requested_at",
                    ):
                        _cv = order_data.get(_cf)
                        if _cv is not None and _cv != getattr(existing, _cf, None):
                            update_fields[_cf] = _cv
                    if update_fields:
                        await svc.update_order(existing.id, update_fields, commit=False)
                        _pending += 1
                        if _pending >= _PERSIST_CHUNK:
                            await session.commit()
                            _pending = 0
                    continue
                # savepoint로 감싸 중복주문(uq_order_tenant_number_seq) 1건이
                # 청크 전체를 롤백시켜 계정 주문 전부 유실되던 버그 방지.
                # asyncpg는 tx 중 IntegrityError 발생 시 이후 쿼리가 전부 abort되므로
                # begin_nested(SAVEPOINT)로 그 1건만 격리 롤백하고 나머지는 저장.
                from sqlalchemy.exc import IntegrityError as _IntegrityError  # noqa: F811

                try:
                    async with session.begin_nested():
                        await svc.create_order(order_data, commit=False)
                    synced += 1
                    _pending += 1
                    if _pending >= _PERSIST_CHUNK:
                        await session.commit()
                        _pending = 0
                except _IntegrityError:
                    logger.warning(
                        f"[주문동기화] {label}: 중복주문 스킵 "
                        f"order_number={order_data.get('order_number')}"
                    )

            # 루프 끝 잔여 청크 일괄 commit (issue #401) — continue 분기와 무관하게 항상 실행
            if _pending:
                await session.commit()
                _pending = 0

            # 롯데홈쇼핑: deliver_list가 교체한 index-format 레코드(K72118:0 등) DB에서 삭제
            if _lh_replaced_old_keys:
                try:
                    _orp_ph = ", ".join(
                        f":orp_{i}" for i in range(len(_lh_replaced_old_keys))
                    )
                    _orp_prm: dict = {
                        f"orp_{i}": v for i, v in enumerate(_lh_replaced_old_keys)
                    }
                    _orp_prm["cid"] = account["id"]
                    await session.execute(
                        _sa_text(
                            f"DELETE FROM samba_order "
                            f"WHERE source = 'lottehome' "
                            f"AND channel_id = :cid "
                            f"AND order_number IN ({_orp_ph})"
                        ),
                        _orp_prm,
                    )
                    await session.commit()
                except Exception as _orp_e:
                    await session.rollback()
                    logger.warning(
                        f"[주문동기화] 롯데홈쇼핑 교체 레코드 삭제 실패(무시): {_orp_e}"
                    )

            total_synced += synced
            if market_type == "smartstore":
                confirmed_count = len(unconfirmed_ids)
            elif market_type == "lotteon":
                confirmed_count = lotteon_confirmed_count
            elif market_type == "11st":
                confirmed_count = _confirmed if _confirm_targets else 0
            else:
                confirmed_count = 0

            # ── 클레임(취소/반품/교환) → SambaReturn 자동 생성 ──────────────
            returns_synced = 0
            claim_statuses = {
                "취소요청",
                "취소처리중",
                "취소완료",
                "반품요청",
                "반품완료",
                "반품거부",
                "교환요청",
                "교환회수완료",
                "교환재배송",
                "교환완료",
            }
            claim_orders = [
                od for od in orders_data if od.get("shipping_status") in claim_statuses
            ]
            if claim_orders:
                from backend.domain.samba.returns.service import SambaReturnService
                from backend.domain.samba.returns.repository import (
                    SambaReturnRepository,
                )
                from backend.domain.samba.returns.model import SambaReturn
                from sqlmodel import select as _sel

                return_svc = SambaReturnService(SambaReturnRepository(session))

                claim_type_map = {
                    "취소요청": "cancel",
                    "취소처리중": "cancel",
                    "취소완료": "cancel",
                    "반품요청": "return",
                    "반품완료": "return",
                    "반품거부": "return",
                    "교환요청": "exchange",
                    "교환회수완료": "exchange",
                    "교환재배송": "exchange",
                    "교환완료": "exchange",
                }
                claim_return_status_map = {
                    "취소완료": "completed",
                    "반품완료": "completed",
                    "교환완료": "completed",
                    "반품거부": "rejected",
                }
                claim_completion_detail_map = {
                    "취소완료": "취소",
                    "반품완료": "반품",
                    "교환완료": "교환",
                    "반품거부": "거부",
                }
                for od in claim_orders:
                    order_no = od.get("order_number", "")
                    if not order_no:
                        continue
                    shipping_status = od.get("shipping_status", "")
                    ret_type = claim_type_map.get(shipping_status, "return")
                    return_status = claim_return_status_map.get(shipping_status)
                    completion_detail = claim_completion_detail_map.get(shipping_status)
                    # 중복 체크
                    existing_ret_result = await session.execute(
                        _sel(SambaReturn).where(SambaReturn.order_number == order_no)
                    )
                    existing_ret = existing_ret_result.scalars().first()
                    if existing_ret:
                        update_fields: dict[str, Any] = {
                            "type": ret_type,
                            "market_order_status": shipping_status,
                        }
                        if return_status:
                            update_fields["status"] = return_status
                        if completion_detail:
                            update_fields["completion_detail"] = completion_detail
                        if return_status in ("completed", "rejected"):
                            from datetime import UTC, datetime as _dt

                            update_fields["completion_date"] = _dt.now(UTC)
                        await return_svc.repo.update_async(
                            existing_ret.id, **update_fields
                        )
                        continue
                    # 연결 주문 조회
                    linked_order = await svc.repo.find_by_async(order_number=order_no)
                    if not linked_order:
                        logger.warning(
                            f"[주문동기화] 클레임 연결 주문 없음: order_number={order_no!r} "
                            f"shipping_status={shipping_status}"
                        )
                        continue
                    ret = await return_svc.create_return(
                        {
                            "order_id": linked_order.id,
                            "order_number": order_no,
                            "type": ret_type,
                            "market": label,
                            "market_order_status": shipping_status,
                            "product_name": od.get("product_name", ""),
                            "product_image": od.get("product_image", ""),
                            "customer_name": od.get("customer_name", ""),
                            "customer_phone": od.get("customer_phone", ""),
                            "customer_address": od.get("customer_address", ""),
                            "requested_amount": od.get("sale_price", 0),
                        }
                    )
                    if return_status or completion_detail:
                        update_fields: dict[str, Any] = {}
                        if return_status:
                            update_fields["status"] = return_status
                        if completion_detail:
                            update_fields["completion_detail"] = completion_detail
                        if return_status in ("completed", "rejected"):
                            from datetime import UTC, datetime as _dt

                            update_fields["completion_date"] = _dt.now(UTC)
                        await return_svc.repo.update_async(ret.id, **update_fields)
                    returns_synced += 1
                logger.info(
                    f"[주문동기화] {label}: 클레임 {len(claim_orders)}건 중 {returns_synced}건 반품교환 생성"
                )

            cancel_requested = sum(
                1 for od in orders_data if od.get("shipping_status") == "취소요청"
            )
            results.append(
                {
                    "account": label,
                    "status": "success",
                    "fetched": len(orders_data),
                    "synced": synced,
                    "confirmed": confirmed_count,
                    "cancel_requested": cancel_requested,
                    "returns_synced": returns_synced,
                }
            )
            logger.info(
                f"[주문동기화] {label}: {len(orders_data)}건 조회, {synced}건 저장, {confirmed_count}건 발주확인"
            )

            # ── paid_at 백필 — 스마트스토어 NULL paid_at 주문 직접 재조회 ──
            if market_type == "smartstore":
                try:
                    _null_rows = await session.execute(
                        _sa_text(
                            "SELECT order_number FROM samba_order "
                            "WHERE paid_at IS NULL AND source = 'smartstore' "
                            "AND channel_id = :cid LIMIT 100"
                        ),
                        {"cid": account["id"]},
                    )
                    _null_po_ids = [r[0] for r in _null_rows.fetchall()]
                    if _null_po_ids:
                        _details = await client.get_product_orders_by_ids(_null_po_ids)
                        _backfilled = 0
                        for _d in _details:
                            _po = _d.get("productOrder", _d)
                            _oi = _d.get("order", {})
                            _paid = _parse_iso_datetime(
                                _oi.get("paymentDate") or _po.get("paymentDate")
                            )
                            if _paid:
                                _poid = _po.get("productOrderId", "")
                                await session.execute(
                                    _sa_text(
                                        "UPDATE samba_order SET paid_at = :paid "
                                        "WHERE order_number = :on AND paid_at IS NULL"
                                    ),
                                    {"paid": _paid, "on": _poid},
                                )
                                _backfilled += 1
                        if _backfilled:
                            await session.commit()
                            logger.info(
                                f"[주문동기화] {label}: paid_at 백필 {_backfilled}건"
                            )
                except Exception as _bf_err:
                    logger.warning(
                        f"[주문동기화] {label}: paid_at 백필 실패 — {_bf_err}"
                    )

            # ── paid_at 백필 — 플레이오토 NULL paid_at 주문 → 동기화 데이터에서 매칭 ──
            elif market_type == "playauto":
                try:
                    # 현재 동기화에서 paid_at이 유효한 주문의 order_number → paid_at 매핑
                    _pa_paid_map: dict[str, datetime] = {}
                    for od in orders_data:
                        if od.get("paid_at") and od.get("order_number"):
                            _pa_paid_map[od["order_number"]] = od["paid_at"]
                    if _pa_paid_map:
                        _null_rows = await session.execute(
                            _sa_text(
                                "SELECT order_number FROM samba_order "
                                "WHERE paid_at IS NULL AND source = 'playauto' "
                                "AND channel_id = :cid LIMIT 200"
                            ),
                            {"cid": account["id"]},
                        )
                        _null_ons = [r[0] for r in _null_rows.fetchall()]
                        _backfilled = 0
                        for _on in _null_ons:
                            _paid = _pa_paid_map.get(_on)
                            if _paid:
                                await session.execute(
                                    _sa_text(
                                        "UPDATE samba_order SET paid_at = :paid "
                                        "WHERE order_number = :on AND paid_at IS NULL"
                                    ),
                                    {"paid": _paid, "on": _on},
                                )
                                _backfilled += 1
                        if _backfilled:
                            await session.commit()
                            logger.info(
                                f"[주문동기화] {label}: 플레이오토 paid_at 백필 {_backfilled}건"
                            )
                except Exception as _bf_err:
                    logger.warning(
                        f"[주문동기화] {label}: 플레이오토 paid_at 백필 실패 — {_bf_err}"
                    )

            # ── paid_at 백필 — 롯데ON NULL paid_at 주문 → 동기화 데이터에서 매칭 ──
            # order_number = "{od_no}_{od_seq}_{proc_seq}" 합성키 기반 (order.py:3406)
            elif market_type == "lotteon":
                try:
                    _lo_paid_map: dict[str, datetime] = {}
                    for od in orders_data:
                        if od.get("paid_at") and od.get("order_number"):
                            _lo_paid_map[od["order_number"]] = od["paid_at"]
                    if _lo_paid_map:
                        _null_rows = await session.execute(
                            _sa_text(
                                "SELECT order_number FROM samba_order "
                                "WHERE paid_at IS NULL AND source = 'lotteon' "
                                "AND channel_id = :cid LIMIT 200"
                            ),
                            {"cid": account["id"]},
                        )
                        _null_ons = [r[0] for r in _null_rows.fetchall()]
                        _backfilled = 0
                        for _on in _null_ons:
                            _paid = _lo_paid_map.get(_on)
                            if _paid:
                                await session.execute(
                                    _sa_text(
                                        "UPDATE samba_order SET paid_at = :paid "
                                        "WHERE order_number = :on AND paid_at IS NULL"
                                    ),
                                    {"paid": _paid, "on": _on},
                                )
                                _backfilled += 1
                        if _backfilled:
                            await session.commit()
                            logger.info(
                                f"[주문동기화] {label}: 롯데ON paid_at 백필 {_backfilled}건"
                            )
                except Exception as _bf_err:
                    logger.warning(
                        f"[주문동기화] {label}: 롯데ON paid_at 백필 실패 — {_bf_err}"
                    )

        except Exception as e:
            await session.rollback()  # 세션 복구 — 다음 계정 연쇄 실패 방지
            logger.error(f"[주문동기화] {label} 실패: {e}")
            results.append({"account": label, "status": "error", "message": str(e)})
        finally:
            # 마켓 클라이언트 httpx keepalive 좀비 정리 — 다음 계정 hang 도미노 차단.
            # CancelledError(상위 wait_for timeout) 시에도 이 finally 가 먼저 실행되므로
            # connection pool 즉시 회수됨.
            for _c in _clients_to_close:
                try:
                    _aclose = getattr(_c, "aclose", None)
                    if _aclose is not None:
                        await _aclose()
                except Exception as _ce:
                    logger.warning(
                        f"[주문동기화] {label} 클라이언트 aclose 실패(무시): {_ce}"
                    )

    # DB 기반 원주문 shipping_status 일괄 동기화
    # samba_return 레코드가 있고 진행 중인 주문의 shipping_status를 강제 업데이트
    try:
        from sqlalchemy import text as _sa_text_upd

        await session.execute(
            _sa_text_upd(
                """
            UPDATE samba_order o
            SET shipping_status = CASE
                WHEN r.type = 'exchange' THEN '교환요청'
                WHEN r.type = 'return' THEN '반품요청'
                WHEN r.type = 'cancel' THEN '취소요청'
                ELSE o.shipping_status
            END
            FROM samba_return r
            WHERE r.order_id = o.id
              AND r.status NOT IN ('completed', 'cancelled', 'rejected')
              AND o.shipping_status NOT IN (
                  '교환요청', '교환회수완료', '교환재배송', '교환완료',
                  '반품요청', '반품완료', '반품거부',
                  -- 취소 라벨은 마켓 종결/진행 신호. samba_return type=return/exchange
                  -- 활성 stale 레코드가 남아있어도 마켓 취소 상태를 반품/교환요청으로
                  -- 덮지 않음 (issue #224, status=cancelled + ship='교환요청' 사고)
                  '취소요청', '취소처리중', '취소완료',
                  -- 마켓이 송장/배송 단계로 진행한 주문은 좀비 cancel return으로
                  -- 되돌리지 않음 (송장출력→배송대기중 단계에선 마켓이 이미 셀러
                  -- 수락 후 처리 진행 중이라 취소요청 표시 부적절)
                  '주문접수', '배송대기중', '송장전송완료', '국내배송중',
                  '배송완료', '구매확정'
              )
        """
            )
        )
        await session.commit()
        logger.info(
            "[주문동기화] 반품/교환/취소 진행 중 원주문 shipping_status 일괄 업데이트 완료"
        )
    except Exception as _upd_err:
        logger.warning(f"[주문동기화] 원주문 일괄 업데이트 실패: {_upd_err}")

    # PlayAuto/롯데홈쇼핑 미매칭 주문 자동 백필 — 동기화 후 collected_product_id IS NULL 잔존 해소.
    # 현대H몰 등 PlayAuto 경유 마켓 주문은 style_code 매칭이 실패해도 DB에 저장은 됨.
    # 롯데홈쇼핑도 인입 당시 수집상품이 없거나 다중후보면 NULL로 남는데, 과거엔
    # playauto 전용이라 재시도 루프가 없어 누적됐다(2026-06-29 1,957건 적체 확인).
    # 매 sync 완료 시 재시도해 누적 미매칭 해소.
    try:
        from sqlalchemy import text as _pa_bf_text

        _pa_null = (
            await session.execute(
                _pa_bf_text(
                    "SELECT id, product_name FROM samba_order "
                    "WHERE source IN ('playauto', 'lottehome') "
                    "AND collected_product_id IS NULL "
                    "AND product_name IS NOT NULL AND product_name != '' "
                    "LIMIT 500"
                )
            )
        ).fetchall()
        if _pa_null:
            _pa_all_tokens: set[str] = set()
            _pa_order_tokens: list[tuple[str, list[str]]] = []
            for _poid, _ppname in _pa_null:
                _ptoks = _lh_style_tokens(str(_ppname or ""))
                _pa_order_tokens.append((str(_poid), _ptoks))
                _pa_all_tokens.update(_ptoks)
            if _pa_all_tokens:
                _pa_cp_rows = (
                    await session.execute(
                        _pa_bf_text(
                            "SELECT id, style_code FROM samba_collected_product "
                            "WHERE style_code = ANY(:t)"
                        ),
                        {"t": list(_pa_all_tokens)},
                    )
                ).fetchall()
                _pa_tok_cp: dict[str, list[str]] = {}
                for _pcr in _pa_cp_rows:
                    _psc = str(_pcr[1] or "")
                    if _psc:
                        _pa_tok_cp.setdefault(_psc, []).append(str(_pcr[0]))
                _pa_linked = 0
                for _poid, _ptoks in _pa_order_tokens:
                    if not _ptoks:
                        continue
                    _pcpid: str | None = None
                    for _ptok in sorted(_ptoks, key=len, reverse=True):
                        _pcands = _pa_tok_cp.get(_ptok, [])
                        if len(_pcands) == 1:
                            _pcpid = _pcands[0]
                            break
                        elif _pcands:
                            break  # ambiguous — skip
                    if _pcpid:
                        await session.execute(
                            _pa_bf_text(
                                "UPDATE samba_order SET collected_product_id = :cpid "
                                "WHERE id = :oid AND collected_product_id IS NULL"
                            ),
                            {"cpid": _pcpid, "oid": _poid},
                        )
                        _pa_linked += 1
                if _pa_linked:
                    await session.commit()
                    logger.info(
                        f"[주문동기화] PlayAuto/롯데홈 미매칭 자동 백필 {_pa_linked}건 완료"
                    )
    except Exception as _pa_bf_err:
        logger.warning(f"[주문동기화] PlayAuto/롯데홈 백필 실패(무시): {_pa_bf_err}")

    if total_synced > 0:
        from backend.utils.kakao_notify import send_kakao_message

        synced_lines = [
            f"  {r['account']}: {r.get('synced', 0)}건"
            for r in results
            if r.get("synced", 0) > 0
        ]
        msg = f"🛒 주문 {total_synced}건 동기화 완료"
        if synced_lines:
            msg += "\n" + "\n".join(synced_lines)
        asyncio.create_task(send_kakao_message(msg))

    return {"total_synced": total_synced, "results": results}


def _parse_iso_datetime(val: str | None) -> datetime | None:
    """ISO 8601 문자열 → datetime 변환. 실패 시 None."""
    if not val:
        return None
    try:
        return datetime.fromisoformat(val.replace("Z", "+00:00"))
    except (ValueError, AttributeError):
        return None


def _parse_smartstore_order(
    po: dict,
    order_info: dict,
    account_id: str,
    account_label: str,
    claim_info: dict | None = None,
) -> dict[str, Any]:
    """스마트스토어 productOrder + order → SambaOrder 데이터 변환."""
    status_map = {
        "PAYED": "pending",
        "DELIVERING": "shipped",
        "DELIVERED": "delivered",
        "PURCHASE_DECIDED": "delivered",
        "EXCHANGED": "delivered",
        "CANCELED": "cancelled",
        "RETURNED": "returned",
        "CANCEL_REQUESTED": "pending",
    }
    naver_status = po.get("productOrderStatus", "")
    place_status = po.get("placeOrderStatus", "")
    sale_price = po.get("totalPaymentAmount", 0) or po.get("unitPrice", 0) or 0
    quantity = po.get("quantity", 1) or 1

    # 클레임 상태 (취소/반품/교환 요청)
    # 우선순위: 호출자가 전달한 claim 서브 객체 → productOrder 최상위 순으로 fallback
    _ci = claim_info or {}
    claim_status = _ci.get("claimStatus") or po.get("claimStatus", "") or ""

    claim_status_map = {
        "CANCEL_REQUEST": "취소요청",
        "CANCELING": "취소처리중",
        "CANCEL_DONE": "취소완료",
        "CANCEL_REJECT": "취소거부",
        "RETURN_REQUEST": "반품요청",
        "COLLECTING": "수거중",
        "COLLECT_DONE": "수거완료",
        "RETURN_DONE": "반품완료",
        "RETURN_REJECT": "반품거부",
        "EXCHANGE_REQUEST": "교환요청",
        "EXCHANGING": "교환처리중",
        "EXCHANGE_DONE": "교환완료",
        "EXCHANGE_REJECT": "교환거부",
    }

    # 정산금액: API에서 직접 가져오기
    expected_settlement = po.get("expectedSettlementAmount")
    if expected_settlement and sale_price > 0:
        fee_rate = round((1 - expected_settlement / sale_price) * 100, 2)
    else:
        expected_settlement = None
        fee_rate = 0

    # 마켓 주문상태 한글 변환
    market_status_map: dict[str, str] = {
        "PAYED": "결제완료",
        "DELIVERING": "국내배송중",
        "DELIVERED": "배송완료",
        "PURCHASE_DECIDED": "구매확정",
        "EXCHANGED": "교환완료",
        "CANCELED": "취소완료",
        "RETURNED": "반품완료",
        "CANCEL_REQUESTED": "취소요청",
        "RETURN_REQUESTED": "반품요청",
        "EXCHANGE_REQUESTED": "교환요청",
    }
    # 클레임이 있으면 클레임 상태 우선
    if claim_status and claim_status in claim_status_map:
        market_order_status = claim_status_map[claim_status]
    elif place_status == "NOT_YET" and naver_status == "PAYED":
        market_order_status = "발주미확인"
    elif naver_status == "PAYED":
        market_order_status = "발송대기"
    else:
        market_order_status = market_status_map.get(naver_status, naver_status)

    # 배송지 정보
    shipping = po.get("shippingAddress", {})
    # 우편번호 후보 키 모두 비어있으면 1회 INFO 로그 (실제 응답 키 진단용)
    if shipping and not (
        shipping.get("zipCode")
        or shipping.get("zipcode")
        or shipping.get("postCode")
        or shipping.get("zipNo")
    ):
        logger.info(
            f"[스마트스토어][zip진단] po={po.get('productOrderId')} "
            f"keys={list(shipping.keys())}"
        )
    # 수령인(배송지) 우선 — 선물하기 주문은 주문자(보내는 사람) ≠ 수령인(받는 사람)이므로
    # CS/배송 단위에서 의미있는 customer는 수령인. 일반 주문은 둘이 동일하므로 영향 없음.
    customer_name = shipping.get("name", "") or order_info.get("ordererName", "")
    customer_tel = (
        shipping.get("tel1", "")
        or shipping.get("tel2", "")
        or order_info.get("ordererTel", "")
    )

    # 마켓 상품번호 (구매페이지 URL 생성용 + 수집상품 매칭 키)
    # 우선순위: channelProductNo > originalProductId > productId
    # - 다른 정상 케이스는 channelProductNo가 있어 그대로 동작
    # - 선물하기/위탁판매 옵션 상품은 channelProductNo 누락 + productId가 옵션별로 별도 발급되어
    #   수집상품 매칭 실패 사고가 있었음(2026-05-12 이종영 주문). 등록은 originalProductId로
    #   되어있는 경우가 많아 fallback 키로 활용.
    channel_product_no = str(
        po.get("channelProductNo", "")
        or po.get("originalProductId", "")
        or po.get("productId", "")
        or ""
    )

    return {
        "order_number": po.get("productOrderId", ""),
        "shipment_id": order_info.get("orderId", ""),
        "channel_id": account_id,
        "channel_name": account_label,
        "product_id": channel_product_no,
        "product_name": po.get("productName", ""),
        "product_option": po.get("productOption", "") or "",
        "product_image": po.get("imageUrl", ""),
        "customer_name": customer_name,
        "orderer_name": order_info.get("ordererName", "") or "",
        "customer_phone": customer_tel,
        "customer_address": (shipping.get("baseAddress", "") or "").strip(),
        "customer_address_detail": (shipping.get("detailedAddress", "") or "").strip(),
        # 우편번호 — 화면 확인용 (복사 버튼 분리). 네이버 응답 케이스 변형 흡수 fallback chain
        "customer_postal_code": (
            str(
                shipping.get("zipCode")
                or shipping.get("zipcode")
                or shipping.get("postCode")
                or shipping.get("zipNo")
                or ""
            ).strip()
            or None
        ),
        "customer_note": po.get("shippingMemo", "") or "",
        "quantity": quantity,
        "sale_price": sale_price,
        "cost": 0,
        "fee_rate": fee_rate,
        "revenue": expected_settlement if expected_settlement else sale_price,
        # 내부 status도 클레임 반영
        "status": (
            "cancel_requested"
            if claim_status in ("CANCEL_REQUEST", "CANCELING")
            else (
                "cancelled"
                if claim_status == "CANCEL_DONE"
                else (
                    "return_requested"
                    if claim_status in ("RETURN_REQUEST", "COLLECTING", "COLLECT_DONE")
                    else (
                        "returned"
                        if claim_status == "RETURN_DONE"
                        else status_map.get(naver_status, "pending")
                    )
                )
            )
        ),
        "shipping_status": market_order_status,
        "shipping_company": po.get("deliveryCompany", ""),
        "tracking_number": po.get("trackingNumber", ""),
        "paid_at": _parse_iso_datetime(
            order_info.get("paymentDate") or po.get("paymentDate")
        ),
        "source": "smartstore",
    }


def _coupang_paid_to_utc(val: str | None) -> datetime | None:
    """쿠팡 paidAt(KST naive ISO) → UTC tz-aware datetime.

    쿠팡 ordersheet 응답의 paidAt/orderedAt은 timezone 정보 없는 KST 문자열이라
    그대로 사용하면 SambaOrder.paid_at(DateTime(timezone=True))과 비교 시
    'can't compare offset-naive and offset-aware datetimes' 에러 발생.
    naive 면 KST 부여, aware 면 그대로 UTC astimezone.
    """
    from datetime import timezone
    from zoneinfo import ZoneInfo

    dt = _parse_iso_datetime(val)
    if dt is None:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=ZoneInfo("Asia/Seoul"))
    return dt.astimezone(timezone.utc)


def _is_safe_phone(v: Any) -> bool:
    """안심번호(050x) 판별 — 0503/0504/0505/0507/0508 등 050 으로 시작.

    #536 — 마켓이 수령자 안심번호(050…)를 내려주면 판매자가 직접 연락 불가.
    """
    digits = re.sub(r"[^0-9]", "", str(v or ""))
    return digits.startswith("050") and len(digits) >= 8


def _pick_real_phone(primary: Any, real: Any) -> str:
    """수령자 전화가 안심번호(050x)면 주문자 실번호로 대체.

    #536 — primary(수령자, 안심 가능) 가 안심번호이고 real(주문자 실번호)이
    실번호면 real 반환. 아니면 primary 우선(무해). 둘 다 안심이면 primary 유지.
    """
    p = str(primary or "").strip()
    r = str(real or "").strip()
    if _is_safe_phone(p) and r and not _is_safe_phone(r):
        return r
    return p or r


def _parse_coupang_order(
    order: dict,
    account_id: str,
    account_label: str,
    cancel_info: Optional[dict] = None,
) -> dict[str, Any]:
    """쿠팡 ordersheet 1건 → SambaOrder 데이터 변환 (#246).

    cancel_info: returnRequests v6 API 응답에서 매칭된 1건. 없으면 None.
      필드: receiptId, receiptType(CANCEL/RETURN), faultByType,
            reasonCode, reasonCodeText, cancelReasonCategory1/2,
            releaseStatus(Y/N/S/A), releaseStopStatus, createdAt
    """
    status_map = {
        "ACCEPT": "pending",
        "INSTRUCT": "pending",
        "DEPARTURE": "shipped",
        "DELIVERING": "shipped",
        "FINAL_DELIVERY": "delivered",
        "CANCEL": "cancelled",
    }
    market_status_map = {
        "ACCEPT": "결제완료",
        "INSTRUCT": "상품준비중",
        "DEPARTURE": "국내배송중",
        "DELIVERING": "국내배송중",
        "FINAL_DELIVERY": "배송완료",
        "CANCEL": "취소완료",
    }

    coupang_status = (order.get("status") or "").upper()
    shipment_box_id = order.get("shipmentBoxId") or 0
    order_id = order.get("orderId") or 0

    # 클레임 (취소/반품 요청) — returnRequests v6 API 응답으로 판단 (#246)
    # 과거: order["cancelRequests"]/["returnRequests"] 의존했으나 ordersheets v5에 존재 X
    receipt_type = (
        ((cancel_info or {}).get("receiptType") or "").upper() if cancel_info else ""
    )
    # [#599] 취소/반품 판정 — receiptType 단독 판단 금지.
    #   쿠팡 returnRequests v6 는 출고중지 취소(상품준비중 고객취소)도
    #   receiptType=RETURN 으로 내려줌. 출고 여부(releaseStatus)가 권위 신호:
    #     Y=출고완료(진짜 반품) / S=출고중지 / N=미출고.
    #   RETURN 이라도 releaseStatus∈{S,N}(미출고) 또는 releaseStopStatus '출고중지'
    #   표기면 실제로는 '취소'. (SSG classify_ssg_completion 의 shpmtQty 판정과 동일 철학)
    #   receiptStatus 로 요청/완료 구분 — 실측 RETURNS_COMPLETED → 완료.
    _ci = cancel_info or {}
    _release_status = ""
    _ci_return_items = _ci.get("returnItems") or []
    if (
        isinstance(_ci_return_items, list)
        and _ci_return_items
        and isinstance(_ci_return_items[0], dict)
    ):
        _release_status = (_ci_return_items[0].get("releaseStatus") or "").upper()
    _release_stop = _ci.get("releaseStopStatus") or ""
    _receipt_status = (_ci.get("receiptStatus") or "").upper()
    _is_completed = "COMPLETED" in _receipt_status
    # RETURN 인데 미출고(S/N) 또는 출고중지 표기 → 취소로 재분류
    _return_is_actually_cancel = receipt_type == "RETURN" and (
        _release_status in ("S", "N") or "출고중지" in _release_stop
    )

    if receipt_type == "CANCEL" or _return_is_actually_cancel:
        if _is_completed:
            market_order_status = "취소완료"
            internal_status = "cancelled"
        else:
            market_order_status = "취소요청"
            internal_status = "cancel_requested"
    elif receipt_type == "RETURN":
        if _is_completed:
            market_order_status = "반품완료"
            internal_status = "returned"
        else:
            market_order_status = "반품요청"
            internal_status = "return_requested"
    else:
        market_order_status = market_status_map.get(coupang_status, coupang_status)
        internal_status = status_map.get(coupang_status, "pending")

    order_items = order.get("orderItems") or []
    first_item = order_items[0] if order_items else {}
    product_name = first_item.get("sellerProductName", "") or ""
    # 쿠팡 옵션 없음 placeholder 패턴 (대소문자/공백/구두점 변형 허용)
    _NO_OPTION_PATTERNS = ("옵션없음", "no option")

    option_name = (
        first_item.get("sellerProductItemName", "")
        or first_item.get("firstSellerProductItemName", "")
        or ""
    ).strip()

    # placeholder 텍스트 정규화 (예: "옵션없음. 옵션없음." → "FREE")
    _normalized = option_name.lower().replace(" ", "").replace(".", "")
    if not option_name or any(
        p.replace(" ", "") in _normalized for p in _NO_OPTION_PATTERNS
    ):
        option_name = "FREE"
    sales_price = int(first_item.get("salesPrice", 0) or 0)
    # 쿠팡 수량 필드는 shippingCount (orderQuantity 키는 응답에 없음)
    quantity = int(first_item.get("shippingCount", 1) or 1)
    shipping_price = int(order.get("shippingPrice", 0) or 0)
    # orderPrice = 라인 총액(단가×수량). 멀티수량 결제총액 정상화 폴백 salesPrice×quantity
    line_total = int(first_item.get("orderPrice", 0) or 0) or (sales_price * quantity)
    sale_price = line_total + shipping_price

    # 쿠팡 정률 수수료 10.5% + VAT 10% = 실효 11.55%
    fee_rate = 11.55
    revenue = round(sale_price * (1 - fee_rate / 100))

    # 쿠팡 ordersheet 응답은 receiver/orderer를 nested object로 내려줌.
    # 과거 flat key (receiverAddr1 등) 사용 코드가 빈값을 만들었음.
    receiver = order.get("receiver") or {}
    orderer = order.get("orderer") or {}

    receiver_addr = (
        receiver.get("addr1")
        or order.get("receiverAddr1", "")
        or order.get("receiverAddress", "")
        or ""
    )
    receiver_addr_detail = (
        receiver.get("addr2")
        or order.get("receiverAddr2", "")
        or order.get("receiverAddrDetail", "")
        or ""
    )
    customer_address = receiver_addr.strip()
    customer_address_detail = receiver_addr_detail.strip()
    # 우편번호 — 화면 확인용 (복사 버튼 분리)
    customer_postal_code = (
        str(receiver.get("postCode") or order.get("receiverPostCode") or "").strip()
        or None
    )

    # 수령인/주문자 분리 — 쿠팡 ordersheet는 receiver(수취인)와 orderer(주문자)를 별도 제공.
    # 과거: 단일 필드에 합쳐 customer_name에만 박아 orderer_name NULL → 화면에서 동일 표시 버그.
    customer_name = (
        receiver.get("name")
        or orderer.get("name")
        or order.get("receiverName", "")
        or order.get("ordererName", "")
        or ""
    )
    orderer_name = (
        orderer.get("name")
        or receiver.get("name")
        or order.get("ordererName", "")
        or order.get("receiverName", "")
        or ""
    )
    # 연락처 — 스펙: orderer.safeNumber 는 "수취인 안심번호"(orderer 객체 안이지만 의미상 수취인용).
    # 우선순위: orderer 안심 > receiver 안심/실번 > orderer 실번호 폴백.
    customer_phone = (
        orderer.get("safeNumber")
        or receiver.get("safeNumber")
        or receiver.get("receiverNumber")
        or orderer.get("ordererNumber")
        or order.get("receiverPhoneNumber", "")
        or order.get("ordererPhoneNumber", "")
        or order.get("orderPhoneNumber", "")
        or ""
    )
    # #536 — 수령자 안심번호(050x)면 주문자 실번호로 대체. 해외구매대행은
    # overseaShippingInfoDto.ordererPhoneNumber(통관용 실번호)를 실번호로 사용.
    _oversea = order.get("overseaShippingInfoDto") or {}
    _real_phone = (
        orderer.get("ordererNumber")
        or _oversea.get("ordererPhoneNumber")
        or order.get("ordererPhoneNumber", "")
        or ""
    )
    customer_phone = _pick_real_phone(customer_phone, _real_phone)

    if not customer_name and not customer_address:
        logger.warning(
            f"[쿠팡][주문파싱] customer 빈값 — keys={list(order.keys())[:25]} "
            f"receiver_keys={list(receiver.keys()) if isinstance(receiver, dict) else 'NA'} "
            f"orderer_keys={list(orderer.keys()) if isinstance(orderer, dict) else 'NA'}"
        )

    # shipmentBoxId 우선 (배송단위 안정 ID), orderId fallback
    order_number = str(shipment_box_id or order_id or "")

    # 쿠팡 옵션 ID — 송장업로드 API(/orders/invoices) body 필수 파라미터
    vendor_item_id = str(first_item.get("vendorItemId") or "") or None

    # 취소·반품 사유 필드 (#246) — cancel_info 매칭된 returnRequests v6 응답에서 추출
    cancel_receipt_id: Optional[int] = None
    cancel_reason_code: Optional[str] = None
    cancel_reason_text: Optional[str] = None
    cancel_reason_category1: Optional[str] = None
    cancel_reason_category2: Optional[str] = None
    cancel_fault_by: Optional[str] = None
    cancel_release_status: Optional[str] = None
    cancel_release_stop_status: Optional[str] = None
    cancel_requested_at = None
    if cancel_info:
        _rid = cancel_info.get("receiptId")
        if _rid is not None:
            try:
                cancel_receipt_id = int(_rid)
            except (TypeError, ValueError):
                cancel_receipt_id = None
        cancel_reason_code = cancel_info.get("reasonCode") or None
        cancel_reason_text = cancel_info.get("reasonCodeText") or None
        cancel_reason_category1 = cancel_info.get("cancelReasonCategory1") or None
        cancel_reason_category2 = cancel_info.get("cancelReasonCategory2") or None
        cancel_fault_by = cancel_info.get("faultByType") or None
        cancel_release_stop_status = cancel_info.get("releaseStopStatus") or None
        # returnItems[].releaseStatus — 첫 항목 기준 (Y/N/S/A 단일값 가정)
        return_items = cancel_info.get("returnItems") or []
        if isinstance(return_items, list) and return_items:
            first_ri = return_items[0] if isinstance(return_items[0], dict) else {}
            cancel_release_status = first_ri.get("releaseStatus") or None
        cancel_requested_at = _coupang_paid_to_utc(cancel_info.get("createdAt"))

    return {
        "order_number": order_number,
        "shipment_id": str(order_id) if order_id else "",
        "ext_order_number": str(order_id) if order_id else "",
        "vendor_item_id": vendor_item_id,
        "channel_id": account_id,
        "channel_name": account_label,
        "product_id": str(
            first_item.get("productId", "")
            or first_item.get("sellerProductId", "")
            or ""
        ),
        # sellerProductId 별도 보존 (#408) — 다중옵션 리스팅의 비대표 옵션 주문은
        # productId/vendorItemId 가 인덱스에 없어 미등록 → 상품당 1개·옵션무관 안정키인
        # sellerProductId 로 폴백 매칭. product_id 에 합쳐 버리면 위 fallback 이 못 씀.
        "seller_product_id": str(first_item.get("sellerProductId", "") or ""),
        "product_name": product_name,
        "coupang_display_name": first_item.get("vendorItemPackageName", "") or "",
        "product_option": option_name,
        "product_image": "",
        "customer_name": customer_name,
        "orderer_name": orderer_name,
        "customer_phone": customer_phone,
        "customer_address": customer_address,
        "customer_address_detail": customer_address_detail,
        "customer_postal_code": customer_postal_code,
        "customer_note": (
            order.get("parcelPrintMessage", "")
            or order.get("shippingMessage", "")
            or ""
        ),
        "quantity": quantity,
        "sale_price": sale_price,
        "cost": 0,
        "fee_rate": fee_rate,
        "revenue": revenue,
        "status": internal_status,
        "shipping_status": market_order_status,
        "shipping_company": order.get("deliveryCompanyName", "") or "",
        "tracking_number": order.get("invoiceNumber", "") or "",
        "paid_at": _coupang_paid_to_utc(order.get("paidAt") or order.get("orderedAt")),
        "source": "coupang",
        # 쿠팡 취소/반품 사유 (#246)
        "cancel_receipt_id": cancel_receipt_id,
        "cancel_reason_code": cancel_reason_code,
        "cancel_reason_text": cancel_reason_text,
        "cancel_reason_category1": cancel_reason_category1,
        "cancel_reason_category2": cancel_reason_category2,
        "cancel_fault_by": cancel_fault_by,
        "cancel_release_status": cancel_release_status,
        "cancel_release_stop_status": cancel_release_stop_status,
        "cancel_requested_at": cancel_requested_at,
    }


def _coerce_lotteon_quantity(item: dict) -> int:
    """롯데ON 주문 수량 안전 파싱 — odQty 우선, float/str 모두 처리 (issue #213)."""
    for key in ("odQty", "slQty"):
        v = item.get(key)
        if v in (None, "", 0, "0"):
            continue
        try:
            return max(1, int(float(v)))
        except (TypeError, ValueError):
            continue
    return 1


def _parse_lotteon_order(item: dict, account_id: str, label: str) -> dict:
    """롯데ON 주문 데이터 → SambaOrder dict 변환."""

    # 주문 진행 단계 코드 → 내부 status/shipping_status 매핑
    step_cd = str(item.get("odPrgsStepCd", "") or "")
    status_map = {
        "10": "pending",  # 발주확인대기
        "11": "preparing",  # 발주확인완료(출고지시) — sync에서 자동 ifCplYN=Y 호출되어 12로 전이
        "12": "preparing",  # 상품준비
        "13": "shipping",  # 발송완료
        "14": "delivered",  # 배송완료
        "20": "pending",  # 발주확인
        "21": "return_requested",  # 교환회수중
        "22": "return_requested",  # 교환회수완료
        "23": "return_requested",  # 교환회수완료확인
        "24": "shipping",  # 교환재배송
        "25": "delivered",  # 교환배송완료
        "30": "shipping",  # 배송중
        "40": "delivered",  # 배송완료
        "50": "confirmed",  # 구매확정
        "90": "cancelled",  # 취소
    }
    shipping_map = {
        "10": "발주확인대기",
        "11": "출고지시",
        "12": "상품준비",
        "13": "발송완료",
        "14": "배송완료",
        "20": "출고지시",
        "21": "교환요청",
        "22": "교환회수완료",
        "23": "교환회수완료",
        "24": "교환재배송",
        "25": "교환완료",
        "30": "국내배송중",
        "40": "배송완료",
        "50": "구매확정",
        "90": "취소완료",
    }
    status = status_map.get(step_cd, "pending")
    shipping_status = shipping_map.get(step_cd, "출고지시")

    # 롯데ON 반품 사유코드(200/300번대)인데 교환 stepCd(21~25)로 들어온 경우
    # → 실제로는 반품이므로 반품 상태로 재매핑
    clm_rsn_cd = str(item.get("clmRsnCd", "") or "")
    if clm_rsn_cd.startswith(("2", "3")) and step_cd in ("21", "22", "23", "24", "25"):
        status = "return_requested"
        shipping_status = "반품요청"
        logger.info(
            f"[롯데ON][주문파싱] 반품 사유코드({clm_rsn_cd}) 교환 stepCd({step_cd}) "
            f"→ 반품요청으로 재매핑: odNo={item.get('odNo')}"
        )

    # 결제일시 파싱 — 롯데ON 응답 실측 키는 odCmptDttm (yyyymmddHHmmss, KST)
    # 참고: owhoDttm(발주확인, ISO 포맷)은 결제 이후 시각이라 결제시각 폴백으로 부적합
    from backend.utils import kst_str_to_utc

    order_dttm_str = item.get("odCmptDttm") or ""
    paid_at = kst_str_to_utc(order_dttm_str)
    if not paid_at:
        logger.warning(
            f"[롯데ON][주문파싱] 결제일시 키 없음 odNo={item.get('odNo')} "
            f"odCmptDttm={item.get('odCmptDttm')!r} "
            f"키후보={[k for k in item.keys() if 'tt' in k.lower() or 'dt' in k.lower()]}"
        )

    # 배송지 주소 분리 저장 (dvpStnmZipAddr=도로명기본주소, dvpStnmDtlAddr=상세주소)
    # 롯데ON API 특성: dvpStnmDtlAddr이 "번지수, 상세주소" 형태로 내려옴.
    # 번지수는 도로명기본주소에 포함되어야 함 → "123, 101동 305호" → base += " 123", detail = "101동 305호"
    addr_base = (item.get("dvpStnmZipAddr") or "").strip()
    addr_detail = (item.get("dvpStnmDtlAddr") or "").strip()
    _lot_match = re.match(r"^(\d+),\s*(.*)", addr_detail)
    if _lot_match:
        addr_base = f"{addr_base} {_lot_match.group(1)}"
        addr_detail = _lot_match.group(2).strip()
    # 우편번호 — 화면 확인용 (복사 버튼 분리). 롯데ON 응답 키 변형 흡수 fallback chain
    postal_code = (
        str(
            item.get("dvpZpcd")
            or item.get("dvpZipNo")
            or item.get("dvpStnmZpcd")
            or item.get("dvpJbngZpcd")
            or item.get("zipNo")
            or ""
        ).strip()
        or None
    )
    # 모든 후보 비어있으면 1회 키 후보 로그 (실제 응답 키 진단용)
    if not postal_code:
        _zip_keys = [k for k in item.keys() if "zp" in k.lower() or "zip" in k.lower()]
        if _zip_keys:
            logger.info(f"[롯데ON][zip진단] od={item.get('odNo')} zip_keys={_zip_keys}")

    _od_no = str(item.get("odNo", "") or "")
    _od_seq = str(item.get("odSeq", "1") or "1")
    _proc_seq = str(item.get("procSeq", "1") or "1")
    _sitm_no = str(item.get("sitmNo", "") or "")

    return {
        "channel_id": account_id,
        "channel_name": label,
        "source": "lotteon",
        # 합성 키: (odNo, odSeq) — procSeq는 처리 단계에 따라 변하므로 제외
        "order_number": f"{_od_no}_{_od_seq}" if _od_no else "",
        "od_no": _od_no,
        "od_seq": _od_seq,
        "proc_seq": _proc_seq,
        "sitm_no": _sitm_no,
        "shipment_id": _sitm_no,
        "product_id": str(item.get("spdNo", "") or ""),
        "product_name": item.get("spdNm", "") or "",
        "product_option": item.get("sitmNm", "") or "",
        # issue #213 — odQty/slQty 응답이 float("5.0") 또는 str로 올 수 있어 int(float()) 사용
        # SellerDeliveryProgressStateSearch/SellerDeliveryOrdersSearch는 odQty, getSROrderList는 둘 다 부재
        "quantity": _coerce_lotteon_quantity(item),
        "sale_price": int(item.get("slAmt", 0) or item.get("slPrc", 0) or 0),
        "cost": 0,
        "status": status,
        "shipping_status": shipping_status,
        "customer_name": item.get("dvpCustNm", "") or "",
        "orderer_name": item.get("odrNm", "") or "",
        # #536 — 수령자(dvpMphnNo)가 안심번호(050x)면 주문자 실번호(mphnNo)로 대체.
        "customer_phone": _pick_real_phone(
            item.get("dvpMphnNo", "") or item.get("dvpTelNo", ""),
            item.get("mphnNo", ""),
        ),
        "customer_address": addr_base,
        "customer_address_detail": addr_detail,
        "customer_postal_code": postal_code,
        "customer_note": item.get("dvMsg", "") or "",
        "paid_at": paid_at,
        # created_at은 명시 X — DB default_factory(now)가 실제 삽입 시각 기록
    }


def _parse_poison_order(item: dict, account_id: str, label: str) -> dict:
    """POIZON(得物) 주문 데이터 → SambaOrder dict 변환."""
    from backend.utils import kst_str_to_utc

    # 주문 상태 코드(order_status, int) → 내부 status 매핑
    # 1000 결제대기, 2000 발송준비, 2100~3040 배송/검수, 2800/4000 완료, 7000~ 취소
    order_status = item.get("order_status")
    try:
        _status_code = int(order_status) if order_status is not None else 0
    except (TypeError, ValueError):
        _status_code = 0
    poison_status_map = {
        1000: "pending",
        2000: "preparing",
        2100: "shipping",
        2200: "shipping",
        2500: "shipping",
        2550: "shipping",
        2600: "shipping",
        2650: "shipping",
        2700: "shipping",
        3040: "shipping",
        2800: "delivered",
        4000: "delivered",
        7000: "cancelled",
        8000: "cancelled",
        8010: "cancelled",
        8080: "cancelled",
    }
    status = poison_status_map.get(_status_code, "preparing")

    # 결제일시 — "yyyy-MM-dd HH:mm:ss" (셀러 타임존 KST 가정) → UTC
    paid_at = kst_str_to_utc(item.get("pay_time") or "")

    # 수량 안전 파싱
    try:
        quantity = max(1, int(item.get("qty") or 1))
    except (TypeError, ValueError):
        quantity = 1

    # 결제금액 — pay_amount(통화 최소단위 정수). TODO: currency(item.get("currency"))
    # 가 KRW가 아닌 경우 환율 환산 필요. 현재는 원본 값을 그대로 저장.
    try:
        product_price = int(item.get("pay_amount") or 0)
    except (TypeError, ValueError):
        product_price = 0

    # 배송지(delivery_address_platform) — 수취인/주소 분리 저장
    dap = item.get("delivery_address_platform") or {}
    if not isinstance(dap, dict):
        dap = {}
    customer_name = (dap.get("name") or "").strip()
    customer_phone = (dap.get("mobile") or "").strip()
    _addr_parts = [
        (dap.get("province") or "").strip(),
        (dap.get("city") or "").strip(),
        (dap.get("district") or "").strip(),
    ]
    customer_address = " ".join(p for p in _addr_parts if p)
    customer_address_detail = (dap.get("address_detail") or "").strip()

    _order_no = str(item.get("order_no", "") or "")
    _currency = str(item.get("currency", "") or "")

    return {
        "channel_id": account_id,
        "channel_name": label,
        "source": "poison",
        "order_number": _order_no,
        "od_no": _order_no,
        # 원본 식별자 보존 (메모 컬럼)
        "shipment_id": str(item.get("seller_bidding_no", "") or ""),
        "product_id": str(item.get("spu_id", "") or item.get("sku_id", "") or ""),
        "product_name": item.get("title", "") or "",
        "product_option": item.get("properties", "") or "",
        "quantity": quantity,
        "sale_price": product_price,
        "cost": 0,
        "status": status,
        "shipping_status": "",
        "customer_name": customer_name,
        "customer_phone": customer_phone,
        "customer_address": customer_address,
        "customer_address_detail": customer_address_detail,
        # currency != KRW 인 경우 환산 TODO. 메모에 원본 통화/품번 보존.
        "customer_note": " / ".join(
            p
            for p in (
                f"통화:{_currency}" if _currency else "",
                f"품번:{item.get('article_number', '')}"
                if item.get("article_number")
                else "",
            )
            if p
        ),
        "paid_at": paid_at,
        # created_at은 명시 X — DB default_factory(now)가 실제 삽입 시각 기록
    }


def _normalize_playauto_alias_code(value: Any) -> str:
    return normalize_playauto_alias_code(value)


def _normalize_synced_order_status(order_data: dict[str, Any]) -> None:
    """Market sync must only drive shipping_status; status stays user-managed.

    예외: 플레이오토 미등록 주문의 취소요청/취소완료 상태는 status도 동기화해야
    UI 드롭다운이 어긋나지 않음 (cancel_requested/cancelled 보존).
    """
    preserved = {
        "cancel_requested",
        "cancelled",
        "cancelling",
        "return_requested",
        "returning",
        "returned",
        "exchanging",
        "exchanged",
        "return_completed",
    }
    cur_status = str(order_data.get("status") or "")
    if (
        order_data.get("source") == "playauto"
        and not order_data.get("collected_product_id")
        and not order_data.get("source_url")
        and not order_data.get("product_image")
        and cur_status in preserved
    ):
        return
    # issue #393 — 롯데홈쇼핑 반품/취소 클레임 상태는 신규 insert(원주문 미매칭) 시에도
    # 보존. 안 그러면 반품 주문이 pending 으로 리셋됨. 정상 배송 상태는 status_map/
    # update 경로가 관리하므로 여기서 pending 으로 떨어뜨려도 무방.
    if order_data.get("source") == "lottehome" and cur_status in preserved:
        return
    # 결제금액 1,000원 이하 주문은 수집 즉시 배송완료 처리 (서비스 비용·증정 등)
    _pamt = float(
        order_data.get("total_payment_amount") or order_data.get("sale_price") or 0
    )
    if 0 < _pamt <= 1000:
        order_data["status"] = "delivered"
    else:
        order_data["status"] = "pending"


def _can_override_source_site_from_sourcing(order_data: dict[str, Any]) -> bool:
    """매칭된 collected_product 의 source_site 로 order.source_site 를 덮어써도 되는지.

    과거: PlayAuto 주문은 source_site 에 별칭("GS이숍(캐논)" 등)을 넣어서 매칭으로 덮어쓰면 안 됐음.
    현재(sales_channel_alias 분리 후): PlayAuto 도 source_site="" 로 임포트되므로 비어 있으면 채워야 정상.
    별칭은 이제 sales_channel_alias 컬럼에 별도 보관됨.
    """
    raw = str(order_data.get("source_site") or "").strip()
    # 비어 있으면 항상 채움. 이미 값이 있으면 (소싱처 코드든 별칭이든) 보존.
    return not raw


def _normalize_carrier_name(value: Any) -> str:
    raw = str(value or "").strip().upper()
    if not raw:
        return ""
    normalized = re.sub(r"[\s()\-_/]", "", raw)
    normalized = normalized.replace("주식회사", "").replace("(주)", "")
    return normalized


def _parse_playauto_order(
    ro: dict,
    account_id: str,
    account_label: str,
    alias_map: dict[str, str] | None = None,
) -> dict[str, Any]:
    """플레이오토 EMP 주문 → SambaOrder 데이터 변환."""

    # spec 진단용 — SiteId(별칭)별 첫 1건씩 raw 로깅. MasterCode/MyCateName 등 키별 값 확인.
    _logged_sites = getattr(_parse_playauto_order, "_logged_sites", set())
    _site_raw = str(ro.get("SiteId", "")).strip()
    if _site_raw and _site_raw not in _logged_sites:
        try:
            import json as _json

            sample = {
                k: str(ro.get(k, ""))[:80]
                for k in (
                    "SiteId",
                    "SiteName",
                    "ProdCode",
                    "MasterCode",
                    "MyCateName",
                    "SellerCode",
                    "Groupkey",
                    "ProdName",
                    "OrderCode",
                    "Number",
                )
            }
            logger.info(
                f"[플레이오토 raw site={_site_raw}] {_json.dumps(sample, ensure_ascii=False)}"
            )
            _logged_sites.add(_site_raw)
            _parse_playauto_order._logged_sites = _logged_sites  # type: ignore[attr-defined]
        except Exception:
            pass

    # MasterCode 추출 (응답에 있으면 매칭에 활용 — Phase 4)
    # SellerCode fallback: 일부 사이트(SiteId=1054236 등)에서 MasterCode="" + SellerCode=AM... 패턴
    master_code = (
        ro.get("MasterCode")
        or ro.get("master_code")
        or ro.get("masterCode")
        or ro.get("SellerCode")
        or ""
    )

    status_map = {
        "신규주문": "pending",
        "송장출력": "wait_ship",
        "송장입력": "processing",
        # shipping_status 가 "국내배송중"일 때 status 드롭다운도 "국내배송중"(shipping)으로 보이도록 동기화.
        # 과거에 "shipped"로 매핑되어 프론트 STATUS_MAP 에 없는 enum 으로 저장되던 버그도 같이 닫힘.
        "출고": "shipping",
        "배송중": "shipping",
        "국내배송중": "shipping",
        "수취확인": "delivered",
        "정산완료": "delivered",
        "주문확인": "pending",
        "취소": "cancelled",
        "취소마감": "cancelled",
        "반품요청": "return_requested",
        "반품마감": "returned",
        "교환요청": "exchange_requested",
        "교환마감": "exchanged",
        "보류": "pending",
    }

    # shipping_status 매핑 (스킬 가이드 기준)
    shipping_status_map = {
        "신규주문": "주문접수",
        "송장출력": "배송대기중",
        "송장입력": "송장전송완료",
        "출고": "국내배송중",
        "배송중": "국내배송중",
        "주문확인": "취소중",
        "취소마감": "취소완료",
        "수취확인": "배송완료",
        "정산완료": "배송완료",
    }

    order_state = ro.get("OrderState", "")
    sale_price = int(ro.get("Price", 0) or 0)
    quantity = int(ro.get("Count", 1) or 1)

    site_name = str(ro.get("SiteName", "") or "").strip()
    site_id = _normalize_playauto_alias_code(ro.get("SiteId", ""))
    supply_price = int(ro.get("SupplyPrice", 0) or 0)

    # 결제일 파싱 — 플레이오토는 KST 기준
    from backend.utils import kst_str_to_utc

    order_date_raw = ro.get("OrderDate", "") or ""
    paid_at = kst_str_to_utc(order_date_raw)

    # 주소 분리 — 플레이오토는 RecipientAddress 한 필드에 도로명+상세를 통째로 내려줌
    # (openapi.json 확인: 별도 상세주소 필드 없음). 휴리스틱으로 기본/상세 분리.
    # 우선순위 (프론트 splitCustomerAddress 와 동일 — 괄호 안 콤마로 잘리지 않도록):
    #  패턴A: 끝 메타괄호 `(법정동/건물명)` + 그 앞 `동/호/층/호실` 패턴
    #         → base = 도로주소 + 메타괄호, detail = 동/호 토큰
    #  패턴B: 마지막 `)` 뒤에 내용이 있으면 그 지점으로 split (괄호 안 콤마 무시)
    #         (예) "...압구정로 403(압구정동, 한양아파트) 81동 1207호"
    #             → base="...압구정로 403(압구정동, 한양아파트)", detail="81동 1207호"
    #  패턴C: 괄호가 없으면 ", " 명시 구분 ("디지털로26길 123, 14층 플레이오토")
    #  패턴D: 도로명(...대로/로/길) + 본번 뒤 공백 기준 분리
    import re as _re_addr

    _addr_full = str(ro.get("RecipientAddress", "") or "").strip()
    _addr_base = _addr_full
    _addr_detail = ""
    if _addr_full:
        _matched = False
        # 패턴A: 끝 메타괄호 + 동/호 패턴 (전체가 `(...)$` 로 끝나는 경우)
        _meta_m = _re_addr.match(r"^(.*?)\s*(\([^)]*\))\s*$", _addr_full)
        if _meta_m:
            _before_meta = _meta_m.group(1).strip()
            _meta = _meta_m.group(2)
            # 옵션 prefix: 건물명(숫자로 시작하지 않는 토큰). 본번 "218"·"1462-14"가
            # detail로 빨려들지 않도록 첫 글자에 숫자 금지.
            _dongho_m = _re_addr.match(
                r"^(.+?)\s+((?:[^\d\s]\S*\s+)?(?:\d+\s*동\s+)?\d+\s*(?:호|층|호실))$",
                _before_meta,
            )
            if _dongho_m:
                _addr_base = f"{_dongho_m.group(1).strip()} {_meta}".strip()
                _addr_detail = _dongho_m.group(2).strip()
                _matched = True
        # 패턴B: 마지막 `)` 기준 분리 — `, ` 보다 우선.
        # 괄호 안 콤마("(압구정동, 한양아파트)")로 base/detail 가 잘못 잘리지 않도록.
        if not _matched:
            _last_paren = _addr_full.rfind(")")
            if 0 < _last_paren < len(_addr_full) - 1:
                _after = _addr_full[_last_paren + 1 :].strip()
                if _after:
                    _addr_base = _addr_full[: _last_paren + 1].strip()
                    _addr_detail = _after
                    _matched = True
        if not _matched:
            # 패턴C: 괄호 없는 도로명주소 — ", " 단순 분리
            if "(" not in _addr_full and ", " in _addr_full:
                _b, _, _d = _addr_full.partition(", ")
                _addr_base, _addr_detail = _b.strip(), _d.strip()
            else:
                # 패턴D: 도로명 + 본번 뒤 공백 기준
                _m = _re_addr.match(
                    r"^(.+?(?:대로|로|길)\s+\d+(?:-\d+)?)\s+(.+)$", _addr_full
                )
                if _m:
                    _addr_base = _m.group(1).strip()
                    _addr_detail = _m.group(2).strip()

    # ── 배송메시지(customer_note) ──
    # 플레이오토 EMP 응답의 배송메시지 필드명이 공개문서에 없어 불명확.
    # (a) 알려진 후보 키 우선 → (b) 값 휴리스틱(배송메시지 특유 문구)으로 포착.
    # 배송사고 방지용(문앞/공동현관 출입번호 등). 실제 필드명 확인되면 단순화.
    _pa_note = ""
    for _cand in (
        "ShipMsg",
        "DlvMsg",
        "DeliveryMsg",
        "OrderMsg",
        "GiftMsg",
        "Msg",
        "Memo",
        "Message",
        "ShippingMessage",
        "DeliveryMessage",
        "ShipMessage",
        "OrderMemo",
        "DlvMemo",
    ):
        _cv = ro.get(_cand)
        if isinstance(_cv, str) and _cv.strip():
            _pa_note = _cv.strip()
            break
    if not _pa_note:
        _NOTE_HINTS = (
            "놓아",
            "출입번호",
            "부재",
            "경비",
            "문 앞",
            "문앞",
            "배송전",
            "직접 받",
            "안심번호",
            "파손",
            "취급주의",
            "부탁드립니다",
            "요청드립니다",
            "요청합니다",
        )
        _SKIP_KEYS = {"ProdName", "RecipientName", "OrderName", "SiteName", "ProdCode"}
        for _k, _v in ro.items():
            if _k in _SKIP_KEYS:
                continue
            if (
                isinstance(_v, str)
                and 2 <= len(_v.strip()) <= 200
                and any(h in _v for h in _NOTE_HINTS)
            ):
                _pa_note = _v.strip()
                logger.info(f"[플레이오토 배송메시지] 휴리스틱 포착 필드='{_k}'")
                break

    return {
        "order_number": ro.get("OrderCode", ""),
        "shipment_id": str(ro.get("Number", "")),
        # 라인 유니크키(Number)를 ord_prd_seq에 넣어 멀티라인 구분.
        # 같은 OrderCode에 상품 여러 개일 때 order_number만으론 uq_order_tenant_number_seq
        # (NULLS NOT DISTINCT)에서 NULL끼리 충돌 → 2번째 라인부터 유실되던 버그 수정.
        # 기존주문(ord_prd_seq=NULL)은 shipment_id 폴백으로 매칭돼 중복 안 생김.
        "ord_prd_seq": str(ro.get("Number", "")).strip() or None,
        "channel_id": account_id,
        "channel_name": account_label,
        "product_id": ro.get("ProdCode", ""),
        "product_name": ro.get("ProdName", ""),
        "product_option": ro.get("Option", ""),
        "product_image": "",
        "customer_name": ro.get("RecipientName", "") or ro.get("OrderName", ""),
        "customer_phone": ro.get("RecipientHtel", "")
        or ro.get("RecipientTel", "")
        or ro.get("OrderHtel", "")
        or ro.get("OrderTel", ""),
        "customer_address": _addr_base,
        "customer_address_detail": _addr_detail,
        # 우편번호 — 화면 확인용 (복사 버튼 분리). 플레이오토 EMP는 RecipientZipCode 필드 사용.
        "customer_postal_code": str(ro.get("RecipientZipCode") or "").strip() or None,
        # 배송메시지 — 다른 마켓처럼 customer_note 매핑 (위 _pa_note 참고). 직배주문 시 소싱처 주문서에 입력.
        "customer_note": _pa_note,
        "quantity": quantity,
        "sale_price": sale_price,
        "cost": 0,
        "fee_rate": 0,
        "revenue": supply_price if supply_price else sale_price,
        "status": status_map.get(order_state, "pending"),
        "shipping_status": shipping_status_map.get(order_state, order_state),
        "shipping_company": ro.get("Sender", ""),
        "tracking_number": ro.get("SenderNo", ""),
        "paid_at": paid_at,
        "source": "playauto",
        # 별칭 단위 매칭 검증용 — DB 저장 전 pop. site_id가 cp의 등록된 site_ids에
        # 포함될 때만 매칭 허용 (기존 cp는 site_ids 미저장이라 호환 매칭).
        "_pa_site_id": site_id,
        # 매칭용 임시 키 — DB 저장 전 pop. plapro 응답에 MasterCode 있으면 추출해
        # _mpn_cache 매칭에 ProdCode와 함께 시도. 매칭 우선순위: master_code > product_id.
        "_pa_master_code": master_code,
        # 판매처(사업자) 별칭 — PlayAuto 1 채널 × 다 site_id 구조 (예: "GS이숍(캐논)").
        # source_site 와 분리 — source_site 는 진짜 소싱처 코드 전용.
        "sales_channel_alias": (
            f"{site_name}({alias_map[site_id]})"
            if alias_map and site_id in alias_map and site_name
            else f"{site_name}({site_id})"
            if site_name
            else ""
        ),
        # source_site 는 collected_product 매칭 후 자동 채워짐 — 임포트 시점엔 빈 값.
        "source_site": "",
    }


def _parse_elevenst_order(item: dict, account_id: str, label: str) -> dict:
    """11번가 주문 데이터를 SambaOrder dict로 변환."""
    from datetime import datetime, timedelta, timezone

    KST = timezone(timedelta(hours=9))

    def _to_int(value, default: int = 0) -> int:
        """콤마, None, 빈 문자열 안전하게 int 변환."""
        try:
            if value in (None, ""):
                return default
            return int(str(value).replace(",", "").strip())
        except (TypeError, ValueError):
            return default

    # ordPrdStat 상태 코드 맵핑
    stat_code = str(item.get("ordPrdStat", "") or "")
    status_map = {
        "200": "pending",  # 결제완료
        "202": "pending",  # 처리중 (배송완료 이전 단계)
        "301": "wait_ship",  # 발주확인(배송대기)
        "400": "shipping",  # 출고완료
        "500": "shipping",  # 배송중
        "600": "delivered",  # 배송완료
        "700": "confirmed",  # 구매확정
        "900": "cancelled",  # 취소완료
        "1000": "returned",  # 반품완료
    }
    shipping_map = {
        "200": "결제완료",
        "202": "결제완료",  # 11번가 내부 처리중 상태 (결제완료와 동일 단계)
        "301": "배송대기중",  # 발주확인 완료
        "400": "출고완료",
        "500": "국내배송중",
        "600": "배송완료",
        "700": "구매확정",
        "900": "취소완료",
        "1000": "반품완료",
    }
    status = status_map.get(stat_code, "pending")
    shipping_status = shipping_map.get(stat_code, "처리중" if stat_code else "결제완료")

    # 주문일 파싱 (API 응답: "YYYY-MM-DD HH:MM:SS" 또는 "YYYYMMDDhhmm", KST)
    ord_dt = str(item.get("ordDt", "") or "").strip()
    try:
        if "-" in ord_dt:
            paid_at = (
                datetime.strptime(ord_dt, "%Y-%m-%d %H:%M:%S")
                .replace(tzinfo=KST)
                .astimezone(timezone.utc)
            )
        else:
            paid_at = (
                datetime.strptime(ord_dt[:12], "%Y%m%d%H%M")
                .replace(tzinfo=KST)
                .astimezone(timezone.utc)
            )
    except Exception:
        paid_at = datetime.now(timezone.utc)

    # 수령인 주소 분리 저장 (실제 API 필드: rcvrBaseAddr=기본, rcvrDtlsAddr=상세)
    addr_base = str(item.get("rcvrBaseAddr", "") or "").strip()
    addr_detail = str(item.get("rcvrDtlsAddr", "") or "").strip()
    # 우편번호 — 화면 확인용 (복사 버튼 분리). 11번가 API 우편번호 필드: rcvrMlmtNo
    postal_code = str(item.get("rcvrMlmtNo") or "").strip() or None

    # 판매금액: selPrc(단가) 우선, 없으면 ordAmt(주문총액)을 수량으로 나눠 단가 환산
    quantity = max(1, _to_int(item.get("ordQty"), 1))
    sel_prc = _to_int(item.get("selPrc"))
    ord_amt = _to_int(item.get("ordAmt"))
    if sel_prc > 0:
        sale_price = sel_prc
    elif ord_amt > 0 and quantity > 0:
        sale_price = ord_amt // quantity
    else:
        sale_price = 0

    # 결제금액(주문 총액) — ordAmt(단가×수량+옵션가) 우선, 폴백 sale_price×quantity
    # 멀티수량 주문에서 결제 컬럼이 단가로 표시되는 회귀 방지 (2026-05-18)
    total_payment_amount = ord_amt if ord_amt > 0 else sale_price * quantity

    # 정산예정금액: stlPlnAmt (라인 총액 — 수량 포함)
    revenue = _to_int(item.get("stlPlnAmt"), total_payment_amount)

    # 수수료율 = (1 - 정산예정금액 / 결제금액) × 100
    # 음수/이상값 방지: revenue가 total_payment_amount보다 크면 0으로 처리
    if total_payment_amount > 0 and 0 < revenue <= total_payment_amount:
        fee_rate = round((1 - revenue / total_payment_amount) * 100, 2)
    else:
        fee_rate = 0.0

    return {
        "channel_id": account_id,
        "channel_name": label,
        "source": "11st",
        "order_number": str(item.get("ordNo", "") or ""),
        # 빈 문자열이면 None으로 정규화 — unique (order_number, ord_prd_seq) 인덱스에서
        # 빈 문자열은 distinct 안 되어 중복 위반, NULL은 distinct로 취급됨 (issue #208).
        "ord_prd_seq": (str(item.get("ordPrdSeq", "") or "").strip() or None),
        "shipment_id": str(item.get("dlvNo", "") or ""),
        "product_id": str(item.get("prdNo", "") or ""),
        "product_name": str(item.get("prdNm", "") or ""),
        "product_option": str(item.get("slctPrdOptNm", "") or ""),
        "quantity": quantity,
        "sale_price": sale_price,
        "total_payment_amount": total_payment_amount,
        "cost": 0,
        "revenue": revenue,
        "fee_rate": fee_rate,
        "status": status,
        "shipping_status": shipping_status,
        "customer_name": str(item.get("rcvrNm", "") or item.get("ordNm", "") or ""),
        # 주문자명 — 11번가 API ordNm (수령인 rcvrNm과 다를 수 있음: 선물하기 등)
        "orderer_name": str(item.get("ordNm", "") or item.get("rcvrNm", "") or ""),
        "customer_phone": str(
            item.get("rcvrPrtblNo", "") or item.get("ordPrtblTel", "") or ""
        ),
        "customer_address": addr_base,
        "customer_address_detail": addr_detail,
        "customer_postal_code": postal_code,
        "customer_note": str(
            item.get("ordDlvReqCont", "") or item.get("dlvMsg", "") or ""
        ),
        "paid_at": paid_at,
        "created_at": paid_at,
    }


def _parse_ebay_datetime(val) -> Optional[datetime]:
    """eBay 날짜 필드는 문자열 또는 {"value": "..."} dict 형태."""
    if val is None:
        return None
    if isinstance(val, dict):
        val = val.get("value", "")
    return _parse_iso_datetime(val if isinstance(val, str) else None)


def _parse_ebay_order(
    o: dict,
    account_id: str,
    account_label: str,
    exchange_rate: float = 1400.0,
) -> dict[str, Any]:
    """eBay Fulfillment API 주문 dict → SambaOrder 필드 매핑.

    eBay는 USD 결제이므로 ``exchange_rate``(USD→KRW)로 변환해 KRW로 저장한다.
    다른 마켓(스마트스토어/롯데ON)과 통일된 KRW 체계 유지.
    """
    order_id = o.get("orderId", "") or ""
    legacy_id = o.get("legacyOrderId", "") or order_id

    line_items = o.get("lineItems") or []
    first_item: dict[str, Any] = line_items[0] if line_items else {}

    # 배송지
    ship_to: dict[str, Any] = {}
    for inst in o.get("fulfillmentStartInstructions") or []:
        step = inst.get("shippingStep") or {}
        ship_to = step.get("shipTo") or {}
        if ship_to:
            break
    contact = ship_to.get("contactAddress") or {}
    # 우편번호 — 화면 확인용으로 별도 컬럼에 저장 (복사 버튼 분리)
    ebay_postal_code = str(contact.get("postalCode", "") or "").strip() or None
    addr_parts = [
        contact.get("addressLine1", ""),
        contact.get("addressLine2", ""),
        contact.get("city", ""),
        contact.get("stateOrProvince", ""),
        contact.get("countryCode", ""),
    ]
    customer_address = ", ".join([p for p in addr_parts if p])

    # 가격 (USD → KRW 변환)
    pricing = o.get("pricingSummary") or {}
    total = pricing.get("total") or {}
    sale_price_usd = float(total.get("value", 0) or 0)
    sale_price_krw = int(round(sale_price_usd * exchange_rate))

    # 수수료 (eBay 마켓플레이스 수수료, USD → KRW 변환)
    marketplace_fee_usd = float(
        (o.get("totalMarketplaceFee") or {}).get("value", 0) or 0
    )
    marketplace_fee_krw = int(round(marketplace_fee_usd * exchange_rate))
    try:
        fee_rate = (
            round(marketplace_fee_usd / sale_price_usd * 100, 2)
            if sale_price_usd > 0
            else 0
        )
    except Exception:
        fee_rate = 0
    revenue = sale_price_krw - marketplace_fee_krw

    # 상태 매핑
    ff_status = o.get("orderFulfillmentStatus", "") or ""
    cancel_state = (o.get("cancelStatus") or {}).get(
        "cancelState", "NONE_REQUESTED"
    ) or "NONE_REQUESTED"
    if cancel_state != "NONE_REQUESTED":
        status = "cancel_requested"
        shipping_status = "취소요청"
    elif ff_status == "FULFILLED":
        status = "pending"
        shipping_status = "국내배송중"
    elif ff_status == "IN_PROGRESS":
        status = "pending"
        shipping_status = "발송대기"
    else:
        status = "pending"
        shipping_status = "발주확인"

    buyer_username = (o.get("buyer") or {}).get("username", "") or ""

    return {
        "order_number": legacy_id,
        "ext_order_number": order_id,
        "shipment_id": first_item.get("sku", ""),
        "channel_id": account_id,
        "channel_name": account_label,
        "product_id": first_item.get("legacyItemId", "") or first_item.get("sku", ""),
        "product_name": first_item.get("title", ""),
        "product_option": first_item.get("legacyVariationId", "") or "",
        "product_image": "",
        "customer_name": ship_to.get("fullName", "") or buyer_username,
        "customer_phone": (ship_to.get("primaryPhone") or {}).get("phoneNumber", "")
        or "",
        "customer_address": customer_address,
        "customer_postal_code": ebay_postal_code,
        "quantity": int(first_item.get("quantity", 1) or 1),
        "sale_price": sale_price_krw,
        "cost": 0,
        "fee_rate": fee_rate,
        "revenue": revenue,
        "status": status,
        "shipping_status": shipping_status,
        "shipping_company": "",
        "tracking_number": "",
        "paid_at": _parse_ebay_datetime(o.get("creationDate")),
        "source": "ebay",
        "notes": f"USD {sale_price_usd:.2f} @ {exchange_rate:.2f}원/USD",
    }


def _apply_ebay_claims_to_orders(
    orders_data: list[dict[str, Any]],
    returns_raw: list[dict[str, Any]],
    cancellations_raw: list[dict[str, Any]],
) -> None:
    """eBay 반품/취소 데이터로 orders_data의 shipping_status 덮어쓰기.

    return.state / cancellation.cancelState 를 기준으로 상태 매핑.
    orders_data에 없는 주문이면 추가하지 않음 (sync 범위 내 주문만 반영).
    """
    # 반품
    return_state_map = {
        "OPEN": "반품요청",
        "ESCALATED": "반품요청",
        "CLOSED": "반품완료",
    }
    for r in returns_raw or []:
        order_id = (
            r.get("orderId")
            or (r.get("itemInfo") or {}).get("orderId")
            or (r.get("creationInfo") or {}).get("orderId")
            or ""
        )
        state = (r.get("status") or {}).get("state", "") or ""
        ss = return_state_map.get(state, "반품요청")
        for od in orders_data:
            if od.get("ext_order_number") == order_id or od.get("order_number") == str(
                order_id
            ):
                od["shipping_status"] = ss
                od["status"] = "returned" if ss == "반품완료" else "return_requested"
                break

    # 취소
    cancel_state_map = {
        "IN_PROGRESS": "취소요청",
        "CANCEL_PENDING": "취소요청",
        "CANCEL_CLOSED": "취소완료",
        "CANCEL_CLOSED_FOR_COMMITMENT": "취소요청",
    }
    # 배송 진행 단계 보호 — '취소요청'은 송장출력 이후 상태를 덮어쓰지 않음
    # ('취소완료'는 실제 종결 상태이므로 그대로 반영)
    _ebay_shipped_guard = {
        "송장전송완료",
        "국내배송중",
        "배송완료",
        "구매확정",
    }
    for c in cancellations_raw or []:
        legacy_order_id = c.get("legacyOrderId", "") or ""
        state = c.get("cancelState", "") or ""
        ss = cancel_state_map.get(state, "취소요청")
        for od in orders_data:
            if od.get("order_number") == legacy_order_id:
                if (
                    ss == "취소요청"
                    and od.get("shipping_status") in _ebay_shipped_guard
                ):
                    break
                od["shipping_status"] = ss
                od["status"] = "cancelled" if ss == "취소완료" else "cancel_requested"
                break


# ──────────────────────────────────────────────────────────────────────────
# 롯데홈쇼핑 중복 goods_no 미매칭 보강 (issue #365)
# 같은 상품이 롯데홈에 여러 goods_no로 중복등록 → 고객은 구(판매중) 번호로 주문하나
# 삼바 DB market_product_nos에는 신(품절) 번호만 저장돼 goods_no 정확매칭 실패.
# 해결: 주문 product_name에 박힌 제조사 모델코드(style_code)로 cp를 매칭한다.
#   - goods_no는 재등록마다 바뀌고 상품명은 AI가공/재수집으로 drift 하지만
#     제조사 style_code(HF9375-010, CRS212095 등)는 불변 → 안정적 연결키.
#   - 토큰 필터: 숫자 1개 이상 AND 길이 6 이상 (색상/일반어 SILVER/KIDS/GIORDANO 배제).
#   - 단일 후보만 연결. 다중 후보(cp 중복등록 포함)는 수동으로 넘긴다(오매칭 0).
#   - cp.style_code 는 DB 컬럼이라 외부 API 호출 없음(순수 DB 조인). 폭주/IP차단 무관.
# 프로덕션 실측(미등록 791건): 복구 368(46.5%) / 다중후보 skip 69 / cp없음 354,
# 다중후보 자동링크 0, 복구쌍 스팟체크 전수 동일상품 확인.
# ──────────────────────────────────────────────────────────────────────────

# 언더스코어 포함 — cp.style_code 가 사이즈 접미사를 '_'로 붙이는 경우
# (예: 'NBRJGS140P_25')와 정확 매칭. '-' 코드(HF9375-010)는 그대로 흡수.
_LH_STYLE_TOKEN_RE = re.compile(r"[A-Za-z0-9][A-Za-z0-9_-]*")


def _lh_style_tokens(name: str) -> list[str]:
    """롯데홈/플레이오토 상품명에서 모델코드(style_code) 후보 추출.

    조건: 길이 6+, 숫자 포함, 영문자 포함.
    순수 숫자 토큰(예: 6166973, 4974058) 제외 — 무관한 상품 style_code와 오매칭으로
    ambiguous 판정을 유발해 진짜 코드(KMM26249N3 등) 매칭을 차단하는 버그 방지.
    """
    return [
        t
        for t in _LH_STYLE_TOKEN_RE.findall(name or "")
        if len(t) >= 6 and any(c.isdigit() for c in t) and any(c.isalpha() for c in t)
    ]


# 색상토큰 재조합용 — 영숫자 단어 분리(하이픈/언더스코어 경계).
_LH_WORD_RE = re.compile(r"[A-Za-z0-9]+")


def _lh_reconstructed(name: str) -> list[str]:
    """상품명의 base 스타일코드 + 인접 색상숫자를 결합한 style_code 후보 생성(#365 확장).

    롯데홈 상품명은 코드와 색상을 공백으로 분리하는 경우가 많다
    (예: '나이키 IF2737 100 ...' — cp 저장 style_code 는 'IF2737-100').
    base 코드(하이픈/언더스코어 없는 순수 코드) 바로 앞/뒤의 색상숫자(2~4자리)를
    '-','_','' 로 결합한 후보를 반환한다. 색상이 글자(BLACK 등)면 재조합하지 않는다.
    정확매칭 + 단일후보 가드로 잘못된 결합은 자연히 0건 매칭되어 오매칭이 없다.
    """
    bare = {b for b in _lh_style_tokens(name) if "-" not in b and "_" not in b}
    if not bare:
        return []
    words = _LH_WORD_RE.findall(name or "")
    out: set[str] = set()
    for i, w in enumerate(words):
        if w not in bare:
            continue
        for j in (i - 1, i + 1):
            if 0 <= j < len(words):
                c = words[j]
                if c.isdigit() and 2 <= len(c) <= 4:
                    out.add(f"{w}-{c}")
                    out.add(f"{w}_{c}")
                    out.add(f"{w}{c}")
    return list(out)


async def _lh_resolve_by_style_code(
    product_name: str, channel_id: str, cache: dict
) -> dict | None:
    """미등록 롯데홈 주문을 product_name의 style_code로 cp 단일후보 매칭(순수 DB).

    채널 등록 cp 우선, 없으면 글로벌 단일후보(등록기록만 끊긴 orphan 구제 — 같은
    style_code 단일이면 같은 물리상품이라 원가/소싱 보강 유효). 다중후보는 None(수동).
    반환: _matched entry dict(_mpn 캐시와 동일 형식) | None
    """
    import json as _json

    tokens = _lh_style_tokens(product_name)
    if not tokens:
        return None
    # 캐시 키 — 상품명 단위(재조합 후보는 상품명 인접관계에 의존하므로 name 기준).
    key = (channel_id, product_name)
    if key in cache:
        return cache[key]
    res: dict | None = None
    try:
        from sqlalchemy import text as _sa_text2

        _cols = "id, source_site, source_url, (images->>0) AS thumb, category, style_code, cost"

        async def _run(_s, cands: list[str]):
            """주어진 style_code 후보로 채널>글로벌>개별토큰 순 단일후보 매칭.

            반환: (picked_row, route) | (None, "")
            """
            ch_rows = (
                await _s.execute(
                    _sa_text2(
                        f"SELECT {_cols} FROM samba_collected_product "
                        "WHERE registered_accounts @> CAST(:a AS jsonb) "
                        "AND style_code = ANY(:t)"
                    ),
                    {"a": _json.dumps([channel_id]), "t": cands},
                )
            ).fetchall()
            # 단일후보 판정 — distinct cp id 1개, 또는 distinct style_code 1개
            # (같은 style_code 중복등록 cp는 동일 물리상품이라 아무거나 연결 안전).
            # 서로 다른 style_code 가 섞이면(다른 상품 오매칭) 여전히 거부.
            _ch_ids = {str(r[0]) for r in ch_rows}
            _ch_styles = {str(r[5]) for r in ch_rows}
            if len(_ch_ids) == 1 or (ch_rows and len(_ch_styles) == 1):
                return ch_rows[0], "channel"
            if not _ch_ids:
                gl_rows = (
                    await _s.execute(
                        _sa_text2(
                            f"SELECT {_cols} FROM samba_collected_product "
                            "WHERE style_code = ANY(:t)"
                        ),
                        {"t": cands},
                    )
                ).fetchall()
                _gl_ids = {str(r[0]) for r in gl_rows}
                _gl_styles = {str(r[5]) for r in gl_rows}
                if len(_gl_ids) == 1 or (gl_rows and len(_gl_styles) == 1):
                    return gl_rows[0], "global"
                if len(_gl_ids) > 1 and len(cands) > 1:
                    # 복수 후보가 여러 style CP 히트(ambiguous) → 개별 후보 단독 재시도.
                    # 가장 긴(=가장 구체적인) 후보부터 시도해 고유 style CP 1개면 매칭.
                    for _tok in sorted(cands, key=len, reverse=True):
                        _gl2 = (
                            await _s.execute(
                                _sa_text2(
                                    f"SELECT {_cols} FROM samba_collected_product "
                                    "WHERE style_code = :t"
                                ),
                                {"t": _tok},
                            )
                        ).fetchall()
                        if len({str(r[0]) for r in _gl2}) == 1 or (
                            _gl2 and len({str(r[5]) for r in _gl2}) == 1
                        ):
                            return _gl2[0], f"global-single({_tok})"
            # 서로 다른 style 다중후보(채널>1 또는 글로벌>1)는 자동연결 금지 → 수동
            return None, ""

        async with get_read_session() as _s:
            _picked, _route = await _run(_s, tokens)
            # 1차(기본 토큰) 실패 시 색상숫자 재조합 후보로 재시도(#365 확장).
            # 예: '나이키 IF2737 100' → 'IF2737-100' 정확매칭. 기본 경로는 그대로라 회귀 없음.
            if _picked is None:
                _recon = _lh_reconstructed(product_name)
                if _recon:
                    _picked, _route = await _run(_s, _recon)
                    if _picked is not None:
                        _route = f"recon/{_route}"
            if _picked is not None:
                res = {
                    "collected_product_id": str(_picked[0]),
                    "source_site": _picked[1] or "",
                    "product_image": _picked[3] or "",
                    "original_link": _picked[2] or "",
                    "category": _picked[4] or "",
                    "cost": float(_picked[6]) if _picked[6] else 0.0,
                    "site_ids_by_account": {},
                }
                logger.info(
                    f"[주문매칭/롯데홈] style_code 보강({_route}): ch={channel_id} "
                    f"name={product_name!r} → cp {_picked[0]}(style={_picked[5]})"
                )
    except Exception as e:
        logger.warning(f"[주문매칭/롯데홈] style_code 매칭 실패 ch={channel_id}: {e}")
    cache[key] = res
    return res


def _parse_lottehome_order_multi(
    item: dict,
    account_id: str,
    label: str,
    force_status: str = "",
    prefer_org_dtl_sn: bool = False,
) -> list[dict]:
    """취소/반품처럼 ProdInfo가 리스트인 롯데홈쇼핑 주문 → 상품별 SambaOrder dict 리스트 반환.

    prefer_org_dtl_sn: 취소/반품 응답의 OrdDtlSn 은 재발급 클레임 라인번호라 원주문과
        어긋난다. True 면 OrgOrdDtlSn(원주문 라인번호) 우선으로 원주문과 매칭(#528/#393).
    """
    _shipping_status_map = {
        "cancelled": "취소완료",
        "return_requested": "반품요청",
        "return_completed": "회수확정",
    }
    prod_info_raw = item.get("ProdInfo", [])
    if isinstance(prod_info_raw, dict):
        prod_info_raw = [prod_info_raw]
    if not prod_info_raw:
        prod_info_raw = [{}]
    results = []
    for i, prod in enumerate(prod_info_raw):
        flat = dict(item)
        flat["ProdInfo"] = prod
        flat["_lh_prod_idx"] = i
        parsed = _parse_lottehome_order(
            flat, account_id, label, prefer_org_dtl_sn=prefer_org_dtl_sn
        )
        if force_status:
            parsed["status"] = force_status
            parsed["shipping_status"] = _shipping_status_map.get(
                force_status, force_status
            )
        results.append(parsed)
    return results


def _parse_lottehome_order(
    item: dict,
    account_id: str,
    label: str,
    force_status: str = "",
    force_shipping_status: str = "",
    prefer_org_dtl_sn: bool = False,
) -> dict:
    """롯데홈쇼핑 주문 데이터 → SambaOrder dict 변환.

    prefer_org_dtl_sn: 취소/반품 조회 응답은 OrdDtlSn 에 새 클레임 라인번호를
        발급하므로 order_number/shipment_id 가 원주문과 어긋난다.
        True 면 OrgOrdDtlSn(원주문 라인번호)을 우선 사용해 원주문과 매칭되도록 통일.
        반품(#393)·취소(#528) 경로 모두 라이브 응답에서 OrgOrdDtlSn 존재 검증 완료.
    """
    from datetime import datetime, timezone

    def _lh_str(*vals) -> str:
        for v in vals:
            s = str(v or "").strip()
            if s and s.lower() not in ("null", "none", "0"):
                return s
        return ""

    prod_info = (
        item.get("ProdInfo", {}) if isinstance(item.get("ProdInfo"), dict) else {}
    )
    delv_info = (
        item.get("DelvInfo", {}) if isinstance(item.get("DelvInfo"), dict) else {}
    )

    order_no = str(item.get("OrdNo", "") or "")
    sub_ord_no = str(item.get("SubOrdNo") or "")

    # 송장전송(registDeliver.lotte)에 ord_no + ord_dtl_sn 둘 다 필수.
    # ext_order_number 에 "ord_no:ord_dtl_sn" 형식으로 합쳐 저장한다.
    # issue #216 — 신규주문 API(searchNewOrdLstOpenApi.lotte)는 OrdDtlSn/DlvUnitSn 키 없음.
    # OrgOrdDtlSn(=같은 값) 또는 ProdSeq/ProdCode 폴백 — ProdInfo 리스트 내 상품 구분에도 사용.
    # issue #393 — 반품 응답은 OrdDtlSn 에 새 클레임 라인번호를 줘서 원주문과 어긋남.
    # prefer_org_dtl_sn=True 면 OrgOrdDtlSn(원주문 라인번호)을 맨 앞으로 옮겨 통일.
    # OrgOrdDtlSn 누락 시 OrdDtlSn 으로 폴백되어 기존 동작과 동일(안전).
    if prefer_org_dtl_sn:
        ord_dtl_sn = str(
            prod_info.get("OrgOrdDtlSn")
            or prod_info.get("OrdDtlSn")
            or prod_info.get("DlvUnitSn")
            or prod_info.get("ProdSeq")
            or prod_info.get("ProdCode")
            or item.get("_lh_prod_idx", "")
            or ""
        )
    else:
        ord_dtl_sn = str(
            prod_info.get("OrdDtlSn")
            or prod_info.get("DlvUnitSn")
            or prod_info.get("OrgOrdDtlSn")
            or prod_info.get("ProdSeq")
            or prod_info.get("ProdCode")
            or item.get("_lh_prod_idx", "")
            or ""
        )
    ext_order_number = (
        f"{order_no}:{ord_dtl_sn}" if (order_no and ord_dtl_sn) else order_no
    )
    # ord_dtl_sn이 있으면 "ord_no:ord_dtl_sn" 형식으로 상품별 고유 식별.
    # 없으면 sub_ord_no(상품주문번호)가 이미 상품별 고유값이므로 그대로 사용.
    order_number = ext_order_number if ord_dtl_sn else (sub_ord_no or order_no)

    proc_stat = str(item.get("OrdProcStat", "") or "")
    is_deliver_api = bool(prod_info.get("DlvUnitSn") or prod_info.get("GoodsNo"))
    status_map = {
        "업체지시": "pending",
        "정상": "pending",
        "출고확정": "shipping",
        "배송완료": "delivered",
        "구매확정": "confirmed",
        "취소": "cancelled",
        "반품진행": "return_requested",
        "회수확정": "return_requested",
        "발송불가": "undeliverable",
    }
    if force_status:
        status = force_status
        shipping_status = force_shipping_status or proc_stat or "출고지시"
    elif is_deliver_api and not proc_stat:
        status = "shipping"
        shipping_status = "배송대기중"
    else:
        status = status_map.get(proc_stat, "pending")
        shipping_status = proc_stat or "출고지시"
        if shipping_status == "출고확정":
            shipping_status = "배송대기중"

    product_name = str(prod_info.get("ProdName") or prod_info.get("GoodsNm") or "")
    product_option = str(
        prod_info.get("prodOption") or prod_info.get("GoodsDesc") or ""
    )
    # #528 — 취소조회(searchCnclList) 응답은 상품번호 키가 GoodNo(단수형)라
    # ProdCode/GoodsNo 폴백에 안 잡혀 product_id 가 비었다. GoodNo 폴백 추가.
    product_id = str(
        prod_info.get("ProdCode")
        or prod_info.get("GoodsNo")
        or prod_info.get("GoodNo")
        or ""
    )
    # product_id 빈 lottehome 주문 — 미등록 주문 발생 원인 진단용 raw 키 로그
    if not product_id:
        logger.warning(
            f"[주문동기화] lottehome product_id 누락 — "
            f"OrdNo={order_no}, SubOrdNo={sub_ord_no}, "
            f"ProdInfo keys={sorted(prod_info.keys())}, "
            f"item keys={sorted(item.keys())}"
        )
    sale_price = int(float(prod_info.get("ordPrice") or prod_info.get("SalePrc") or 0))
    buy_real_price = int(float(prod_info.get("buyRealPrice", 0) or 0))
    qty = int(prod_info.get("ordQty") or prod_info.get("OrdQty") or 1)

    recv_name = str(
        delv_info.get("recvName")
        or delv_info.get("RmitNm")
        or item.get("OrderName")
        or ""
    )
    recv_addr = str(
        delv_info.get("recvAddr1", "")
        or delv_info.get("Addr", "")
        or item.get("OrderAddr1", "")
    )
    recv_addr2 = str(delv_info.get("recvAddr2", "") or item.get("OrderAddr2", ""))
    recv_tel = str(
        delv_info.get("recvTel")
        or delv_info.get("recvHp")
        or item.get("OrderTelNo")
        or ""
    )
    shipping_company = str(delv_info.get("delvName") or delv_info.get("HdcNm") or "")
    tracking_number = _lh_str(delv_info.get("invoiceNo"), delv_info.get("InvNo"))

    trd_date = str(item.get("TrdDate", "") or "")
    paid_at = None
    if trd_date:
        try:
            paid_at = datetime.strptime(trd_date, "%Y-%m-%d").replace(
                tzinfo=timezone.utc
            )
        except ValueError:
            pass
    if paid_at is None and len(order_no) >= 8:
        try:
            paid_at = datetime.strptime(order_no[:8], "%Y%m%d").replace(
                tzinfo=timezone.utc
            )
        except ValueError:
            pass
    if paid_at is None:
        paid_at = datetime.now(timezone.utc)

    return {
        "order_number": order_number,
        "channel_id": account_id,
        "channel_name": label,
        "product_id": product_id,
        "product_name": product_name,
        "product_option": product_option,
        "customer_name": recv_name,
        "customer_phone": recv_tel,
        "customer_address": f"{recv_addr} {recv_addr2}".strip(),
        # 우편번호 — 화면 확인용 (복사 버튼 분리). 롯데홈쇼핑 API 필드: recvZipCd
        "customer_postal_code": (
            str(delv_info.get("recvZipCd") or delv_info.get("ZipCd") or "").strip()
            or None
        ),
        # 배송메시지(customer_note) — 롯데홈 주문 최상위 `DlvMemoCont`(배송메모).
        # (2026-07 실측: "해외배송이면 주문취소해요"가 DlvMemoCont 로 옴. 값 없으면
        #  "null" 문자열로 오는데 _lh_str 가 걸러줌.) 선물카드 메시지 CardMemoCont 는 백업.
        # 다른 마켓(스마트스토어/롯데ON/11번가/쿠팡/GS/플레이오토)처럼 customer_note 매핑
        # — 롯데홈 파서에만 누락돼 삼바 화면 고객메모가 항상 비어있던 문제 수정.
        "customer_note": _lh_str(item.get("DlvMemoCont"), item.get("CardMemoCont")),
        "quantity": qty,
        "sale_price": sale_price,
        "total_payment_amount": sale_price * qty,
        "cost": 0,
        # buy_real_price는 단가 기준 정산금액 → quantity 곱해 라인 총액으로 저장
        "fee_rate": (
            round((1 - buy_real_price / sale_price) * 100, 2)
            if (sale_price > 0 and buy_real_price > 0)
            else 0
        ),
        "revenue": buy_real_price * qty if buy_real_price > 0 else 0,
        "status": status,
        "shipping_status": shipping_status,
        "shipping_company": shipping_company,
        "tracking_number": tracking_number,
        "paid_at": paid_at,
        "source": "lottehome",
        # 주문번호(shipment_id) = 상품상세번호(OrdProdCode/OrgOrdDtlSn = ord_dtl_sn).
        # 과거엔 OrdNo(예: 20260529C08552)를 넣어 상품주문번호(order_number)와 중복됐음.
        # ord_dtl_sn은 ext_order_number 콜론 뒤 부분과 동일 → 백필값과 일치 보장.
        "shipment_id": ord_dtl_sn or order_no,
        "ext_order_number": ext_order_number,
    }


def _parse_esmplus_order(
    item: dict,
    account_id: str,
    label: str,
    market_type: str,
) -> dict[str, Any]:
    """ESM Plus(G마켓/옥션) RequestOrders 응답 item → SambaOrder dict.

    응답 키 PascalCase. 주요 필드:
      OrderNo, OutOrderNo, OrderStatus(1~5), OrderDate, PayDate(KST naive),
      SiteGoodsNo, GoodsName, SalePrice(string), OrderAmount, ContrAmount,
      ServiceFee, ShippingFee, BuyerName, ReceiverName, HpNo, TelNo,
      ZipCode, DelFrontAddress, DelBackAddress, DelMemo,
      TakbaeName, NoSongjang, ItemOptionSelectList[]
    """
    from datetime import datetime as _dt, timezone as _tz
    from zoneinfo import ZoneInfo as _ZI

    def _s(v: Any) -> str:
        return str(v or "").strip()

    def _f(v: Any) -> float:
        try:
            return float(str(v or "0"))
        except (ValueError, TypeError):
            return 0.0

    def _i(v: Any, default: int = 0) -> int:
        try:
            return int(_f(v))
        except (ValueError, TypeError):
            return default

    def _kst_to_utc(val: str | None) -> datetime | None:
        if not val:
            return None
        try:
            s = str(val).strip()
            if "." in s:
                s = s.split(".")[0]
            dt = _dt.fromisoformat(s)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=_ZI("Asia/Seoul"))
            return dt.astimezone(_tz.utc)
        except (ValueError, TypeError):
            return None

    # 내부 status / shipping_status 매핑
    # ESM OrderStatus: 1=결제완료, 2=배송준비, 3=배송중, 4=배송완료, 5=구매결정
    _status = _i(item.get("OrderStatus"), 1)
    status_map = {
        1: ("pending", "결제완료"),
        2: ("pending", "배송준비중"),
        3: ("shipped", "국내배송중"),
        4: ("delivered", "배송완료"),
        5: ("delivered", "구매확정"),
    }
    internal_status, shipping_status = status_map.get(_status, ("pending", "결제완료"))

    # 옵션 문자열 — ItemOptionSelectList[{ItemOptionValue, ItemOptionOrderCnt}]
    options = item.get("ItemOptionSelectList") or []
    opt_parts: list[str] = []
    if isinstance(options, list):
        for opt in options:
            if isinstance(opt, dict):
                ov = _s(opt.get("ItemOptionValue"))
                if ov:
                    opt_parts.append(ov)
    product_option = " / ".join(opt_parts)

    # 가격 / 수량
    sale_price = _f(item.get("SalePrice"))
    quantity = _i(item.get("ContrAmount"), 1) or 1
    order_amount = _f(item.get("OrderAmount"))
    service_fee = _f(item.get("ServiceFee"))
    fee_rate = round(service_fee / order_amount * 100, 2) if order_amount > 0 else 0.0
    revenue = order_amount - service_fee if order_amount > 0 else sale_price * quantity

    # 주소
    front_addr = _s(item.get("DelFrontAddress"))
    back_addr = _s(item.get("DelBackAddress"))
    full_addr = _s(item.get("DelFullAddress")) or f"{front_addr} {back_addr}".strip()

    return {
        "order_number": _s(item.get("OrderNo")),
        "shipment_id": _s(item.get("OrderNo")),
        "channel_id": account_id,
        "channel_name": label,
        "product_id": _s(item.get("SiteGoodsNo")) or _s(item.get("OutGoodsNo")),
        "product_name": _s(item.get("GoodsName")),
        "product_option": product_option,
        "product_image": "",
        "customer_name": _s(item.get("ReceiverName")) or _s(item.get("BuyerName")),
        "orderer_name": _s(item.get("BuyerName")),
        "customer_phone": _s(item.get("HpNo")) or _s(item.get("TelNo")),
        "customer_address": front_addr or full_addr,
        "customer_address_detail": back_addr,
        "customer_postal_code": _s(item.get("ZipCode")) or None,
        "customer_note": _s(item.get("DelMemo")),
        "quantity": quantity,
        "sale_price": sale_price,
        "total_payment_amount": order_amount or (sale_price * quantity),
        "cost": 0,
        "fee_rate": fee_rate,
        "revenue": revenue,
        "status": internal_status,
        "shipping_status": shipping_status,
        "shipping_company": _s(item.get("TakbaeName")),
        "tracking_number": _s(item.get("NoSongjang")),
        "paid_at": _kst_to_utc(item.get("PayDate") or item.get("OrderDate")),
        "source": market_type,
        "ext_order_number": _s(item.get("OutOrderNo")),
    }


# ═══════════════════════════════════════════════════════════════
# 롯데ON 선물하기 — 카톡 알림 기반 송장 자동입력 + 마켓전송
# 카톡(다른 PC)이 읽은 {이름, 품번, 송장번호}를 받아 주문을 찾아 처리한다.
# 안전규칙: 이름+품번으로 '송장 없는' 주문이 정확히 1건일 때만 처리.
#           0건/2건↑ 은 건너뜀(로그만). dry_run 이면 실제 전송 안 함.
# ═══════════════════════════════════════════════════════════════
class ShipByKakaoRequest(BaseModel):
    customer_name: str
    product_code: str
    shipping_company: str = "롯데택배"
    tracking_number: str
    tenant_id: str
    dry_run: bool = False


class KakaoNameCandidatesRequest(BaseModel):
    product_code: str
    tenant_id: str


def _extract_product_code(text: Optional[str]) -> Optional[str]:
    """상품명 안에서 품번(YMM24377Z1 형태) 추출."""
    if not text:
        return None
    m = re.search(r"[A-Z]{2,4}\d[A-Z0-9]{5,9}", text.upper())
    return m.group(0) if m else None


def _kakao_code_matches(kakao_code, product_name):
    """카톡 품번이 이 주문 상품명과 일치하는지.
    1) 기존 방식: 상품명에서 추출한 품번과 정확히 같음
    2) 포함 방식: 카톡 품번(7자 이상)이 상품명 안에 그대로 들어있음
    둘 중 하나라도 맞으면 매칭. 짧은 품번(<7)은 오매칭 위험이 커서 포함매칭 제외."""
    code = (kakao_code or "").upper().replace("-", "").replace(" ", "")
    if not code:
        return False
    # 1) 기존 추출-동일 방식 (하위호환)
    if _extract_product_code(product_name) == code:
        return True
    # 2) 포함 방식 (7자 이상만)
    if len(code) >= 7:
        norm = (product_name or "").upper().replace("-", "").replace(" ", "")
        if code in norm:
            return True
    return False


def _validate_invoice(inv: str) -> tuple[bool, str]:
    """송장 형식 검증. 3184 하드코딩 안 함(번호 바뀔 수 있음).
    필수: 숫자 + 자릿수(10~14). 3184 미시작은 막지 않고 '경고'만."""
    inv = (inv or "").strip()
    if not inv.isdigit():
        return False, "송장번호가 숫자가 아님"
    if not (10 <= len(inv) <= 14):
        return False, f"송장 자릿수 비정상({len(inv)}자리)"
    warn = "" if inv.startswith("3184") else "송장 패턴이 평소와 다름(확인 권장)"
    return True, warn


async def _verify_kakao_secret(
    x_kakao_secret: str = Header(default="", alias="X-Kakao-Secret"),
) -> None:
    """카톡 송장 자동입력 전용 인증. JWT 대신 전용 시크릿 키로 검증."""
    import secrets
    from backend.core.config import settings

    expected = (settings.kakao_ship_secret or "").strip()
    if not expected:
        raise HTTPException(status_code=503, detail="서버 인증 키 미설정")
    if not x_kakao_secret:
        raise HTTPException(status_code=401, detail="인증 키 없음")
    if not secrets.compare_digest(x_kakao_secret, expected):
        raise HTTPException(status_code=403, detail="인증 키 불일치")


@public_router.post("/ship-by-kakao", dependencies=[Depends(_verify_kakao_secret)])
async def ship_by_kakao(
    body: ShipByKakaoRequest,
    session: AsyncSession = Depends(get_write_session_dependency),
):
    """카톡 알림(이름+품번+송장)으로 주문을 찾아 송장입력 + 마켓전송."""
    tenant_id = (body.tenant_id or "").strip()
    name = (body.customer_name or "").strip()
    code = (body.product_code or "").strip().upper()
    inv = (body.tracking_number or "").strip()

    # 1) 송장 형식 검증
    ok, warn = _validate_invoice(inv)
    if not ok:
        logger.warning(
            "[ship-by-kakao] 송장검증실패 name=%s code=%s inv=%s (%s)",
            name,
            code,
            inv,
            warn,
        )
        return {"ok": False, "action": "rejected", "reason": warn}

    # 2) 이름 일치 + 송장 없는 주문 후보 조회
    svc = _write_service(session)
    stmt = select(SambaOrder).where(
        SambaOrder.customer_name == name,
        SambaOrder.source_site == "LOTTEON",  # 소싱처가 롯데ON
        # 선물하기 건만 (마켓 무관). action_tag 는 콤마 다중태그라
        # 경계매칭 헬퍼로 부분일치(regift 등) 오매칭 방지
        _build_action_tag_filter("gift"),
    )
    if tenant_id is not None:
        stmt = stmt.where(SambaOrder.tenant_id == tenant_id)
    result = await session.execute(stmt)
    candidates = result.scalars().all()

    # 3) 품번 일치 + 아직 송장 없는 것만 필터
    matched = [
        o
        for o in candidates
        if not (o.tracking_number or "").strip()
        and _kakao_code_matches(code, o.product_name)
    ]

    # 4) 안전규칙: 정확히 1건일 때만 처리
    if len(matched) == 0:
        logger.info("[ship-by-kakao] 매칭 0건 — 건너뜀 name=%s code=%s", name, code)
        return {
            "ok": False,
            "action": "skipped",
            "reason": "매칭 주문 없음(미수집/이미처리)",
        }
    if len(matched) > 1:
        logger.warning(
            "[ship-by-kakao] 매칭 %d건 — 건너뜀(사람확인) name=%s code=%s ids=%s",
            len(matched),
            name,
            code,
            [o.id for o in matched],
        )
        return {
            "ok": False,
            "action": "skipped",
            "reason": f"매칭 {len(matched)}건(사람 확인 필요)",
            "order_ids": [o.id for o in matched],
        }

    order = matched[0]

    # 5) dry_run: 실제 전송 안 하고 '이렇게 보낼 것'만 반환
    if body.dry_run:
        return {
            "ok": True,
            "action": "dry_run",
            "order_id": order.id,
            "would_send": {
                "shipping_company": body.shipping_company,
                "tracking_number": inv,
            },
            "warning": warn,
        }

    # 6) 실제 처리 — 기존 ship_order 와 동일 로직 재사용
    await svc.update_order(
        order.id,
        {"shipping_company": body.shipping_company, "tracking_number": inv},
    )
    from backend.domain.samba.order.dispatch_service import send_invoice_to_market

    market_sent, market_msg = await send_invoice_to_market(
        order, body.shipping_company, inv, session
    )
    if market_sent:
        await svc.update_order(
            order.id,
            {"shipping_status": "송장전송완료", "status": "shipping"},
        )
    logger.info(
        "[ship-by-kakao] 처리완료 order=%s sent=%s name=%s code=%s%s",
        order.id,
        market_sent,
        name,
        code,
        f" / {warn}" if warn else "",
    )
    return {
        "ok": True,
        "action": "shipped",
        "order_id": order.id,
        "market_sent": market_sent,
        "message": market_msg,
        "warning": warn,
    }


@public_router.post(
    "/kakao-name-candidates", dependencies=[Depends(_verify_kakao_secret)]
)
async def kakao_name_candidates(
    body: KakaoNameCandidatesRequest,
    session: AsyncSession = Depends(get_read_session_dependency),
):
    """카톡 OCR 이름 깨짐 대응 — 품번으로 송장입력 가능한 후보 이름 목록 조회.

    OCR이 고객 이름을 깨먹으면 ship-by-kakao 의 이름 매칭이 0건이 되어 처리 불가.
    이때 품번으로 ship-by-kakao 와 동일 기준(롯데ON 선물 + 송장없음 + 품번일치)의
    후보 주문을 찾아 '이름 목록'만 반환한다(읽기 전용, 송장 처리 안 함).
    카톡PC가 OCR 이름과 후보를 비교/선택해 ship-by-kakao 를 재호출하는 용도.
    """
    tenant_id = (body.tenant_id or "").strip()
    code = (body.product_code or "").strip().upper()

    if not code:
        return {"ok": False, "reason": "품번 없음", "count": 0, "candidates": []}
    # 테넌트 격리: tenant_id 없으면 전 테넌트 후보(타사 고객명)가 노출되므로 거부
    if not tenant_id:
        return {"ok": False, "reason": "tenant_id 없음", "count": 0, "candidates": []}

    # ship-by-kakao 와 동일 기준: 롯데ON 선물 + 해당 테넌트 후보 조회.
    # action_tag 는 콤마 다중태그라 경계매칭 헬퍼로 regift/gifted 등 오매칭 방지.
    stmt = select(SambaOrder).where(
        SambaOrder.source_site == "LOTTEON",  # 소싱처가 롯데ON
        _build_action_tag_filter("gift"),  # 선물하기 건만 (마켓 무관)
        SambaOrder.tenant_id == tenant_id,
    )
    result = await session.execute(stmt)
    rows = result.scalars().all()

    # 품번 일치 + 아직 송장 없는 것만 (ship-by-kakao 3단계와 동일 필터)
    matched = [
        o
        for o in rows
        if not (o.tracking_number or "").strip()
        and _kakao_code_matches(code, o.product_name)
    ]

    candidates = [
        {
            "order_id": o.id,
            "customer_name": o.customer_name,
            "product_name": o.product_name,
        }
        for o in matched
    ]
    # 고객명(PII)은 로그에 남기지 않음 — 건수만 기록
    logger.info(
        "[kakao-name-candidates] code=%s tenant=%s 후보=%d건",
        code,
        tenant_id,
        len(candidates),
    )
    return {"ok": True, "count": len(candidates), "candidates": candidates}


async def _kream_cost_backfill_from_shopmine(
    ws, session: AsyncSession, tenant_id: Optional[str]
) -> dict:
    """마스터 엑셀 '샵마인' 시트 → 크림주문 실구매가(cost)+소싱주문번호 백필 후
    스니덩크 해외송장 자동수집.

    컬럼(0-based): B(1)/C(2)=쇼핑몰·별칭(크림 필터), H(7)=오픈마켓주문번호(매칭키),
    P(15)=소싱주문번호, Q(16)=매입금액(실구매가). profit·수익률은 화면에서 자동계산되므로
    저장하지 않는다.
    """
    import asyncio
    from datetime import datetime, timezone

    from sqlalchemy import func, select

    # 1) 샵마인 크림행 파싱 → {오픈마켓주문번호: (소싱주문번호, 실구매가)}
    sheet_map: dict[str, tuple[str, float]] = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or len(r) < 17:
            continue
        if "크림" not in (str(r[1] or "") + str(r[2] or "")):
            continue
        h = str(r[7] or "").strip()  # 앞 개행 포함될 수 있어 strip 필수
        if not h:
            continue
        p, q = r[15], r[16]
        sono = str(p).strip() if p is not None else ""
        try:
            cost = float(round(float(q))) if q not in (None, "") else None
        except (TypeError, ValueError):
            cost = None
        if cost is None:
            continue
        sheet_map[h] = (sono, cost)

    # 2) 크림주문 매칭 → cost + sourcing_order_number 갱신
    stmt = select(SambaOrder).where(
        func.upper(func.coalesce(SambaOrder.source_site, "")) == "KREAM"
    )
    if tenant_id is not None:
        stmt = stmt.where(SambaOrder.tenant_id == tenant_id)
    kream_orders = (await session.execute(stmt)).scalars().all()

    filled = 0
    unmatched = 0
    for o in kream_orders:
        key = (o.order_number or "").strip()
        if key not in sheet_map:
            unmatched += 1
            continue
        sono, cost = sheet_map[key]
        changed = False
        if float(o.cost or 0) != cost:
            o.cost = cost
            changed = True
        if sono and (o.sourcing_order_number or "") != sono:
            o.sourcing_order_number = sono
            changed = True
        # 소싱주문번호 있으면 상태 '배송대기중'(wait_ship) — 이미 진행된 상태는 유지(역행 방지)
        _advanced = {
            "shipping",
            "delivered",
            "confirmed",
            "cancelled",
            "returned",
            "cancel_requested",
            "return_requested",
            "ship_failed",
        }
        if (
            (o.sourcing_order_number or "")
            and o.status not in _advanced
            and o.status != "wait_ship"
        ):
            o.status = "wait_ship"
            changed = True
        if changed:
            o.updated_at = datetime.now(timezone.utc)
            filled += 1
    await session.commit()

    # 3) 스니덩크 해외송장 자동수집 (확장앱 세션쿠키 필요, 소싱주문번호 有 & 송장 空)
    tracking_checked = 0
    tracking_shipped = 0
    cookie = await _get_snkr_session_cookie(session)
    if cookie:
        tstmt = select(SambaOrder).where(
            func.upper(func.coalesce(SambaOrder.source_site, "")) == "KREAM",
            SambaOrder.sourcing_order_number.is_not(None),
            SambaOrder.sourcing_order_number != "",
            (SambaOrder.overseas_tracking_number.is_(None))
            | (SambaOrder.overseas_tracking_number == ""),
        )
        if tenant_id is not None:
            tstmt = tstmt.where(SambaOrder.tenant_id == tenant_id)
        tstmt = tstmt.limit(500)
        targets = (await session.execute(tstmt)).scalars().all()
        for o in targets:
            tracking_checked += 1
            res = await _apply_snkr_overseas_tracking(session, o, cookie)
            if res.get("shipped"):
                tracking_shipped += 1
            await asyncio.sleep(0.3)  # SNKRDUNK 레이트리밋 보수값

    # 4) 허브넷 택배번호 자동기입 (해외송장 보유 주문 전체)
    hubnet = await _push_hubnet_tracking(session)

    return {
        "ok": True,
        "mode": "cost_backfill",
        "filled": filled,
        "unmatched": unmatched,
        "tracking_checked": tracking_checked,
        "tracking_shipped": tracking_shipped,
        "hubnet_updated": hubnet.get("updated", 0),
        "hubnet_error": hubnet.get("error"),
        "cookie_missing": not cookie,
    }


@router.post("/kream-excel")
async def import_kream_excel(
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_write_session_dependency),
    tenant_id: Optional[str] = Depends(get_optional_tenant_id),
):
    """KREAM 발송완료내역 엑셀 업로드 → 주문 생성."""
    import openpyxl  # noqa: F811
    from datetime import timezone
    from io import BytesIO

    from sqlalchemy import text as sa_text

    from backend.domain.samba.account.model import SambaMarketAccount

    # KREAM 계정 조회 (channel_id 용)
    acc_stmt = select(SambaMarketAccount).where(
        SambaMarketAccount.market_type == "kream"
    )
    if tenant_id is not None:
        acc_stmt = acc_stmt.where(SambaMarketAccount.tenant_id == tenant_id)
    acc_row = await session.execute(acc_stmt)
    kream_acc = acc_row.scalars().first()
    kream_channel_id = kream_acc.id if kream_acc else None

    # 크림 주문의 실제 소싱처는 SNKRDUNK(성희 계정) — 주문계정 자동 선택.
    # 기본 로그인 계정 우선, 없으면 활성 계정 중 최초.
    snkr_sourcing_account_id = None
    try:
        from sqlalchemy import func as _sfunc

        from backend.domain.samba.sourcing_account.model import SambaSourcingAccount

        _snkr_base = (
            select(SambaSourcingAccount.id)
            .where(
                _sfunc.upper(SambaSourcingAccount.site_name) == "SNKRDUNK",
                SambaSourcingAccount.is_active.is_(True),
            )
            .order_by(
                SambaSourcingAccount.is_login_default.desc(),
                SambaSourcingAccount.created_at,
            )
        )
        if tenant_id is not None:
            snkr_sourcing_account_id = (
                (
                    await session.execute(
                        _snkr_base.where(SambaSourcingAccount.tenant_id == tenant_id)
                    )
                )
                .scalars()
                .first()
            )
        if not snkr_sourcing_account_id:
            snkr_sourcing_account_id = (
                (await session.execute(_snkr_base)).scalars().first()
            )
    except Exception as _e:
        logger.warning(f"[KREAM엑셀] SNKRDUNK 주문계정 조회 실패(무시): {_e}")
        snkr_sourcing_account_id = None

    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    # 마스터 엑셀('샵마인' 시트) 감지 → 발송완료 주문생성 대신 실구매가/소싱주문번호 백필 +
    # 스니덩크 해외송장 자동수집 모드로 분기.
    _shopmine_sheet = next((s for s in wb.sheetnames if "샵마인" in s), None)
    if _shopmine_sheet is not None:
        _sm_result = await _kream_cost_backfill_from_shopmine(
            wb[_shopmine_sheet], session, tenant_id
        )
        wb.close()
        return _sm_result
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    # kream product_id → collected_product_id + 한글 상품명 역매칭
    kream_pids = [str(r[3]) for r in rows if r and len(r) > 3 and r[3]]
    cp_map: dict[str, str] = {}
    cp_name_map: dict[str, str] = {}
    if kream_pids:
        tid_cond = "AND tenant_id = :tid" if tenant_id is not None else ""
        bind = {"pids": kream_pids}
        if tenant_id is not None:
            bind["tid"] = tenant_id
        cp_rows = await session.execute(
            sa_text(f"""
                SELECT id, name, resell_matches->'kream'->>'product_id' AS kream_pid
                FROM samba_collected_product
                WHERE source_site = 'SNKRDUNK'
                  AND resell_matches->'kream'->>'product_id' = ANY(:pids)
                  {tid_cond}
            """),
            bind,
        )
        for cp_row in cp_rows.mappings():
            pid = str(cp_row["kream_pid"])
            cp_map[pid] = str(cp_row["id"])
            cp_name_map[pid] = cp_row["name"] or ""

    def _parse_dt(val):
        if val is None:
            return None
        if isinstance(val, datetime):
            return val.replace(tzinfo=timezone.utc) if val.tzinfo is None else val
        try:
            return datetime.fromisoformat(str(val).replace(" ", "T")).replace(
                tzinfo=timezone.utc
            )
        except Exception:
            return None

    created = 0
    skipped = 0
    for row in rows:
        if not row or not row[0]:
            continue
        order_number = str(row[0]).strip()
        paid_at_raw = row[2] if len(row) > 2 else None
        kream_pid = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        product_name = cp_name_map.get(kream_pid, "")
        option_name = str(row[6]).strip() if len(row) > 6 and row[6] else ""
        sale_price = float(row[7]) if len(row) > 7 and row[7] else 0.0
        tracking_number = str(row[9]).strip() if len(row) > 9 and row[9] else ""
        shipped_at_raw = row[10] if len(row) > 10 else None

        # 중복 체크
        dup_stmt = select(SambaOrder.id).where(SambaOrder.order_number == order_number)
        if tenant_id is not None:
            dup_stmt = dup_stmt.where(SambaOrder.tenant_id == tenant_id)
        dup = await session.execute(dup_stmt)
        if dup.scalar():
            skipped += 1
            continue

        order = SambaOrder(
            tenant_id=tenant_id,
            order_number=order_number,
            channel_id=kream_channel_id,
            channel_name="KREAM",
            source_site="KREAM",
            product_id=kream_pid or None,
            product_name=product_name,
            product_option=option_name,
            sale_price=sale_price,
            # 정산금액 = 결제금액과 동일 표시 (크림 해외판매 — 마켓수수료 별도)
            revenue=sale_price,
            cost=0.0,
            # 배송비 기본 8,000원 자동 입력 (크림 해외배송 고정)
            shipping_fee=8000.0,
            profit=0.0,
            tracking_number=tracking_number or None,
            shipped_at=_parse_dt(shipped_at_raw),
            paid_at=_parse_dt(paid_at_raw),
            status="pending",
            shipping_status="결제완료",
            shipping_company="허브넷로지스틱스",
            collected_product_id=cp_map.get(kream_pid) if kream_pid else None,
            sourcing_account_id=snkr_sourcing_account_id,
        )
        session.add(order)
        created += 1

    await session.commit()

    # 발송완료 업로드 후에도 스니덩크 해외송장 수집 + 허브넷 기입 자동 수행
    # (기존 주문 중 소싱주문번호 있고 송장 없는 것 대상 — 방금 생성분은 소싱번호 없어 스킵됨)
    from sqlalchemy import func as _kfunc

    tracking_checked = 0
    tracking_shipped = 0
    snkr_cookie = await _get_snkr_session_cookie(session)
    if snkr_cookie:
        tstmt = (
            select(SambaOrder)
            .where(
                _kfunc.upper(_kfunc.coalesce(SambaOrder.source_site, "")) == "KREAM",
                SambaOrder.sourcing_order_number.is_not(None),
                SambaOrder.sourcing_order_number != "",
                (SambaOrder.overseas_tracking_number.is_(None))
                | (SambaOrder.overseas_tracking_number == ""),
            )
            .limit(500)
        )
        if tenant_id is not None:
            tstmt = tstmt.where(SambaOrder.tenant_id == tenant_id)
        for o in (await session.execute(tstmt)).scalars().all():
            tracking_checked += 1
            res = await _apply_snkr_overseas_tracking(session, o, snkr_cookie)
            if res.get("shipped"):
                tracking_shipped += 1
            await asyncio.sleep(0.3)

    hubnet = await _push_hubnet_tracking(session)

    return {
        "ok": True,
        "created": created,
        "skipped": skipped,
        "tracking_checked": tracking_checked,
        "tracking_shipped": tracking_shipped,
        "hubnet_updated": hubnet.get("updated", 0),
        "hubnet_error": hubnet.get("error"),
        "cookie_missing": not snkr_cookie,
    }
